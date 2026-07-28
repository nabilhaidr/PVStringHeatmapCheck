"""Smoke Specific_Yield_Sibling_Ratio.ipynb + Specific_Yield_Cleaning_Impact.ipynb.

Offline: download Drive di-mock, memakai workbook specific yield sintetis
dan rekap cleaning sintetis dalam format asli (sheet STS n, checklist TRUE).
Memverifikasi angka -- cuaca harus tercoret sehingga string kotor 0.85x
terbaca sebagai defisit yang benar dan uplift ~17.6% setelah dibersihkan.
"""
from __future__ import annotations

import json
import sys
import tempfile
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

SIBLING_NB = ROOT / "output_string/Specific_Yield_Sibling_Ratio.ipynb"
CLEANING_NB = ROOT / "output_string/Specific_Yield_Cleaning_Impact.ipynb"

WEATHER = [1.0, 0.42, 0.95, 1.12, 0.38, 0.88, 1.05, 0.55, 0.97, 1.10,
           0.45, 1.02, 0.91, 0.36, 1.08, 0.99, 0.62, 1.14, 0.87, 0.51]
BASE_YF = 4.2
DIRTY = "WB01-INV01-PV1"
CLEAN_DAY = pd.Timestamp("2026-06-11")


def _source(notebook: dict, index: int) -> str:
    return "".join(notebook["cells"][index]["source"])


def _exec(notebook: dict, indexes, scope: dict) -> None:
    for index in indexes:
        with patch.dict(sys.modules, {"IPython": None, "IPython.display": None}):
            exec(_source(notebook, index), scope)


def _write_specific_yield(path: Path) -> None:
    """20 hari x 2 inverter x 6 string; PV1 INV01 kotor 0.85x sampai 10 Juni."""
    days = pd.date_range("2026-06-01", periods=20, freq="D")
    rows = []
    for index, day in enumerate(days):
        weather = WEATHER[index]
        for inverter in ("WB01-INV01", "WB01-INV02"):
            for pv in range(1, 7):
                pv_string = f"{inverter}-PV{pv}"
                factor = 0.85 if (pv_string == DIRTY and index < 10) else 1.0
                yf = BASE_YF * weather * factor
                rows.append({
                    "date": day,
                    "pv_string": pv_string,
                    "inverter_id": inverter,
                    "pv_label": f"PV{pv}",
                    "string_yield_kwh": yf * 15.0,
                    "capacity_kwp": 15.0,
                    "specific_yield_kwh_per_kwp": yf,
                    "valid_power_samples": 140,
                    "expected_samples": 288,
                    "coverage_pct": 48.6,
                    "missing_power_samples": 148,
                    "source_csv": f"{day:%Y%m%d}.csv",
                    "status": "PARTIAL",
                })
    pd.DataFrame(rows).to_excel(path, sheet_name="Detail_Harian", index=False)


def _write_cleaning_report(path: Path) -> None:
    """Format asli: sheet 'STS 1', kolom C 'String', tanggal dari kolom D."""
    dates = [CLEAN_DAY, CLEAN_DAY + pd.Timedelta(days=1)]
    grid = [["", "", "String", *dates]]
    for st in range(1, 7):
        marks = ["TRUE" if st == 1 else "FALSE", "FALSE"]
        grid.append(["", "INV-101", f"ST {st}", *marks])
    pd.DataFrame(grid).to_excel(
        path, sheet_name="STS 1", index=False, header=False,
    )


def _fake_download(paths):
    def _download(folder_url, destination, *, basename="", pattern="", pick="last"):
        if basename and basename in paths:
            return paths[basename]
        if pattern:
            return paths["specific_yield"]
        raise FileNotFoundError(f"mock: {basename or pattern}")
    return _download


def _run_sibling(temp: Path, specific_yield: Path, output_dir: Path) -> Path:
    notebook = json.loads(SIBLING_NB.read_text(encoding="utf-8"))
    scope = {"REPO_DIR": ROOT, "OUTPUT_DIR": output_dir, "INPUT_DIR": temp}
    with patch(
        "pv_pipeline.yf_ratio_report.download_drive_file",
        side_effect=_fake_download({"specific_yield": specific_yield}),
    ):
        _exec(notebook, (2, 3, 4, 5), scope)

    output = Path(scope["OUTPUT_XLSX"])
    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == [
        "Ranking_Sibling", "Rasio_Harian", "Metadata",
    ], workbook.sheetnames
    workbook.close()

    ranking = pd.read_excel(output, sheet_name="Ranking_Sibling")
    top = ranking.iloc[0]
    assert top["pv_string"] == DIRTY, top["pv_string"]
    assert top["status"] == "CLEANING_CANDIDATE"
    # 10 hari kotor 0.85x + 10 hari bersih -> median rasio harian = 0.925.
    assert abs(top["ratio_vs_inverter"] - 0.925) < 1e-6, top["ratio_vs_inverter"]
    assert abs(top["deficit_vs_inverter_pct"] - 7.5) < 1e-4
    daily = pd.read_excel(output, sheet_name="Rasio_Harian")
    assert daily.columns[0] == "pv_string"
    assert len(daily.columns) == 21, len(daily.columns)  # pv_string + 20 hari
    print("[smoke] sibling ratio OK ->", output.name)
    return output


def _run_cleaning(temp: Path, paths: dict, output_dir: Path) -> Path:
    notebook = json.loads(CLEANING_NB.read_text(encoding="utf-8"))
    scope = {"REPO_DIR": ROOT, "OUTPUT_DIR": output_dir, "INPUT_DIR": temp}
    empty_cable_map = pd.DataFrame(columns=["wb", "inv", "st", "pv", "mppt"])
    with patch(
        "pv_pipeline.yf_ratio_report.download_drive_file",
        side_effect=_fake_download(paths),
    ), patch(
        "pv_pipeline.m2a.cleaning_report.load_dc_cable_map",
        return_value=empty_cable_map,   # WB01: nomor String == nomor PV
    ):
        _exec(notebook, (2, 3, 4, 5, 6), scope)

    output = Path(scope["OUTPUT_XLSX"])
    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == [
        "Dampak_Cleaning", "Rekap_Campaign", "Metadata",
    ], workbook.sheetnames
    workbook.close()

    impact = pd.read_excel(output, sheet_name="Dampak_Cleaning")
    assert len(impact) == 1, len(impact)
    row = impact.iloc[0]
    assert row["pv_string"] == DIRTY
    assert row["status"] == "RECOVERED"
    assert row["reference_mode"] == "WB_UNCLEANED"
    assert row["n_reference_strings"] == 11
    # 0.85 -> 1.00 => uplift 17.6%, soiling loss 15%.
    assert abs(row["uplift_pct"] - 15.0 / 85.0 * 100) < 0.5, row["uplift_pct"]
    assert abs(row["soiling_loss_pct"] - 15.0) < 0.5
    assert len(pd.read_excel(output, sheet_name="Rekap_Campaign")) == 1
    print("[smoke] cleaning impact OK ->", output.name)
    return output


def main() -> None:
    with tempfile.TemporaryDirectory(prefix="yf_ratio_smoke_") as name:
        temp = Path(name)
        output_dir = temp / "output_string"
        output_dir.mkdir()

        specific_yield = temp / "specific_yield_20260601_20260620.xlsx"
        cleaning_report = temp / "Report & Schedule Cleaning PLTS IKN.xlsx"
        dc_cable = temp / "List of DC Cables 0411.xls"
        _write_specific_yield(specific_yield)
        _write_cleaning_report(cleaning_report)
        dc_cable.write_bytes(b"")   # tidak dibaca (load_dc_cable_map di-mock)

        paths = {
            "specific_yield": specific_yield,
            "Report & Schedule Cleaning PLTS IKN.xlsx": cleaning_report,
            "List of DC Cables 0411.xls": dc_cable,
        }
        sibling_output = _run_sibling(temp, specific_yield, output_dir)
        cleaning_output = _run_cleaning(temp, paths, output_dir)

        # Idempoten: rerun tidak menghasilkan file kedua.
        _run_sibling(temp, specific_yield, output_dir)
        _run_cleaning(temp, paths, output_dir)
        assert list(output_dir.glob("sibling_ratio_*.xlsx")) == [sibling_output]
        assert list(output_dir.glob("cleaning_impact_yf_*.xlsx")) == [
            cleaning_output
        ]
        print("[smoke] OK")


if __name__ == "__main__":
    main()
