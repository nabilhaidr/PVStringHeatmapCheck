"""Smoke test Specific_Yield_Daily.ipynb (offline, tanpa Drive/Colab).

Menjalankan Cell 2..6 dengan download_csv_inputs di-mock, memakai CSV
sintetis, lalu memverifikasi angka specific yield dan struktur workbook.
"""
from __future__ import annotations

import json
import sys
import tempfile
from datetime import date
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from pv_pipeline.string_yield_report import (  # noqa: E402
    DownloadedInputs,
    DriveItem,
    SourceManifest,
    parse_date_range,
)


NOTEBOOK = ROOT / "output_string/Specific_Yield_Daily.ipynb"


def _source(notebook: dict, index: int) -> str:
    return "".join(notebook["cells"][index]["source"])


def _metadata_values(workbook) -> dict[str, object]:
    sheet = workbook["Metadata"]
    return {
        sheet.cell(row=row, column=1).value:
        sheet.cell(row=row, column=2).value
        for row in range(2, sheet.max_row + 1)
    }


def _execute_cells(notebook: dict, scope: dict) -> None:
    exec(_source(notebook, 2), scope)  # Cell 2 config
    scope.update({
        "START_DATE": "2026-06-01",
        "END_DATE": "2026-06-02",
        "DATES": parse_date_range("2026-06-01", "2026-06-02"),
        "EXCLUDE_EMPTY_SLOTS": False,
    })
    for index in (3, 4, 5, 6):  # download, energi, specific, export
        if index in (4, 5):
            with patch.dict(
                sys.modules,
                {"IPython": None, "IPython.display": None},
            ):
                exec(_source(notebook, index), scope)
        else:
            exec(_source(notebook, index), scope)


def main() -> None:
    notebook = json.loads(NOTEBOOK.read_text(encoding="utf-8"))
    with tempfile.TemporaryDirectory(prefix="specific_yield_smoke_") as tmp:
        temp = Path(tmp)
        input_dir = temp / "inputs"
        output_dir = temp / "output_string"
        input_dir.mkdir()
        output_dir.mkdir()

        # WB01-INV01-PV1 (fase 1, 15 kWp) & WB03-INV02-PV5 (fase 2, 16.25 kWp).
        # 12 x 5min = 1 jam; daya konstan -> energi = daya (kWh).
        stamps_one = pd.date_range("2026-06-01 00:00", periods=12, freq="5min")
        day_one = pd.concat([
            pd.DataFrame({
                "Start Time": stamps_one,
                "Inverter_ID": "WB01-INV01",
                "PV1 Power(kW)": 15.0,   # -> 15 kWh / 15 kWp = 1.0
            }),
            pd.DataFrame({
                "Start Time": stamps_one,
                "Inverter_ID": "WB03-INV02",
                "PV5 Power(kW)": 16.25,  # -> 16.25 kWh / 16.25 kWp = 1.0
            }),
        ], ignore_index=True)
        csv_one = input_dir / "20260601.csv"
        day_one.to_csv(csv_one, index=False)

        csv_two = input_dir / "20260602.csv"
        pd.DataFrame({
            "Start Time": pd.date_range(
                "2026-06-02 00:00", periods=12, freq="5min",
            ),
            "Inverter_ID": "WB01-INV01",
            "PV1 Power(kW)": 7.5,        # -> 7.5 / 15 = 0.5
        }).to_csv(csv_two, index=False)

        manifest = SourceManifest(
            csv_by_date={
                date(2026, 6, 1): DriveItem("local-one", csv_one.name),
                date(2026, 6, 2): DriveItem("local-two", csv_two.name),
            },
            poa_by_year={},
            missing_csv_dates=[],
            missing_poa_years=[],
            url_csv="https://drive.google.com/drive/folders/csv-smoke",
            csv_inventory_count=2,
        )
        inputs = DownloadedInputs(
            csv_by_date={
                date(2026, 6, 1): csv_one,
                date(2026, 6, 2): csv_two,
            },
            poa_by_year={},
            download_errors={},
        )
        scope = {
            "REPO_DIR": ROOT,
            "OUTPUT_DIR": output_dir,
            "INPUT_DIR": input_dir,
            "STRINGS_YAML": str(ROOT / "config" / "strings.yaml"),
        }

        with patch(
            "pv_pipeline.all_string_yield_report.download_csv_inputs",
            return_value=(manifest, inputs),
        ):
            _execute_cells(notebook, scope)

        output_xlsx = Path(scope["OUTPUT_XLSX"])
        workbook = load_workbook(output_xlsx, data_only=False)
        assert workbook.sheetnames == [
            "Rekap_SpecificYield", "Detail_Harian", "Metadata",
        ], workbook.sheetnames
        recap = workbook["Rekap_SpecificYield"]
        assert recap["A1"].value == "pv_string"
        assert [
            recap.cell(row=row, column=1).value
            for row in range(2, recap.max_row + 1)
        ] == ["WB01-INV01-PV1", "WB03-INV02-PV5"]
        # Fase 1 & fase 2 keduanya ternormalkan ke 1.0 kWh/kWp hari pertama.
        assert abs(recap["B2"].value - 1.0) < 1e-9, recap["B2"].value
        assert abs(recap["B3"].value - 1.0) < 1e-9, recap["B3"].value
        # Hari kedua WB01 = 0.5; WB03 tidak ada data -> kosong.
        assert abs(recap["C2"].value - 0.5) < 1e-9, recap["C2"].value
        assert recap["C3"].value is None

        detail = workbook["Detail_Harian"]
        assert [cell.value for cell in detail[1]] == [
            "date", "pv_string", "inverter_id", "pv_label",
            "string_yield_kwh", "capacity_kwp", "specific_yield_kwh_per_kwp",
            "valid_power_samples", "expected_samples", "coverage_pct",
            "missing_power_samples", "source_csv", "status",
        ]
        metadata = _metadata_values(workbook)
        assert metadata["standard"] == "IEC 61724-1"
        assert metadata["module_wp"] == 625.0
        assert metadata["specific_yield_string_count"] == 2
        workbook.close()

        # Idempoten: jalankan ulang, workbook tetap satu file.
        with patch(
            "pv_pipeline.all_string_yield_report.download_csv_inputs",
            return_value=(manifest, inputs),
        ):
            _execute_cells(notebook, scope)
        assert list(output_dir.glob("specific_yield_*.xlsx")) == [output_xlsx]
        print("[smoke] OK")


if __name__ == "__main__":
    main()
