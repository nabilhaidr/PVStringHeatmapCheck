from __future__ import annotations

import json
from pathlib import Path


OUT = Path(__file__).parent / "String_Yield_Power_Irradiance.ipynb"

MD_INTRO = '''# Laporan String Yield, Power, dan Irradiance

Notebook ini memeriksa satu PV string pada rentang tanggal inklusif.

- CSV string: folder Google Drive publik CSV Export PV String.
- POA: file `POA PLTS IKN YYYY.xlsx` dari folder raw data input.
- Yield harian: `Σ(power_kw_valid × 5/60)`; power hilang tidak diisi atau diestimasi.
- Output: workbook empat-sheet di `output_string/`, lalu dapat diunduh dari Cell 8.

Jalankan Cell 1 sampai Cell 8 berurutan. Edit hanya lima nilai di Cell 2.
'''

CODE_SETUP = '''# Cell 1 — Setup Colab/repo/output
import subprocess, sys
subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "--upgrade", "gdown>=6.0.0"])
from pathlib import Path
import os, tempfile

def find_repo_root(start=None):
    path = Path(start or os.getcwd()).resolve()
    for candidate in [path, *path.parents]:
        if (candidate / "pv_pipeline").is_dir() and (candidate / "config").is_dir():
            return candidate
    raise RuntimeError("Repo root tidak ditemukan; jalankan notebook dari clone SolarYieldPro.")

PUBLIC_REPO_URL = "https://github.com/nabilhaidr/PVStringHeatmapCheck.git"
DEFAULT_REPO_DIR = Path.cwd() / "PVStringHeatmapCheck"
try:
    REPO_DIR = find_repo_root()
except RuntimeError:
    if not DEFAULT_REPO_DIR.exists():
        subprocess.check_call([
            "git", "clone", "--depth", "1", PUBLIC_REPO_URL,
            str(DEFAULT_REPO_DIR),
        ])
    REPO_DIR = find_repo_root(DEFAULT_REPO_DIR)
if str(REPO_DIR) not in sys.path:
    sys.path.insert(0, str(REPO_DIR))
OUTPUT_DIR = REPO_DIR / "output_string"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
INPUT_DIR = Path(tempfile.mkdtemp(prefix="string_yield_inputs_"))
print("OUTPUT_DIR:", OUTPUT_DIR)
'''

CODE_CONFIG = '''# Cell 2 — Konfigurasi input (edit lima nilai ini)
from pv_pipeline.string_yield_report import parse_date_range, parse_string_selection, validate_drive_folder_url
URL_CSV = "https://drive.google.com/drive/folders/1f_KrPuqfZJTE5I9cVQiyp65QrHbkF3Iw?usp=sharing"
URL_RAW_DATA_INPUT = "https://drive.google.com/drive/folders/1y37AsVViI7IL1tCRhvAGEjiq9wWnfu_e?usp=drive_link"
PV_STRING = "WB05-INV01-PV03"
START_DATE = "2026-05-01"
END_DATE = "2026-05-14"
SELECTION = parse_string_selection(PV_STRING)
DATES = parse_date_range(START_DATE, END_DATE)
validate_drive_folder_url(URL_CSV)
validate_drive_folder_url(URL_RAW_DATA_INPUT)
print(SELECTION, DATES[0].date(), "s.d.", DATES[-1].date())
'''

CODE_DOWNLOAD = '''# Cell 3 — Inventaris dan download selektif
from pv_pipeline.string_yield_report import download_report_inputs
if "DATES" not in globals():
    raise RuntimeError("Jalankan Cell 2 terlebih dahulu.")
MANIFEST, INPUTS = download_report_inputs(URL_CSV, URL_RAW_DATA_INPUT, DATES, INPUT_DIR)
print("CSV inventory:", MANIFEST.csv_inventory_count, "dipilih:", len(MANIFEST.csv_by_date), "berhasil:", len(INPUTS.csv_by_date), "missing:", MANIFEST.missing_csv_dates)
print("POA inventory:", MANIFEST.poa_inventory_count, "dipilih:", len(MANIFEST.poa_by_year), "berhasil:", len(INPUTS.poa_by_year), "missing years:", MANIFEST.missing_poa_years)
print("Download errors:", INPUTS.download_errors)
'''

CODE_PROCESS = '''# Cell 4 — Load, grid 5-menit, POA, dan yield harian
try:
    from IPython.display import display
except ImportError:
    display = print
from pv_pipeline.string_yield_report import build_report_data
if "INPUTS" not in globals():
    raise RuntimeError("Jalankan Cell 3 terlebih dahulu.")
REPORT = build_report_data(
    INPUTS.csv_by_date, INPUTS.poa_by_year, SELECTION, DATES,
    REPO_DIR / "config/site_geometry.yaml",
    source_manifest=MANIFEST, download_errors=INPUTS.download_errors,
)
display(REPORT.daily)
print(REPORT.metadata)
STATUS_COUNTS = REPORT.daily["status"].value_counts()
print({
    "requested_days": len(DATES),
    "loaded_csv_days": len(REPORT.metadata.get("loaded_csv_files", [])),
    "missing_csv_days": int((REPORT.daily["status"] == "MISSING_CSV").sum()),
    "complete_days": int(STATUS_COUNTS.get("COMPLETE", 0)),
    "partial_days": int(STATUS_COUNTS.get("PARTIAL", 0)),
})
'''

CODE_DAILY_PLOT = '''# Cell 5 — Grafik yield harian per tanggal
from pv_pipeline.string_yield_report import plot_daily_yield
if "REPORT" not in globals():
    raise RuntimeError("Jalankan Cell 4 terlebih dahulu.")
FIG_DAILY, AX_DAILY = plot_daily_yield(REPORT.daily, SELECTION)
import matplotlib.pyplot as plt
plt.show()
'''

CODE_POWER_PLOT = '''# Cell 6 — Grafik power vs irradiance, sumbu-Y sekunder
from pv_pipeline.string_yield_report import plot_power_vs_poa
if "REPORT" not in globals():
    raise RuntimeError("Jalankan Cell 4 terlebih dahulu.")
FIG_POWER_POA, (AX_POWER, AX_POA) = plot_power_vs_poa(REPORT.five_minute, SELECTION, DATES[0].date(), DATES[-1].date())
import matplotlib.pyplot as plt
plt.show()
'''

CODE_EXPORT = '''# Cell 7 — Ekspor dan verifikasi workbook
from openpyxl import load_workbook
from pv_pipeline.string_yield_report import build_output_path, verify_report_workbook, write_report_workbook
if "REPORT" not in globals():
    raise RuntimeError("Jalankan Cell 4 terlebih dahulu.")
OUTPUT_XLSX = build_output_path(OUTPUT_DIR, SELECTION, DATES[0].date(), DATES[-1].date())
write_report_workbook(OUTPUT_XLSX, REPORT)
verify_report_workbook(OUTPUT_XLSX)
CHECK_WB = load_workbook(OUTPUT_XLSX, read_only=False, data_only=False)
print("Workbook:", OUTPUT_XLSX, "sheets:", CHECK_WB.sheetnames, "bytes:", OUTPUT_XLSX.stat().st_size)
CHECK_WB.close()
'''

CODE_DOWNLOAD_XLSX = '''# Cell 8 — Download dari Colab; lokal hanya menampilkan path
if "OUTPUT_XLSX" not in globals() or not OUTPUT_XLSX.exists():
    raise RuntimeError("Jalankan Cell 7 terlebih dahulu.")
try:
    from google.colab import files
except ImportError:
    print("Bukan Colab; workbook tersedia di", OUTPUT_XLSX)
else:
    files.download(str(OUTPUT_XLSX))
'''

CELLS = [
    ("markdown", MD_INTRO),
    ("code", CODE_SETUP),
    ("code", CODE_CONFIG),
    ("code", CODE_DOWNLOAD),
    ("code", CODE_PROCESS),
    ("code", CODE_DAILY_PLOT),
    ("code", CODE_POWER_PLOT),
    ("code", CODE_EXPORT),
    ("code", CODE_DOWNLOAD_XLSX),
]


def _cell(kind: str, source: str, index: int) -> dict:
    cell = {
        "cell_type": kind,
        "id": f"string-yield-{index:02d}",
        "metadata": {},
        "source": source.splitlines(keepends=True),
    }
    if kind == "code":
        cell["execution_count"] = None
        cell["outputs"] = []
    return cell


def build(out: Path = OUT) -> Path:
    target = Path(out)
    target.parent.mkdir(parents=True, exist_ok=True)
    notebook = {
        "cells": [
            _cell(kind, source, index)
            for index, (kind, source) in enumerate(CELLS)
        ],
        "metadata": {
            "kernelspec": {
                "display_name": "Python 3",
                "language": "python",
                "name": "python3",
            },
            "language_info": {"name": "python"},
        },
        "nbformat": 4,
        "nbformat_minor": 5,
    }
    target.write_text(
        json.dumps(notebook, indent=1, ensure_ascii=False),
        encoding="utf-8",
    )
    print(f"Wrote {target} ({len(CELLS)} cells)")
    return target


if __name__ == "__main__":
    build()
