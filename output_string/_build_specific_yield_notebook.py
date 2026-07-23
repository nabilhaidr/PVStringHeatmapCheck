"""Builder Specific_Yield_Daily.ipynb.

Notebook meng-generate specific yield harian per PV string (kWh/kWp/hari,
IEC 61724-1) untuk rentang tanggal, meniru alur All_String_Daily_Yield:
download CSV -> energi harian (all_string_yield_report, satu sumber
kebenaran) -> normalisasi kapasitas per string (specific_yield_report).

Edit builder ini lalu jalankan:
    python output_string/_build_specific_yield_notebook.py
"""
from __future__ import annotations

import json
from pathlib import Path


OUT = Path(__file__).parent / "Specific_Yield_Daily.ipynb"

MD_INTRO = '''# Rekap Specific Yield Harian Semua PV String (IEC 61724-1)

Notebook ini menghitung **specific yield** harian tiap PV string pada rentang
tanggal inklusif.

- Specific yield = `energi_string_kWh / kapasitas_string_kWp`, satuan
  **kWh/kWp/hari** (setara *equivalent peak sun hours*).
- Kapasitas string: fase 1 (WB01-WB02) = 24 x 625 Wp = 15,00 kWp;
  fase 2 (WB03-WB10) = 26 x 625 Wp = 16,25 kWp.
- Energi harian dipakai apa adanya dari pipeline all-string daily yield
  (`sum(power_kw_valid * 5/60)`) supaya definisi energi konsisten satu sumber.
- Sumber: folder Google Drive publik CSV Export PV String.
- Output: workbook tiga-sheet di `output_string/`
  (`Rekap_SpecificYield`, `Detail_Harian`, `Metadata`).
- Rekap Excel: PV string sebagai baris, tanggal sebagai kolom.
- Cell terakhir menyediakan download workbook saat dijalankan di Colab.

Catatan: specific yield BUKAN Performance Ratio -- nilainya ikut naik-turun
mengikuti cuaca (belum dinormalkan insolasi POA). Untuk membandingkan antar
string pada hari yang sama, metrik ini memadai.

Jalankan Cell 1 sampai Cell 7 berurutan. Edit hanya nilai di Cell 2.

1.   Ubah START_DATE dan END_DATE sesuai tanggal yang diinginkan.
2.   EXCLUDE_EMPTY_SLOTS=True membuang slot PV kosong by design (strings.yaml).
3.   Klik Run All dan tunggu sampai proses selesai; file Excel akan otomatis
     terunduh, lalu upload ke folder Rekap Specific Yield.
'''

CODE_SETUP = '''# Cell 1 - Setup Colab/repo/output
import subprocess, sys
subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "--upgrade", "gdown>=6.0.0"])
from pathlib import Path
import os, tempfile

def find_repo_root(start=None):
    path = Path(start or os.getcwd()).resolve()
    for candidate in [path, *path.parents]:
        if (candidate / "pv_pipeline").is_dir() and (candidate / "config").is_dir():
            return candidate
    raise RuntimeError("Repo root tidak ditemukan; jalankan notebook dari clone SolarYieldPro.")

PUBLIC_REPO_URL = "https://github.com/nabilhaidr/PVStringHeatmapCheck.git"
DEFAULT_REPO_DIR = Path.cwd() / "PVStringHeatmapCheck"
try:
    REPO_DIR = find_repo_root()
except RuntimeError:
    if not DEFAULT_REPO_DIR.exists():
        subprocess.check_call([
            "git", "clone", "--depth", "1", PUBLIC_REPO_URL,
            str(DEFAULT_REPO_DIR),
        ])
    REPO_DIR = find_repo_root(DEFAULT_REPO_DIR)
if str(REPO_DIR) not in sys.path:
    sys.path.insert(0, str(REPO_DIR))
OUTPUT_DIR = REPO_DIR / "output_string"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
INPUT_DIR = Path(tempfile.mkdtemp(prefix="specific_yield_inputs_"))
STRINGS_YAML = str(REPO_DIR / "config" / "strings.yaml")
print("OUTPUT_DIR:", OUTPUT_DIR)
'''

CODE_CONFIG = '''# Cell 2 - Konfigurasi input (edit nilai di sini)
from pv_pipeline.string_yield_report import parse_date_range, validate_drive_folder_url
URL_CSV = "https://drive.google.com/drive/folders/1f_KrPuqfZJTE5I9cVQiyp65QrHbkF3Iw?usp=sharing"
START_DATE = "2026-06-01"
END_DATE = "2026-06-30"
EXCLUDE_EMPTY_SLOTS = True  # buang slot PV kosong by design (config/strings.yaml)
DATES = parse_date_range(START_DATE, END_DATE)
validate_drive_folder_url(URL_CSV)
print(DATES[0].date(), "s.d.", DATES[-1].date(), "| exclude_empty_slots:", EXCLUDE_EMPTY_SLOTS)
'''

CODE_DOWNLOAD = '''# Cell 3 - Inventaris dan download CSV selektif
from pv_pipeline.all_string_yield_report import download_csv_inputs
if any(name not in globals() for name in ("DATES", "URL_CSV", "INPUT_DIR")):
    raise RuntimeError("Jalankan Cell 2 terlebih dahulu.")
MANIFEST, INPUTS = download_csv_inputs(URL_CSV, DATES, INPUT_DIR)
print(
    "CSV inventory:", MANIFEST.csv_inventory_count,
    "dipilih:", len(MANIFEST.csv_by_date),
    "berhasil:", len(INPUTS.csv_by_date),
    "missing:", MANIFEST.missing_csv_dates,
)
print("Download errors:", INPUTS.download_errors)
'''

CODE_ENERGY = '''# Cell 4 - Hitung energi harian seluruh string (sumber kebenaran energi)
try:
    from IPython.display import display
except ImportError:
    display = print
from pv_pipeline.all_string_yield_report import build_all_string_daily_yield
if any(name not in globals() for name in ("INPUTS", "MANIFEST", "DATES")):
    raise RuntimeError("Jalankan Cell 3 terlebih dahulu.")
ENERGY_REPORT = build_all_string_daily_yield(
    INPUTS.csv_by_date,
    DATES,
    source_manifest=MANIFEST,
    download_errors=INPUTS.download_errors,
)
print({
    "requested_days": len(DATES),
    "detected_strings": ENERGY_REPORT.metadata["detected_string_count"],
    "detail_rows": len(ENERGY_REPORT.daily),
    "status_counts": ENERGY_REPORT.daily["status"].value_counts().to_dict(),
})
print("Warnings:", ENERGY_REPORT.metadata["warnings"])
'''

CODE_SPECIFIC = '''# Cell 5 - Normalisasi ke specific yield (kWh/kWp/hari, IEC 61724-1)
from pv_pipeline.specific_yield_report import build_specific_yield
from pv_pipeline.string_config import get_empty_pv_map
if any(name not in globals() for name in ("ENERGY_REPORT", "STRINGS_YAML")):
    raise RuntimeError("Jalankan Cell 4 terlebih dahulu.")

EMPTY_PV_MAP = None
if globals().get("EXCLUDE_EMPTY_SLOTS", False):
    EMPTY_PV_MAP = get_empty_pv_map(STRINGS_YAML)

REPORT = build_specific_yield(ENERGY_REPORT, empty_pv_map=EMPTY_PV_MAP)
display(REPORT.summary)
_sy = REPORT.daily["specific_yield_kwh_per_kwp"].dropna()
print({
    "strings": REPORT.metadata["specific_yield_string_count"],
    "excluded_empty_slots": REPORT.metadata["excluded_empty_slot_strings"],
    "specific_yield_median": round(float(_sy.median()), 3) if len(_sy) else None,
    "specific_yield_p5_p95": [
        round(float(_sy.quantile(0.05)), 3),
        round(float(_sy.quantile(0.95)), 3),
    ] if len(_sy) else None,
    "unit": REPORT.metadata["specific_yield_unit"],
})
'''

CODE_EXPORT = '''# Cell 6 - Ekspor dan verifikasi workbook
from openpyxl import load_workbook
from pv_pipeline.specific_yield_report import (
    build_specific_yield_output_path,
    verify_specific_yield_workbook,
    write_specific_yield_workbook,
)
OUTPUT_VERIFIED = False
if any(name not in globals() for name in ("REPORT", "DATES", "OUTPUT_DIR")):
    raise RuntimeError("Jalankan Cell 5 terlebih dahulu.")
OUTPUT_XLSX = build_specific_yield_output_path(
    OUTPUT_DIR,
    DATES[0].date(),
    DATES[-1].date(),
)
write_specific_yield_workbook(OUTPUT_XLSX, REPORT)
verify_specific_yield_workbook(OUTPUT_XLSX)
CHECK_WB = load_workbook(OUTPUT_XLSX, read_only=False, data_only=False)
print(
    "Workbook:", OUTPUT_XLSX,
    "sheets:", CHECK_WB.sheetnames,
    "bytes:", OUTPUT_XLSX.stat().st_size,
    "days:", REPORT.metadata["requested_days"],
    "strings:", REPORT.metadata["specific_yield_string_count"],
)
CHECK_WB.close()
OUTPUT_VERIFIED = True
'''

CODE_DOWNLOAD_XLSX = '''# Cell 7 - Download dari Colab; lokal hanya menampilkan path
if (
    not globals().get("OUTPUT_VERIFIED", False)
    or "OUTPUT_XLSX" not in globals()
    or not OUTPUT_XLSX.exists()
):
    raise RuntimeError("Jalankan Cell 6 terlebih dahulu.")
try:
    from google.colab import files
except ImportError:
    print("Bukan Colab; workbook tersedia di", OUTPUT_XLSX)
else:
    files.download(str(OUTPUT_XLSX))
'''

CELLS = [
    ("markdown", MD_INTRO),
    ("code", CODE_SETUP),
    ("code", CODE_CONFIG),
    ("code", CODE_DOWNLOAD),
    ("code", CODE_ENERGY),
    ("code", CODE_SPECIFIC),
    ("code", CODE_EXPORT),
    ("code", CODE_DOWNLOAD_XLSX),
]


def _cell(kind: str, source: str, index: int) -> dict:
    cell = {
        "cell_type": kind,
        "id": f"specific-yield-{index:02d}",
        "metadata": {},
        "source": source.splitlines(keepends=True),
    }
    if kind == "code":
        cell["execution_count"] = None
        cell["outputs"] = []
    return cell


def build(out: Path = OUT) -> Path:
    target = Path(out)
    target.parent.mkdir(parents=True, exist_ok=True)
    notebook = {
        "cells": [
            _cell(kind, source, index)
            for index, (kind, source) in enumerate(CELLS)
        ],
        "metadata": {
            "kernelspec": {
                "display_name": "Python 3",
                "language": "python",
                "name": "python3",
            },
            "language_info": {"name": "python"},
        },
        "nbformat": 4,
        "nbformat_minor": 5,
    }
    target.write_text(
        json.dumps(notebook, indent=1, ensure_ascii=False),
        encoding="utf-8",
    )
    print(f"Wrote {target} ({len(CELLS)} cells)")
    return target


if __name__ == "__main__":
    build()
