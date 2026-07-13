from __future__ import annotations

import json
import sys
import tempfile
from datetime import date
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from pv_pipeline.string_yield_report import (
    DownloadedInputs,
    DriveItem,
    SourceManifest,
    parse_date_range,
)


NOTEBOOK = ROOT / "output_string/All_String_Daily_Yield.ipynb"


def _source(notebook: dict, index: int) -> str:
    return "".join(notebook["cells"][index]["source"])


def _metadata_values(workbook) -> dict[str, object]:
    sheet = workbook["Metadata"]
    return {
        sheet.cell(row=row, column=1).value:
        sheet.cell(row=row, column=2).value
        for row in range(2, sheet.max_row + 1)
    }


def _execute_report_cells(notebook: dict, scope: dict) -> None:
    exec(_source(notebook, 2), scope)
    scope.update({
        "START_DATE": "2026-05-01",
        "END_DATE": "2026-05-02",
        "DATES": parse_date_range("2026-05-01", "2026-05-02"),
    })
    for index in range(3, 6):
        if index == 4:
            with patch.dict(
                sys.modules,
                {"IPython": None, "IPython.display": None},
            ):
                exec(_source(notebook, index), scope)
        else:
            exec(_source(notebook, index), scope)


def main() -> None:
    notebook = json.loads(NOTEBOOK.read_text(encoding="utf-8"))
    with tempfile.TemporaryDirectory(
        prefix="all_string_yield_smoke_",
    ) as temp_name:
        temp = Path(temp_name)
        input_dir = temp / "inputs"
        output_dir = temp / "output_string"
        input_dir.mkdir()
        output_dir.mkdir()

        stamps_one = pd.date_range(
            "2026-05-01 00:00",
            periods=12,
            freq="5min",
        )
        day_one = pd.concat([
            pd.DataFrame({
                "Start Time": stamps_one,
                "Inverter_ID": "WB01-INV01",
                "PV1 Power(kW)": 12.0,
            }),
            pd.DataFrame({
                "Start Time": stamps_one,
                "Inverter_ID": "WB01-INV02",
                "PV1 Power(kW)": 6.0,
            }),
        ], ignore_index=True)
        csv_one = input_dir / "20260501.csv"
        day_one.to_csv(csv_one, index=False)

        csv_two = input_dir / "20260502.csv"
        pd.DataFrame({
            "Start Time": pd.date_range(
                "2026-05-02 00:00",
                periods=12,
                freq="5min",
            ),
            "Inverter_ID": "WB01-INV01",
            "PV1 Power(kW)": 10.0,
        }).to_csv(csv_two, index=False)

        manifest = SourceManifest(
            csv_by_date={
                date(2026, 5, 1): DriveItem("local-one", csv_one.name),
                date(2026, 5, 2): DriveItem("local-two", csv_two.name),
            },
            poa_by_year={},
            missing_csv_dates=[],
            missing_poa_years=[],
            url_csv=(
                "https://drive.google.com/drive/folders/csv-smoke"
            ),
            csv_inventory_count=2,
        )
        inputs = DownloadedInputs(
            csv_by_date={
                date(2026, 5, 1): csv_one,
                date(2026, 5, 2): csv_two,
            },
            poa_by_year={},
            download_errors={},
        )
        scope = {
            "REPO_DIR": ROOT,
            "OUTPUT_DIR": output_dir,
            "INPUT_DIR": input_dir,
        }
        download_calls = []

        def fake_download_csv_inputs(
            url_csv,
            dates,
            destination,
        ):
            download_calls.append((
                url_csv,
                tuple(timestamp.date() for timestamp in dates),
                Path(destination),
            ))
            return manifest, inputs

        with patch(
            "pv_pipeline.all_string_yield_report.download_csv_inputs",
            side_effect=fake_download_csv_inputs,
        ):
            _execute_report_cells(notebook, scope)
            output_xlsx = Path(scope["OUTPUT_XLSX"])
            workbook = load_workbook(output_xlsx, data_only=False)
            assert workbook.sheetnames == [
                "Rekap_Yield_kWh",
                "Detail_Harian",
                "Metadata",
            ]
            recap = workbook["Rekap_Yield_kWh"]
            assert [cell.value for cell in recap[1]] == [
                "date",
                "WB01-INV01-PV1",
                "WB01-INV02-PV1",
            ]
            assert recap["B2"].value == 12.0
            assert recap["C2"].value == 6.0
            assert recap["B3"].value == 10.0
            assert recap["C3"].value is None
            assert workbook["Detail_Harian"].max_row == 5
            metadata = _metadata_values(workbook)
            assert metadata["requested_days"] == 2
            assert metadata["detected_string_count"] == 2
            row_counts = (
                recap.max_row,
                workbook["Detail_Harian"].max_row,
            )
            workbook.close()

            _execute_report_cells(notebook, scope)

        expected_call = (
            "https://drive.google.com/drive/folders/"
            "1f_KrPuqfZJTE5I9cVQiyp65QrHbkF3Iw?usp=sharing",
            (date(2026, 5, 1), date(2026, 5, 2)),
            input_dir,
        )
        assert download_calls == [expected_call, expected_call]
        workbooks = list(output_dir.glob("all_string_yield_*.xlsx"))
        assert workbooks == [output_xlsx]
        rerun = load_workbook(output_xlsx, data_only=False)
        assert (
            rerun["Rekap_Yield_kWh"].max_row,
            rerun["Detail_Harian"].max_row,
        ) == row_counts
        rerun.close()
        print("[smoke] OK")


if __name__ == "__main__":
    main()
