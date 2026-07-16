from __future__ import annotations

import json
from pathlib import Path


OUT = Path(__file__).parent / "All_String_Daily_Yield.ipynb"

MD_INTRO = '''# Rekap Yield Harian Semua PV String

Notebook ini menghitung yield harian seluruh PV string yang terdeteksi pada
rentang tanggal inklusif.

- Sumber: folder Google Drive publik CSV Export PV String.
- Yield: `sum(power_kw_valid * 5/60)` tanpa mengisi data yang hilang.
- Output: workbook tiga-sheet di `output_string/`.
- Rekap Excel: PV string sebagai baris dan tanggal sebagai kolom.
- Jika detail melampaui batas baris Excel, pendekkan rentang tanggal sesuai pesan error.
- Cell terakhir menyediakan download workbook saat dijalankan di Colab.

Jalankan Cell 1 sampai Cell 6 berurutan. Edit hanya tiga nilai di Cell 2.
'''

CODE_SETUP = '''# Cell 1 - Setup Colab/repo/output
import subprocess, sys
subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "--upgrade", "gdown>=6.0.0"])
from pathlib import Path
import os, tempfile

def find_repo_root(start=None):
    path = Path(start or os.getcwd()).resolve()
    for candidate in [path, *path.parents]:
        if (candidate / "pv_pipeline").is_dir() and (candidate / "config").is_dir():
            return candidate
    raise RuntimeError("Repo root tidak ditemukan; jalankan notebook dari clone SolarYieldPro.")

PUBLIC_REPO_URL = "https://github.com/nabilhaidr/PVStringHeatmapCheck.git"
DEFAULT_REPO_DIR = Path.cwd() / "PVStringHeatmapCheck"
try:
    REPO_DIR = find_repo_root()
except RuntimeError:
    if not DEFAULT_REPO_DIR.exists():
        subprocess.check_call([
            "git", "clone", "--depth", "1", PUBLIC_REPO_URL,
            str(DEFAULT_REPO_DIR),
        ])
    REPO_DIR = find_repo_root(DEFAULT_REPO_DIR)
if str(REPO_DIR) not in sys.path:
    sys.path.insert(0, str(REPO_DIR))
OUTPUT_DIR = REPO_DIR / "output_string"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
INPUT_DIR = Path(tempfile.mkdtemp(prefix="all_string_yield_inputs_"))
print("OUTPUT_DIR:", OUTPUT_DIR)
'''

CODE_CONFIG = '''# Cell 2 - Konfigurasi input (edit tiga nilai ini)
from pv_pipeline.string_yield_report import parse_date_range, validate_drive_folder_url
URL_CSV = "https://drive.google.com/drive/folders/1f_KrPuqfZJTE5I9cVQiyp65QrHbkF3Iw?usp=sharing"
START_DATE = "2026-05-01"
END_DATE = "2026-05-14"
DATES = parse_date_range(START_DATE, END_DATE)
validate_drive_folder_url(URL_CSV)
print(DATES[0].date(), "s.d.", DATES[-1].date())
'''

CODE_DOWNLOAD = '''# Cell 3 - Inventaris dan download CSV selektif
from pv_pipeline.all_string_yield_report import download_csv_inputs
if any(name not in globals() for name in ("DATES", "URL_CSV", "INPUT_DIR")):
    raise RuntimeError("Jalankan Cell 2 terlebih dahulu.")
MANIFEST, INPUTS = download_csv_inputs(URL_CSV, DATES, INPUT_DIR)
print(
    "CSV inventory:", MANIFEST.csv_inventory_count,
    "dipilih:", len(MANIFEST.csv_by_date),
    "berhasil:", len(INPUTS.csv_by_date),
    "missing:", MANIFEST.missing_csv_dates,
)
print("Download errors:", INPUTS.download_errors)
'''

CODE_PROCESS = '''# Cell 4 - Hitung yield harian seluruh string
try:
    from IPython.display import display
except ImportError:
    display = print
from pv_pipeline.all_string_yield_report import build_all_string_daily_yield
if any(name not in globals() for name in ("INPUTS", "MANIFEST", "DATES")):
    raise RuntimeError("Jalankan Cell 3 terlebih dahulu.")
REPORT = build_all_string_daily_yield(
    INPUTS.csv_by_date,
    DATES,
    source_manifest=MANIFEST,
    download_errors=INPUTS.download_errors,
)
display(REPORT.summary)
STATUS_COUNTS = REPORT.daily["status"].value_counts().to_dict()
print({
    "requested_days": len(DATES),
    "detected_strings": REPORT.metadata["detected_string_count"],
    "detail_rows": len(REPORT.daily),
    "status_counts": STATUS_COUNTS,
})
print("Warnings:", REPORT.metadata["warnings"])
'''

CODE_EXPORT = '''# Cell 5 - Ekspor dan verifikasi workbook
from openpyxl import load_workbook
from pv_pipeline.all_string_yield_report import (
    build_all_string_output_path,
    verify_all_string_workbook,
    write_all_string_workbook,
)
OUTPUT_VERIFIED = False
if any(name not in globals() for name in ("REPORT", "DATES", "OUTPUT_DIR")):
    raise RuntimeError("Jalankan Cell 4 terlebih dahulu.")
OUTPUT_XLSX = build_all_string_output_path(
    OUTPUT_DIR,
    DATES[0].date(),
    DATES[-1].date(),
)
write_all_string_workbook(OUTPUT_XLSX, REPORT)
verify_all_string_workbook(OUTPUT_XLSX)
CHECK_WB = load_workbook(OUTPUT_XLSX, read_only=False, data_only=False)
print(
    "Workbook:", OUTPUT_XLSX,
    "sheets:", CHECK_WB.sheetnames,
    "bytes:", OUTPUT_XLSX.stat().st_size,
    "days:", REPORT.metadata["requested_days"],
    "strings:", REPORT.metadata["detected_string_count"],
)
CHECK_WB.close()
OUTPUT_VERIFIED = True
'''

CODE_DOWNLOAD_XLSX = '''# Cell 6 - Download dari Colab; lokal hanya menampilkan path
if (
    not globals().get("OUTPUT_VERIFIED", False)
    or "OUTPUT_XLSX" not in globals()
    or not OUTPUT_XLSX.exists()
):
    raise RuntimeError("Jalankan Cell 5 terlebih dahulu.")
try:
    from google.colab import files
except ImportError:
    print("Bukan Colab; workbook tersedia di", OUTPUT_XLSX)
else:
    files.download(str(OUTPUT_XLSX))
'''

CELLS = [
    ("markdown", MD_INTRO),
    ("code", CODE_SETUP),
    ("code", CODE_CONFIG),
    ("code", CODE_DOWNLOAD),
    ("code", CODE_PROCESS),
    ("code", CODE_EXPORT),
    ("code", CODE_DOWNLOAD_XLSX),
]


def _cell(kind: str, source: str, index: int) -> dict:
    cell = {
        "cell_type": kind,
        "id": f"all-string-yield-{index:02d}",
        "metadata": {},
        "source": source.splitlines(keepends=True),
    }
    if kind == "code":
        cell["execution_count"] = None
        cell["outputs"] = []
    return cell


def build(out: Path = OUT) -> Path:
    target = Path(out)
    target.parent.mkdir(parents=True, exist_ok=True)
    notebook = {
        "cells": [
            _cell(kind, source, index)
            for index, (kind, source) in enumerate(CELLS)
        ],
        "metadata": {
            "kernelspec": {
                "display_name": "Python 3",
                "language": "python",
                "name": "python3",
            },
            "language_info": {"name": "python"},
        },
        "nbformat": 4,
        "nbformat_minor": 5,
    }
    target.write_text(
        json.dumps(notebook, indent=1, ensure_ascii=False),
        encoding="utf-8",
    )
    print(f"Wrote {target} ({len(CELLS)} cells)")
    return target


if __name__ == "__main__":
    build()
