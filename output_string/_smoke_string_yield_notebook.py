from __future__ import annotations

import json
import sys
import tempfile
from datetime import date
from pathlib import Path
from unittest.mock import patch

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import pandas as pd


plt.show = lambda *args, **kwargs: None

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from pv_pipeline.string_yield_report import (
    DownloadedInputs,
    DriveItem,
    SourceManifest,
    parse_date_range,
    parse_string_selection,
)


NOTEBOOK = ROOT / "output_string/String_Yield_Power_Irradiance.ipynb"


def _source(notebook: dict, index: int) -> str:
    return "".join(notebook["cells"][index]["source"])


def _metadata_values(workbook) -> dict[str, object]:
    sheet = workbook["Metadata"]
    return {
        sheet.cell(row=row, column=1).value:
        sheet.cell(row=row, column=2).value
        for row in range(2, sheet.max_row + 1)
    }


def _execute_report_cells(notebook: dict, scope: dict) -> None:
    exec(_source(notebook, 2), scope)
    scope.update({
        "PV_STRING": "WB05-INV01-PV03",
        "START_DATE": "2026-05-01",
        "END_DATE": "2026-05-01",
        "SELECTION": parse_string_selection("WB05-INV01-PV03"),
        "DATES": parse_date_range("2026-05-01", "2026-05-01"),
    })
    for index in range(3, 8):
        if index == 4:
            with patch.dict(
                sys.modules,
                {"IPython": None, "IPython.display": None},
            ):
                exec(_source(notebook, index), scope)
        else:
            exec(_source(notebook, index), scope)


def main() -> None:
    notebook = json.loads(NOTEBOOK.read_text(encoding="utf-8"))
    with tempfile.TemporaryDirectory(
        prefix="string_yield_smoke_",
    ) as temp_name:
        temp = Path(temp_name)
        input_dir = temp / "inputs"
        output_dir = temp / "output_string"
        input_dir.mkdir()
        output_dir.mkdir()

        # 13 nominal slots 00:00..01:00; delete 00:05 so 12 x 10 kW = 10 kWh.
        nominal = pd.date_range("2026-05-01 00:00", periods=13, freq="5min")
        observed = nominal.delete(1)
        csv_path = input_dir / "20260501.csv"
        pd.DataFrame({
            "Start Time": observed,
            "Inverter_ID": "WB05-INV01",
            "PV3 Power(kW)": 10.0,
        }).to_csv(csv_path, index=False)

        poa_path = input_dir / "POA PLTS IKN 2026.xlsx"
        poa_data = {
            "Date time": nominal,
            "Rata-rata WS 1 - WS 5": 450.0,
        }
        for ws_number in range(1, 6):
            poa_data[
                f"POA Irradiance (W/m2) WS {ws_number}"
            ] = 500.0
        poa_data["POA Irradiance (W/m2) WS 2"] = np.array(
            [500.0, np.nan] + [500.0] * 11,
        )
        with pd.ExcelWriter(poa_path, engine="openpyxl") as writer:
            pd.DataFrame(poa_data).to_excel(
                writer,
                sheet_name="POA PLTS IKN",
                index=False,
            )

        selected_csv = DriveItem("local-csv", "20260501.csv")
        selected_poa = DriveItem("local-poa", "POA PLTS IKN 2026.xlsx")
        manifest = SourceManifest(
            csv_by_date={date(2026, 5, 1): selected_csv},
            poa_by_year={2026: selected_poa},
            missing_csv_dates=[],
            missing_poa_years=[],
            url_csv="https://drive.google.com/drive/folders/csv-smoke",
            url_poa="https://drive.google.com/drive/folders/poa-smoke",
            csv_inventory_count=1,
            poa_inventory_count=1,
        )
        inputs = DownloadedInputs(
            csv_by_date={date(2026, 5, 1): csv_path},
            poa_by_year={2026: poa_path},
            download_errors={},
        )
        scope = {
            "REPO_DIR": ROOT,
            "OUTPUT_DIR": output_dir,
            "INPUT_DIR": input_dir,
        }

        with patch(
            "pv_pipeline.string_yield_report.download_report_inputs",
            return_value=(manifest, inputs),
        ):
            _execute_report_cells(notebook, scope)

            output_xlsx = Path(scope["OUTPUT_XLSX"])
            workbook = load_workbook(output_xlsx, data_only=False)
            assert workbook.sheetnames == [
                "Ringkasan_Harian",
                "Data_5Menit",
                "Grafik",
                "Metadata",
            ]
            assert workbook["Ringkasan_Harian"]["B2"].value == 10.0
            headers = {
                cell.value: cell.column
                for cell in workbook["Data_5Menit"][1]
            }
            assert workbook["Data_5Menit"].cell(
                row=3,
                column=headers["power_kw"],
            ).value is None
            assert workbook["Data_5Menit"].cell(
                row=3,
                column=headers["poa_source"],
            ).value == "avg"
            metadata = _metadata_values(workbook)
            assert metadata["poa_fallback_samples"] == 1
            assert len(workbook["Ringkasan_Harian"]._charts) == 1
            assert len(workbook["Grafik"]._charts) == 2
            row_counts = (
                workbook["Ringkasan_Harian"].max_row,
                workbook["Data_5Menit"].max_row,
            )
            workbook.close()

            plt.close("all")
            _execute_report_cells(notebook, scope)

        workbooks = list(output_dir.glob("string_yield_*.xlsx"))
        assert workbooks == [output_xlsx]
        rerun = load_workbook(output_xlsx, data_only=False)
        assert (
            rerun["Ringkasan_Harian"].max_row,
            rerun["Data_5Menit"].max_row,
        ) == row_counts
        assert len(rerun["Ringkasan_Harian"]._charts) == 1
        assert len(rerun["Grafik"]._charts) == 2
        rerun.close()
        print("[smoke] OK")


if __name__ == "__main__":
    main()
