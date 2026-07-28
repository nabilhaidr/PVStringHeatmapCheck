"""Builder Specific_Yield_Sibling_Ratio.ipynb.

Notebook membandingkan Yf tiap string terhadap tetangga se-inverter pada
hari yang sama (Jalur B) -- tanpa data POA sama sekali. Sumber: workbook
specific yield di folder Drive "Rekap String Specific Yield".

Edit builder ini lalu jalankan:
    python output_string/_build_sibling_ratio_notebook.py
"""
from __future__ import annotations

import json
from pathlib import Path


OUT = Path(__file__).parent / "Specific_Yield_Sibling_Ratio.ipynb"

MD_INTRO = '''# Rasio String vs Tetangga (Jalur B) -- tanpa data POA

Notebook ini mencari **string mana yang kotor** dengan membandingkan
specific yield (Yf) tiap string terhadap **median tetangga se-inverter pada
hari yang sama**.

Kenapa tanpa POA bisa: semua string menerima irradiance dan suhu yang sama
pada hari yang sama, jadi keduanya tercoret saat dibagi:

```
Yf_string / Yf_tetangga = (SR_string x Yr x Tf) / (SR_tetangga x Yr x Tf)
                        = SR_string / SR_tetangga
```

Yang tersisa murni rasio soiling relatif. Rasio dihitung **per hari dulu**,
baru diagregasi dengan median lintas hari -- urutan ini yang membuat cuaca
benar-benar hilang.

**Titik buta:** soiling yang SERAGAM se-inverter tidak terlihat (semua turun
bersama, rasionya tetap 1). Kolom `ratio_vs_wb` sedikit menutupi ini dengan
membandingkan terhadap seluruh WB. Untuk loss absolut (%) tetap pakai
M2aSoiling SRR.

Sumber: workbook `specific_yield_*.xlsx` di folder Drive
**Cek PV String/output_string/Rekap String Specific Yield**.

Output workbook tiga sheet di `output_string/`:
- `Ranking_Sibling` -- satu baris per string, terurut paling kotor dulu.
- `Rasio_Harian` -- rasio harian (string x tanggal) untuk plotting tren.
- `Metadata` -- rumus, ambang, dan jumlah baris yang disaring.

Jalankan Cell 1 sampai Cell 6 berurutan. Edit hanya nilai di Cell 2.
'''

CODE_SETUP = '''# Cell 1 - Setup Colab/repo/output
import subprocess, sys
subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "--upgrade", "gdown>=6.0.0"])
from pathlib import Path
import os, tempfile

def find_repo_root(start=None):
    path = Path(start or os.getcwd()).resolve()
    for candidate in [path, *path.parents]:
        if (candidate / "pv_pipeline").is_dir() and (candidate / "config").is_dir():
            return candidate
    raise RuntimeError("Repo root tidak ditemukan; jalankan notebook dari clone SolarYieldPro.")

PUBLIC_REPO_URL = "https://github.com/nabilhaidr/PVStringHeatmapCheck.git"
DEFAULT_REPO_DIR = Path.cwd() / "PVStringHeatmapCheck"
try:
    REPO_DIR = find_repo_root()
except RuntimeError:
    if not DEFAULT_REPO_DIR.exists():
        subprocess.check_call([
            "git", "clone", "--depth", "1", PUBLIC_REPO_URL,
            str(DEFAULT_REPO_DIR),
        ])
    REPO_DIR = find_repo_root(DEFAULT_REPO_DIR)
if str(REPO_DIR) not in sys.path:
    sys.path.insert(0, str(REPO_DIR))
OUTPUT_DIR = REPO_DIR / "output_string"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
INPUT_DIR = Path(tempfile.mkdtemp(prefix="sibling_ratio_inputs_"))
print("OUTPUT_DIR:", OUTPUT_DIR)
'''

CODE_CONFIG = '''# Cell 2 - Konfigurasi input (edit nilai di sini)
from pv_pipeline.yf_ratio_report import validate_drive_folder_url

# Folder Drive: Cek PV String/output_string/Rekap String Specific Yield
URL_SPECIFIC_YIELD = "https://drive.google.com/drive/folders/16FRx87xG5G9tcyS3cfViQw1wYMgMEMbF"
# Kosongkan untuk memakai workbook specific_yield_*.xlsx TERBARU di folder;
# atau tulis nama persis, mis. "specific_yield_20260601_20260630.xlsx".
SPECIFIC_YIELD_FILENAME = ""

# Ambang analisis.
MIN_DAYS = 5                 # minimal hari data valid per string
MIN_SIBLINGS = 3             # minimal string per inverter per hari
MIN_RELATIVE_COVERAGE = 0.8  # buang string-day yg cakupannya < 0.8x tetangga
DEAD_RATIO = 0.10            # rasio < ini -> DEAD_OR_OFFLINE (bukan kotor)
CANDIDATE_DEFICIT_PCT = 5.0  # defisit >= ini -> CLEANING_CANDIDATE

validate_drive_folder_url(URL_SPECIFIC_YIELD)
print("Sumber:", URL_SPECIFIC_YIELD)
print("File   :", SPECIFIC_YIELD_FILENAME or "(otomatis: yang terbaru)")
'''

CODE_DOWNLOAD = '''# Cell 3 - Unduh workbook specific yield dari Drive
from pv_pipeline.yf_ratio_report import download_drive_file
if any(name not in globals() for name in ("URL_SPECIFIC_YIELD", "INPUT_DIR")):
    raise RuntimeError("Jalankan Cell 2 terlebih dahulu.")
SPECIFIC_YIELD_XLSX = download_drive_file(
    URL_SPECIFIC_YIELD,
    INPUT_DIR,
    basename=SPECIFIC_YIELD_FILENAME,
    pattern="" if SPECIFIC_YIELD_FILENAME else r"^specific_yield_.*\\.xlsx$",
)
print("Workbook sumber:", SPECIFIC_YIELD_XLSX.name)
'''

CODE_BUILD = '''# Cell 4 - Hitung rasio string vs tetangga
try:
    from IPython.display import display
except ImportError:
    display = print
from pv_pipeline.yf_ratio_report import build_sibling_ratio, load_specific_yield_long
if "SPECIFIC_YIELD_XLSX" not in globals():
    raise RuntimeError("Jalankan Cell 3 terlebih dahulu.")

LONG_DF = load_specific_yield_long(SPECIFIC_YIELD_XLSX)
DATES = sorted(LONG_DF["date"].dt.date.unique())
print(f"{len(LONG_DF):,} baris | {LONG_DF['pv_string'].nunique():,} string | "
      f"{DATES[0]} s.d. {DATES[-1]}")

REPORT = build_sibling_ratio(
    LONG_DF,
    min_days=MIN_DAYS,
    min_siblings=MIN_SIBLINGS,
    min_relative_coverage=MIN_RELATIVE_COVERAGE,
    dead_ratio=DEAD_RATIO,
    candidate_deficit_pct=CANDIDATE_DEFICIT_PCT,
    source_metadata={"source_workbook": SPECIFIC_YIELD_XLSX.name},
)
print("Status:", REPORT.metadata["status_counts"])
print("\\n20 string paling kotor (kandidat cleaning):")
display(
    REPORT.ranking[REPORT.ranking["status"] == "CLEANING_CANDIDATE"]
    .head(20)[[
        "rank", "pv_string", "n_days", "yf_median", "sibling_yf_median",
        "deficit_vs_inverter_pct", "deficit_vs_wb_pct",
    ]].round(3)
)
_dead = REPORT.ranking[REPORT.ranking["status"] == "DEAD_OR_OFFLINE"]
if len(_dead):
    print(f"\\n{len(_dead)} string terindikasi MATI/OFFLINE (rasio < {DEAD_RATIO}) "
          "-> perlu inspeksi teknis, BUKAN cleaning. Contoh:",
          list(_dead["pv_string"].head(5)))
'''

CODE_EXPORT = '''# Cell 5 - Ekspor dan verifikasi workbook
from openpyxl import load_workbook
from pv_pipeline.yf_ratio_report import (
    build_sibling_ratio_output_path,
    verify_sibling_ratio_workbook,
    write_sibling_ratio_workbook,
)
OUTPUT_VERIFIED = False
if any(name not in globals() for name in ("REPORT", "DATES", "OUTPUT_DIR")):
    raise RuntimeError("Jalankan Cell 4 terlebih dahulu.")
OUTPUT_XLSX = build_sibling_ratio_output_path(OUTPUT_DIR, DATES[0], DATES[-1])
write_sibling_ratio_workbook(OUTPUT_XLSX, REPORT)
verify_sibling_ratio_workbook(OUTPUT_XLSX)
CHECK_WB = load_workbook(OUTPUT_XLSX, read_only=False, data_only=False)
print(
    "Workbook:", OUTPUT_XLSX,
    "sheets:", CHECK_WB.sheetnames,
    "bytes:", OUTPUT_XLSX.stat().st_size,
    "strings:", REPORT.metadata["string_count"],
    "days:", REPORT.metadata["day_count"],
)
CHECK_WB.close()
OUTPUT_VERIFIED = True
'''

CODE_DOWNLOAD_XLSX = '''# Cell 6 - Download dari Colab; lokal hanya menampilkan path
if (
    not globals().get("OUTPUT_VERIFIED", False)
    or "OUTPUT_XLSX" not in globals()
    or not OUTPUT_XLSX.exists()
):
    raise RuntimeError("Jalankan Cell 5 terlebih dahulu.")
try:
    from google.colab import files
except ImportError:
    print("Bukan Colab; workbook tersedia di", OUTPUT_XLSX)
else:
    files.download(str(OUTPUT_XLSX))
'''

CELLS = [
    ("markdown", MD_INTRO),
    ("code", CODE_SETUP),
    ("code", CODE_CONFIG),
    ("code", CODE_DOWNLOAD),
    ("code", CODE_BUILD),
    ("code", CODE_EXPORT),
    ("code", CODE_DOWNLOAD_XLSX),
]


def _cell(kind: str, source: str, index: int) -> dict:
    cell = {
        "cell_type": kind,
        "id": f"sibling-ratio-{index:02d}",
        "metadata": {},
        "source": source.splitlines(keepends=True),
    }
    if kind == "code":
        cell["execution_count"] = None
        cell["outputs"] = []
    return cell


def build(out: Path = OUT) -> Path:
    target = Path(out)
    target.parent.mkdir(parents=True, exist_ok=True)
    notebook = {
        "cells": [
            _cell(kind, source, index)
            for index, (kind, source) in enumerate(CELLS)
        ],
        "metadata": {
            "kernelspec": {
                "display_name": "Python 3",
                "language": "python",
                "name": "python3",
            },
            "language_info": {"name": "python"},
        },
        "nbformat": 4,
        "nbformat_minor": 5,
    }
    target.write_text(
        json.dumps(notebook, indent=1, ensure_ascii=False),
        encoding="utf-8",
    )
    print(f"Wrote {target} ({len(CELLS)} cells)")
    return target


if __name__ == "__main__":
    build()
