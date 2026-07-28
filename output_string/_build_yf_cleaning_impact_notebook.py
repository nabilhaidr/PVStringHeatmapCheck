"""Builder Specific_Yield_Cleaning_Impact.ipynb.

Notebook mengukur dampak cleaning (pre/post) di atas specific yield yang
ternormalisasi string kontrol -- tanpa data POA. Sumber: workbook specific
yield + rekap cleaning + DC cable list, semuanya dari Google Drive.

Edit builder ini lalu jalankan:
    python output_string/_build_yf_cleaning_impact_notebook.py
"""
from __future__ import annotations

import json
from pathlib import Path


OUT = Path(__file__).parent / "Specific_Yield_Cleaning_Impact.ipynb"

MD_INTRO = '''# Dampak Pre/Post Cleaning pada Yf -- tanpa data POA

Notebook ini menjawab **"apakah cleaning benar-benar memulihkan performa,
dan berapa besar?"** dengan membandingkan specific yield (Yf) tiap string
sebelum vs sesudah tanggal cleaning-nya sendiri.

Kenapa tanpa POA bisa: Yf tiap string dibagi **median harian STRING
KONTROL** -- yaitu string di WB yang sama yang **TIDAK dibersihkan** pada
jendela campaign. Cuaca dan drift musiman tercoret karena kontrol mengalami
cuaca yang sama.

Kontrol ini penting: kalau referensinya ikut dibersihkan, referensi ikut
naik dan uplift akan tampak nol. Kolom `reference_mode` menyatakan referensi
mana yang dipakai:
- `WB_UNCLEANED` -- kontrol di WB yang sama (terbaik).
- `SITE_UNCLEANED` -- kontrol dari WB lain (WB itu dibersihkan menyeluruh).
- `RAW_NO_REFERENCE` -- tidak ada kontrol; angka masih terkontaminasi cuaca,
  perlakukan sebagai indikatif saja.

Sumber data (semua dari Google Drive):
1. `specific_yield_*.xlsx` -- folder **output_string/Rekap String Specific Yield**.
2. `Report & Schedule Cleaning PLTS IKN.xlsx` -- folder **raw data input**.
   Sheet `STS 1`..`STS 10` = WB-01..WB-10; checklist `TRUE` per tanggal.
3. `List of DC Cables 0411.xls` -- folder **raw data input**. Kolom G =
   nomor String (dipakai di rekap cleaning), kolom H = nomor PV sisi
   inverter (dipakai di data Huawei). Mapping ini wajib untuk WB03-WB10;
   WB01-WB02 nomor String = nomor PV.

Output workbook tiga sheet di `output_string/`:
- `Dampak_Cleaning` -- satu baris per string per campaign, uplift terbesar dulu.
- `Rekap_Campaign` -- ringkasan per campaign.
- `Metadata` -- rumus, ambang, dan diagnostik.

Jalankan Cell 1 sampai Cell 7 berurutan. Edit hanya nilai di Cell 2.
'''

CODE_SETUP = '''# Cell 1 - Setup Colab/repo/output
import subprocess, sys
subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "--upgrade", "gdown>=6.0.0"])
from pathlib import Path
import os, tempfile

def find_repo_root(start=None):
    path = Path(start or os.getcwd()).resolve()
    for candidate in [path, *path.parents]:
        if (candidate / "pv_pipeline").is_dir() and (candidate / "config").is_dir():
            return candidate
    raise RuntimeError("Repo root tidak ditemukan; jalankan notebook dari clone SolarYieldPro.")

PUBLIC_REPO_URL = "https://github.com/nabilhaidr/PVStringHeatmapCheck.git"
DEFAULT_REPO_DIR = Path.cwd() / "PVStringHeatmapCheck"
try:
    REPO_DIR = find_repo_root()
except RuntimeError:
    if not DEFAULT_REPO_DIR.exists():
        subprocess.check_call([
            "git", "clone", "--depth", "1", PUBLIC_REPO_URL,
            str(DEFAULT_REPO_DIR),
        ])
    REPO_DIR = find_repo_root(DEFAULT_REPO_DIR)
if str(REPO_DIR) not in sys.path:
    sys.path.insert(0, str(REPO_DIR))
OUTPUT_DIR = REPO_DIR / "output_string"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
INPUT_DIR = Path(tempfile.mkdtemp(prefix="cleaning_impact_inputs_"))
print("OUTPUT_DIR:", OUTPUT_DIR)
'''

CODE_CONFIG = '''# Cell 2 - Konfigurasi input (edit nilai di sini)
from pv_pipeline.yf_ratio_report import validate_drive_folder_url

# Folder Drive: Cek PV String/output_string/Rekap String Specific Yield
URL_SPECIFIC_YIELD = "https://drive.google.com/drive/folders/16FRx87xG5G9tcyS3cfViQw1wYMgMEMbF"
# Folder Drive: Cek PV String/raw data input
URL_RAW_INPUT = "https://drive.google.com/drive/folders/1y37AsVViI7IL1tCRhvAGEjiq9wWnfu_e"

# Kosongkan untuk memakai workbook specific_yield_*.xlsx TERBARU di folder.
SPECIFIC_YIELD_FILENAME = ""
CLEANING_REPORT_FILENAME = "Report & Schedule Cleaning PLTS IKN.xlsx"
DC_CABLE_FILENAME = "List of DC Cables 0411.xls"

# Ambang analisis.
WINDOW_DAYS = 7              # rata-rata Yf N hari sebelum/sesudah campaign
GAP_DAYS = 7                 # jarak antar tanggal cleaning -> campaign beda
MIN_WINDOW_DAYS = 2          # minimal hari valid di tiap sisi
MIN_REFERENCE_STRINGS = 5    # minimal string kontrol (tidak dibersihkan)
MIN_RELATIVE_COVERAGE = 0.8  # buang string-day yg cakupannya < 0.8x tetangga

for _url in (URL_SPECIFIC_YIELD, URL_RAW_INPUT):
    validate_drive_folder_url(_url)
print("Specific yield:", URL_SPECIFIC_YIELD)
print("Raw data input:", URL_RAW_INPUT)
'''

CODE_DOWNLOAD = '''# Cell 3 - Unduh ketiga file sumber dari Drive
from pv_pipeline.yf_ratio_report import download_drive_file
if any(name not in globals() for name in ("URL_SPECIFIC_YIELD", "URL_RAW_INPUT", "INPUT_DIR")):
    raise RuntimeError("Jalankan Cell 2 terlebih dahulu.")

SPECIFIC_YIELD_XLSX = download_drive_file(
    URL_SPECIFIC_YIELD,
    INPUT_DIR,
    basename=SPECIFIC_YIELD_FILENAME,
    pattern="" if SPECIFIC_YIELD_FILENAME else r"^specific_yield_.*\\.xlsx$",
)
CLEANING_REPORT_XLSX = download_drive_file(
    URL_RAW_INPUT, INPUT_DIR, basename=CLEANING_REPORT_FILENAME,
)
DC_CABLE_XLS = download_drive_file(
    URL_RAW_INPUT, INPUT_DIR, basename=DC_CABLE_FILENAME,
)
print("Specific yield :", SPECIFIC_YIELD_XLSX.name)
print("Cleaning report:", CLEANING_REPORT_XLSX.name)
print("DC cable list  :", DC_CABLE_XLS.name)
'''

CODE_EVENTS = '''# Cell 4 - Baca rekap cleaning + mapping ST->PV
from pv_pipeline.m2a.cleaning_report import (
    build_st_to_pv, load_cleaning_report, load_dc_cable_map,
)
if "CLEANING_REPORT_XLSX" not in globals():
    raise RuntimeError("Jalankan Cell 3 terlebih dahulu.")

CABLE_MAP = load_dc_cable_map(str(DC_CABLE_XLS))
ST_TO_PV = build_st_to_pv(CABLE_MAP)
print(f"Mapping ST->PV: {len(ST_TO_PV):,} string (WB03-WB10); "
      "WB01-WB02 memakai identitas nomor String = nomor PV.")

EVENTS = load_cleaning_report(str(CLEANING_REPORT_XLSX), ST_TO_PV)
_mapped = int(EVENTS["pv"].notna().sum())
print(f"Event cleaning: {len(EVENTS):,} baris; termapping ke nomor PV: "
      f"{_mapped:,} ({_mapped / max(len(EVENTS), 1) * 100:.1f}%)")
if len(EVENTS):
    print("Rentang tanggal cleaning:",
          EVENTS["date"].min().date(), "s.d.", EVENTS["date"].max().date())
'''

CODE_BUILD = '''# Cell 5 - Hitung uplift pre/post pada Yf ternormalisasi kontrol
try:
    from IPython.display import display
except ImportError:
    display = print
from pv_pipeline.yf_ratio_report import (
    build_yf_cleaning_impact, load_specific_yield_long,
)
if any(name not in globals() for name in ("SPECIFIC_YIELD_XLSX", "EVENTS")):
    raise RuntimeError("Jalankan Cell 4 terlebih dahulu.")

LONG_DF = load_specific_yield_long(SPECIFIC_YIELD_XLSX)
DATES = sorted(LONG_DF["date"].dt.date.unique())
print(f"Yield: {len(LONG_DF):,} baris | {LONG_DF['pv_string'].nunique():,} string | "
      f"{DATES[0]} s.d. {DATES[-1]}")

REPORT = build_yf_cleaning_impact(
    LONG_DF,
    EVENTS,
    window_days=WINDOW_DAYS,
    gap_days=GAP_DAYS,
    min_window_days=MIN_WINDOW_DAYS,
    min_reference_strings=MIN_REFERENCE_STRINGS,
    min_relative_coverage=MIN_RELATIVE_COVERAGE,
    source_metadata={
        "source_workbook": SPECIFIC_YIELD_XLSX.name,
        "cleaning_report": CLEANING_REPORT_XLSX.name,
        "dc_cable_list": DC_CABLE_XLS.name,
    },
)
print("Status:", REPORT.metadata["status_counts"])
print("Referensi:", REPORT.impact["reference_mode"].value_counts().to_dict())
print("\\nRekap per campaign:")
display(REPORT.campaigns.round(3))
print("\\n20 string dengan uplift terbesar (paling kotor sebelum dibersihkan):")
display(
    REPORT.impact.head(20)[[
        "rank_uplift", "pv_string", "st", "cleaning_start", "cleaning_end",
        "rel_before", "rel_after", "uplift_pct", "soiling_loss_pct",
        "reference_mode",
    ]].round(3)
)
'''

CODE_EXPORT = '''# Cell 6 - Ekspor dan verifikasi workbook
from openpyxl import load_workbook
from pv_pipeline.yf_ratio_report import (
    build_cleaning_impact_output_path,
    verify_cleaning_impact_workbook,
    write_cleaning_impact_workbook,
)
OUTPUT_VERIFIED = False
if any(name not in globals() for name in ("REPORT", "DATES", "OUTPUT_DIR")):
    raise RuntimeError("Jalankan Cell 5 terlebih dahulu.")
OUTPUT_XLSX = build_cleaning_impact_output_path(OUTPUT_DIR, DATES[0], DATES[-1])
write_cleaning_impact_workbook(OUTPUT_XLSX, REPORT)
verify_cleaning_impact_workbook(OUTPUT_XLSX)
CHECK_WB = load_workbook(OUTPUT_XLSX, read_only=False, data_only=False)
print(
    "Workbook:", OUTPUT_XLSX,
    "sheets:", CHECK_WB.sheetnames,
    "bytes:", OUTPUT_XLSX.stat().st_size,
    "baris:", REPORT.metadata["evaluated_rows"],
    "campaign:", REPORT.metadata["campaign_count"],
)
CHECK_WB.close()
OUTPUT_VERIFIED = True
'''

CODE_DOWNLOAD_XLSX = '''# Cell 7 - Download dari Colab; lokal hanya menampilkan path
if (
    not globals().get("OUTPUT_VERIFIED", False)
    or "OUTPUT_XLSX" not in globals()
    or not OUTPUT_XLSX.exists()
):
    raise RuntimeError("Jalankan Cell 6 terlebih dahulu.")
try:
    from google.colab import files
except ImportError:
    print("Bukan Colab; workbook tersedia di", OUTPUT_XLSX)
else:
    files.download(str(OUTPUT_XLSX))
'''

CELLS = [
    ("markdown", MD_INTRO),
    ("code", CODE_SETUP),
    ("code", CODE_CONFIG),
    ("code", CODE_DOWNLOAD),
    ("code", CODE_EVENTS),
    ("code", CODE_BUILD),
    ("code", CODE_EXPORT),
    ("code", CODE_DOWNLOAD_XLSX),
]


def _cell(kind: str, source: str, index: int) -> dict:
    cell = {
        "cell_type": kind,
        "id": f"cleaning-impact-{index:02d}",
        "metadata": {},
        "source": source.splitlines(keepends=True),
    }
    if kind == "code":
        cell["execution_count"] = None
        cell["outputs"] = []
    return cell


def build(out: Path = OUT) -> Path:
    target = Path(out)
    target.parent.mkdir(parents=True, exist_ok=True)
    notebook = {
        "cells": [
            _cell(kind, source, index)
            for index, (kind, source) in enumerate(CELLS)
        ],
        "metadata": {
            "kernelspec": {
                "display_name": "Python 3",
                "language": "python",
                "name": "python3",
            },
            "language_info": {"name": "python"},
        },
        "nbformat": 4,
        "nbformat_minor": 5,
    }
    target.write_text(
        json.dumps(notebook, indent=1, ensure_ascii=False),
        encoding="utf-8",
    )
    print(f"Wrote {target} ({len(CELLS)} cells)")
    return target


if __name__ == "__main__":
    build()
