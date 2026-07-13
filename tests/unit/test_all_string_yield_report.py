from __future__ import annotations

import ast
from datetime import date
import importlib
import importlib.util
import json
from pathlib import Path
import subprocess
import sys
import tempfile

import numpy as np
from openpyxl import load_workbook
import pandas as pd
import pytest

from pv_pipeline.string_yield_report import (
    DownloadedInputs,
    DriveItem,
    SourceManifest,
)


CSV_URL = (
    "https://drive.google.com/drive/folders/"
    "1f_KrPuqfZJTE5I9cVQiyp65QrHbkF3Iw?usp=sharing"
)


def _write_csv(path, frame):
    frame.to_csv(path, index=False)
    return path


def test_download_csv_inputs_inventories_one_folder_and_selects_dates(
    monkeypatch,
    tmp_path,
):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    calls = []
    items = [
        DriveItem(
            "https://drive.google.com/uc?id=a",
            "nested/20260501.csv",
        ),
        DriveItem("https://drive.google.com/uc?id=b", "20260502.csv"),
        DriveItem("https://drive.google.com/uc?id=c", "20260503.csv"),
    ]

    def fake_inventory(url):
        calls.append(("inventory", url))
        return items

    def fake_download(manifest, destination):
        calls.append(("download", manifest, Path(destination)))
        return DownloadedInputs({}, {}, {})

    monkeypatch.setattr(
        report_module,
        "inventory_drive_folder",
        fake_inventory,
    )
    monkeypatch.setattr(
        report_module,
        "download_manifest",
        fake_download,
    )
    dates = pd.date_range("2026-05-01", "2026-05-02", freq="D")

    manifest, inputs = report_module.download_csv_inputs(
        CSV_URL,
        dates,
        tmp_path,
    )

    assert list(manifest.csv_by_date) == [
        date(2026, 5, 1),
        date(2026, 5, 2),
    ]
    assert manifest.poa_by_year == {}
    assert manifest.missing_poa_years == []
    assert manifest.url_poa == ""
    assert manifest.csv_inventory_count == 3
    assert inputs == DownloadedInputs({}, {}, {})
    assert [entry[0] for entry in calls] == ["inventory", "download"]


def test_daily_yield_builds_union_once_and_leaves_missing_values_blank(
    tmp_path,
):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    stamps = pd.date_range(
        "2026-05-01 00:00",
        periods=12,
        freq="5min",
    )
    day_one = pd.concat([
        pd.DataFrame({
            "Start Time": stamps,
            "Inverter_ID": "WB02-INV10",
            "PV1 Power(kW)": 20.0,
            "PV2 Voltage(V)": np.nan,
            "PV2 Current(A)": np.nan,
        }),
        pd.DataFrame({
            "Start Time": stamps,
            "Inverter_ID": "WB02-INV02",
            "PV1 Power(kW)": 10.0,
            "PV2 Voltage(V)": 500.0,
            "PV2 Current(A)": 10.0,
        }),
    ], ignore_index=True)
    day_two = day_one.loc[
        day_one["Inverter_ID"] == "WB02-INV02"
    ].copy()
    day_two["Start Time"] = pd.date_range(
        "2026-05-02 00:00",
        periods=12,
        freq="5min",
    )
    paths = {
        date(2026, 5, 1): _write_csv(
            tmp_path / "20260501.csv",
            day_one,
        ),
        date(2026, 5, 2): _write_csv(
            tmp_path / "20260502.csv",
            day_two,
        ),
    }
    dates = pd.date_range("2026-05-01", "2026-05-03", freq="D")

    result = report_module.build_all_string_daily_yield(paths, dates)

    expected_strings = [
        "WB02-INV02-PV1",
        "WB02-INV02-PV2",
        "WB02-INV10-PV1",
    ]
    assert list(result.summary.columns) == ["date", *expected_strings]
    assert len(result.daily) == 9
    indexed = result.daily.set_index(["date", "pv_string"])
    assert indexed.loc[
        (date(2026, 5, 1), "WB02-INV02-PV1"),
        "string_yield_kwh",
    ] == pytest.approx(10.0)
    assert indexed.loc[
        (date(2026, 5, 1), "WB02-INV02-PV2"),
        "string_yield_kwh",
    ] == pytest.approx(5.0)
    assert indexed.loc[
        (date(2026, 5, 2), "WB02-INV10-PV1"),
        "status",
    ] == "NO_STRING_DATA"
    assert indexed.loc[
        (date(2026, 5, 3), "WB02-INV02-PV1"),
        "status",
    ] == "MISSING_CSV"
    assert pd.isna(result.summary.loc[2, "WB02-INV02-PV1"])


def test_direct_power_does_not_fall_back_when_values_are_empty(tmp_path):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    day_one = pd.DataFrame({
        "Start Time": ["2026-05-01 00:00"],
        "Inverter_ID": ["WB01-INV01"],
        "PV1 Power(kW)": [4.0],
        "PV1 Voltage(V)": [500.0],
        "PV1 Current(A)": [10.0],
    })
    day_two = day_one.copy()
    day_two["Start Time"] = "2026-05-02 00:00"
    day_two["PV1 Power(kW)"] = np.nan
    paths = {
        date(2026, 5, 1): _write_csv(
            tmp_path / "20260501.csv",
            day_one,
        ),
        date(2026, 5, 2): _write_csv(
            tmp_path / "20260502.csv",
            day_two,
        ),
    }

    result = report_module.build_all_string_daily_yield(
        paths,
        pd.date_range("2026-05-01", "2026-05-02", freq="D"),
    )

    second = result.daily.loc[
        result.daily["date"] == date(2026, 5, 2)
    ].iloc[0]
    assert second["status"] == "NO_STRING_DATA"
    assert pd.isna(second["string_yield_kwh"])


def test_read_error_is_distinct_and_no_usable_range_fails_loudly(
    tmp_path,
):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    valid = _write_csv(
        tmp_path / "20260501.csv",
        pd.DataFrame({
            "Start Time": ["2026-05-01 00:00"],
            "Inverter_ID": ["WB01-INV01"],
            "PV1 Power(kW)": [6.0],
        }),
    )
    broken = _write_csv(
        tmp_path / "20260502.csv",
        pd.DataFrame({"bad": [1]}),
    )
    dates = pd.date_range("2026-05-01", "2026-05-02", freq="D")

    result = report_module.build_all_string_daily_yield(
        {
            date(2026, 5, 1): valid,
            date(2026, 5, 2): broken,
        },
        dates,
    )

    assert result.daily.loc[
        result.daily["date"] == date(2026, 5, 2),
        "status",
    ].tolist() == ["CSV_READ_ERROR"]
    assert "20260502.csv" in result.metadata["csv_read_errors"]
    assert result.metadata["missing_csv_dates"] == []
    with pytest.raises(
        RuntimeError,
        match="No requested CSV could be read",
    ):
        report_module.build_all_string_daily_yield(
            {date(2026, 5, 2): broken},
            dates[1:],
        )


def test_quality_diagnostics_drop_off_grid_and_keep_negative_power(
    tmp_path,
):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    frame = pd.DataFrame({
        "Start Time": [
            "2026-05-01 00:00:00",
            "2026-05-01 00:00:00",
            "2026-05-01 00:05:00",
            "2026-05-01 00:10:30",
            "2026-05-02 00:10:00",
        ],
        "Inverter_ID": "WB01-INV01",
        "PV1 Power(kW)": [5.0, 7.0, -3.0, 10.0, 99.0],
    })
    path = _write_csv(tmp_path / "20260501.csv", frame)

    result = report_module.build_all_string_daily_yield(
        {date(2026, 5, 1): path},
        pd.date_range("2026-05-01", periods=1),
    )

    row = result.daily.iloc[0]
    assert row["valid_power_samples"] == 2
    assert row["string_yield_kwh"] == pytest.approx(0.25)
    assert result.metadata["duplicate_rows"] == 1
    assert result.metadata["negative_power_samples"] == 1
    assert result.metadata["wrong_date_rows"]["20260501.csv"] == 1
    assert "Negative power samples were retained." in result.metadata[
        "warnings"
    ]
    assert (
        "Rows outside their YYYYMMDD.csv date were dropped."
        in result.metadata["warnings"]
    )


def test_complete_day_uses_manage_object_fallback(tmp_path):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    path = _write_csv(
        tmp_path / "20260501.csv",
        pd.DataFrame({
            "Start Time": pd.date_range(
                "2026-05-01 00:00",
                periods=288,
                freq="5min",
            ),
            "ManageObject": "WB01-INV01",
            "PV1 Power(kW)": 1.0,
        }),
    )

    result = report_module.build_all_string_daily_yield(
        {date(2026, 5, 1): path},
        pd.date_range("2026-05-01", periods=1),
    )

    row = result.daily.iloc[0]
    assert row["pv_string"] == "WB01-INV01-PV1"
    assert row["valid_power_samples"] == 288
    assert row["status"] == "COMPLETE"


def test_manage_object_fallback_normalizes_mixed_case(tmp_path):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    path = _write_csv(
        tmp_path / "20260501.csv",
        pd.DataFrame({
            "Start Time": ["2026-05-01 00:00"],
            "manageobject": ["root/inv_a_234_ikn"],
            "PV3 Power(kW)": [4.0],
        }),
    )

    result = report_module.build_all_string_daily_yield(
        {date(2026, 5, 1): path},
        pd.date_range("2026-05-01", periods=1),
    )

    assert result.daily["pv_string"].tolist() == ["WB02-INV34-PV3"]


def test_metadata_preserves_inventory_and_download_diagnostics(tmp_path):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    path = _write_csv(
        tmp_path / "20260501.csv",
        pd.DataFrame({
            "Start Time": ["2026-05-01 00:00"],
            "Inverter_ID": ["WB01-INV01"],
            "PV1 Power(kW)": [1.0],
        }),
    )
    missing_day = date(2026, 5, 2)
    manifest = SourceManifest(
        csv_by_date={
            date(2026, 5, 1): DriveItem("local", "20260501.csv"),
        },
        poa_by_year={},
        missing_csv_dates=[missing_day],
        missing_poa_years=[],
        url_csv="https://drive.google.com/drive/folders/csv-source",
        csv_inventory_count=472,
    )

    result = report_module.build_all_string_daily_yield(
        {date(2026, 5, 1): path},
        pd.date_range("2026-05-01", "2026-05-02", freq="D"),
        source_manifest=manifest,
        download_errors={"20260502.csv": "RuntimeError: failed"},
    )

    assert result.metadata["csv_inventory_count"] == 472
    assert result.metadata["inventory_missing_csv_dates"] == [
        "2026-05-02"
    ]
    assert result.metadata["download_errors"] == {
        "20260502.csv": "RuntimeError: failed"
    }


def test_no_valid_string_sample_fails_loudly(tmp_path):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    path = _write_csv(
        tmp_path / "20260501.csv",
        pd.DataFrame({
            "Start Time": ["2026-05-01 00:00"],
            "Inverter_ID": ["WB01-INV01"],
            "PV1 Power(kW)": [np.nan],
        }),
    )

    with pytest.raises(
        RuntimeError,
        match="No valid PV string power sample",
    ):
        report_module.build_all_string_daily_yield(
            {date(2026, 5, 1): path},
            pd.date_range("2026-05-01", periods=1),
        )


def test_csv_reader_disables_chunked_type_inference(tmp_path, monkeypatch):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    path = _write_csv(
        tmp_path / "20260501.csv",
        pd.DataFrame({
            "Start Time": ["2026-05-01 00:00"],
            "Inverter_ID": ["WB01-INV01"],
            "PV1 Power(kW)": [1.0],
        }),
    )
    original_read_csv = pd.read_csv
    calls = []

    def tracked_read_csv(*args, **kwargs):
        calls.append(kwargs.copy())
        return original_read_csv(*args, **kwargs)

    monkeypatch.setattr(report_module.pd, "read_csv", tracked_read_csv)

    report_module.build_all_string_daily_yield(
        {date(2026, 5, 1): path},
        pd.date_range("2026-05-01", periods=1),
    )

    assert calls == [{"low_memory": False}]


def test_build_does_not_concat_power_across_days(monkeypatch):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    original_concat = pd.concat

    monkeypatch.setattr(
        report_module.pd,
        "read_csv",
        lambda path, **kwargs: pd.DataFrame({"path": [str(path)]}),
    )

    def fake_extract(frame, requested_day):
        return pd.DataFrame({
            "timestamp": [pd.Timestamp(requested_day)],
            "inverter_id": ["WB01-INV01"],
            "pv_string": ["WB01-INV01-PV1"],
            "pv_label": ["PV1"],
            "power_kw": [1.0],
        }), {
            "wrong_date_rows": 0,
            "duplicate_rows": 0,
            "negative_samples": 0,
            "non_finite_samples": 0,
            "power_sources": ["PV1:direct"],
        }

    def reject_cross_day_concat(objects, *args, **kwargs):
        frames = list(objects)
        if len(frames) > 1 and all(
            "pv_string" in frame.columns
            for frame in frames
        ):
            raise AssertionError("cross-day power concat is not allowed")
        return original_concat(frames, *args, **kwargs)

    monkeypatch.setattr(
        report_module,
        "_extract_all_string_power",
        fake_extract,
    )
    monkeypatch.setattr(
        report_module.pd,
        "concat",
        reject_cross_day_concat,
    )

    result = report_module.build_all_string_daily_yield(
        {
            date(2026, 5, 1): Path("20260501.csv"),
            date(2026, 5, 2): Path("20260502.csv"),
        },
        pd.date_range("2026-05-01", "2026-05-02", freq="D"),
    )

    assert result.summary["WB01-INV01-PV1"].tolist() == [
        pytest.approx(5 / 60),
        pytest.approx(5 / 60),
    ]


def test_detail_row_limit_fails_before_materializing_product(
    tmp_path,
    monkeypatch,
):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    path = _write_csv(
        tmp_path / "20260501.csv",
        pd.DataFrame({
            "Start Time": ["2026-05-01 00:00"] * 2,
            "Inverter_ID": ["WB01-INV01", "WB01-INV02"],
            "PV1 Power(kW)": [1.0, 1.0],
        }),
    )
    monkeypatch.setattr(
        report_module,
        "EXCEL_MAX_ROWS",
        2,
        raising=False,
    )

    with pytest.raises(ValueError, match="Excel row limit"):
        report_module.build_all_string_daily_yield(
            {date(2026, 5, 1): path},
            pd.date_range("2026-05-01", periods=1),
        )


def test_non_finite_direct_and_derived_power_are_dropped(tmp_path):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    first = _write_csv(
        tmp_path / "20260501.csv",
        pd.DataFrame({
            "Start Time": ["2026-05-01 00:00"],
            "Inverter_ID": ["WB01-INV01"],
            "PV1 Power(kW)": [1.0],
            "PV2 Voltage(V)": [500.0],
            "PV2 Current(A)": [10.0],
        }),
    )
    second = _write_csv(
        tmp_path / "20260502.csv",
        pd.DataFrame({
            "Start Time": ["2026-05-02 00:00"],
            "Inverter_ID": ["WB01-INV01"],
            "PV1 Power(kW)": [np.inf],
            "PV2 Voltage(V)": [1e308],
            "PV2 Current(A)": [1e308],
        }),
    )

    result = report_module.build_all_string_daily_yield(
        {
            date(2026, 5, 1): first,
            date(2026, 5, 2): second,
        },
        pd.date_range("2026-05-01", "2026-05-02", freq="D"),
    )

    second_day = result.daily.loc[
        result.daily["date"] == date(2026, 5, 2)
    ]
    assert second_day["status"].tolist() == [
        "NO_STRING_DATA",
        "NO_STRING_DATA",
    ]
    assert second_day["string_yield_kwh"].isna().all()
    assert result.metadata["non_finite_power_samples"] == 2
    assert "Non-finite power samples were dropped." in result.metadata[
        "warnings"
    ]


@pytest.fixture
def all_string_result(tmp_path):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    csv_path = _write_csv(
        tmp_path / "20260501.csv",
        pd.DataFrame({
            "Start Time": ["2026-05-01 00:00"],
            "Inverter_ID": ["WB01-INV01"],
            "PV1 Power(kW)": [0.0],
        }),
    )
    return report_module.build_all_string_daily_yield(
        {date(2026, 5, 1): csv_path},
        pd.date_range("2026-05-01", "2026-05-03", freq="D"),
    )


def test_all_string_workbook_has_wide_detail_and_metadata_contract(
    tmp_path,
    all_string_result,
):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    output = report_module.build_all_string_output_path(
        tmp_path,
        date(2026, 5, 1),
        date(2026, 5, 3),
    )
    assert output.name == "all_string_yield_20260501_20260503.xlsx"

    report_module.write_all_string_workbook(output, all_string_result)
    report_module.verify_all_string_workbook(output)

    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == [
        "Rekap_Yield_kWh",
        "Detail_Harian",
        "Metadata",
    ]
    recap = workbook["Rekap_Yield_kWh"]
    assert [cell.value for cell in recap[1]] == list(
        all_string_result.summary.columns
    )
    assert recap.freeze_panes == "B2"
    assert recap["B2"].number_format == "0.000"
    assert recap["B2"].value == 0
    assert recap["B3"].value is None
    detail = workbook["Detail_Harian"]
    assert [cell.value for cell in detail[1]] == (
        report_module.DETAIL_COLUMNS
    )
    assert detail.freeze_panes == "A2"
    assert detail.max_row == len(all_string_result.daily) + 1
    metadata = workbook["Metadata"]
    assert metadata["A1"].value == "key"
    assert metadata["B1"].value == "value"
    assert metadata.freeze_panes == "A2"
    workbook.close()


def test_workbook_verifier_rejects_noncanonical_recap_header(
    tmp_path,
    all_string_result,
):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    output = tmp_path / "tampered.xlsx"
    report_module.write_all_string_workbook(output, all_string_result)
    workbook = load_workbook(output)
    workbook["Rekap_Yield_kWh"]["B1"] = "BROKEN"
    workbook.save(output)
    workbook.close()

    with pytest.raises(RuntimeError, match="recap PV string headers"):
        report_module.verify_all_string_workbook(output)


def test_workbook_verifier_rejects_duplicate_recap_date(
    tmp_path,
    all_string_result,
):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    output = tmp_path / "duplicate-recap-date.xlsx"
    report_module.write_all_string_workbook(output, all_string_result)
    workbook = load_workbook(output)
    workbook["Rekap_Yield_kWh"]["A3"] = workbook[
        "Rekap_Yield_kWh"
    ]["A2"].value
    workbook.save(output)
    workbook.close()

    with pytest.raises(RuntimeError, match="recap dates"):
        report_module.verify_all_string_workbook(output)


def test_workbook_verifier_rejects_duplicate_detail_combination(
    tmp_path,
    all_string_result,
):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    output = tmp_path / "duplicate-detail-row.xlsx"
    report_module.write_all_string_workbook(output, all_string_result)
    workbook = load_workbook(output)
    detail = workbook["Detail_Harian"]
    detail["A3"] = detail["A2"].value
    detail["B3"] = detail["B2"].value
    workbook.save(output)
    workbook.close()

    with pytest.raises(RuntimeError, match="detail combinations"):
        report_module.verify_all_string_workbook(output)


def test_writer_does_not_publish_unverified_workbook(
    tmp_path,
    all_string_result,
    monkeypatch,
):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    output = tmp_path / "unverified.xlsx"

    def reject_verification(path):
        raise RuntimeError("forced verification failure")

    monkeypatch.setattr(
        report_module,
        "verify_all_string_workbook",
        reject_verification,
    )

    with pytest.raises(RuntimeError, match="forced verification failure"):
        report_module.write_all_string_workbook(output, all_string_result)

    assert not output.exists()
    assert list(tmp_path.glob("*.tmp.xlsx")) == []


@pytest.mark.parametrize(
    "sensitive_key",
    ["api_token", "session-cookie", "clientSecret", "credential_id", "db password"],
)
def test_workbook_rejects_sensitive_metadata_before_saving(
    tmp_path,
    all_string_result,
    sensitive_key,
):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    all_string_result.metadata[sensitive_key] = "must-not-be-written"
    output = tmp_path / "sensitive.xlsx"

    with pytest.raises(ValueError, match="Sensitive metadata key"):
        report_module.write_all_string_workbook(output, all_string_result)

    assert not output.exists()


def test_workbook_canonicalizes_source_url(tmp_path, all_string_result):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    all_string_result.metadata["source_url_csv"] = CSV_URL
    output = tmp_path / "canonical.xlsx"

    report_module.write_all_string_workbook(output, all_string_result)

    workbook = load_workbook(output, data_only=False)
    metadata = {
        workbook["Metadata"].cell(row=row, column=1).value:
        workbook["Metadata"].cell(row=row, column=2).value
        for row in range(2, workbook["Metadata"].max_row + 1)
    }
    workbook.close()
    assert metadata["source_url_csv"] == (
        "https://drive.google.com/drive/folders/"
        "1f_KrPuqfZJTE5I9cVQiyp65QrHbkF3Iw"
    )


def test_workbook_chunks_long_metadata_without_truncation(
    tmp_path,
    all_string_result,
):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    diagnostic = {"20260501.csv": "x" * 40_000}
    all_string_result.metadata["csv_read_errors"] = diagnostic
    output = tmp_path / "chunked-metadata.xlsx"

    report_module.write_all_string_workbook(output, all_string_result)

    workbook = load_workbook(output, data_only=False)
    rows = [
        (
            workbook["Metadata"].cell(row=row, column=1).value,
            workbook["Metadata"].cell(row=row, column=2).value,
        )
        for row in range(2, workbook["Metadata"].max_row + 1)
    ]
    workbook.close()
    chunks = [
        value
        for key, value in rows
        if str(key).startswith("csv_read_errors[")
    ]
    assert len(chunks) == 2
    assert all(len(chunk) <= 30_000 for chunk in chunks)
    assert "".join(chunks) == json.dumps(
        diagnostic,
        ensure_ascii=False,
        default=str,
    )


def test_workbook_redacts_drive_query_from_diagnostics(
    tmp_path,
    all_string_result,
):
    report_module = importlib.import_module(
        "pv_pipeline.all_string_yield_report"
    )
    all_string_result.metadata["download_errors"] = {
        "20260502.csv": (
            "failed https://drive.google.com/uc?"
            "id=secret-file-id&resourcekey=secret-resource-key"
        )
    }
    output = tmp_path / "redacted-diagnostic.xlsx"

    report_module.write_all_string_workbook(output, all_string_result)

    workbook = load_workbook(output, data_only=False)
    persisted = "\n".join(
        str(workbook["Metadata"].cell(row=row, column=2).value)
        for row in range(2, workbook["Metadata"].max_row + 1)
    )
    workbook.close()
    assert "secret-file-id" not in persisted
    assert "secret-resource-key" not in persisted
    assert "drive.google.com/uc?" not in persisted


def test_builder_writes_nbformat_45_with_seven_expected_cells(tmp_path):
    path = Path("output_string/_build_all_string_yield_notebook.py")
    spec = importlib.util.spec_from_file_location(
        "all_string_yield_nb_builder",
        path,
    )
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    target = module.build(tmp_path / "all-string.ipynb")
    notebook = json.loads(target.read_text(encoding="utf-8"))
    committed = json.loads(
        Path("output_string/All_String_Daily_Yield.ipynb").read_text(
            encoding="utf-8"
        )
    )

    assert committed == notebook
    assert (notebook["nbformat"], notebook["nbformat_minor"]) == (4, 5)
    assert [cell["cell_type"] for cell in notebook["cells"]] == [
        "markdown",
        *("code" for _ in range(6)),
    ]
    markers = [
        "gdown>=6.0.0",
        "URL_CSV",
        "download_csv_inputs",
        "build_all_string_daily_yield",
        "write_all_string_workbook",
        "google.colab",
    ]
    for cell, marker in zip(notebook["cells"][1:], markers):
        source = "".join(cell["source"])
        assert marker in source
        ast.parse(source)

    config_source = "".join(notebook["cells"][2]["source"])
    tree = ast.parse(config_source)
    literal_assignments = []
    for node in ast.walk(tree):
        if not isinstance(node, (ast.Assign, ast.AnnAssign)):
            continue
        try:
            value = ast.literal_eval(node.value)
        except (TypeError, ValueError):
            continue
        targets = node.targets if isinstance(node, ast.Assign) else [node.target]
        assert len(targets) == 1
        assert isinstance(targets[0], ast.Name)
        literal_assignments.append((targets[0].id, value))
    assert literal_assignments == [
        ("URL_CSV", CSV_URL),
        ("START_DATE", "2026-05-01"),
        ("END_DATE", "2026-05-14"),
    ]
    all_source = "\n".join(
        "".join(cell["source"])
        for cell in notebook["cells"]
    )
    assert "PV_STRING" not in all_source
    assert "URL_RAW_DATA_INPUT" not in all_source
    assert "POA" not in all_source
    export_source = "".join(notebook["cells"][5]["source"])
    download_source = "".join(notebook["cells"][6]["source"])
    assert "OUTPUT_VERIFIED = False" in export_source
    assert "OUTPUT_VERIFIED = True" in export_source
    assert 'globals().get("OUTPUT_VERIFIED", False)' in download_source


def test_notebook_setup_cell_auto_clones_public_repo_offline(
    tmp_path,
    monkeypatch,
):
    notebook = json.loads(
        Path("output_string/All_String_Daily_Yield.ipynb").read_text(
            encoding="utf-8"
        )
    )
    source = "".join(notebook["cells"][1]["source"])
    clean_dir = tmp_path / "clean"
    clone_dir = clean_dir / "PVStringHeatmapCheck"
    input_dir = tmp_path / "inputs"
    clean_dir.mkdir()
    calls = []

    def fake_check_call(command):
        calls.append(command)
        if command[:4] == ["git", "clone", "--depth", "1"]:
            assert command[-1] == str(clone_dir)
            (clone_dir / "pv_pipeline").mkdir(parents=True)
            (clone_dir / "config").mkdir()

    def fake_mkdtemp(*, prefix):
        assert prefix == "all_string_yield_inputs_"
        input_dir.mkdir()
        return str(input_dir)

    monkeypatch.chdir(clean_dir)
    monkeypatch.setattr(sys, "path", list(sys.path))
    monkeypatch.setattr(subprocess, "check_call", fake_check_call)
    monkeypatch.setattr(tempfile, "mkdtemp", fake_mkdtemp)

    scope = {}
    exec(source, scope)

    assert calls == [
        [
            sys.executable,
            "-m",
            "pip",
            "install",
            "-q",
            "--upgrade",
            "gdown>=6.0.0",
        ],
        [
            "git",
            "clone",
            "--depth",
            "1",
            "https://github.com/nabilhaidr/PVStringHeatmapCheck.git",
            str(clone_dir),
        ],
    ]
    assert scope["REPO_DIR"] == clone_dir.resolve()
    assert scope["OUTPUT_DIR"] == clone_dir / "output_string"
    assert scope["INPUT_DIR"] == input_dir
