import ast
from datetime import date, datetime
import importlib.util
import json
from pathlib import Path
import sys
from types import SimpleNamespace
import xml.etree.ElementTree as ET
from zipfile import ZipFile

import matplotlib
import numpy as np
from openpyxl import load_workbook
import pandas as pd
import pytest

matplotlib.use("Agg")
from matplotlib import pyplot as plt

from pv_pipeline.string_yield_report import (
    DriveItem,
    ReportData,
    SourceManifest,
    _extract_string_power,
    build_output_path,
    build_report_data,
    download_manifest,
    inventory_drive_folder,
    parse_date_range,
    parse_inventory_json,
    parse_string_selection,
    plot_daily_yield,
    plot_power_vs_poa,
    select_source_manifest,
    validate_drive_folder_url,
    verify_report_workbook,
    write_report_workbook,
)


@pytest.fixture
def selection():
    return parse_string_selection("WB05-INV01-PV03")


@pytest.fixture
def report_fixture():
    grid = pd.date_range("2026-05-01", "2026-05-02 23:55", freq="5min")
    power = pd.Series(np.nan, index=grid, dtype="float64")
    power.iloc[pd.Index(range(13)).difference([1])] = 10.0
    poa = pd.Series(500.0, index=grid)
    return ReportData(
        daily=pd.DataFrame({
            "date": [date(2026, 5, 1), date(2026, 5, 2)],
            "string_yield_kwh": [10.0, np.nan],
            "valid_power_samples": [12, 0],
            "expected_samples": [288, 288],
            "coverage_pct": [12 / 288 * 100, 0.0],
            "missing_power_samples": [276, 288],
            "poa_valid_samples": [288, 288],
            "source_csv": ["20260501.csv", None],
            "status": ["PARTIAL", "MISSING_CSV"],
        }),
        five_minute=pd.DataFrame({
            "timestamp": grid,
            "inverter_id": "WB05-INV01",
            "pv_string": "PV3",
            "power_kw": power.to_numpy(),
            "poa_wm2": poa.to_numpy(),
            "source_csv": [
                "20260501.csv" if ts.date() == date(2026, 5, 1) else None
                for ts in grid
            ],
            "poa_source": "WS-2",
            "data_status": np.where(power.notna(), "POWER_POA", "POA_ONLY"),
        }),
        metadata={
            "source_url_csv": "https://drive.google.com/drive/folders/csv-test",
            "source_url_poa": "https://drive.google.com/drive/folders/poa-test",
            "missing_csv_dates": ["2026-05-02"],
            "poa_fallback_samples": 1,
            "warnings": [],
        },
    )


def _serialized_combo_axes(path):
    namespaces = {
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
    }
    with ZipFile(path) as archive:
        for name in archive.namelist():
            if not name.startswith("xl/charts/chart") or not name.endswith(".xml"):
                continue
            root = ET.fromstring(archive.read(name))
            title_text = "".join(
                node.text or "" for node in root.findall(".//a:t", namespaces)
            )
            if "Power String vs POA Irradiance" not in title_text:
                continue
            plot_area = root.find("c:chart/c:plotArea", namespaces)
            axes = []
            for tag in ("catAx", "dateAx", "valAx"):
                axes.extend((tag, node) for node in plot_area.findall(f"c:{tag}", namespaces))
            return axes, namespaces
    raise AssertionError("Serialized power/POA combo chart was not found.")


def test_parse_string_selection_normalizes_case_and_pv_number():
    got = parse_string_selection("wb05-inv01-pv03")
    assert (got.canonical, got.wb_id, got.inverter_id, got.pv_number, got.pv_label) == (
        "WB05-INV01-PV3",
        "WB05",
        "WB05-INV01",
        3,
        "PV3",
    )


@pytest.mark.parametrize("value", ["WB5-INV01-PV03", "WB05-INV01-PV0", "WB05-INV01-PV29"])
def test_parse_string_selection_rejects_invalid_format_and_range(value):
    with pytest.raises(ValueError):
        parse_string_selection(value)


def test_parse_date_range_is_inclusive_and_rejects_reverse():
    dates = parse_date_range("2026-12-31", "2027-01-02")
    assert [d.date() for d in dates] == [date(2026, 12, 31), date(2027, 1, 1), date(2027, 1, 2)]
    with pytest.raises(ValueError, match="START_DATE"):
        parse_date_range("2027-01-02", "2026-12-31")


def test_drive_url_and_inventory_json_contract():
    url = "https://drive.google.com/drive/folders/folder-id?usp=sharing"
    assert validate_drive_folder_url(url) == url
    items = parse_inventory_json('[{"url":"u1","path":"root/20260501.csv"}]')
    assert items == [DriveItem(url="u1", path="root/20260501.csv")]
    with pytest.raises(ValueError):
        validate_drive_folder_url("https://example.com/folder-id")
    with pytest.raises(ValueError):
        parse_inventory_json('[{"url":"u1"}]')


def test_drive_folder_validation_preserves_access_query_and_fragment():
    url = (
        "https://drive.google.com/drive/folders/folder-id"
        "?resourcekey=drive-access-key&usp=sharing#frag"
    )

    validated = validate_drive_folder_url(f"  {url}  ")

    assert validated == url


def test_inventory_subprocess_receives_resourcekey_access_url(monkeypatch):
    calls = []

    def fake_run(command, *, check, capture_output, text):
        calls.append((command, check, capture_output, text))
        return SimpleNamespace(stdout="[]")

    monkeypatch.setattr(
        "pv_pipeline.string_yield_report.subprocess.run", fake_run
    )
    url = (
        "https://drive.google.com/drive/folders/folder-id"
        "?resourcekey=drive-access-key"
    )

    assert inventory_drive_folder(url) == []
    command, check, capture_output, text = calls[0]
    assert command[3] == url
    assert (check, capture_output, text) == (True, True, True)


def test_select_inventory_matches_only_requested_csv_dates_and_poa_years():
    dates = parse_date_range("2026-12-31", "2027-01-01")
    csv_items = [
        DriveItem("csv-old", "root/20261230.csv"),
        DriveItem("csv-a", "root/20261231.csv"),
    ]
    poa_items = [
        DriveItem("poa-26", "root/POA PLTS IKN 2026.xlsx"),
        DriveItem("poa-27", "root/POA PLTS IKN 2027.xlsx"),
        DriveItem("other", "root/other.xlsx"),
    ]
    got = select_source_manifest(
        csv_items,
        poa_items,
        dates,
        url_csv=(
            "https://drive.google.com/drive/folders/csv-folder"
            "?resourcekey=csv-access#frag"
        ),
        url_poa=(
            "https://drive.google.com/drive/folders/poa-folder?usp=sharing"
        ),
    )
    assert got.csv_by_date == {date(2026, 12, 31): csv_items[1]}
    assert got.missing_csv_dates == [date(2027, 1, 1)]
    assert got.poa_by_year == {2026: poa_items[0], 2027: poa_items[1]}
    assert got.missing_poa_years == []
    assert got.url_csv == "https://drive.google.com/drive/folders/csv-folder"
    assert got.url_poa == "https://drive.google.com/drive/folders/poa-folder"


def test_select_inventory_rejects_duplicate_requested_basename():
    dates = parse_date_range("2026-05-01", "2026-05-01")
    csv_items = [
        DriveItem("csv-a", "root-a/20260501.csv"),
        DriveItem("csv-b", "root-b/20260501.csv"),
    ]

    with pytest.raises(ValueError, match="Duplicate requested basename"):
        select_source_manifest(csv_items, [], dates)


def test_download_manifest_downloads_only_selected_items(monkeypatch, tmp_path):
    calls = []

    def fake_download(*, url, output, quiet):
        calls.append((url, Path(output).name, quiet))
        Path(output).write_text("synthetic", encoding="utf-8")
        return output

    monkeypatch.setitem(sys.modules, "gdown", SimpleNamespace(download=fake_download))
    manifest = SourceManifest(
        csv_by_date={date(2026, 5, 1): DriveItem("csv-url", "root/20260501.csv")},
        poa_by_year={2026: DriveItem("poa-url", "root/POA PLTS IKN 2026.xlsx")},
        missing_csv_dates=[],
        missing_poa_years=[],
    )
    got = download_manifest(manifest, tmp_path)
    assert calls == [
        ("csv-url", "20260501.csv", False),
        ("poa-url", "POA PLTS IKN 2026.xlsx", False),
    ]
    assert got.download_errors == {}
    assert got.csv_by_date[date(2026, 5, 1)].is_file()
    assert got.poa_by_year[2026].is_file()


def _write_geometry(tmp_path, mapping="  WS-2: [WB05]"):
    path = tmp_path / "site_geometry.yaml"
    path.write_text(
        f"ws_to_wb:\n{mapping}\npyranometer:\n  sheet: POA PLTS IKN\n",
        encoding="utf-8",
    )
    return path


def _write_poa(path, timestamps, ws2, avg):
    frame = pd.DataFrame({
        "Date time": timestamps,
        **{
            f"POA Irradiance (W/m2) WS {number}":
                ws2 if number == 2 else np.full(len(timestamps), 100.0)
            for number in range(1, 6)
        },
        "Rata-rata WS 1 - WS 5": avg,
    })
    frame.to_excel(path, sheet_name="POA PLTS IKN", index=False)


def test_extract_power_prefers_explicit_power_column():
    df = pd.DataFrame({
        "Start Time": ["2026-05-01 00:00"],
        "Inverter_ID": ["wb05-inv01"],
        "PV3 Power(kW)": [7.5],
        "PV3 Input Voltage(V)": [1000],
        "PV3 Input Current(A)": [10],
    })

    series, diagnostics = _extract_string_power(
        df, parse_string_selection("WB05-INV01-PV03")
    )

    assert series.iloc[0] == pytest.approx(7.5)
    assert diagnostics["power_source"] == "direct"


def test_extract_power_falls_back_case_insensitively_and_deduplicates_mean():
    df = pd.DataFrame({
        "start time": ["2026-05-01 00:00", "2026-05-01 00:00", "2026-05-01 00:10"],
        "inverter_id": ["WB05-INV01"] * 3,
        "pv03 INPUT voltage(v)": [1000, 1000, 1000],
        "PV3 input CURRENT(a)": [8, 12, 10],
    })

    series, diagnostics = _extract_string_power(
        df, parse_string_selection("WB05-INV01-PV3")
    )

    assert list(series) == pytest.approx([10.0, 10.0])
    assert diagnostics["duplicate_rows"] == 1
    assert diagnostics["power_source"] == "voltage_current"


def test_daily_yield_uses_only_observed_slots_and_distinguishes_statuses(tmp_path):
    dates = parse_date_range("2026-05-01", "2026-05-03")
    csv1 = tmp_path / "20260501.csv"
    csv2 = tmp_path / "20260502.csv"
    pd.DataFrame({
        "Start Time": pd.date_range("2026-05-01", periods=12, freq="5min"),
        "Inverter_ID": "WB05-INV01",
        "PV3 Power(kW)": 10.0,
    }).to_csv(csv1, index=False)
    pd.DataFrame({
        "Start Time": ["2026-05-02 00:00"],
        "Inverter_ID": ["WB05-INV02"],
        "PV3 Power(kW)": [10.0],
    }).to_csv(csv2, index=False)

    report = build_report_data(
        {date(2026, 5, 1): csv1, date(2026, 5, 2): csv2},
        {},
        parse_string_selection("WB05-INV01-PV03"),
        dates,
        Path("config/site_geometry.yaml"),
    )

    assert len(report.five_minute) == 3 * 288
    assert report.five_minute["power_kw"].notna().sum() == 12
    assert report.daily["status"].tolist() == ["PARTIAL", "NO_STRING_DATA", "MISSING_CSV"]
    assert report.daily.loc[0, "string_yield_kwh"] == pytest.approx(10.0)
    assert report.daily.loc[1:, "string_yield_kwh"].isna().all()
    assert list(report.five_minute.columns) == [
        "timestamp", "inverter_id", "pv_string", "power_kw", "poa_wm2",
        "source_csv", "poa_source", "data_status",
    ]
    assert list(report.daily.columns) == [
        "date", "string_yield_kwh", "valid_power_samples", "expected_samples",
        "coverage_pct", "missing_power_samples", "poa_valid_samples", "source_csv", "status",
    ]


def test_complete_day_requires_exactly_288_valid_samples(tmp_path):
    csv_path = tmp_path / "20260501.csv"
    pd.DataFrame({
        "Start Time": pd.date_range("2026-05-01", periods=288, freq="5min"),
        "Inverter_ID": "WB05-INV01",
        "PV3 Power(kW)": 1.0,
    }).to_csv(csv_path, index=False)

    report = build_report_data(
        {date(2026, 5, 1): csv_path}, {},
        parse_string_selection("WB05-INV01-PV3"),
        parse_date_range("2026-05-01", "2026-05-01"),
        Path("config/site_geometry.yaml"),
    )

    assert report.daily.loc[0, "status"] == "COMPLETE"
    assert report.daily.loc[0, "coverage_pct"] == pytest.approx(100.0)


def test_wrong_file_date_rows_are_dropped_and_warned(tmp_path):
    csv_path = tmp_path / "20260501.csv"
    pd.DataFrame({
        "Start Time": ["2026-05-01 00:00", "2026-05-02 00:00"],
        "Inverter_ID": "WB05-INV01",
        "PV3 Power(kW)": [2.0, 99.0],
    }).to_csv(csv_path, index=False)

    report = build_report_data(
        {date(2026, 5, 1): csv_path}, {},
        parse_string_selection("WB05-INV01-PV3"),
        parse_date_range("2026-05-01", "2026-05-01"),
        Path("config/site_geometry.yaml"),
    )

    assert report.five_minute["power_kw"].notna().sum() == 1
    assert report.metadata["wrong_date_rows"] == {"20260501.csv": 1}
    assert "Rows outside their YYYYMMDD.csv date were dropped." in report.metadata["warnings"]


def test_negative_power_is_retained_and_warned(tmp_path):
    csv_path = tmp_path / "20260501.csv"
    pd.DataFrame({
        "Start Time": ["2026-05-01 00:00"],
        "Inverter_ID": ["WB05-INV01"],
        "PV3 Power(kW)": [-3.0],
    }).to_csv(csv_path, index=False)

    report = build_report_data(
        {date(2026, 5, 1): csv_path}, {},
        parse_string_selection("WB05-INV01-PV3"),
        parse_date_range("2026-05-01", "2026-05-01"),
        Path("config/site_geometry.yaml"),
    )

    assert report.five_minute.loc[0, "power_kw"] == pytest.approx(-3.0)
    assert report.metadata["negative_power_samples"] == 1
    assert "Negative power samples were retained." in report.metadata["warnings"]


def test_missing_five_minute_slot_stays_nan(tmp_path):
    csv_path = tmp_path / "20260501.csv"
    pd.DataFrame({
        "Start Time": ["2026-05-01 00:00", "2026-05-01 00:10"],
        "Inverter_ID": "WB05-INV01",
        "PV3 Power(kW)": [1.0, 1.0],
    }).to_csv(csv_path, index=False)

    report = build_report_data(
        {date(2026, 5, 1): csv_path}, {},
        parse_string_selection("WB05-INV01-PV3"),
        parse_date_range("2026-05-01", "2026-05-01"),
        Path("config/site_geometry.yaml"),
    )

    missing = report.five_minute.loc[
        report.five_minute["timestamp"] == pd.Timestamp("2026-05-01 00:05"), "power_kw"
    ]
    assert missing.isna().all()


def test_manage_object_fallback_normalizes_mixed_case_before_conversion():
    frame = pd.DataFrame({
        "Start Time": ["2026-05-01 00:00"],
        "manageobject": ["root/inv_a_234_ikn"],
        "PV3 Power(kW)": [4.0],
    })

    series, diagnostics = _extract_string_power(
        frame, parse_string_selection("WB02-INV34-PV3")
    )

    assert series.iloc[0] == pytest.approx(4.0)
    assert diagnostics["power_source"] == "direct"


def test_no_readable_csv_raises_loudly(tmp_path):
    with pytest.raises(RuntimeError, match="No requested CSV could be read"):
        build_report_data(
            {date(2026, 5, 1): tmp_path / "missing.csv"}, {},
            parse_string_selection("WB05-INV01-PV3"),
            parse_date_range("2026-05-01", "2026-05-01"),
            Path("config/site_geometry.yaml"),
        )


def test_no_valid_selected_string_sample_across_readable_csvs_raises(tmp_path):
    csv_path = tmp_path / "20260501.csv"
    pd.DataFrame({
        "Start Time": ["2026-05-01 00:00"],
        "Inverter_ID": ["WB05-INV02"],
        "PV3 Power(kW)": [4.0],
    }).to_csv(csv_path, index=False)

    with pytest.raises(RuntimeError, match="No valid power sample found for WB05-INV01-PV3"):
        build_report_data(
            {date(2026, 5, 1): csv_path}, {},
            parse_string_selection("WB05-INV01-PV3"),
            parse_date_range("2026-05-01", "2026-05-01"),
            Path("config/site_geometry.yaml"),
        )


def test_poa_uses_mapped_ws_and_labels_only_gap_fallback_as_avg(tmp_path):
    csv_path = tmp_path / "20260501.csv"
    poa_path = tmp_path / "POA PLTS IKN 2026.xlsx"
    pd.DataFrame({
        "Start Time": ["2026-05-01 00:00"],
        "Inverter_ID": ["WB05-INV01"],
        "PV3 Power(kW)": [4.0],
    }).to_csv(csv_path, index=False)
    _write_poa(
        poa_path,
        pd.date_range("2026-05-01", periods=2, freq="5min"),
        [500.0, np.nan],
        [600.0, 650.0],
    )
    manifest = SourceManifest(
        csv_by_date={}, poa_by_year={},
        missing_csv_dates=[date(2026, 5, 2)], missing_poa_years=[2027],
        url_csv="https://drive.google.com/drive/folders/csv",
        url_poa="https://drive.google.com/drive/folders/poa",
        csv_inventory_count=4, poa_inventory_count=2,
    )

    report = build_report_data(
        {date(2026, 5, 1): csv_path}, {2026: poa_path},
        parse_string_selection("WB05-INV01-PV3"),
        parse_date_range("2026-05-01", "2026-05-01"),
        _write_geometry(tmp_path),
        source_manifest=manifest,
        download_errors={"failed.csv": "HTTPError: denied"},
    )

    assert report.five_minute.loc[:1, "poa_wm2"].tolist() == pytest.approx([500.0, 650.0])
    assert report.five_minute.loc[:1, "poa_source"].tolist() == ["WS-2", "avg"]
    assert report.metadata["mapped_ws"] == "WS-2"
    assert report.metadata["poa_fallback_samples"] == 1
    assert report.metadata["loaded_poa_files"] == ["POA PLTS IKN 2026.xlsx"]
    assert report.metadata["source_url_csv"] == manifest.url_csv
    assert report.metadata["source_url_poa"] == manifest.url_poa
    assert report.metadata["csv_inventory_count"] == 4
    assert report.metadata["inventory_missing_csv_dates"] == ["2026-05-02"]
    assert report.metadata["download_errors"] == {"failed.csv": "HTTPError: denied"}


def test_unmapped_wb_does_not_receive_wholesale_average_poa(tmp_path):
    csv_path = tmp_path / "20260501.csv"
    poa_path = tmp_path / "POA PLTS IKN 2026.xlsx"
    pd.DataFrame({
        "Start Time": ["2026-05-01 00:00"],
        "Inverter_ID": ["WB06-INV01"],
        "PV3 Power(kW)": [4.0],
    }).to_csv(csv_path, index=False)
    _write_poa(poa_path, [pd.Timestamp("2026-05-01 00:00")], [np.nan], [700.0])

    report = build_report_data(
        {date(2026, 5, 1): csv_path}, {2026: poa_path},
        parse_string_selection("WB06-INV01-PV3"),
        parse_date_range("2026-05-01", "2026-05-01"),
        _write_geometry(tmp_path),
    )

    assert report.five_minute["poa_wm2"].isna().all()
    assert report.metadata["poa_fallback_samples"] == 0


def test_unreadable_poa_year_isolated_from_power_yield(tmp_path):
    csv_path = tmp_path / "20260501.csv"
    missing_poa = tmp_path / "POA PLTS IKN 2026.xlsx"
    pd.DataFrame({
        "Start Time": ["2026-05-01 00:00"],
        "Inverter_ID": ["WB05-INV01"],
        "PV3 Power(kW)": [6.0],
    }).to_csv(csv_path, index=False)

    report = build_report_data(
        {date(2026, 5, 1): csv_path}, {2026: missing_poa},
        parse_string_selection("WB05-INV01-PV3"),
        parse_date_range("2026-05-01", "2026-05-01"),
        _write_geometry(tmp_path),
    )

    assert report.daily.loc[0, "string_yield_kwh"] == pytest.approx(0.5)
    assert report.five_minute["poa_wm2"].isna().all()
    assert "POA PLTS IKN 2026.xlsx" in report.metadata["poa_read_errors"]


def test_output_path_and_workbook_contract(tmp_path, report_fixture, selection):
    path = build_output_path(
        tmp_path, selection, date(2026, 5, 1), date(2026, 5, 2)
    )

    assert path.name == "string_yield_WB05-INV01_PV3_20260501_20260502.xlsx"
    written = write_report_workbook(path, report_fixture)
    verify_report_workbook(written)

    workbook = load_workbook(written, data_only=False)
    try:
        assert workbook.sheetnames == [
            "Ringkasan_Harian", "Data_5Menit", "Grafik", "Metadata",
        ]
        daily = workbook["Ringkasan_Harian"]
        five_minute = workbook["Data_5Menit"]
        graph = workbook["Grafik"]
        metadata_sheet = workbook["Metadata"]

        assert daily.max_row == len(report_fixture.daily) + 1
        assert five_minute.max_row == len(report_fixture.five_minute) + 1
        assert daily.freeze_panes == "A2"
        assert five_minute.freeze_panes == "A2"
        assert metadata_sheet.freeze_panes == "A2"
        assert daily.auto_filter.ref == daily.dimensions
        assert five_minute.auto_filter.ref == five_minute.dimensions
        assert metadata_sheet.auto_filter.ref == metadata_sheet.dimensions

        assert isinstance(daily["A2"].value, (date, datetime))
        assert daily["A2"].number_format == "yyyy-mm-dd"
        assert daily["B2"].value == pytest.approx(10.0)
        assert daily["B3"].value is None
        assert daily["B2"].number_format == "0.000"
        assert daily["E2"].value == pytest.approx(12 / 288 * 100)
        assert daily["E2"].number_format == '0.0"%"'
        assert isinstance(five_minute["A2"].value, datetime)
        assert five_minute["A2"].number_format == "yyyy-mm-dd hh:mm"
        assert five_minute["D2"].value == pytest.approx(10.0)
        assert five_minute["D3"].value is None
        assert five_minute["D2"].number_format == "0.000"
        assert five_minute["E2"].number_format == "0.0"

        assert all(cell.font.bold for cell in daily[1])
        assert all(cell.font.bold for cell in five_minute[1])
        assert all(cell.font.bold for cell in metadata_sheet[1])
        assert daily.column_dimensions["A"].width >= 12
        assert five_minute.column_dimensions["A"].width >= 18
        assert metadata_sheet.column_dimensions["B"].width >= 40

        assert len(daily._charts) == 1
        assert len(graph._charts) == 2
        summary_chart = daily._charts[0]
        graph_daily_chart, combo_chart = graph._charts
        assert summary_chart.display_blanks == "gap"
        assert "Ringkasan_Harian" in summary_chart.ser[0].val.numRef.f
        assert "Ringkasan_Harian" in graph_daily_chart.ser[0].val.numRef.f
        assert len(combo_chart._charts) == 2
        power_chart, poa_chart = combo_chart._charts
        assert "Data_5Menit" in power_chart.ser[0].val.numRef.f
        assert "Data_5Menit" in poa_chart.ser[0].val.numRef.f
        axis_ids = {
            power_chart.x_axis.axId,
            power_chart.y_axis.axId,
            poa_chart.y_axis.axId,
        }
        assert power_chart.x_axis.tagname in {"catAx", "dateAx"}
        assert len(axis_ids) == 3
        assert power_chart.x_axis.crossAx == power_chart.y_axis.axId
        assert power_chart.y_axis.crossAx == power_chart.x_axis.axId
        assert poa_chart.y_axis.crossAx == power_chart.x_axis.axId
        assert poa_chart.y_axis.crosses == "max"
        assert power_chart.y_axis.title.tx.rich.p[0].r[0].t == "Power string (kW)"
        assert poa_chart.y_axis.title.tx.rich.p[0].r[0].t == "POA irradiance (W/m²)"

        serialized_axes, namespaces = _serialized_combo_axes(written)
        assert [tag for tag, _ in serialized_axes].count("catAx") + [
            tag for tag, _ in serialized_axes
        ].count("dateAx") == 1
        assert [tag for tag, _ in serialized_axes].count("valAx") == 2
        serialized_ids = [
            int(node.find("c:axId", namespaces).get("val"))
            for _, node in serialized_axes
        ]
        assert len(serialized_ids) == len(set(serialized_ids)) == 3
        serialized_crossings = {
            int(node.find("c:axId", namespaces).get("val")):
            int(node.find("c:crossAx", namespaces).get("val"))
            for _, node in serialized_axes
        }
        assert set(serialized_crossings.values()).issubset(serialized_crossings)
        category_id = next(
            int(node.find("c:axId", namespaces).get("val"))
            for tag, node in serialized_axes if tag in {"catAx", "dateAx"}
        )
        value_axes = [node for tag, node in serialized_axes if tag == "valAx"]
        secondary_axis = next(
            node for node in value_axes
            if node.find("c:crosses", namespaces) is not None
            and node.find("c:crosses", namespaces).get("val") == "max"
        )
        primary_axis = next(node for node in value_axes if node is not secondary_axis)
        primary_id = int(primary_axis.find("c:axId", namespaces).get("val"))
        secondary_id = int(secondary_axis.find("c:axId", namespaces).get("val"))
        assert serialized_crossings[category_id] == primary_id
        assert serialized_crossings[primary_id] == category_id
        assert serialized_crossings[secondary_id] == category_id

        metadata = dict(metadata_sheet.iter_rows(min_row=2, values_only=True))
        assert metadata["source_url_csv"] == report_fixture.metadata["source_url_csv"]
        assert metadata["source_url_poa"] == report_fixture.metadata["source_url_poa"]
        assert metadata["poa_fallback_samples"] == 1
        assert json.loads(metadata["missing_csv_dates"]) == ["2026-05-02"]
        sensitive_terms = ("token", "cookie", "secret", "credential", "password")
        for key in metadata:
            normalized = "".join(char for char in str(key).casefold() if char.isalpha())
            assert not any(term in normalized for term in sensitive_terms)
    finally:
        workbook.close()


@pytest.mark.parametrize(
    "sensitive_key",
    ["api_token", "session-cookie", "clientSecret", "credential_id", "db password"],
)
def test_workbook_rejects_sensitive_metadata_before_saving(
    tmp_path, report_fixture, sensitive_key
):
    report = ReportData(
        daily=report_fixture.daily,
        five_minute=report_fixture.five_minute,
        metadata={**report_fixture.metadata, sensitive_key: "must-not-leak"},
    )
    path = tmp_path / "sensitive.xlsx"

    with pytest.raises(ValueError, match="Sensitive metadata key"):
        write_report_workbook(path, report)

    assert not path.exists()


def test_workbook_rejects_sensitive_keys_nested_inside_structures(
    tmp_path, report_fixture
):
    report = ReportData(
        daily=report_fixture.daily,
        five_minute=report_fixture.five_minute,
        metadata={
            **report_fixture.metadata,
            "diagnostics": [{"request": {"api_token": "must-not-leak"}}],
        },
    )
    path = tmp_path / "nested-sensitive.xlsx"

    with pytest.raises(ValueError, match="Sensitive metadata key.*api_token"):
        write_report_workbook(path, report)

    assert not path.exists()


def test_workbook_canonicalizes_source_urls_before_metadata_serialization(
    tmp_path, report_fixture
):
    report = ReportData(
        daily=report_fixture.daily,
        five_minute=report_fixture.five_minute,
        metadata={
            **report_fixture.metadata,
            "source_url_csv": (
                "https://drive.google.com/drive/folders/csv-test"
                "?access_token=abc#frag"
            ),
            "source_url_poa": (
                "https://drive.google.com/drive/folders/poa-test?usp=sharing"
            ),
        },
    )

    written = write_report_workbook(tmp_path / "canonical-urls.xlsx", report)
    workbook = load_workbook(written, data_only=False)
    try:
        metadata = dict(
            workbook["Metadata"].iter_rows(min_row=2, values_only=True)
        )
        assert metadata["source_url_csv"] == (
            "https://drive.google.com/drive/folders/csv-test"
        )
        assert metadata["source_url_poa"] == (
            "https://drive.google.com/drive/folders/poa-test"
        )
        assert "access_token" not in metadata["source_url_csv"]
        assert "abc" not in metadata["source_url_csv"]
    finally:
        workbook.close()


def test_plot_contract_uses_gaps_and_secondary_axis(report_fixture, selection):
    fig_yield, ax_yield = plot_daily_yield(report_fixture.daily, selection)
    fig_power, (ax_power, ax_poa) = plot_power_vs_poa(
        report_fixture.five_minute,
        selection,
        date(2026, 5, 1),
        date(2026, 5, 2),
    )
    try:
        assert np.isnan(ax_yield.lines[0].get_ydata()).any()
        assert ax_yield.get_ylabel() == "String yield (kWh)"
        assert np.isnan(ax_power.lines[0].get_ydata()).any()
        assert ax_power.get_ylabel() == "Power string (kW)"
        assert ax_poa.get_ylabel() == "POA irradiance (W/m²)"
        assert ax_power is not ax_poa
    finally:
        plt.close(fig_yield)
        plt.close(fig_power)


def test_verify_report_workbook_rejects_missing_file(tmp_path):
    with pytest.raises(RuntimeError, match="Workbook was not created"):
        verify_report_workbook(tmp_path / "missing.xlsx")


def test_builder_writes_nbformat_45_with_nine_expected_cells(tmp_path):
    path = Path("output_string/_build_string_yield_notebook.py")
    spec = importlib.util.spec_from_file_location("string_yield_nb_builder", path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    target = module.build(tmp_path / "report.ipynb")
    nb = json.loads(target.read_text(encoding="utf-8"))
    committed = json.loads(
        Path("output_string/String_Yield_Power_Irradiance.ipynb").read_text(
            encoding="utf-8"
        )
    )
    assert committed == nb
    drifted = json.loads(json.dumps(committed))
    drifted["cells"][0]["source"].append("temporary drift")
    with pytest.raises(AssertionError):
        assert drifted == nb
    assert (nb["nbformat"], nb["nbformat_minor"]) == (4, 5)
    assert [cell["cell_type"] for cell in nb["cells"]] == [
        "markdown",
        *(["code"] * 8),
    ]
    markers = [
        "gdown>=6.0.0",
        "URL_CSV",
        "download_report_inputs",
        "build_report_data",
        "plot_daily_yield",
        "plot_power_vs_poa",
        "write_report_workbook",
        "google.colab",
    ]
    for cell, marker in zip(nb["cells"][1:], markers):
        source = "".join(cell["source"])
        assert marker in source
        ast.parse(source)


def test_notebook_cell_2_has_exactly_five_approved_literal_defaults():
    notebook = json.loads(
        Path("output_string/String_Yield_Power_Irradiance.ipynb").read_text(
            encoding="utf-8"
        )
    )
    tree = ast.parse("".join(notebook["cells"][2]["source"]))
    literal_assignments = []
    for node in ast.walk(tree):
        if not isinstance(node, (ast.Assign, ast.AnnAssign)):
            continue
        try:
            value = ast.literal_eval(node.value)
        except (TypeError, ValueError):
            continue
        targets = node.targets if isinstance(node, ast.Assign) else [node.target]
        assert len(targets) == 1
        assert isinstance(targets[0], ast.Name)
        literal_assignments.append((targets[0].id, value))

    assert literal_assignments == [
        (
            "URL_CSV",
            "https://drive.google.com/drive/folders/"
            "1f_KrPuqfZJTE5I9cVQiyp65QrHbkF3Iw?usp=sharing",
        ),
        (
            "URL_RAW_DATA_INPUT",
            "https://drive.google.com/drive/folders/"
            "1y37AsVViI7IL1tCRhvAGEjiq9wWnfu_e?usp=drive_link",
        ),
        ("PV_STRING", "WB05-INV01-PV03"),
        ("START_DATE", "2026-05-01"),
        ("END_DATE", "2026-05-14"),
    ]
