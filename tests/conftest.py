"""Shared pytest fixtures untuk pv_pipeline tests.

Fixtures yang disediakan:
- ``synthetic_combined_df`` : DataFrame mirip ``combined_df`` output Cell 2,
  1 hari × 3 inverters × 10 PV strings @ 5-min interval.
- ``synthetic_combined_df_with_outlier`` : sama tapi PV3 R abnormal (high_R).
- ``mock_poa`` : MockPOA dengan sin-curve POA (peak 1000 W/m^2 @ noon).
- ``mock_panel`` : MockPanel Jinko JKM625N (Voc_STC=55.72, 26 modules/string).
- ``mock_cell_temp`` : MockTcell constant 30 C.
"""
from __future__ import annotations

import numpy as np
import pandas as pd
import pytest


# ---------- Synthetic combined_df ----------


@pytest.fixture
def synthetic_combined_df():
    """3 inverters × 10 PV strings × 145 timestamps (2026-05-14 06:00..18:00 @ 5min)."""
    rng = np.random.default_rng(42)
    n = 145
    t = pd.date_range("2026-05-14 06:00", "2026-05-14 18:00", freq="5min")[:n]
    hours = np.linspace(0, 12, n)
    sun = np.sin(np.pi * hours / 12) ** 2
    I_base = 13.0 * sun
    V_base = 1200.0 + 200.0 * np.exp(-3 * sun)
    rows = []
    for inv in ["WB05-INV01", "WB05-INV02", "WB02-INV05"]:
        for ts_i, ts in enumerate(t):
            row = {"Inverter_ID": inv, "Start Time": ts}
            for pv_n in range(1, 11):
                row[f"PV{pv_n} input voltage(V)"] = V_base[ts_i] + rng.normal(0, 2)
                row[f"PV{pv_n} input current(A)"] = I_base[ts_i] + rng.normal(0, 0.05)
            row["Voltage between PV– and the ground(V)"] = rng.normal(0, 5)
            rows.append(row)
    return pd.DataFrame(rows)


@pytest.fixture
def synthetic_combined_df_with_outlier(synthetic_combined_df):
    """Same as synthetic_combined_df but PV3 di WB05-INV01 punya high_R signature.

    PV3: I × 0.40 (60% drop), V × 1.03 -> R_str ~2.5x normal.
    Expected: M2bPeerZScore flag PV3 dengan |z| > 2.5, voc_ratio > 0.95.
    """
    df = synthetic_combined_df.copy()
    mask = df["Inverter_ID"] == "WB05-INV01"
    df.loc[mask, "PV3 input voltage(V)"] = df.loc[mask, "PV3 input voltage(V)"] * 1.03
    df.loc[mask, "PV3 input current(A)"] = df.loc[mask, "PV3 input current(A)"] * 0.40
    return df


# ---------- Mock infrastructure providers ----------


class _MockPOA:
    """Mock POAProvider: sin-curve POA peak 1000 W/m^2 @ noon, elev peak ~85 @ noon."""

    def get_poa(self, timestamps, wb_id, source="auto"):
        ts = pd.DatetimeIndex(timestamps)
        hrs = (ts.hour - 6) + (ts.minute / 60.0)
        poa_arr = np.where(
            (hrs >= 0) & (hrs <= 12),
            1000.0 * np.sin(np.pi * hrs / 12) ** 2,
            0.0,
        )
        return pd.Series(poa_arr, index=ts)

    def get_poa_all_sources(self, timestamps, wb_id):
        s = self.get_poa(timestamps, wb_id)
        return pd.DataFrame({"auto": s})

    def get_solar_elevation(self, timestamps):
        """Synthetic sun elevation: peak ~85 deg at noon, 0 at sunrise/sunset,
        negative (below horizon) outside 06:00..18:00 jam-of-day."""
        ts = pd.DatetimeIndex(timestamps)
        hrs = (ts.hour - 6) + (ts.minute / 60.0)
        elev = np.where(
            (hrs >= 0) & (hrs <= 12),
            85.0 * np.sin(np.pi * hrs / 12),
            -45.0,
        )
        return pd.Series(elev, index=ts, name="solar_elevation_deg")


class _MockPanel:
    """Mock PanelSpec: Jinko JKM625N (Voc STC = 55.72V, 26 modules/string)."""

    stc = type("X", (), {"voc_v": 55.72, "isc_a": 14.27, "pmax_w": 625.0})
    max_system_voltage_v = 1500
    panel_model = "Mock Jinko JKM625N"

    def voc_at_cell_temp(self, t, base="stc"):
        # Apply -0.25%/C temp coef from 25C reference.
        return 55.72 * (1.0 + (-0.25 / 100.0) * (float(t) - 25.0))

    def modules_per_string(self, wb_id):
        wb = str(wb_id).upper()
        if wb in {"WB01", "WB02"}:
            return 24
        return 26

    def voc_string_stc(self, wb_id):
        return 55.72 * self.modules_per_string(wb_id)

    def voc_string_nominal(self, t, wb_id, base="stc"):
        return self.voc_at_cell_temp(t, base=base) * self.modules_per_string(wb_id)

    def voc_string_at_design_min_temp(self, wb_id, min_cell_temp_c=10.0):
        return self.voc_string_nominal(min_cell_temp_c, wb_id)


class _MockCellTemp:
    """Mock CellTempProvider: constant 30 C."""

    def get_tcell(self, timestamps, wb_id, source="auto"):
        ts = pd.DatetimeIndex(timestamps)
        return pd.Series([30.0] * len(ts), index=ts)


@pytest.fixture
def mock_poa():
    return _MockPOA()


@pytest.fixture
def mock_panel():
    return _MockPanel()


@pytest.fixture
def mock_cell_temp():
    return _MockCellTemp()


# ---------- Synthetic xlsx fixtures (for loaders) ----------


@pytest.fixture
def synthetic_pyranometer_xlsx(tmp_path):
    """Synthetic POA xlsx mirip ``POA PLTS IKN 2026.xlsx`` layout.

    Sheet ``POA PLTS IKN``, cols: Date time + WS-1..5 + Rata-rata.
    288 timestamps (1 day @ 5-min) dengan sin-curve POA.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "POA PLTS IKN"
    headers = ["Date time"] + [
        f"POA Irradiance (W/m2) WS {i}" for i in range(1, 6)
    ] + ["Rata-rata WS 1 - WS 5"]
    ws.append(headers)

    n = 288  # 1 day @ 5-min
    ts = pd.date_range("2026-05-14 00:00", periods=n, freq="5min")
    for t in ts:
        hour = t.hour + t.minute / 60.0
        sun = np.sin(np.pi * (hour - 6.0) / 12.0) ** 2 if 6.0 <= hour <= 18.0 else 0.0
        poa = 1000.0 * sun
        # WS-2 sengaja sebagian NaN untuk test fallback behavior
        ws_values = [poa, np.nan if 8 <= hour <= 14 else poa, poa, poa, poa]
        avg = float(np.nanmean(ws_values))
        ws.append([t.to_pydatetime()] + ws_values + [avg])

    path = tmp_path / "synthetic_pyranometer.xlsx"
    wb.save(str(path))
    return str(path)


@pytest.fixture
def synthetic_tcell_xlsx(tmp_path):
    """Synthetic Tcell xlsx mirip ``PV Module Temperature PLTS IKN.xlsx`` layout.

    Sheet ``PV Module Temp``, 18 cols. Row 0 = sub-headers (sensor names),
    row 1+ = data (5-min naive WITA).
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PV Module Temp"

    # Header row: Datetime + 4 WS groups (3 sensors + 1 avg) + Overall
    # Pakai nama generic karena loader pakai column INDEX, bukan nama.
    header = ["Datetime"]
    for ws_n in range(1, 5):
        header += [f"Weather Station {ws_n} (WS-{ws_n})"] + [f"Unnamed: {i}" for i in range(1, 4)]
    header += ["Overall"]
    ws.append(header)

    # Sub-header row (akan di-drop saat load)
    sub_header = [None]
    for ws_n in range(1, 5):
        for sensor in range(1, 4):
            sub_header.append(f"PV Module Temperature 0{sensor} WS-{ws_n} (oC)")
        sub_header.append(f"Average PV Module Temperature WS-{ws_n} (oC)")
    sub_header.append("Average PV Module Temperature (oC)")
    ws.append(sub_header)

    # Data rows: 100 timestamps @ 5-min
    n = 100
    ts = pd.date_range("2026-05-14 06:00", periods=n, freq="5min")
    rng = np.random.default_rng(42)
    for t in ts:
        hour = t.hour + t.minute / 60.0
        # Tcell mirror temperature sun-driven (25 C dawn -> 45 C noon -> 30 C evening)
        base = 25.0 + 20.0 * np.sin(np.pi * max(0, (hour - 6.0) / 12.0)) if 6.0 <= hour <= 18.0 else 25.0
        row = [t.to_pydatetime()]
        for ws_n in range(1, 5):
            sensor_vals = [base + rng.normal(0, 0.5) for _ in range(3)]
            avg_ws = float(np.mean(sensor_vals))
            row += sensor_vals + [avg_ws]
        # Overall = avg of 4 WS averages
        overall = float(np.mean([row[4], row[8], row[12], row[16]]))
        row.append(overall)
        ws.append(row)

    path = tmp_path / "synthetic_tcell.xlsx"
    wb.save(str(path))
    return str(path)


@pytest.fixture
def synthetic_albedo_xlsx(tmp_path):
    """Synthetic albedo xlsx mirip ``Surface Albedo Forecast TMY NSRDB PLTS IKN.xlsx``.

    Sheet ``Sheet1``, cols Date/Time + Surface Albedo. 48 timestamps (1 day @ 30-min).
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Date/Time", "Surface Albedo"])

    n = 48
    ts = pd.date_range("2026-05-14 00:00", periods=n, freq="30min")
    for t in ts:
        # NSRDB TMY albedo typical 0.13-0.18 tropical mixed terrain
        ws.append([t.to_pydatetime(), 0.15])

    path = tmp_path / "synthetic_albedo.xlsx"
    wb.save(str(path))
    return str(path)


@pytest.fixture
def m2_config_minimal():
    """Minimal cfg dict untuk panggil detector.run(combined_df, cfg)."""
    return {
        "m2b": {
            "poa_threshold_wm2": 300.0,
            "z_threshold": 2.5,
            "voc_ratio_threshold": 0.95,
            "stat_method": "median",
            "pv_max": 10,
            "min_peer_strings": 3,
            "min_daylight_samples": 10,
        },
        "m2b_open_circuit": {
            "poa_threshold_wm2": 200.0,
            "i_ratio_threshold": 0.05,
            "debounce_consecutive_steps": 2,
            "confidence_pct": 95.0,
            "pv_max": 10,
            "min_peer_strings": 3,
            "min_daylight_samples": 5,
        },
        "m2b_ground_fault": {
            "poa_threshold_wm2": 200.0,
            "v_to_ground_abs_threshold_v": 50.0,
            "adaptive_z_threshold": 3.0,
            "voc_ratio_threshold": 0.85,
            "i_high_z_threshold": 2.0,
            "pv_max": 10,
            "min_daylight_samples": 5,
        },
        "poa": {
            "emit_all_sources": False,
            "default_source": "auto",
            "site_geometry_path": "config/site_geometry.yaml",
        },
        "panel": {"spec_path": "config/panel_spec.yaml"},
    }
