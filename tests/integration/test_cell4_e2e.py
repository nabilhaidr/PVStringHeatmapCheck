"""End-to-end integration test untuk Cell 4 M2 Pipeline (Wave 10).

Compose M2Engine dengan 3 M2b detectors (peer_zscore + open_circuit +
ground_fault), run pada synthetic combined_df dengan PV3 high_R signature,
write xlsx output, verify sheet structure + content.

Fixture reuse:
- synthetic_combined_df_with_outlier (conftest.py): PV3 di WB05-INV01 punya
  high_R signature (voltage *1.03, current *0.40 -> R_str ~2.5x normal).
- mock_poa, mock_panel, mock_cell_temp (conftest.py): synthetic providers.
- m2_config_minimal (conftest.py): cfg dict tanpa real yaml.
"""
from __future__ import annotations

import openpyxl
import pandas as pd
import pytest

from pv_pipeline.core import M2Engine, Severity
from pv_pipeline.ground_fault import M2bGroundFault
from pv_pipeline.open_circuit import M2bOpenCircuit
from pv_pipeline.peer_zscore import M2bPeerZScore


def _build_engine(mock_poa, mock_panel, mock_cell_temp):
    """3-detector M2Engine with shared mock providers."""
    return M2Engine([
        M2bPeerZScore(poa=mock_poa, panel=mock_panel, cell_temp=mock_cell_temp),
        M2bOpenCircuit(poa=mock_poa),
        M2bGroundFault(poa=mock_poa, panel=mock_panel, cell_temp=mock_cell_temp),
    ])


# ---------- Cell 4 e2e ----------


def test_cell4_pipeline_emits_findings_for_pv3_outlier(
    synthetic_combined_df_with_outlier,
    mock_poa, mock_panel, mock_cell_temp,
    m2_config_minimal,
):
    """End-to-end: PV3 di WB05-INV01 dengan R abnormal harus surface high_R finding."""
    engine = _build_engine(mock_poa, mock_panel, mock_cell_temp)
    findings = engine.run_all(synthetic_combined_df_with_outlier, m2_config_minimal)
    high_r = [f for f in findings if f.fault_type == "high_R"]
    assert high_r, "PV3 high_R not detected"
    pv3 = [f for f in high_r if f.inverter_id == "WB05-INV01" and f.pv_string == "PV3"]
    assert pv3, "PV3 specifically not flagged"
    assert pv3[0].severity in (Severity.HIGH, Severity.MEDIUM)


def test_cell4_xlsx_output_contains_string_status_sheets(
    synthetic_combined_df_with_outlier,
    mock_poa, mock_panel, mock_cell_temp,
    m2_config_minimal,
    tmp_path,
):
    """write_xlsx_multi: verify Findings + 3 StringStatus sheets emitted."""
    engine = _build_engine(mock_poa, mock_panel, mock_cell_temp)
    findings = engine.run_all(synthetic_combined_df_with_outlier, m2_config_minimal)

    out_path = str(tmp_path / "m2_findings_e2e.xlsx")
    M2Engine.write_xlsx_multi(findings, engine.submodules, out_path)

    wb = openpyxl.load_workbook(out_path, read_only=True)
    try:
        sheets = list(wb.sheetnames)
    finally:
        wb.close()

    assert "Findings" in sheets
    expected_status_sheets = {
        "M2b_peer_zscore_StringStatus",
        "M2b_open_circuit_StringStatus",
        "M2b_ground_fault_StringStatus",
    }
    actual_status = {s for s in sheets if "StringStatus" in s}
    assert actual_status == expected_status_sheets, (
        f"Status sheets mismatch.\n  expected: {expected_status_sheets}\n"
        f"  actual: {actual_status}"
    )


def test_cell4_string_status_sheets_contain_status_column(
    synthetic_combined_df_with_outlier,
    mock_poa, mock_panel, mock_cell_temp,
    m2_config_minimal,
    tmp_path,
):
    """Each StringStatus sheet has 'status' col + valid values per detector."""
    engine = _build_engine(mock_poa, mock_panel, mock_cell_temp)
    findings = engine.run_all(synthetic_combined_df_with_outlier, m2_config_minimal)

    out_path = str(tmp_path / "m2_findings_e2e.xlsx")
    M2Engine.write_xlsx_multi(findings, engine.submodules, out_path)

    valid_statuses = {
        "M2b_peer_zscore_StringStatus": {"NORMAL", "high_R", "EMPTY"},
        "M2b_open_circuit_StringStatus": {"NORMAL", "open_circuit", "EMPTY"},
        "M2b_ground_fault_StringStatus": {"NORMAL", "ground_fault", "EMPTY"},
    }
    for sheet_name, allowed in valid_statuses.items():
        df = pd.read_excel(out_path, sheet_name=sheet_name)
        assert "status" in df.columns, f"{sheet_name}: status col missing"
        assert set(df["status"].unique()).issubset(allowed), (
            f"{sheet_name}: unexpected status values {df['status'].unique()}"
        )


def test_cell4_pipeline_with_preprocessing_enabled_emits_audit(
    synthetic_combined_df_with_outlier,
    mock_poa, mock_panel, mock_cell_temp,
    m2_config_minimal,
    tmp_path,
):
    """Wave 9 wire: enable preprocessing in e2e pipeline + verify audit sheets."""
    cfg = dict(m2_config_minimal)
    cfg["preprocessing"] = {"enabled": True, "window": 7, "max_deviation": 3.0}

    engine = _build_engine(mock_poa, mock_panel, mock_cell_temp)
    findings = engine.run_all(synthetic_combined_df_with_outlier, cfg)

    out_path = str(tmp_path / "m2_findings_with_preprocessing.xlsx")
    M2Engine.write_xlsx_multi(findings, engine.submodules, out_path)

    wb = openpyxl.load_workbook(out_path, read_only=True)
    try:
        sheets = list(wb.sheetnames)
    finally:
        wb.close()

    audit_sheets = {s for s in sheets if "PreprocessingAudit" in s}
    # At least 1 detector emits audit (synthetic noise triggers Hampel false positives).
    assert len(audit_sheets) >= 1, f"No PreprocessingAudit sheets: {sheets}"


def test_cell4_pipeline_clean_data_few_findings(
    synthetic_combined_df,
    mock_poa, mock_panel, mock_cell_temp,
    m2_config_minimal,
):
    """Clean synthetic data (no outlier): low false positive rate."""
    engine = _build_engine(mock_poa, mock_panel, mock_cell_temp)
    findings = engine.run_all(synthetic_combined_df, m2_config_minimal)
    # 3 inverters x 3 detectors = max 9 false positive tolerance.
    assert len(findings) < 10, f"Too many false positives: {len(findings)}"
