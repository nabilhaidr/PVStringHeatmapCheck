from __future__ import annotations

from dataclasses import dataclass
from datetime import date
import json
import os
from pathlib import Path
import re
from urllib.parse import urlsplit, urlunsplit

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pandas as pd

from pv_pipeline.string_yield_report import (
    DownloadedInputs,
    SourceManifest,
    _append_dataframe,
    _canonicalize_drive_folder_url_for_persistence,
    _excel_value,
    _normalize_manage_object_value,
    _reject_sensitive_metadata_keys,
    _set_column_widths,
    download_manifest,
    inventory_drive_folder,
    select_source_manifest,
    validate_drive_folder_url,
)
from pv_pipeline.transformations import add_inverter_id


DETAIL_COLUMNS = [
    "date",
    "pv_string",
    "inverter_id",
    "pv_label",
    "string_yield_kwh",
    "valid_power_samples",
    "expected_samples",
    "coverage_pct",
    "missing_power_samples",
    "source_csv",
    "status",
]
SHEET_ORDER = ["Rekap_Yield_kWh", "Detail_Harian", "Metadata"]
EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_COLUMNS = 16_384
EXCEL_MAX_CELL_CHARACTERS = 32_767
METADATA_CHUNK_SIZE = 30_000
INVERTER_RE = re.compile(r"^WB(\d{2})-INV(\d{2})$", re.I)
POWER_RE = re.compile(r"^PV0*(\d{1,2}) Power\(kW\)$", re.I)
VOLTAGE_RE = re.compile(r"^PV0*(\d{1,2}) Voltage\(V\)$", re.I)
CURRENT_RE = re.compile(r"^PV0*(\d{1,2}) Current\(A\)$", re.I)
DRIVE_URL_RE = re.compile(
    r"https://drive\.google\.com/[^\s\"']+",
    re.I,
)


@dataclass
class AllStringYieldData:
    summary: pd.DataFrame
    daily: pd.DataFrame
    metadata: dict[str, object]


def download_csv_inputs(
    url_csv: str,
    dates: pd.DatetimeIndex,
    destination: Path,
) -> tuple[SourceManifest, DownloadedInputs]:
    url_csv = validate_drive_folder_url(url_csv)
    csv_items = inventory_drive_folder(url_csv)
    manifest = select_source_manifest(
        csv_items,
        [],
        dates,
        url_csv=url_csv,
        include_poa=False,
    )
    return manifest, download_manifest(manifest, Path(destination))


def _find_column(frame: pd.DataFrame, expected: str) -> str | None:
    return next(
        (
            str(column)
            for column in frame.columns
            if str(column).casefold() == expected.casefold()
        ),
        None,
    )


def _natural_string_key(value: str) -> tuple[int, int, int]:
    match = re.fullmatch(
        r"WB(\d{2})-INV(\d{2})-PV(\d{1,2})",
        value,
    )
    if not match:
        raise ValueError(f"Invalid canonical PV string: {value}")
    return tuple(int(part) for part in match.groups())


def _power_source_columns(frame: pd.DataFrame):
    direct = {}
    voltage = {}
    current = {}
    for column in frame.columns:
        name = str(column)
        for pattern, target in (
            (POWER_RE, direct),
            (VOLTAGE_RE, voltage),
            (CURRENT_RE, current),
        ):
            match = pattern.fullmatch(name)
            if match and 1 <= int(match.group(1)) <= 28:
                target[int(match.group(1))] = name
    sources = {}
    for pv_number in sorted(set(direct) | (set(voltage) & set(current))):
        if pv_number in direct:
            sources[pv_number] = ("direct", direct[pv_number], None)
        else:
            sources[pv_number] = (
                "voltage_current",
                voltage[pv_number],
                current[pv_number],
            )
    return sources


def _extract_all_string_power(
    frame: pd.DataFrame,
    requested_day: date,
):
    timestamp_column = _find_column(frame, "Start Time")
    if timestamp_column is None:
        raise KeyError("CSV missing Start Time column.")

    inverter_column = _find_column(frame, "Inverter_ID")
    if inverter_column is None:
        manage_column = _find_column(frame, "ManageObject")
        if manage_column is None:
            raise KeyError("CSV missing Inverter_ID and ManageObject columns.")
        normalized = frame.rename(columns={manage_column: "ManageObject"})
        normalized["ManageObject"] = normalized["ManageObject"].map(
            _normalize_manage_object_value
        )
        frame = add_inverter_id(normalized)
        inverter_column = "Inverter_ID"

    sources = _power_source_columns(frame)
    if not sources:
        raise KeyError("CSV has no usable PV power columns.")

    timestamps = pd.to_datetime(frame[timestamp_column], errors="coerce")
    wrong_date_mask = timestamps.notna() & (
        timestamps.dt.date != requested_day
    )
    aligned_mask = (
        timestamps.notna()
        & timestamps.dt.minute.mod(5).eq(0)
        & timestamps.dt.second.eq(0)
        & timestamps.dt.microsecond.eq(0)
    )
    inverter_ids = (
        frame[inverter_column]
        .astype("string")
        .str.strip()
        .str.upper()
    )
    valid_inverter = inverter_ids.str.fullmatch(INVERTER_RE, na=False)
    parts = []
    source_labels = set()
    diagnostics_non_finite = 0
    for pv_number, (source, first_column, second_column) in sources.items():
        first = pd.to_numeric(frame[first_column], errors="coerce")
        if source == "direct":
            power = first
        else:
            second = pd.to_numeric(frame[second_column], errors="coerce")
            with np.errstate(over="ignore", invalid="ignore"):
                power = first * second / 1000
        pv_label = f"PV{pv_number}"
        part = pd.DataFrame({
            "timestamp": timestamps,
            "inverter_id": inverter_ids,
            "pv_label": pv_label,
            "power_kw": power,
        })
        candidate_mask = (
            ~wrong_date_mask
            & aligned_mask
            & valid_inverter
            & part["timestamp"].notna()
            & part["power_kw"].notna()
        )
        finite_mask = np.isfinite(part["power_kw"])
        non_finite_samples = int((candidate_mask & ~finite_mask).sum())
        part = part.loc[candidate_mask & finite_mask].copy()
        part["pv_string"] = part["inverter_id"] + f"-{pv_label}"
        parts.append(part)
        source_labels.add(f"{pv_label}:{source}")
        diagnostics_non_finite += non_finite_samples

    tidy = pd.concat(parts, ignore_index=True)
    duplicate_rows = int(
        len(tidy)
        - len(tidy.drop_duplicates(["timestamp", "pv_string"]))
    )
    tidy = (
        tidy.groupby(
            ["timestamp", "inverter_id", "pv_string", "pv_label"],
            as_index=False,
        )["power_kw"]
        .mean()
        .sort_values(["timestamp", "pv_string"])
    )
    return tidy, {
        "wrong_date_rows": int(wrong_date_mask.sum()),
        "duplicate_rows": duplicate_rows,
        "negative_samples": int((tidy["power_kw"] < 0).sum()),
        "non_finite_samples": diagnostics_non_finite,
        "power_sources": sorted(source_labels),
    }


def build_all_string_daily_yield(
    csv_by_date,
    dates,
    *,
    source_manifest=None,
    download_errors=None,
) -> AllStringYieldData:
    if len(dates) == 0:
        raise ValueError("dates must contain at least one requested day.")

    requested_days = [pd.Timestamp(value).date() for value in dates]
    source_by_day = {}
    read_errors = {}
    wrong_date_rows = {}
    duplicate_rows = 0
    negative_samples = 0
    non_finite_samples = 0
    power_sources = set()
    stats_by_day = {}
    string_info = {}
    readable_days = set()

    for requested_day in requested_days:
        raw_path = csv_by_date.get(requested_day)
        if raw_path is None:
            continue
        path = Path(raw_path)
        try:
            frame = pd.read_csv(path, low_memory=False)
            tidy, diagnostics = _extract_all_string_power(
                frame,
                requested_day,
            )
        except Exception as exc:
            read_errors[path.name] = f"{type(exc).__name__}: {exc}"
            continue
        readable_days.add(requested_day)
        source_by_day[requested_day] = path.name
        wrong_date_rows[path.name] = diagnostics["wrong_date_rows"]
        duplicate_rows += diagnostics["duplicate_rows"]
        negative_samples += diagnostics["negative_samples"]
        non_finite_samples += diagnostics["non_finite_samples"]
        power_sources.update(diagnostics["power_sources"])
        if not tidy.empty:
            stats = (
                tidy.groupby(
                    ["inverter_id", "pv_string", "pv_label"],
                    as_index=False,
                )["power_kw"]
                .agg(["count", "sum"])
                .reset_index()
                .rename(columns={"count": "valid_power_samples"})
            )
            stats["string_yield_kwh"] = stats.pop("sum") * 5 / 60
            stats = stats.set_index("pv_string")
            stats_by_day[requested_day] = stats
            string_info.update({
                str(pv_string): (str(inverter_id), str(pv_label))
                for pv_string, inverter_id, pv_label in stats.reset_index()[
                    ["pv_string", "inverter_id", "pv_label"]
                ].itertuples(index=False, name=None)
            })

    if not readable_days:
        raise RuntimeError("No requested CSV could be read and validated.")
    if not string_info:
        raise RuntimeError(
            "No valid PV string power sample found in requested range."
        )

    strings = sorted(string_info, key=_natural_string_key)
    detail_row_count = len(requested_days) * len(strings)
    if detail_row_count + 1 > EXCEL_MAX_ROWS:
        maximum_days = (EXCEL_MAX_ROWS - 1) // len(strings)
        raise ValueError(
            "Detail_Harian would exceed the Excel row limit "
            f"({detail_row_count + 1:,} > {EXCEL_MAX_ROWS:,}); "
            f"shorten the range to at most {maximum_days} days for "
            f"{len(strings):,} detected strings."
        )
    if len(strings) + 1 > EXCEL_MAX_COLUMNS:
        raise ValueError(
            "Rekap_Yield_kWh would exceed the Excel column limit "
            f"({len(strings) + 1:,} > {EXCEL_MAX_COLUMNS:,})."
        )

    daily_rows = []
    for requested_day in requested_days:
        raw_path = csv_by_date.get(requested_day)
        path_name = Path(raw_path).name if raw_path is not None else None
        day_stats = stats_by_day.get(requested_day)
        for pv_string in strings:
            if day_stats is not None and pv_string in day_stats.index:
                stats = day_stats.loc[pv_string]
                valid = int(stats["valid_power_samples"])
                yield_kwh = float(stats["string_yield_kwh"])
            else:
                valid = 0
                yield_kwh = np.nan
            if raw_path is None:
                status = "MISSING_CSV"
            elif path_name in read_errors:
                status = "CSV_READ_ERROR"
            elif valid == 0:
                status = "NO_STRING_DATA"
            elif valid == 288:
                status = "COMPLETE"
            else:
                status = "PARTIAL"
            inverter_id, pv_label = string_info[pv_string]
            daily_rows.append({
                "date": requested_day,
                "pv_string": pv_string,
                "inverter_id": inverter_id,
                "pv_label": pv_label,
                "string_yield_kwh": yield_kwh,
                "valid_power_samples": valid,
                "expected_samples": 288,
                "coverage_pct": valid / 288 * 100,
                "missing_power_samples": 288 - valid,
                "source_csv": path_name,
                "status": status,
            })

    daily = pd.DataFrame(daily_rows, columns=DETAIL_COLUMNS)
    summary = daily.pivot(
        index="date",
        columns="pv_string",
        values="string_yield_kwh",
    ).reindex(index=requested_days, columns=strings)
    summary.columns.name = None
    summary = summary.reset_index()

    missing_dates = [
        value.isoformat()
        for value in requested_days
        if csv_by_date.get(value) is None
    ]
    inventory_missing_dates = [
        pd.Timestamp(value).date().isoformat()
        for value in (
            source_manifest.missing_csv_dates if source_manifest else []
        )
    ]
    warnings_list = []
    if any(wrong_date_rows.values()):
        warnings_list.append(
            "Rows outside their YYYYMMDD.csv date were dropped."
        )
    if negative_samples:
        warnings_list.append("Negative power samples were retained.")
    if non_finite_samples:
        warnings_list.append("Non-finite power samples were dropped.")

    metadata = {
        "start_date": requested_days[0].isoformat(),
        "end_date": requested_days[-1].isoformat(),
        "generated_at_wita": pd.Timestamp.now(
            tz="Asia/Makassar"
        ).isoformat(),
        "source_url_csv": (
            source_manifest.url_csv if source_manifest else ""
        ),
        "csv_inventory_count": (
            source_manifest.csv_inventory_count if source_manifest else 0
        ),
        "requested_days": len(requested_days),
        "detected_string_count": len(strings),
        "loaded_csv_files": sorted(source_by_day.values()),
        "missing_csv_dates": missing_dates,
        "inventory_missing_csv_dates": inventory_missing_dates,
        "download_errors": dict(download_errors or {}),
        "csv_read_errors": read_errors,
        "wrong_date_rows": wrong_date_rows,
        "duplicate_rows": duplicate_rows,
        "negative_power_samples": negative_samples,
        "non_finite_power_samples": non_finite_samples,
        "power_sources": sorted(power_sources),
        "yield_formula": "sum(power_kw_valid * 5/60)",
        "interval_minutes": 5,
        "warnings": warnings_list,
    }
    return AllStringYieldData(
        summary=summary,
        daily=daily,
        metadata=metadata,
    )


def build_all_string_output_path(output_dir, start, end) -> Path:
    return Path(output_dir) / (
        f"all_string_yield_{start:%Y%m%d}_{end:%Y%m%d}.xlsx"
    )


def _without_drive_query(match: re.Match) -> str:
    parts = urlsplit(match.group(0))
    return urlunsplit((parts.scheme, parts.netloc, parts.path, "", ""))


def _metadata_rows(metadata):
    for key, value in metadata.items():
        if key == "source_url_csv" and value:
            value = _canonicalize_drive_folder_url_for_persistence(value)
        if isinstance(value, (dict, list, tuple)):
            value = json.dumps(value, ensure_ascii=False, default=str)
        value = _excel_value(value)
        if isinstance(value, str):
            value = DRIVE_URL_RE.sub(_without_drive_query, value)
            if len(value) > METADATA_CHUNK_SIZE:
                chunks = [
                    value[index:index + METADATA_CHUNK_SIZE]
                    for index in range(0, len(value), METADATA_CHUNK_SIZE)
                ]
                for index, chunk in enumerate(chunks, start=1):
                    yield f"{key}[{index}/{len(chunks)}]", chunk
                continue
        yield str(key), value


def write_all_string_workbook(
    output_path: Path,
    report: AllStringYieldData,
) -> Path:
    if list(report.daily.columns) != DETAIL_COLUMNS:
        raise ValueError(
            f"Unexpected detail columns: {list(report.daily.columns)!r}"
        )
    summary_columns = list(report.summary.columns)
    if not summary_columns or summary_columns[0] != "date":
        raise ValueError(
            f"Unexpected summary columns: {summary_columns!r}"
        )
    pv_strings = summary_columns[1:]
    if pv_strings != sorted(pv_strings, key=_natural_string_key):
        raise ValueError("Summary PV string columns are not naturally sorted.")
    _reject_sensitive_metadata_keys(report.metadata)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    recap_sheet = workbook.create_sheet("Rekap_Yield_kWh")
    detail_sheet = workbook.create_sheet("Detail_Harian")
    metadata_sheet = workbook.create_sheet("Metadata")

    _append_dataframe(recap_sheet, report.summary)
    recap_sheet.freeze_panes = "B2"
    recap_sheet.auto_filter.ref = recap_sheet.dimensions
    for row in range(2, recap_sheet.max_row + 1):
        recap_sheet.cell(row=row, column=1).number_format = "yyyy-mm-dd"
        for column in range(2, recap_sheet.max_column + 1):
            recap_sheet.cell(row=row, column=column).number_format = "0.000"
    recap_widths = {"A": 12}
    for column in range(2, recap_sheet.max_column + 1):
        recap_widths[get_column_letter(column)] = 20
    _set_column_widths(recap_sheet, recap_widths)

    _append_dataframe(detail_sheet, report.daily[DETAIL_COLUMNS])
    detail_sheet.freeze_panes = "A2"
    detail_sheet.auto_filter.ref = detail_sheet.dimensions
    for row in range(2, detail_sheet.max_row + 1):
        detail_sheet.cell(row=row, column=1).number_format = "yyyy-mm-dd"
        detail_sheet.cell(row=row, column=5).number_format = "0.000"
        detail_sheet.cell(row=row, column=8).number_format = '0.0"%"'
    _set_column_widths(detail_sheet, {
        "A": 12,
        "B": 20,
        "C": 16,
        "D": 12,
        "E": 18,
        "F": 20,
        "G": 18,
        "H": 14,
        "I": 22,
        "J": 18,
        "K": 18,
    })

    metadata_sheet.append(["key", "value"])
    for cell in metadata_sheet[1]:
        cell.font = Font(bold=True)
    for key, value in _metadata_rows(report.metadata):
        metadata_sheet.append([key, value])
    metadata_sheet.freeze_panes = "A2"
    metadata_sheet.auto_filter.ref = metadata_sheet.dimensions
    _set_column_widths(metadata_sheet, {"A": 32, "B": 80})

    temporary_path = output_path.with_name(
        f"{output_path.stem}.tmp{output_path.suffix}"
    )
    temporary_path.unlink(missing_ok=True)
    try:
        workbook.save(temporary_path)
        verify_all_string_workbook(temporary_path)
        os.replace(temporary_path, output_path)
    finally:
        workbook.close()
        temporary_path.unlink(missing_ok=True)
    return output_path


def verify_all_string_workbook(path: Path) -> None:
    path = Path(path)
    if not path.is_file() or path.stat().st_size == 0:
        raise RuntimeError(f"Workbook was not created: {path}")
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        if workbook.sheetnames != SHEET_ORDER:
            raise RuntimeError(
                f"Unexpected sheet order: {workbook.sheetnames!r}"
            )
        recap_sheet = workbook["Rekap_Yield_kWh"]
        detail_sheet = workbook["Detail_Harian"]
        metadata_sheet = workbook["Metadata"]
        recap_headers = [cell.value for cell in recap_sheet[1]]
        if not recap_headers or recap_headers[0] != "date":
            raise RuntimeError(
                f"Unexpected recap headers: {recap_headers!r}"
            )
        recap_strings = recap_headers[1:]
        try:
            naturally_sorted = sorted(
                recap_strings,
                key=_natural_string_key,
            )
        except (TypeError, ValueError) as exc:
            raise RuntimeError(
                "Unexpected recap PV string headers."
            ) from exc
        if recap_strings != naturally_sorted:
            raise RuntimeError(
                "Unexpected recap PV string headers."
            )
        if len(set(recap_strings)) != len(recap_strings):
            raise RuntimeError(
                "Unexpected recap PV string headers."
            )
        detail_headers = [cell.value for cell in detail_sheet[1]]
        if detail_headers != DETAIL_COLUMNS:
            raise RuntimeError(
                f"Unexpected detail headers: {detail_headers!r}"
            )
        if [cell.value for cell in metadata_sheet[1]] != ["key", "value"]:
            raise RuntimeError("Unexpected Metadata headers.")
        metadata = {
            metadata_sheet.cell(row=row, column=1).value:
            metadata_sheet.cell(row=row, column=2).value
            for row in range(2, metadata_sheet.max_row + 1)
        }
        if any(
            isinstance(metadata_sheet.cell(row=row, column=2).value, str)
            and len(metadata_sheet.cell(row=row, column=2).value)
            > EXCEL_MAX_CELL_CHARACTERS
            for row in range(2, metadata_sheet.max_row + 1)
        ):
            raise RuntimeError("Metadata contains an oversized Excel cell.")
        requested_days = int(metadata.get("requested_days", 0))
        detected_strings = int(metadata.get("detected_string_count", 0))
        if requested_days < 1 or detected_strings < 1:
            raise RuntimeError("Metadata has invalid report dimensions.")
        try:
            start_date = pd.Timestamp(metadata["start_date"]).date()
            end_date = pd.Timestamp(metadata["end_date"]).date()
        except (KeyError, TypeError, ValueError) as exc:
            raise RuntimeError("Metadata has invalid report dates.") from exc
        expected_dates = [
            value.date()
            for value in pd.date_range(start_date, end_date, freq="D")
        ]
        if len(expected_dates) != requested_days:
            raise RuntimeError("Metadata has invalid report dates.")
        if recap_sheet.max_row != requested_days + 1:
            raise RuntimeError("Recap row count does not match requested days.")
        if recap_sheet.max_column != detected_strings + 1:
            raise RuntimeError(
                "Recap column count does not match detected strings."
            )
        if detail_sheet.max_row != requested_days * detected_strings + 1:
            raise RuntimeError(
                "Detail row count does not match date-string combinations."
            )
        try:
            recap_dates = [
                pd.Timestamp(
                    recap_sheet.cell(row=row, column=1).value
                ).date()
                for row in range(2, recap_sheet.max_row + 1)
            ]
        except (TypeError, ValueError) as exc:
            raise RuntimeError("Unexpected recap dates.") from exc
        if recap_dates != expected_dates:
            raise RuntimeError("Unexpected recap dates.")
        for offset, (expected_date, expected_string) in enumerate(
            (
                (expected_date, expected_string)
                for expected_date in expected_dates
                for expected_string in recap_strings
            ),
            start=2,
        ):
            try:
                actual_date = pd.Timestamp(
                    detail_sheet.cell(row=offset, column=1).value
                ).date()
            except (TypeError, ValueError) as exc:
                raise RuntimeError(
                    "Unexpected detail combinations."
                ) from exc
            actual_string = detail_sheet.cell(
                row=offset,
                column=2,
            ).value
            if (actual_date, actual_string) != (
                expected_date,
                expected_string,
            ):
                raise RuntimeError("Unexpected detail combinations.")
    finally:
        workbook.close()
