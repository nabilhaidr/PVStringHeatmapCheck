"""Analisis soiling berbasis RASIO specific yield (Yf) -- tanpa data POA.

Dua metode, keduanya bekerja langsung di atas workbook specific yield
(``specific_yield_*.xlsx`` hasil ``specific_yield_report``):

1. ``build_sibling_ratio`` -- Jalur B: rasio Yf tiap string terhadap MEDIAN
   TETANGGA se-inverter pada HARI YANG SAMA. Irradiance dan suhu identik
   antar string sehari itu, jadi keduanya tercoret pada rasio::

       Yf_string / Yf_tetangga = (SR_str x Yr x Tf) / (SR_tet x Yr x Tf)
                               = SR_string / SR_tetangga

   Hasilnya rasio soiling relatif tanpa perlu sensor irradiance sama sekali.
   Titik buta: soiling SERAGAM (semua string turun bersamaan) tidak terlihat
   -- untuk itu tetap perlu SRR/PR atau tes pre/post fisik.

2. ``build_yf_cleaning_impact`` -- pre/post cleaning pada Yf ternormalisasi.
   Yf tiap string dibagi median harian STRING KONTROL (string di WB yang
   sama yang TIDAK dibersihkan pada jendela campaign), lalu dibandingkan
   sebelum vs sesudah cleaning. Kontrol ini penting: kalau referensi ikut
   dibersihkan, uplift akan tampak nol.

Catatan: rasio TIDAK memberi loss absolut (%) -- untuk itu pakai SRR di
``pv_pipeline.m2a.soiling``. Kedua metode di sini menjawab "string mana"
dan "apakah cleaning berdampak", bukan "berapa loss plant".
"""
from __future__ import annotations

from dataclasses import dataclass
import os
from pathlib import Path
import re

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pandas as pd

from pv_pipeline.all_string_yield_report import (
    _append_dataframe,
    _metadata_rows,
    _natural_string_key,
    _reject_sensitive_metadata_keys,
    _set_column_widths,
)
from pv_pipeline.string_yield_report import (
    inventory_drive_folder,
    validate_drive_folder_url,
)


DETAIL_SHEET = "Detail_Harian"
RECAP_SHEET = "Rekap_SpecificYield"
YF_COLUMN = "specific_yield_kwh_per_kwp"
PV_STRING_RE = re.compile(r"^(WB(\d{1,2})-INV\d{1,2})-PV(\d{1,2})$", re.I)

# Default knob (semua bisa dioverride dari notebook).
DEFAULT_MIN_DAYS = 5
DEFAULT_MIN_SIBLINGS = 3
DEFAULT_MIN_RELATIVE_COVERAGE = 0.8
DEFAULT_DEAD_RATIO = 0.10
DEFAULT_CANDIDATE_DEFICIT_PCT = 5.0
DEFAULT_WINDOW_DAYS = 7
DEFAULT_GAP_DAYS = 7
DEFAULT_MIN_WINDOW_DAYS = 2
DEFAULT_MIN_REFERENCE_STRINGS = 5

SIBLING_COLUMNS = [
    "pv_string",
    "inverter_id",
    "pv_label",
    "wb",
    "n_days",
    "yf_median",
    "sibling_yf_median",
    "ratio_vs_inverter",
    "deficit_vs_inverter_pct",
    "ratio_vs_wb",
    "deficit_vs_wb_pct",
    "status",
    "rank",
]
SIBLING_SHEETS = ["Ranking_Sibling", "Rasio_Harian", "Metadata"]

CLEANING_COLUMNS = [
    "pv_string",
    "inverter_id",
    "pv_label",
    "st",
    "cleaning_start",
    "cleaning_end",
    "n_days_before",
    "n_days_after",
    "yf_before",
    "yf_after",
    "rel_before",
    "rel_after",
    "uplift_pct",
    "soiling_loss_pct",
    "reference_mode",
    "n_reference_strings",
    "status",
    "rank_uplift",
]
CAMPAIGN_COLUMNS = [
    "cleaning_start",
    "cleaning_end",
    "n_strings",
    "n_strings_evaluated",
    "median_uplift_pct",
    "median_soiling_loss_pct",
    "reference_mode",
]
CLEANING_SHEETS = ["Dampak_Cleaning", "Rekap_Campaign", "Metadata"]


@dataclass
class SiblingRatioData:
    ranking: pd.DataFrame
    daily_ratio: pd.DataFrame
    metadata: dict[str, object]


@dataclass
class YfCleaningImpactData:
    impact: pd.DataFrame
    campaigns: pd.DataFrame
    metadata: dict[str, object]


# --- Input -----------------------------------------------------------------
def split_pv_string(pv_string: str) -> tuple[str, str, int]:
    """``WB01-INV01-PV3`` -> (``WB01-INV01``, ``PV3``, 1) [inverter, pv, wb]."""
    match = PV_STRING_RE.match(str(pv_string).strip())
    if match is None:
        raise ValueError(f"Unexpected pv_string format: {pv_string!r}")
    return match.group(1), f"PV{int(match.group(3))}", int(match.group(2))


def load_specific_yield_long(path) -> pd.DataFrame:
    """Baca workbook specific yield -> long df.

    Sheet ``Detail_Harian`` dipakai bila ada (membawa ``coverage_pct`` untuk
    penyaringan data bolong); kalau tidak, ``Rekap_SpecificYield`` di-melt.
    Kolom hasil: date, pv_string, inverter_id, pv_label, wb, yf,
    coverage_pct (NaN bila tak tersedia).
    """
    path = Path(path)
    sheets = pd.ExcelFile(path).sheet_names
    if DETAIL_SHEET in sheets:
        frame = pd.read_excel(path, sheet_name=DETAIL_SHEET)
        if YF_COLUMN not in frame.columns:
            raise ValueError(
                f"Sheet {DETAIL_SHEET} tidak punya kolom {YF_COLUMN!r}."
            )
        out = pd.DataFrame({
            "date": pd.to_datetime(frame["date"], errors="coerce"),
            "pv_string": frame["pv_string"].astype(str),
            "yf": pd.to_numeric(frame[YF_COLUMN], errors="coerce"),
            "coverage_pct": pd.to_numeric(
                frame["coverage_pct"], errors="coerce",
            ) if "coverage_pct" in frame.columns else np.nan,
        })
    elif RECAP_SHEET in sheets:
        frame = pd.read_excel(path, sheet_name=RECAP_SHEET)
        melted = frame.melt(
            id_vars="pv_string", var_name="date", value_name="yf",
        )
        out = pd.DataFrame({
            "date": pd.to_datetime(melted["date"], errors="coerce"),
            "pv_string": melted["pv_string"].astype(str),
            "yf": pd.to_numeric(melted["yf"], errors="coerce"),
            "coverage_pct": np.nan,
        })
    else:
        raise ValueError(
            f"{path.name} tidak punya sheet {DETAIL_SHEET!r} maupun "
            f"{RECAP_SHEET!r}."
        )

    out = out.dropna(subset=["date"])
    parsed = out["pv_string"].map(split_pv_string)
    out["inverter_id"] = [value[0] for value in parsed]
    out["pv_label"] = [value[1] for value in parsed]
    out["wb"] = [value[2] for value in parsed]
    out["date"] = out["date"].dt.normalize()
    return out[[
        "date", "pv_string", "inverter_id", "pv_label", "wb",
        "yf", "coverage_pct",
    ]].reset_index(drop=True)


def _filter_relative_coverage(
    long_df: pd.DataFrame,
    min_relative_coverage: float,
) -> tuple[pd.DataFrame, int]:
    """Buang string-day yang cakupan sampelnya jauh di bawah tetangganya.

    Ambang absolut tidak bisa dipakai: coverage_pct dihitung terhadap 288
    slot 24 jam sedangkan siang hanya ~12 jam, jadi ~50% itu NORMAL. Yang
    menandakan data bolong adalah cakupan yang jauh lebih rendah daripada
    tetangga se-inverter pada hari yang sama.
    """
    if min_relative_coverage <= 0 or long_df["coverage_pct"].isna().all():
        return long_df, 0
    reference = long_df.groupby(["date", "inverter_id"])["coverage_pct"].transform(
        "median"
    )
    keep = (
        long_df["coverage_pct"].isna()
        | reference.isna()
        | (reference <= 0)
        | (long_df["coverage_pct"] >= reference * float(min_relative_coverage))
    )
    return long_df.loc[keep].copy(), int((~keep).sum())


# --- Jalur B: rasio string vs tetangga -------------------------------------
def build_sibling_ratio(
    long_df: pd.DataFrame,
    *,
    min_days: int = DEFAULT_MIN_DAYS,
    min_siblings: int = DEFAULT_MIN_SIBLINGS,
    min_relative_coverage: float = DEFAULT_MIN_RELATIVE_COVERAGE,
    dead_ratio: float = DEFAULT_DEAD_RATIO,
    candidate_deficit_pct: float = DEFAULT_CANDIDATE_DEFICIT_PCT,
    source_metadata: dict | None = None,
) -> SiblingRatioData:
    """Ranking string berdasarkan rasio Yf harian terhadap tetangganya.

    Rasio dihitung PER HARI lebih dulu (``yf / median tetangga hari itu``)
    baru diagregasi dengan median lintas hari -- urutan ini penting supaya
    cuaca benar-benar tercoret dan hari yang datanya tidak lengkap tidak
    membiaskan hasil.
    """
    if long_df.empty:
        raise ValueError("long_df kosong; tidak ada data untuk dibandingkan.")

    frame, dropped_coverage = _filter_relative_coverage(
        long_df, min_relative_coverage,
    )
    frame = frame.dropna(subset=["yf"])
    if frame.empty:
        raise ValueError("Tidak ada nilai Yf valid setelah penyaringan.")

    grouped_inverter = frame.groupby(["date", "inverter_id"])["yf"]
    frame["sibling_yf"] = grouped_inverter.transform("median")
    frame["n_siblings"] = grouped_inverter.transform("size")
    frame["wb_yf"] = frame.groupby(["date", "wb"])["yf"].transform("median")

    frame = frame[frame["n_siblings"] >= min_siblings]
    if frame.empty:
        raise ValueError(
            f"Tidak ada inverter dengan >= {min_siblings} string per hari."
        )

    frame["ratio_inverter"] = frame["yf"] / frame["sibling_yf"].replace(0.0, np.nan)
    frame["ratio_wb"] = frame["yf"] / frame["wb_yf"].replace(0.0, np.nan)

    aggregated = frame.groupby(
        ["pv_string", "inverter_id", "pv_label", "wb"], as_index=False,
    ).agg(
        n_days=("date", "nunique"),
        yf_median=("yf", "median"),
        sibling_yf_median=("sibling_yf", "median"),
        ratio_vs_inverter=("ratio_inverter", "median"),
        ratio_vs_wb=("ratio_wb", "median"),
    )
    aggregated = aggregated[aggregated["n_days"] >= min_days]
    if aggregated.empty:
        raise ValueError(
            f"Tidak ada string dengan >= {min_days} hari data valid."
        )

    aggregated["deficit_vs_inverter_pct"] = (
        1.0 - aggregated["ratio_vs_inverter"]
    ) * 100.0
    aggregated["deficit_vs_wb_pct"] = (1.0 - aggregated["ratio_vs_wb"]) * 100.0
    aggregated["status"] = np.where(
        aggregated["ratio_vs_inverter"] < dead_ratio,
        "DEAD_OR_OFFLINE",
        np.where(
            aggregated["deficit_vs_inverter_pct"] >= candidate_deficit_pct,
            "CLEANING_CANDIDATE",
            "NORMAL",
        ),
    )
    # DEAD_OR_OFFLINE bukan kandidat cleaning -> dikeluarkan dari ranking
    # supaya prioritas teratas benar-benar string kotor, bukan string mati.
    rank_basis = aggregated["deficit_vs_inverter_pct"].where(
        aggregated["status"] != "DEAD_OR_OFFLINE"
    )
    aggregated["rank"] = rank_basis.rank(
        ascending=False, method="min",
    ).astype("Int64")
    ranking = aggregated.sort_values(
        ["rank", "pv_string"], na_position="last", ignore_index=True,
    )[SIBLING_COLUMNS]

    daily_ratio = frame.pivot_table(
        index="pv_string", columns="date", values="ratio_inverter",
    )
    daily_ratio = daily_ratio.reindex(
        sorted(daily_ratio.index, key=_natural_string_key)
    )
    daily_ratio.columns = [
        pd.Timestamp(value).date() for value in daily_ratio.columns
    ]
    daily_ratio.columns.name = None
    daily_ratio = daily_ratio.reset_index()

    metadata = dict(source_metadata or {})
    metadata.update({
        "method": "sibling_ratio_yf (Jalur B, tanpa POA)",
        "ratio_formula": (
            "median_harian(yf_string / median_yf_tetangga_se_inverter)"
        ),
        "blind_spot": (
            "Soiling seragam se-inverter tidak terdeteksi; "
            "pakai SRR/PR atau tes pre/post untuk itu."
        ),
        "min_days": min_days,
        "min_siblings": min_siblings,
        "min_relative_coverage": min_relative_coverage,
        "dead_ratio_threshold": dead_ratio,
        "candidate_deficit_pct": candidate_deficit_pct,
        "dropped_low_coverage_rows": dropped_coverage,
        "string_count": int(len(ranking)),
        "day_count": int(frame["date"].nunique()),
        "status_counts": ranking["status"].value_counts().to_dict(),
    })
    return SiblingRatioData(
        ranking=ranking, daily_ratio=daily_ratio, metadata=metadata,
    )


# --- Pre/post cleaning pada Yf ternormalisasi ------------------------------
def _campaigns(dates, gap_days: int) -> list[list[pd.Timestamp]]:
    ordered = sorted(pd.Timestamp(value) for value in dates)
    groups: list[list[pd.Timestamp]] = []
    current = [ordered[0]]
    for value in ordered[1:]:
        if (value - current[-1]).days <= gap_days:
            current.append(value)
        else:
            groups.append(current)
            current = [value]
    groups.append(current)
    return groups


def cleaning_events_to_pv_string(events: pd.DataFrame) -> pd.DataFrame:
    """Rekap cleaning (``load_cleaning_report``) -> [date, pv_string, st].

    Baris tanpa mapping ST->PV (pv NaN, khas WB03-WB10 yang tidak ada di
    DC cable list) dibuang karena tidak bisa dicocokkan ke data yield.
    """
    empty = pd.DataFrame(columns=["date", "pv_string", "inverter_id", "st"])
    if events is None or events.empty or "pv" not in events.columns:
        return empty
    frame = events.dropna(subset=["pv"]).copy()
    if frame.empty:
        return empty
    frame["date"] = pd.to_datetime(frame["date"], errors="coerce").dt.normalize()
    frame = frame.dropna(subset=["date"])
    frame["inverter_id"] = frame["inverter_id"].astype(str).str.upper()
    frame["pv_string"] = (
        frame["inverter_id"] + "-PV" + frame["pv"].astype(int).astype(str)
    )
    return frame[["date", "pv_string", "inverter_id", "st"]].reset_index(drop=True)


def build_yf_cleaning_impact(
    long_df: pd.DataFrame,
    events: pd.DataFrame,
    *,
    window_days: int = DEFAULT_WINDOW_DAYS,
    gap_days: int = DEFAULT_GAP_DAYS,
    min_window_days: int = DEFAULT_MIN_WINDOW_DAYS,
    min_reference_strings: int = DEFAULT_MIN_REFERENCE_STRINGS,
    min_relative_coverage: float = DEFAULT_MIN_RELATIVE_COVERAGE,
    source_metadata: dict | None = None,
) -> YfCleaningImpactData:
    """Uplift pre/post cleaning memakai Yf ternormalisasi string kontrol.

    Untuk tiap campaign cleaning sebuah string, Yf hariannya dibagi median
    harian STRING KONTROL -- string di WB yang sama yang tidak dibersihkan
    dalam jendela campaign. Cuaca dan drift musiman tercoret, dan karena
    kontrol tidak ikut dibersihkan uplift tidak "hilang" seperti bila
    referensinya ikut bersih.
    """
    if long_df.empty:
        raise ValueError("long_df kosong; tidak ada data untuk dibandingkan.")
    cleaned = cleaning_events_to_pv_string(events)
    if cleaned.empty:
        raise ValueError(
            "Tidak ada event cleaning dengan mapping ST->PV; cek DC cable list."
        )

    frame, dropped_coverage = _filter_relative_coverage(
        long_df, min_relative_coverage,
    )
    frame = frame.dropna(subset=["yf"])
    yf_by_string = {
        key: group.set_index("date")["yf"].sort_index()
        for key, group in frame.groupby("pv_string")
    }
    wb_by_string = dict(
        frame[["pv_string", "wb"]].drop_duplicates().itertuples(index=False)
    )
    window = pd.Timedelta(days=window_days)

    rows = []
    skipped_no_yield = 0
    for pv_string, group in cleaned.groupby("pv_string"):
        series = yf_by_string.get(pv_string)
        if series is None or series.empty:
            skipped_no_yield += 1
            continue
        inverter_id, pv_label, wb = split_pv_string(pv_string)
        st_value = group["st"].iloc[0] if "st" in group.columns else None
        for campaign in _campaigns(group["date"].unique(), gap_days):
            start, end = campaign[0], campaign[-1]
            span_start, span_end = start - window, end + window

            # String kontrol: tidak punya event cleaning di jendela campaign.
            cleaned_in_span = set(
                cleaned.loc[
                    (cleaned["date"] >= span_start)
                    & (cleaned["date"] <= span_end),
                    "pv_string",
                ]
            )
            reference_wb = [
                name for name, value in wb_by_string.items()
                if value == wb and name not in cleaned_in_span
            ]
            reference_site = [
                name for name in yf_by_string if name not in cleaned_in_span
            ]
            if len(reference_wb) >= min_reference_strings:
                reference, mode = reference_wb, "WB_UNCLEANED"
            elif len(reference_site) >= min_reference_strings:
                reference, mode = reference_site, "SITE_UNCLEANED"
            else:
                reference, mode = [], "RAW_NO_REFERENCE"

            if reference:
                reference_daily = (
                    frame[frame["pv_string"].isin(reference)]
                    .groupby("date")["yf"].median()
                )
                relative = series / reference_daily.reindex(
                    series.index
                ).replace(0.0, np.nan)
            else:
                relative = series.copy()

            before_mask = (series.index >= span_start) & (series.index < start)
            after_mask = (series.index > end) & (series.index <= span_end)
            rel_before = relative[before_mask].dropna()
            rel_after = relative[after_mask].dropna()
            if len(rel_before) < min_window_days or len(rel_after) < min_window_days:
                continue
            mean_before = float(rel_before.mean())
            mean_after = float(rel_after.mean())
            if not (np.isfinite(mean_before) and np.isfinite(mean_after)):
                continue
            if mean_before <= 0 or mean_after <= 0:
                continue
            rows.append({
                "pv_string": pv_string,
                "inverter_id": inverter_id,
                "pv_label": pv_label,
                "st": st_value,
                "cleaning_start": start,
                "cleaning_end": end,
                "n_days_before": int(len(rel_before)),
                "n_days_after": int(len(rel_after)),
                "yf_before": float(series[before_mask].dropna().mean()),
                "yf_after": float(series[after_mask].dropna().mean()),
                "rel_before": mean_before,
                "rel_after": mean_after,
                "uplift_pct": (mean_after - mean_before) / mean_before * 100.0,
                "soiling_loss_pct": (
                    (mean_after - mean_before) / mean_after * 100.0
                ),
                "reference_mode": mode,
                "n_reference_strings": len(reference),
            })

    impact = pd.DataFrame(rows, columns=[
        column for column in CLEANING_COLUMNS
        if column not in {"status", "rank_uplift"}
    ])
    if impact.empty:
        raise ValueError(
            "Tidak ada campaign dengan data Yf cukup di kedua sisi "
            f"(butuh >= {min_window_days} hari sebelum dan sesudah)."
        )
    impact["status"] = np.where(
        impact["uplift_pct"] >= 2.0,
        "RECOVERED",
        np.where(impact["uplift_pct"] <= -2.0, "WORSE", "NO_CHANGE"),
    )
    impact["rank_uplift"] = impact["uplift_pct"].rank(
        ascending=False, method="min",
    ).astype(int)
    impact = impact.sort_values(
        "uplift_pct", ascending=False, ignore_index=True,
    )[CLEANING_COLUMNS]

    campaigns = impact.groupby(
        ["cleaning_start", "cleaning_end"], as_index=False,
    ).agg(
        n_strings_evaluated=("pv_string", "nunique"),
        median_uplift_pct=("uplift_pct", "median"),
        median_soiling_loss_pct=("soiling_loss_pct", "median"),
        reference_mode=("reference_mode", lambda values: values.mode().iat[0]),
    )
    campaigns["n_strings"] = campaigns["n_strings_evaluated"]
    campaigns = campaigns.sort_values(
        "cleaning_start", ignore_index=True,
    )[CAMPAIGN_COLUMNS]

    metadata = dict(source_metadata or {})
    metadata.update({
        "method": "pre_post_cleaning_yf_ratio (tanpa POA)",
        "normalisation": (
            "yf_string / median_harian(yf string kontrol WB yang sama "
            "yang TIDAK dibersihkan pada jendela campaign)"
        ),
        "uplift_formula": "(rel_after - rel_before) / rel_before * 100",
        "soiling_loss_formula": "(rel_after - rel_before) / rel_after * 100",
        "window_days": window_days,
        "gap_days": gap_days,
        "min_window_days": min_window_days,
        "min_reference_strings": min_reference_strings,
        "min_relative_coverage": min_relative_coverage,
        "dropped_low_coverage_rows": dropped_coverage,
        "cleaning_events": int(len(cleaned)),
        "cleaned_strings": int(cleaned["pv_string"].nunique()),
        "strings_without_yield_data": skipped_no_yield,
        "evaluated_rows": int(len(impact)),
        "campaign_count": int(len(campaigns)),
        "status_counts": impact["status"].value_counts().to_dict(),
    })
    return YfCleaningImpactData(
        impact=impact, campaigns=campaigns, metadata=metadata,
    )


# --- Drive helper ----------------------------------------------------------
def download_drive_file(
    folder_url: str,
    destination,
    *,
    basename: str = "",
    pattern: str = "",
    pick: str = "last",
):
    """Unduh SATU file dari folder Drive publik berdasarkan nama.

    ``basename`` = nama persis; ``pattern`` = regex atas nama file. Bila
    beberapa cocok, ``pick`` menentukan yang diambil setelah diurutkan nama:
    ``"last"`` (default, mis. periode terbaru) atau ``"first"``.
    """
    import gdown

    validate_drive_folder_url(folder_url)
    items = inventory_drive_folder(folder_url)
    if basename:
        matches = [item for item in items if Path(item.path).name == basename]
    elif pattern:
        regex = re.compile(pattern, re.I)
        matches = [
            item for item in items if regex.search(Path(item.path).name)
        ]
    else:
        raise ValueError("Sebutkan basename atau pattern.")
    if not matches:
        available = sorted(Path(item.path).name for item in items)[:10]
        raise FileNotFoundError(
            f"Tidak ada file cocok di folder Drive "
            f"({basename or pattern!r}); contoh isi folder: {available}"
        )
    matches.sort(key=lambda item: Path(item.path).name)
    chosen = matches[-1] if pick == "last" else matches[0]

    destination = Path(destination)
    destination.mkdir(parents=True, exist_ok=True)
    output = destination / Path(chosen.path).name
    gdown.download(url=chosen.url, output=str(output), quiet=False)
    if not output.is_file():
        raise RuntimeError(f"gdown gagal membuat {output.name}")
    return output


# --- Output ----------------------------------------------------------------
def build_sibling_ratio_output_path(output_dir, start, end) -> Path:
    return Path(output_dir) / (
        f"sibling_ratio_{start:%Y%m%d}_{end:%Y%m%d}.xlsx"
    )


def build_cleaning_impact_output_path(output_dir, start, end) -> Path:
    return Path(output_dir) / (
        f"cleaning_impact_yf_{start:%Y%m%d}_{end:%Y%m%d}.xlsx"
    )


def _write_workbook(
    output_path: Path,
    sheets,
    metadata: dict,
    *,
    expected_sheets,
    number_formats=None,
    widths=None,
) -> Path:
    _reject_sensitive_metadata_keys(metadata)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, frame in sheets:
        sheet = workbook.create_sheet(name)
        _append_dataframe(sheet, frame)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column, fmt in (number_formats or {}).get(name, {}).items():
            if column > sheet.max_column:
                continue
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row=row, column=column).number_format = fmt
        sheet_widths = (widths or {}).get(name)
        if sheet_widths is None:
            sheet_widths = {
                get_column_letter(column): 16
                for column in range(1, sheet.max_column + 1)
            }
        _set_column_widths(sheet, sheet_widths)

    metadata_sheet = workbook.create_sheet("Metadata")
    metadata_sheet.append(["key", "value"])
    for cell in metadata_sheet[1]:
        cell.font = Font(bold=True)
    for key, value in _metadata_rows(metadata):
        metadata_sheet.append([key, value])
    metadata_sheet.freeze_panes = "A2"
    metadata_sheet.auto_filter.ref = metadata_sheet.dimensions
    _set_column_widths(metadata_sheet, {"A": 32, "B": 80})

    temporary_path = output_path.with_name(
        f"{output_path.stem}.tmp{output_path.suffix}"
    )
    temporary_path.unlink(missing_ok=True)
    try:
        workbook.save(temporary_path)
        _verify_workbook(temporary_path, expected_sheets)
        os.replace(temporary_path, output_path)
    finally:
        workbook.close()
        temporary_path.unlink(missing_ok=True)
    return output_path


def _verify_workbook(path: Path, expected_sheets) -> None:
    path = Path(path)
    if not path.is_file() or path.stat().st_size == 0:
        raise RuntimeError(f"Workbook was not created: {path}")
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        if workbook.sheetnames != list(expected_sheets):
            raise RuntimeError(
                f"Unexpected sheet order: {workbook.sheetnames!r}"
            )
        if [cell.value for cell in workbook["Metadata"][1]] != ["key", "value"]:
            raise RuntimeError("Unexpected Metadata headers.")
    finally:
        workbook.close()


def write_sibling_ratio_workbook(
    output_path: Path,
    report: SiblingRatioData,
) -> Path:
    if list(report.ranking.columns) != SIBLING_COLUMNS:
        raise ValueError(
            f"Unexpected ranking columns: {list(report.ranking.columns)!r}"
        )
    return _write_workbook(
        output_path,
        [
            (SIBLING_SHEETS[0], report.ranking),
            (SIBLING_SHEETS[1], report.daily_ratio),
        ],
        report.metadata,
        expected_sheets=SIBLING_SHEETS,
        number_formats={
            SIBLING_SHEETS[0]: {
                6: "0.000", 7: "0.000", 8: "0.0000", 9: '0.0"%"',
                10: "0.0000", 11: '0.0"%"',
            },
        },
        widths={
            SIBLING_SHEETS[0]: {
                "A": 20, "B": 16, "C": 10, "D": 6, "E": 9, "F": 12,
                "G": 18, "H": 18, "I": 22, "J": 14, "K": 18, "L": 20, "M": 8,
            },
        },
    )


def verify_sibling_ratio_workbook(path: Path) -> None:
    _verify_workbook(Path(path), SIBLING_SHEETS)


def write_cleaning_impact_workbook(
    output_path: Path,
    report: YfCleaningImpactData,
) -> Path:
    if list(report.impact.columns) != CLEANING_COLUMNS:
        raise ValueError(
            f"Unexpected impact columns: {list(report.impact.columns)!r}"
        )
    return _write_workbook(
        output_path,
        [
            (CLEANING_SHEETS[0], report.impact),
            (CLEANING_SHEETS[1], report.campaigns),
        ],
        report.metadata,
        expected_sheets=CLEANING_SHEETS,
        number_formats={
            CLEANING_SHEETS[0]: {
                5: "yyyy-mm-dd", 6: "yyyy-mm-dd",
                9: "0.000", 10: "0.000", 11: "0.0000", 12: "0.0000",
                13: '0.0"%"', 14: '0.0"%"',
            },
            CLEANING_SHEETS[1]: {
                1: "yyyy-mm-dd", 2: "yyyy-mm-dd", 5: '0.0"%"', 6: '0.0"%"',
            },
        },
        widths={
            CLEANING_SHEETS[0]: {
                "A": 20, "B": 16, "C": 10, "D": 6, "E": 14, "F": 14,
                "G": 14, "H": 14, "I": 12, "J": 12, "K": 12, "L": 12,
                "M": 12, "N": 18, "O": 20, "P": 20, "Q": 14, "R": 12,
            },
        },
    )


def verify_cleaning_impact_workbook(path: Path) -> None:
    _verify_workbook(Path(path), CLEANING_SHEETS)
