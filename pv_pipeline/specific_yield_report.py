"""Specific yield harian per PV string (IEC 61724-1).

Specific yield = energi harian string dibagi kapasitas DC terpasang string
itu, satuan kWh/kWp/hari -- setara "equivalent peak sun hours" karena
1 kWh/kWp = 1 jam pada irradiance 1000 W/m^2. Metrik ini menormalkan beda
kapasitas fase 1 (24 modul/string) vs fase 2 (26 modul/string) sehingga
antar string bisa dibandingkan langsung.

Energi harian TIDAK dihitung ulang di sini: modul ini menerima hasil
``build_all_string_daily_yield`` (integrasi ``sum(power_kw * 5/60)``) lalu
membaginya dengan kapasitas per string, sehingga definisi energi tetap satu
sumber kebenaran di ``all_string_yield_report``.

Catatan interpretasi: specific yield BUKAN Performance Ratio. Nilainya ikut
naik-turun mengikuti cuaca karena belum dinormalkan terhadap insolasi POA
aktual. Untuk membandingkan antar string pada hari yang sama metrik ini
memadai; untuk membandingkan antar hari, bagi lagi dengan insolasi harian.
"""
from __future__ import annotations

from dataclasses import dataclass
import os
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pandas as pd

from pv_pipeline.all_string_yield_report import (
    EXCEL_MAX_CELL_CHARACTERS,
    AllStringYieldData,
    _append_dataframe,
    _metadata_rows,
    _natural_string_key,
    _reject_sensitive_metadata_keys,
    _set_column_widths,
)


# --- Kapasitas string ---------------------------------------------------
# Fase 1 (WB01-WB02): 24 modul/string; fase 2 (WB03-WB10): 26 modul/string.
# Sama dengan MODULES_PER_STRING_BY_WB di pv_pipeline/m2a/soiling.py.
MODULE_WP: float = 625.0
MODULES_PER_STRING_BY_WB: dict[int, int] = {1: 24, 2: 24}
DEFAULT_MODULES_PER_STRING: int = 26

PV_STRING_RE = re.compile(r"^WB(\d{1,2})-INV(\d{1,2})-PV(\d{1,2})$", re.I)

DETAIL_COLUMNS = [
    "date",
    "pv_string",
    "inverter_id",
    "pv_label",
    "string_yield_kwh",
    "capacity_kwp",
    "specific_yield_kwh_per_kwp",
    "valid_power_samples",
    "expected_samples",
    "coverage_pct",
    "missing_power_samples",
    "source_csv",
    "status",
]
SHEET_ORDER = ["Rekap_SpecificYield", "Detail_Harian", "Metadata"]


@dataclass
class SpecificYieldData:
    summary: pd.DataFrame
    daily: pd.DataFrame
    metadata: dict[str, object]


def modules_per_string(pv_string: str) -> int:
    """Jumlah modul per string dari nomor WB pada ``WBxx-INVyy-PVn``."""
    match = PV_STRING_RE.match(str(pv_string).strip())
    if match is None:
        raise ValueError(f"Unexpected pv_string format: {pv_string!r}")
    wb_number = int(match.group(1))
    return MODULES_PER_STRING_BY_WB.get(wb_number, DEFAULT_MODULES_PER_STRING)


def string_capacity_kwp(
    pv_string: str,
    module_wp: float = MODULE_WP,
) -> float:
    """Kapasitas DC satu string (kWp) = modul/string x Wp modul / 1000."""
    return modules_per_string(pv_string) * float(module_wp) / 1000.0


def build_specific_yield(
    report: AllStringYieldData,
    *,
    module_wp: float = MODULE_WP,
    empty_pv_map: dict[str, list[int]] | None = None,
) -> SpecificYieldData:
    """Turunkan specific yield harian dari hasil all-string daily yield.

    ``empty_pv_map`` (opsional, dari ``config/strings.yaml``): slot PV yang
    memang kosong by design dibuang dari kedua sheet supaya rekap tidak
    berisi baris yang selalu kosong.
    """
    daily = report.daily.copy()
    if daily.empty:
        raise ValueError("report.daily is empty; nothing to normalise.")

    excluded = 0
    if empty_pv_map:
        empty_pairs = {
            (str(inverter_id).upper(), int(pv_number))
            for inverter_id, pv_numbers in empty_pv_map.items()
            for pv_number in (pv_numbers or [])
        }
        if empty_pairs:
            pv_numbers = pd.to_numeric(
                daily["pv_label"].str.extract(r"(\d+)", expand=False),
                errors="coerce",
            )
            is_empty_slot = [
                (inverter_id, pv_number) in empty_pairs
                for inverter_id, pv_number in zip(
                    daily["inverter_id"].str.upper(),
                    pv_numbers.astype("Int64"),
                )
            ]
            drop_mask = pd.Series(is_empty_slot, index=daily.index)
            excluded = int(daily.loc[drop_mask, "pv_string"].nunique())
            daily = daily.loc[~drop_mask].copy()
            if daily.empty:
                raise ValueError(
                    "All strings were excluded as empty slots; "
                    "check empty_pv_map."
                )

    daily["capacity_kwp"] = daily["pv_string"].map(
        lambda value: string_capacity_kwp(value, module_wp)
    )
    daily["specific_yield_kwh_per_kwp"] = (
        daily["string_yield_kwh"] / daily["capacity_kwp"]
    )
    daily = daily[DETAIL_COLUMNS]

    strings = sorted(daily["pv_string"].unique(), key=_natural_string_key)
    dates = sorted(daily["date"].unique())
    summary = daily.pivot(
        index="pv_string",
        columns="date",
        values="specific_yield_kwh_per_kwp",
    ).reindex(index=strings, columns=dates)
    summary.columns.name = None
    summary = summary.reset_index()

    metadata = dict(report.metadata)
    metadata.update({
        "specific_yield_formula": (
            "string_yield_kwh / (modules_per_string * module_wp / 1000)"
        ),
        "specific_yield_unit": "kWh/kWp/day (equivalent peak sun hours)",
        "standard": "IEC 61724-1",
        "module_wp": float(module_wp),
        "modules_per_string_wb01_wb02": MODULES_PER_STRING_BY_WB[1],
        "modules_per_string_other_wb": DEFAULT_MODULES_PER_STRING,
        "excluded_empty_slot_strings": excluded,
        "specific_yield_string_count": len(strings),
    })
    return SpecificYieldData(summary=summary, daily=daily, metadata=metadata)


def build_specific_yield_output_path(output_dir, start, end) -> Path:
    return Path(output_dir) / (
        f"specific_yield_{start:%Y%m%d}_{end:%Y%m%d}.xlsx"
    )


def write_specific_yield_workbook(
    output_path: Path,
    report: SpecificYieldData,
) -> Path:
    if list(report.daily.columns) != DETAIL_COLUMNS:
        raise ValueError(
            f"Unexpected detail columns: {list(report.daily.columns)!r}"
        )
    summary_columns = list(report.summary.columns)
    if not summary_columns or summary_columns[0] != "pv_string":
        raise ValueError(f"Unexpected summary columns: {summary_columns!r}")
    _reject_sensitive_metadata_keys(report.metadata)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    recap_sheet = workbook.create_sheet(SHEET_ORDER[0])
    detail_sheet = workbook.create_sheet(SHEET_ORDER[1])
    metadata_sheet = workbook.create_sheet(SHEET_ORDER[2])

    _append_dataframe(recap_sheet, report.summary)
    recap_sheet.freeze_panes = "B2"
    recap_sheet.auto_filter.ref = recap_sheet.dimensions
    for column in range(2, recap_sheet.max_column + 1):
        recap_sheet.cell(row=1, column=column).number_format = "yyyy-mm-dd"
        for row in range(2, recap_sheet.max_row + 1):
            recap_sheet.cell(row=row, column=column).number_format = "0.000"
    recap_widths = {"A": 22}
    for column in range(2, recap_sheet.max_column + 1):
        recap_widths[get_column_letter(column)] = 12
    _set_column_widths(recap_sheet, recap_widths)

    _append_dataframe(detail_sheet, report.daily)
    detail_sheet.freeze_panes = "A2"
    detail_sheet.auto_filter.ref = detail_sheet.dimensions
    for row in range(2, detail_sheet.max_row + 1):
        detail_sheet.cell(row=row, column=1).number_format = "yyyy-mm-dd"
        for column in (5, 6, 7):
            detail_sheet.cell(row=row, column=column).number_format = "0.000"
        detail_sheet.cell(row=row, column=10).number_format = '0.0"%"'
    _set_column_widths(detail_sheet, {
        "A": 12, "B": 20, "C": 16, "D": 12, "E": 18, "F": 14, "G": 26,
        "H": 20, "I": 18, "J": 14, "K": 22, "L": 18, "M": 18,
    })

    metadata_sheet.append(["key", "value"])
    for cell in metadata_sheet[1]:
        cell.font = Font(bold=True)
    for key, value in _metadata_rows(report.metadata):
        metadata_sheet.append([key, value])
    metadata_sheet.freeze_panes = "A2"
    metadata_sheet.auto_filter.ref = metadata_sheet.dimensions
    _set_column_widths(metadata_sheet, {"A": 34, "B": 80})

    temporary_path = output_path.with_name(
        f"{output_path.stem}.tmp{output_path.suffix}"
    )
    temporary_path.unlink(missing_ok=True)
    try:
        workbook.save(temporary_path)
        verify_specific_yield_workbook(temporary_path)
        os.replace(temporary_path, output_path)
    finally:
        workbook.close()
        temporary_path.unlink(missing_ok=True)
    return output_path


def verify_specific_yield_workbook(path: Path) -> None:
    path = Path(path)
    if not path.is_file() or path.stat().st_size == 0:
        raise RuntimeError(f"Workbook was not created: {path}")
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        if workbook.sheetnames != SHEET_ORDER:
            raise RuntimeError(
                f"Unexpected sheet order: {workbook.sheetnames!r}"
            )
        recap_sheet = workbook[SHEET_ORDER[0]]
        detail_sheet = workbook[SHEET_ORDER[1]]
        metadata_sheet = workbook[SHEET_ORDER[2]]

        recap_headers = [cell.value for cell in recap_sheet[1]]
        if not recap_headers or recap_headers[0] != "pv_string":
            raise RuntimeError(f"Unexpected recap headers: {recap_headers!r}")
        recap_strings = [
            recap_sheet.cell(row=row, column=1).value
            for row in range(2, recap_sheet.max_row + 1)
        ]
        try:
            naturally_sorted = sorted(recap_strings, key=_natural_string_key)
        except (TypeError, ValueError) as exc:
            raise RuntimeError("Unexpected recap PV string rows.") from exc
        if recap_strings != naturally_sorted:
            raise RuntimeError("Unexpected recap PV string rows.")
        if len(set(recap_strings)) != len(recap_strings):
            raise RuntimeError("Unexpected recap PV string rows.")

        detail_headers = [cell.value for cell in detail_sheet[1]]
        if detail_headers != DETAIL_COLUMNS:
            raise RuntimeError(
                f"Unexpected detail headers: {detail_headers!r}"
            )
        if [cell.value for cell in metadata_sheet[1]] != ["key", "value"]:
            raise RuntimeError("Unexpected Metadata headers.")
        if any(
            isinstance(metadata_sheet.cell(row=row, column=2).value, str)
            and len(metadata_sheet.cell(row=row, column=2).value)
            > EXCEL_MAX_CELL_CHARACTERS
            for row in range(2, metadata_sheet.max_row + 1)
        ):
            raise RuntimeError("Metadata contains an oversized Excel cell.")
    finally:
        workbook.close()
