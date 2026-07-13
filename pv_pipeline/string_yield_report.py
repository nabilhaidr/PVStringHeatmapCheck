from __future__ import annotations

import json
import re
import subprocess
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Sequence
from urllib.parse import urlparse

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import yaml

from pv_pipeline.physics import compute_active_power_integration_kwh
from pv_pipeline.poa import PyranometerLoader


STRING_RE = re.compile(r"^(WB\d{2})-(INV\d{2})-PV(\d{1,2})$", re.I)
DAILY_COLUMNS = [
    "date", "string_yield_kwh", "valid_power_samples", "expected_samples",
    "coverage_pct", "missing_power_samples", "poa_valid_samples", "source_csv",
    "status",
]
FIVE_MINUTE_COLUMNS = [
    "timestamp", "inverter_id", "pv_string", "power_kw", "poa_wm2",
    "source_csv", "poa_source", "data_status",
]
SHEET_ORDER = ["Ringkasan_Harian", "Data_5Menit", "Grafik", "Metadata"]
SENSITIVE_METADATA_TERMS = ("token", "cookie", "secret", "credential", "password")


@dataclass(frozen=True)
class StringSelection:
    canonical: str
    wb_id: str
    inverter_id: str
    pv_number: int
    pv_label: str


@dataclass(frozen=True)
class DriveItem:
    url: str
    path: str


@dataclass
class SourceManifest:
    csv_by_date: dict[date, DriveItem]
    poa_by_year: dict[int, DriveItem]
    missing_csv_dates: list[date]
    missing_poa_years: list[int]
    url_csv: str = ""
    url_poa: str = ""
    csv_inventory_count: int = 0
    poa_inventory_count: int = 0


@dataclass
class DownloadedInputs:
    csv_by_date: dict[date, Path]
    poa_by_year: dict[int, Path]
    download_errors: dict[str, str]


@dataclass
class ReportData:
    daily: pd.DataFrame
    five_minute: pd.DataFrame
    metadata: dict[str, object]


def parse_string_selection(value: str) -> StringSelection:
    match = STRING_RE.fullmatch(str(value).strip())
    if not match:
        raise ValueError("PV_STRING must match WBxx-INVxx-PVxx.")
    wb_id, inv_part, pv_text = match.groups()
    pv_number = int(pv_text)
    if not 1 <= pv_number <= 28:
        raise ValueError("PV number must be in range 1..28.")
    wb_id, inv_part = wb_id.upper(), inv_part.upper()
    inverter_id = f"{wb_id}-{inv_part}"
    return StringSelection(
        canonical=f"{inverter_id}-PV{pv_number}",
        wb_id=wb_id,
        inverter_id=inverter_id,
        pv_number=pv_number,
        pv_label=f"PV{pv_number}",
    )


def parse_date_range(start: str, end: str) -> pd.DatetimeIndex:
    try:
        start_ts = pd.Timestamp(datetime.strptime(start, "%Y-%m-%d"))
        end_ts = pd.Timestamp(datetime.strptime(end, "%Y-%m-%d"))
    except (TypeError, ValueError) as exc:
        raise ValueError("START_DATE and END_DATE must use yyyy-mm-dd.") from exc
    if start_ts > end_ts:
        raise ValueError("START_DATE must be on or before END_DATE.")
    return pd.date_range(start_ts.normalize(), end_ts.normalize(), freq="D")


def validate_drive_folder_url(url: str) -> str:
    parsed = urlparse(str(url).strip())
    if parsed.scheme != "https" or parsed.netloc != "drive.google.com" or "/drive/folders/" not in parsed.path:
        raise ValueError(f"Expected a public Google Drive folder URL, got {url!r}.")
    return str(url).strip()


def parse_inventory_json(payload: str) -> list[DriveItem]:
    raw = json.loads(payload)
    if not isinstance(raw, list):
        raise ValueError("gdown inventory must be a JSON array.")
    items = []
    for entry in raw:
        if not isinstance(entry, dict) or not entry.get("url") or not entry.get("path"):
            raise ValueError(f"Invalid gdown inventory entry: {entry!r}")
        items.append(DriveItem(url=str(entry["url"]), path=str(entry["path"])))
    return items


def inventory_drive_folder(folder_url: str) -> list[DriveItem]:
    folder_url = validate_drive_folder_url(folder_url)
    result = subprocess.run(
        [sys.executable, "-m", "gdown", folder_url, "--folder", "--json"],
        check=True,
        capture_output=True,
        text=True,
    )
    return parse_inventory_json(result.stdout)


def _index_exact_basename(items: Sequence[DriveItem], targets: set[str]) -> dict[str, DriveItem]:
    found: dict[str, DriveItem] = {}
    for item in items:
        name = Path(item.path).name
        if name not in targets:
            continue
        if name in found:
            raise ValueError(f"Duplicate requested basename in Drive inventory: {name}")
        found[name] = item
    return found


def select_source_manifest(
    csv_items,
    poa_items,
    dates: pd.DatetimeIndex,
    *,
    url_csv="",
    url_poa="",
) -> SourceManifest:
    requested_dates = [ts.date() for ts in dates]
    csv_names = {d.strftime("%Y%m%d") + ".csv" for d in requested_dates}
    years = sorted({d.year for d in requested_dates})
    poa_names = {f"POA PLTS IKN {year}.xlsx" for year in years}
    csv_index = _index_exact_basename(csv_items, csv_names)
    poa_index = _index_exact_basename(poa_items, poa_names)
    csv_by_date = {
        d: csv_index[d.strftime("%Y%m%d") + ".csv"]
        for d in requested_dates
        if d.strftime("%Y%m%d") + ".csv" in csv_index
    }
    poa_by_year = {
        year: poa_index[f"POA PLTS IKN {year}.xlsx"]
        for year in years
        if f"POA PLTS IKN {year}.xlsx" in poa_index
    }
    return SourceManifest(
        csv_by_date=csv_by_date,
        poa_by_year=poa_by_year,
        missing_csv_dates=[d for d in requested_dates if d not in csv_by_date],
        missing_poa_years=[year for year in years if year not in poa_by_year],
        url_csv=url_csv,
        url_poa=url_poa,
        csv_inventory_count=len(csv_items),
        poa_inventory_count=len(poa_items),
    )


def download_manifest(manifest: SourceManifest, destination: Path) -> DownloadedInputs:
    import gdown

    destination = Path(destination)
    destination.mkdir(parents=True, exist_ok=True)
    csv_paths: dict[date, Path] = {}
    poa_paths: dict[int, Path] = {}
    errors: dict[str, str] = {}
    for key, item in manifest.csv_by_date.items():
        path = destination / Path(item.path).name
        try:
            gdown.download(url=item.url, output=str(path), quiet=False)
            if not path.is_file():
                raise RuntimeError(f"gdown did not create {path.name}")
        except Exception as exc:
            errors[path.name] = f"{type(exc).__name__}: {exc}"
        else:
            csv_paths[key] = path
    for key, item in manifest.poa_by_year.items():
        path = destination / Path(item.path).name
        try:
            gdown.download(url=item.url, output=str(path), quiet=False)
            if not path.is_file():
                raise RuntimeError(f"gdown did not create {path.name}")
        except Exception as exc:
            errors[path.name] = f"{type(exc).__name__}: {exc}"
        else:
            poa_paths[key] = path
    return DownloadedInputs(csv_by_date=csv_paths, poa_by_year=poa_paths, download_errors=errors)


def download_report_inputs(url_csv, url_poa, dates, destination):
    csv_items = inventory_drive_folder(url_csv)
    poa_items = inventory_drive_folder(url_poa)
    manifest = select_source_manifest(
        csv_items,
        poa_items,
        dates,
        url_csv=url_csv,
        url_poa=url_poa,
    )
    return manifest, download_manifest(manifest, Path(destination))


def _find_column(df: pd.DataFrame, predicate) -> str | None:
    return next((str(col) for col in df.columns if predicate(str(col))), None)


def _normalize_manage_object_value(value):
    if pd.isna(value):
        return value
    parts = str(value).strip().split("/")
    leaf = re.sub(r"(?i)inv_a_", "Inv_A_", parts[-1])
    leaf = re.sub(r"(?i)inv_b_", "Inv_B_", leaf)
    if re.fullmatch(r"(?i)WB\d{2}-INV\d+", leaf):
        leaf = leaf.upper()
    parts[-1] = leaf
    return "/".join(parts)


def _extract_string_power(df: pd.DataFrame, selection: StringSelection):
    work = df.copy()
    ts_col = _find_column(work, lambda c: c.casefold() == "start time")
    inv_col = _find_column(work, lambda c: c.casefold() == "inverter_id")
    manage_col = _find_column(work, lambda c: c.casefold() == "manageobject")
    if ts_col is None:
        raise KeyError("CSV missing Start Time column.")
    if inv_col is None:
        if manage_col is None:
            raise KeyError("CSV missing Inverter_ID and ManageObject columns.")
        if manage_col != "ManageObject":
            work = work.rename(columns={manage_col: "ManageObject"})
        from pv_pipeline.transformations import add_inverter_id

        work["ManageObject"] = work["ManageObject"].map(_normalize_manage_object_value)
        work = add_inverter_id(work)
        inv_col = "Inverter_ID"
    mask = work[inv_col].astype("string").str.strip().str.upper() == selection.inverter_id
    work = work.loc[mask]
    direct_re = re.compile(rf"^PV0*{selection.pv_number}\s+Power\(kW\)$", re.I)
    direct_col = _find_column(work, lambda c: direct_re.fullmatch(c.strip()) is not None)
    if direct_col is not None:
        power = pd.to_numeric(work[direct_col], errors="coerce")
        source = "direct"
    else:
        voltage_re = re.compile(rf"^PV0*{selection.pv_number}(?!\d).*voltage", re.I)
        current_re = re.compile(rf"^PV0*{selection.pv_number}(?!\d).*current", re.I)
        voltage_col = _find_column(work, lambda c: voltage_re.search(c) is not None)
        current_col = _find_column(work, lambda c: current_re.search(c) is not None)
        if voltage_col is None or current_col is None:
            return pd.Series(dtype="float64", name="power_kw"), {
                "power_source": "missing",
                "duplicate_rows": 0,
                "negative_samples": 0,
            }
        power = (
            pd.to_numeric(work[voltage_col], errors="coerce")
            * pd.to_numeric(work[current_col], errors="coerce")
            / 1000.0
        )
        source = "voltage_current"
    indexed = pd.Series(
        power.to_numpy(),
        index=pd.to_datetime(work[ts_col], errors="coerce"),
        name="power_kw",
    ).dropna(axis="index")
    indexed = indexed[indexed.index.notna()]
    duplicate_rows = int(len(indexed) - indexed.index.nunique())
    indexed = indexed.groupby(level=0).mean().sort_index()
    return indexed, {
        "power_source": source,
        "duplicate_rows": duplicate_rows,
        "negative_samples": int((indexed < 0).sum()),
    }


def build_report_data(
    csv_by_date,
    poa_by_year,
    selection,
    dates,
    geometry_path,
    *,
    source_manifest=None,
    download_errors=None,
):
    if len(dates) == 0:
        raise ValueError("dates must contain at least one requested day.")
    grid = pd.date_range(
        pd.Timestamp(dates[0]).normalize(),
        pd.Timestamp(dates[-1]).normalize()
        + pd.Timedelta(days=1)
        - pd.Timedelta(minutes=5),
        freq="5min",
    )
    source_by_day = {}
    power_parts = []
    csv_read_errors = {}
    wrong_date_rows = {}
    duplicate_rows = 0
    negative_samples = 0
    power_sources = set()

    for requested_ts in dates:
        requested_day = pd.Timestamp(requested_ts).date()
        path = csv_by_date.get(requested_day)
        if path is None:
            continue
        path = Path(path)
        try:
            frame = pd.read_csv(path)
            timestamp_column = _find_column(
                frame, lambda name: name.casefold() == "start time"
            )
            if timestamp_column is None:
                raise KeyError("CSV missing Start Time column.")
            parsed = pd.to_datetime(frame[timestamp_column], errors="coerce")
            wrong_mask = parsed.notna() & (parsed.dt.date != requested_day)
            wrong_date_rows[path.name] = int(wrong_mask.sum())
            frame = frame.loc[~wrong_mask].copy()
            series, diagnostics = _extract_string_power(frame, selection)
        except Exception as exc:
            csv_read_errors[path.name] = f"{type(exc).__name__}: {exc}"
            continue
        source_by_day[requested_day] = path.name
        power_parts.append(series)
        duplicate_rows += int(diagnostics["duplicate_rows"])
        negative_samples += int(diagnostics["negative_samples"])
        power_sources.add(str(diagnostics["power_source"]))

    if not source_by_day:
        raise RuntimeError("No requested CSV could be read and validated.")
    non_empty = [series for series in power_parts if not series.empty]
    if not non_empty:
        raise RuntimeError(f"No valid power sample found for {selection.canonical}.")
    combined = pd.concat(non_empty).sort_index()
    duplicate_rows += int(len(combined) - combined.index.nunique())
    combined = combined.groupby(level=0).mean().sort_index()
    power = combined.reindex(grid)
    if not power.notna().any():
        raise RuntimeError(
            f"No valid power sample found for {selection.canonical} in requested range."
        )

    geometry = yaml.safe_load(Path(geometry_path).read_text(encoding="utf-8")) or {}
    ws_to_wb = geometry.get("ws_to_wb") or {}
    sheet = str((geometry.get("pyranometer") or {}).get("sheet", "POA PLTS IKN"))
    poa = pd.Series(index=grid, dtype="float64", name="poa_wm2")
    poa_source = pd.Series(index=grid, dtype="object", name="poa_source")
    poa_read_errors = {}
    loaded_poa_files = []
    poa_fallback_samples = 0
    mapped_ws = None

    for year, raw_path in sorted(poa_by_year.items()):
        year_grid = grid[grid.year == int(year)]
        if year_grid.empty:
            continue
        path = Path(raw_path)
        try:
            loader = PyranometerLoader(str(path), sheet=sheet, ws_to_wb=ws_to_wb)
            strict = loader.get_per_ws(
                year_grid,
                selection.wb_id,
                fallback_to_avg=False,
            )
            final = loader.get_per_ws(
                year_grid,
                selection.wb_id,
                fallback_to_avg=True,
            )
        except Exception as exc:
            poa_read_errors[path.name] = f"{type(exc).__name__}: {exc}"
            continue
        loaded_poa_files.append(path.name)
        mapped_ws = strict.attrs.get("ws_label") or mapped_ws
        fallback_mask = strict.isna() & final.notna()
        poa.loc[year_grid] = final.to_numpy()
        poa_source.loc[year_grid[strict.notna().to_numpy()]] = strict.attrs.get(
            "ws_label"
        )
        poa_source.loc[year_grid[fallback_mask.to_numpy()]] = "avg"
        poa_fallback_samples += int(fallback_mask.sum())

    has_power = power.notna().to_numpy()
    has_poa = poa.notna().to_numpy()
    data_status = np.select(
        [has_power & has_poa, has_power & ~has_poa, ~has_power & has_poa],
        ["POWER_POA", "POWER_ONLY", "POA_ONLY"],
        default="MISSING_BOTH",
    )
    five_minute = pd.DataFrame({
        "timestamp": grid,
        "inverter_id": selection.inverter_id,
        "pv_string": selection.pv_label,
        "power_kw": power.to_numpy(),
        "poa_wm2": poa.to_numpy(),
        "source_csv": [source_by_day.get(timestamp.date()) for timestamp in grid],
        "poa_source": poa_source.to_numpy(),
        "data_status": data_status,
    })

    daily_rows = []
    for requested_ts in dates:
        requested_day = pd.Timestamp(requested_ts).date()
        day_mask = five_minute["timestamp"].dt.date == requested_day
        day_power = five_minute.loc[day_mask].set_index("timestamp")["power_kw"]
        valid = int(day_power.notna().sum())
        if requested_day not in source_by_day:
            status = "MISSING_CSV"
        elif valid == 0:
            status = "NO_STRING_DATA"
        elif valid == 288:
            status = "COMPLETE"
        else:
            status = "PARTIAL"
        yield_kwh = (
            compute_active_power_integration_kwh(day_power, freq_hours=5 / 60)
            if valid > 0
            else np.nan
        )
        daily_rows.append({
            "date": requested_day,
            "string_yield_kwh": yield_kwh,
            "valid_power_samples": valid,
            "expected_samples": 288,
            "coverage_pct": valid / 288 * 100,
            "missing_power_samples": 288 - valid,
            "poa_valid_samples": int(
                five_minute.loc[day_mask, "poa_wm2"].notna().sum()
            ),
            "source_csv": source_by_day.get(requested_day),
            "status": status,
        })
    daily = pd.DataFrame(daily_rows)

    inventory_missing_csv_dates = [
        pd.Timestamp(value).date().isoformat()
        for value in (source_manifest.missing_csv_dates if source_manifest else [])
    ]
    missing_csv_dates = [
        row["date"].isoformat()
        for row in daily_rows
        if row["status"] == "MISSING_CSV"
    ]
    missing_poa_years = (
        list(source_manifest.missing_poa_years) if source_manifest else []
    )
    warnings_list = []
    if any(wrong_date_rows.values()):
        warnings_list.append("Rows outside their YYYYMMDD.csv date were dropped.")
    if negative_samples:
        warnings_list.append("Negative power samples were retained.")
    if missing_poa_years or poa_read_errors:
        warnings_list.append("POA is unavailable for one or more requested years.")

    metadata = {
        "pv_string_input": selection.canonical,
        "inverter_id": selection.inverter_id,
        "pv_string": selection.pv_label,
        "start_date": pd.Timestamp(dates[0]).date().isoformat(),
        "end_date": pd.Timestamp(dates[-1]).date().isoformat(),
        "generated_at_wita": pd.Timestamp.now(tz="Asia/Makassar").isoformat(),
        "source_url_csv": source_manifest.url_csv if source_manifest else "",
        "source_url_poa": source_manifest.url_poa if source_manifest else "",
        "csv_inventory_count": source_manifest.csv_inventory_count if source_manifest else 0,
        "poa_inventory_count": source_manifest.poa_inventory_count if source_manifest else 0,
        "loaded_csv_files": sorted(source_by_day.values()),
        "missing_csv_dates": missing_csv_dates,
        "inventory_missing_csv_dates": inventory_missing_csv_dates,
        "missing_poa_years": missing_poa_years,
        "download_errors": dict(download_errors or {}),
        "csv_read_errors": csv_read_errors,
        "loaded_poa_files": sorted(loaded_poa_files),
        "poa_read_errors": poa_read_errors,
        "wrong_date_rows": wrong_date_rows,
        "duplicate_rows": duplicate_rows,
        "negative_power_samples": negative_samples,
        "power_sources": sorted(power_sources),
        "mapped_ws": mapped_ws,
        "ws_to_wb": ws_to_wb,
        "poa_fallback_samples": poa_fallback_samples,
        "yield_formula": "sum(power_kw_valid * 5/60)",
        "interval_minutes": 5,
        "warnings": warnings_list,
    }
    return ReportData(daily=daily, five_minute=five_minute, metadata=metadata)


def plot_daily_yield(daily, selection):
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(11, 4))
    ax.plot(
        pd.to_datetime(daily["date"]),
        daily["string_yield_kwh"],
        marker="o",
    )
    ax.set(
        title=f"Yield harian {selection.canonical}",
        xlabel="Tanggal",
        ylabel="String yield (kWh)",
    )
    ax.grid(True, alpha=0.3)
    fig.autofmt_xdate()
    fig.tight_layout()
    return fig, ax


def plot_power_vs_poa(five_minute, selection, start, end):
    import matplotlib.pyplot as plt

    fig, ax_power = plt.subplots(figsize=(14, 5))
    ax_poa = ax_power.twinx()
    line_power = ax_power.plot(
        five_minute["timestamp"],
        five_minute["power_kw"],
        label="Power string",
        color="tab:blue",
    )[0]
    line_poa = ax_poa.plot(
        five_minute["timestamp"],
        five_minute["poa_wm2"],
        label="POA irradiance",
        color="tab:orange",
        alpha=0.75,
    )[0]
    ax_power.set(
        xlabel="Waktu",
        ylabel="Power string (kW)",
        title=f"Power vs POA {selection.canonical} | {start} s.d. {end}",
    )
    ax_poa.set_ylabel("POA irradiance (W/m²)")
    ax_power.legend(
        [line_power, line_poa],
        [line_power.get_label(), line_poa.get_label()],
        loc="upper left",
    )
    ax_power.grid(True, alpha=0.25)
    fig.autofmt_xdate()
    fig.tight_layout()
    return fig, (ax_power, ax_poa)


def _excel_value(value):
    if value is None or value is pd.NA:
        return None
    if isinstance(value, np.generic):
        value = value.item()
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        return value
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


def _append_dataframe(sheet, frame):
    for row in dataframe_to_rows(frame, index=False, header=True):
        sheet.append([_excel_value(value) for value in row])
    for cell in sheet[1]:
        cell.font = Font(bold=True)


def _set_column_widths(sheet, widths):
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _daily_excel_chart(source_sheet):
    chart = LineChart()
    chart.title = "String Yield Harian"
    chart.y_axis.title = "Yield (kWh)"
    chart.x_axis.title = "Tanggal"
    chart.height = 8
    chart.width = 16
    chart.display_blanks = "gap"
    values = Reference(
        source_sheet,
        min_col=2,
        min_row=1,
        max_row=source_sheet.max_row,
    )
    dates = Reference(
        source_sheet,
        min_col=1,
        min_row=2,
        max_row=source_sheet.max_row,
    )
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(dates)
    return chart


def build_output_path(output_dir, selection, start, end):
    return Path(output_dir) / (
        f"string_yield_{selection.inverter_id}_{selection.pv_label}_"
        f"{start:%Y%m%d}_{end:%Y%m%d}.xlsx"
    )


def write_report_workbook(output_path, report):
    if list(report.daily.columns) != DAILY_COLUMNS:
        raise ValueError(f"Unexpected daily columns: {list(report.daily.columns)!r}")
    if list(report.five_minute.columns) != FIVE_MINUTE_COLUMNS:
        raise ValueError(
            "Unexpected five-minute columns: "
            f"{list(report.five_minute.columns)!r}"
        )

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    daily_sheet = workbook.create_sheet("Ringkasan_Harian")
    five_sheet = workbook.create_sheet("Data_5Menit")
    graph_sheet = workbook.create_sheet("Grafik")
    metadata_sheet = workbook.create_sheet("Metadata")

    _append_dataframe(daily_sheet, report.daily[DAILY_COLUMNS])
    daily_sheet.freeze_panes = "A2"
    daily_sheet.auto_filter.ref = daily_sheet.dimensions
    for row in range(2, daily_sheet.max_row + 1):
        daily_sheet.cell(row=row, column=1).number_format = "yyyy-mm-dd"
        daily_sheet.cell(row=row, column=2).number_format = "0.000"
        daily_sheet.cell(row=row, column=5).number_format = '0.0"%"'
    _set_column_widths(daily_sheet, {
        "A": 12, "B": 18, "C": 20, "D": 18, "E": 14, "F": 22,
        "G": 18, "H": 20, "I": 16,
    })
    daily_sheet.add_chart(_daily_excel_chart(daily_sheet), "K2")

    _append_dataframe(five_sheet, report.five_minute[FIVE_MINUTE_COLUMNS])
    five_sheet.freeze_panes = "A2"
    five_sheet.auto_filter.ref = five_sheet.dimensions
    for row in range(2, five_sheet.max_row + 1):
        five_sheet.cell(row=row, column=1).number_format = "yyyy-mm-dd hh:mm"
        five_sheet.cell(row=row, column=4).number_format = "0.000"
        five_sheet.cell(row=row, column=5).number_format = "0.0"
    _set_column_widths(five_sheet, {
        "A": 18, "B": 16, "C": 12, "D": 14, "E": 14, "F": 18,
        "G": 14, "H": 18,
    })

    graph_sheet.sheet_view.showGridLines = False
    graph_sheet.add_chart(_daily_excel_chart(daily_sheet), "A1")
    power_chart = LineChart()
    power_chart.title = "Power String vs POA Irradiance"
    power_chart.y_axis.title = "Power string (kW)"
    power_chart.x_axis.title = "Waktu"
    power_chart.y_axis.axId = 10
    power_chart.height = 12
    power_chart.width = 24
    power_chart.display_blanks = "gap"
    time_values = Reference(
        five_sheet,
        min_col=1,
        min_row=2,
        max_row=five_sheet.max_row,
    )
    power_values = Reference(
        five_sheet,
        min_col=4,
        min_row=1,
        max_row=five_sheet.max_row,
    )
    power_chart.add_data(power_values, titles_from_data=True)
    power_chart.set_categories(time_values)

    poa_chart = LineChart()
    poa_chart.y_axis.title = "POA irradiance (W/m²)"
    poa_chart.y_axis.axId = 200
    poa_chart.y_axis.crosses = "max"
    poa_chart.display_blanks = "gap"
    poa_values = Reference(
        five_sheet,
        min_col=5,
        min_row=1,
        max_row=five_sheet.max_row,
    )
    poa_chart.add_data(poa_values, titles_from_data=True)
    poa_chart.set_categories(time_values)
    power_chart += poa_chart
    graph_sheet.add_chart(power_chart, "A18")

    metadata_sheet.append(["key", "value"])
    for cell in metadata_sheet[1]:
        cell.font = Font(bold=True)
    for key, value in report.metadata.items():
        normalized_key = re.sub(r"[^a-z]", "", str(key).casefold())
        if any(term in normalized_key for term in SENSITIVE_METADATA_TERMS):
            raise ValueError(f"Sensitive metadata key is not allowed: {key!r}")
        if isinstance(value, (dict, list, tuple)):
            value = json.dumps(value, ensure_ascii=False, default=str)
        metadata_sheet.append([str(key), _excel_value(value)])
    metadata_sheet.freeze_panes = "A2"
    metadata_sheet.auto_filter.ref = metadata_sheet.dimensions
    _set_column_widths(metadata_sheet, {"A": 32, "B": 80})

    workbook.save(output_path)
    verify_report_workbook(output_path)
    return output_path


def verify_report_workbook(path):
    path = Path(path)
    if not path.is_file() or path.stat().st_size == 0:
        raise RuntimeError(f"Workbook was not created: {path}")
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        if workbook.sheetnames != SHEET_ORDER:
            raise RuntimeError(f"Unexpected sheet order: {workbook.sheetnames!r}")
        if workbook["Ringkasan_Harian"].max_row < 2:
            raise RuntimeError("Ringkasan_Harian has no data row.")
        if workbook["Data_5Menit"].max_row < 2:
            raise RuntimeError("Data_5Menit has no data row.")
        if len(workbook["Ringkasan_Harian"]._charts) != 1:
            raise RuntimeError("Ringkasan_Harian must contain one chart.")
        if len(workbook["Grafik"]._charts) != 2:
            raise RuntimeError("Grafik must contain daily and combo charts.")
    finally:
        workbook.close()
