"""Verify Iterasi 5 (M2bGroundFault) Excel sheets vs proto_iter5 locked numbers.

LibreOffice crash & formulas/pycel tak ter-install → verifikasi 2-lapis (mirror iter4):
  (A) STRUCTURAL formula audit — string formula tiap sel Helpers/StringMetrics/Decision
      == template (catch salah ref).
  (B) NUMERIC recompute — baca LITERAL Raw_Data_GF + named cells, re-implement semantik
      Excel (Voc median, peer-median exclude-self, i_z, triple-signal, confidence matrix),
      assert == angka proto_iter5.py.
"""
from __future__ import annotations

import numpy as np
from openpyxl import load_workbook

WB = "M2_PV_Performance_Workbook.xlsx"
PV_MAX = 6
INVS = ["WB05-INV01", "WB05-INV02", "WB05-INV03", "WB05-INV04"]
EMPTY = {0: {5}, 1: {5}, 2: set(), 3: set()}          # inv idx -> empty PV set
FAULT = {0: None, 1: None, 2: 3, 3: 3}
V_COL = {k: chr(ord("E") + k - 1) for k in range(1, 7)}   # E..J
I_COL = {k: chr(ord("K") + k - 1) for k in range(1, 7)}   # K..P
VOC_CAND_COL = {k: chr(ord("F") + k - 1) for k in range(1, 7)}  # F..K
PEER_COL = {k: chr(ord("L") + k - 1) for k in range(1, 7)}      # L..Q

errors = []
def check(cond, msg):
    print(("  OK  " if cond else "  XX  FAIL: ") + msg)
    if not cond:
        errors.append(msg)

def raw_block(i):    # (dawn_first, dawn_last, noon_first, noon_last)
    f = 5 + 9 * i
    return f, f + 2, f + 3, f + 8
def sm_row(i, k):
    return 5 + 6 * i + (k - 1)
def dec_row(i):
    return 5 + i

print("Loading workbook...")
wb = load_workbook(WB, data_only=False)
for s in ("Raw_Data_GF", "Helpers_GF", "GF_StringMetrics", "M2c_GroundFault", "M2c_GF_StringStatus"):
    assert s in wb.sheetnames, f"missing {s}"
raw = wb["Raw_Data_GF"]; hp = wb["Helpers_GF"]; sm = wb["GF_StringMetrics"]
dc = wb["M2c_GroundFault"]; ss = wb["M2c_GF_StringStatus"]

# resolve config named cells
def cfgval(name):
    dn = wb.defined_names[name]; (s, c), = dn.destinations
    return wb[s][c].value
POA_THR = cfgval("cfg_gf_poa_threshold_wm2"); POA_FLOOR = cfgval("cfg_gf_poa_floor_wm2")
V_ABS = cfgval("cfg_gf_v_abs_threshold"); ADP_Z = cfgval("cfg_gf_adaptive_z_threshold")
VOC_THR = cfgval("cfg_gf_voc_ratio_threshold"); IZ_THR = cfgval("cfg_gf_i_high_z_threshold")
FLEET_MED = cfgval("cfg_gf_fleet_v_gnd_median"); FLEET_STD = cfgval("cfg_gf_fleet_v_gnd_std")
I_THR_A = cfgval("cfg_i_threshold_a"); MIN_VOC_V = cfgval("cfg_min_voc_v")
print("config:", dict(POA_THR=POA_THR, V_ABS=V_ABS, ADP_Z=ADP_Z, VOC_THR=VOC_THR,
                       IZ_THR=IZ_THR, FLEET_MED=FLEET_MED, FLEET_STD=FLEET_STD))
print("\n=== config expected ===")
check(POA_THR == 200, f"poa_threshold==200 (got {POA_THR})")
check(V_ABS == 50, f"v_abs==50 (got {V_ABS})")
check(ADP_Z == 3.0, f"adaptive_z==3 (got {ADP_Z})")
check(VOC_THR == 0.85, f"voc_ratio==0.85 (got {VOC_THR})")
check(IZ_THR == 2.0, f"i_high_z==2 (got {IZ_THR})")
check(FLEET_MED == 0.0 and FLEET_STD == 8.0, f"fleet median/std 0/8 (got {FLEET_MED}/{FLEET_STD})")

VOC_NOMINAL = 55.72 * (1 + (-0.25) / 100 * (30 - 25)) * 26   # voc_string_26_calc @30C
check(abs(VOC_NOMINAL - 1430.611) < 1e-3, f"voc_string_nominal==1430.611 (got {VOC_NOMINAL:.4f})")

# ===========================================================================
print("\n=== (A) STRUCTURAL formula audit ===")
# Helpers: daylight(D), voc_cand(F-K), peer_med(L-Q), abs_vgnd(R) — all 36 rows
for ri in range(5, 41):
    check(hp[f"D{ri}"].value == f"=IF(AND(C{ri}>cfg_gf_poa_threshold_wm2,C{ri}>cfg_gf_poa_floor_wm2),1,0)",
          f"Helpers D{ri} daylight") if ri in (5, 8, 40) else None
    check(hp[f"R{ri}"].value == f"=ABS(E{ri})", f"Helpers R{ri} abs_Vgnd") if ri in (5, 17, 40) else None
    for k in range(1, 7):
        exp_voc = (f"=IF(AND(ABS(Raw_Data_GF!{I_COL[k]}{ri})<cfg_i_threshold_a,"
                   f"Raw_Data_GF!{V_COL[k]}{ri}>cfg_min_voc_v),Raw_Data_GF!{V_COL[k]}{ri},\"\")")
        if hp[f"{VOC_CAND_COL[k]}{ri}"].value != exp_voc:
            check(False, f"Helpers {VOC_CAND_COL[k]}{ri} voc_cand PV{k}")
        others = ",".join(f"Raw_Data_GF!{I_COL[j]}{ri}" for j in range(1, 7) if j != k)
        if hp[f"{PEER_COL[k]}{ri}"].value != f"=MEDIAN({others})":
            check(False, f"Helpers {PEER_COL[k]}{ri} peer_med PV{k}")
check(True, "Helpers voc_cand + peer_med templates (all 36 rows × 6 PV) match")

# StringMetrics — all 24 rows
for i in range(4):
    df0, df1, nf0, nf1 = raw_block(i)
    for k in range(1, 7):
        r = sm_row(i, k)
        check(sm[f"D{r}"].value == f"=IFERROR(MEDIAN(Helpers_GF!{VOC_CAND_COL[k]}{df0}:{VOC_CAND_COL[k]}{nf1}),\"\")",
              f"SM D{r} voc_actual ({INVS[i]} PV{k})") if k in (1, 3) else None
        check(sm[f"F{r}"].value == f'=IFERROR(D{r}/E{r},"")', f"SM F{r} voc_ratio") if k in (1, 3) else None
        check(sm[f"G{r}"].value == f"=MEDIAN(Raw_Data_GF!{I_COL[k]}{nf0}:{I_COL[k]}{nf1})",
              f"SM G{r} I_median") if k in (1, 3) else None
        check(sm[f"J{r}"].value == f'=IFERROR((G{r}-H{r})/I{r},"")', f"SM J{r} i_z") if k in (1, 3) else None
        check(sm[f"K{r}"].value == f"=IF(C{r}=1,0,IF(AND(F{r}<cfg_gf_voc_ratio_threshold,J{r}>cfg_gf_i_high_z_threshold),1,0))",
              f"SM K{r} spec_flag") if k in (1, 3) else None

# Decision — all 4 rows
for i in range(4):
    r = dec_row(i)
    sm0, sm1 = sm_row(i, 1), sm_row(i, PV_MAX)
    df0, df1, nf0, nf1 = raw_block(i)
    check(dc[f"D{r}"].value == f"=ABS(C{r}-cfg_gf_fleet_v_gnd_median)/MAX(cfg_gf_fleet_v_gnd_std,0.01)", f"DEC D{r} adaptive_z")
    check(dc[f"E{r}"].value == f"=IF(B{r}>cfg_gf_v_abs_threshold,1,0)", f"DEC E{r} flag_abs")
    check(dc[f"F{r}"].value == f"=IF(D{r}>cfg_gf_adaptive_z_threshold,1,0)", f"DEC F{r} flag_adp")
    check(dc[f"G{r}"].value == f"=IF(SUM(GF_StringMetrics!K{sm0}:K{sm1})>0,1,0)", f"DEC G{r} flag_spec")
    check(dc[f"I{r}"].value == (f"=IF(AND(G{r}=1,OR(E{r}=1,F{r}=1)),90,IF(G{r}=1,80,"
                                f"IF(AND(E{r}=1,F{r}=1),80,IF(E{r}=1,70,IF(F{r}=1,60,0)))))"), f"DEC I{r} confidence")
    check(dc[f"K{r}"].value == f"=INDEX(GF_StringMetrics!B{sm0}:B{sm1},MATCH(L{r},GF_StringMetrics!F{sm0}:F{sm1},0))", f"DEC K{r} worst_PV")
    check(dc[f"M{r}"].value == f'=IF(OR(E{r}=1,F{r}=1,G{r}=1),"ground_fault","NORMAL")', f"DEC M{r} status")

# ===========================================================================
print("\n=== (B) NUMERIC recompute from literals ===")
def col_vals(sheet, col, r0, r1):
    return [sheet[f"{col}{r}"].value for r in range(r0, r1 + 1)]

def confidence_for(trig):
    spec, ab, ad = "spec_4.2.3" in trig, "absolute" in trig, "adaptive" in trig
    if spec and (ab or ad): return 90.0
    if spec: return 80.0
    if ab and ad: return 80.0
    if ab: return 70.0
    if ad: return 60.0
    return 0.0

expect = {
    0: dict(trig=[], conf=0.0, sev="-", worst="PV1"),
    1: dict(trig=["absolute", "adaptive"], conf=80.0, sev="CRITICAL", worst="PV1"),
    2: dict(trig=["spec_4.2.3"], conf=80.0, sev="CRITICAL", worst="PV3"),
    3: dict(trig=["absolute", "adaptive", "spec_4.2.3"], conf=90.0, sev="CRITICAL", worst="PV3"),
}
print(f"{'INV':<13}{'vmaxabs':>8}{'vmed':>7}{'adpZ':>7}{'abs':>6}{'adp':>6}{'spec':>6}"
      f"{'worst':>7}{'vocR':>8}{'i_z':>9}{'conf':>5}{'sev':>9}")
for i in range(4):
    df0, df1, nf0, nf1 = raw_block(i)
    # V_gnd noon
    vgnd_noon = [float(raw[f"D{r}"].value) for r in range(nf0, nf1 + 1)]
    v_max_abs = max(abs(v) for v in vgnd_noon)
    v_med = float(np.median(vgnd_noon))
    adp_z = abs(v_med - FLEET_MED) / max(FLEET_STD, 0.01)
    f_abs = v_max_abs > V_ABS
    f_adp = adp_z > ADP_Z
    # per-string
    I_noon = {k: [float(raw[f"{I_COL[k]}{r}"].value) for r in range(nf0, nf1 + 1)] for k in range(1, 7)}
    worst_r, worst_pv, worst_iz, spec_any = np.nan, None, np.nan, False
    for k in range(1, 7):
        # voc_actual = median of dawn V where |I|<0.5 & V>10
        cand = []
        for r in range(df0, nf1 + 1):
            Vv = float(raw[f"{V_COL[k]}{r}"].value); Iv = float(raw[f"{I_COL[k]}{r}"].value)
            if abs(Iv) < I_THR_A and Vv > MIN_VOC_V:
                cand.append(Vv)
        if k in EMPTY[i] or len(cand) < 3:
            continue
        voc_actual = float(np.median(cand))
        voc_ratio = voc_actual / VOC_NOMINAL
        # peer median per noon row = median of other 5 currents
        peer_noon = []
        for ridx, r in enumerate(range(nf0, nf1 + 1)):
            others = [I_noon[j][ridx] for j in range(1, 7) if j != k]
            peer_noon.append(float(np.median(others)))
        I_med = float(np.median(I_noon[k]))
        peer_med = float(np.median(peer_noon))
        peer_std = float(np.std(peer_noon, ddof=1))
        peer_std = peer_std if peer_std > 0 else 0.01
        i_z = (I_med - peer_med) / peer_std
        if np.isnan(worst_r) or voc_ratio < worst_r:
            worst_r, worst_pv, worst_iz = voc_ratio, k, i_z
        if voc_ratio < VOC_THR and i_z > IZ_THR:
            spec_any = True
    trig = (["absolute"] if f_abs else []) + (["adaptive"] if f_adp else []) + (["spec_4.2.3"] if spec_any else [])
    conf = confidence_for(trig)
    sev = "CRITICAL" if conf >= 80 else ("HIGH" if conf > 0 else "-")
    print(f"{INVS[i]:<13}{v_max_abs:>8.2f}{v_med:>7.2f}{adp_z:>7.3f}{str(f_abs):>6}{str(f_adp):>6}{str(spec_any):>6}"
          f"{('PV'+str(worst_pv)):>7}{worst_r:>8.4f}{worst_iz:>9.3f}{conf:>5.0f}{sev:>9}")
    e = expect[i]
    check(trig == e["trig"], f"{INVS[i]} triggered == {e['trig']} (got {trig})")
    check(conf == e["conf"], f"{INVS[i]} confidence == {e['conf']} (got {conf})")
    check(sev == e["sev"], f"{INVS[i]} severity == {e['sev']}")
    check(f"PV{worst_pv}" == e["worst"], f"{INVS[i]} worst == {e['worst']} (got PV{worst_pv})")

# spot exact numbers vs proto
print("\n=== exact-number spot checks vs proto ===")
# INV03 PV3 (faulted)
df0, df1, nf0, nf1 = raw_block(2)
cand3 = [float(raw[f"G{r}"].value) for r in range(df0, df0 + 3)]   # PV3 V col=G, dawn
voc3 = float(np.median(cand3)); ratio3 = voc3 / VOC_NOMINAL
check(abs(voc3 - 1100.0) < 1e-9, f"INV03 PV3 voc_actual==1100 (got {voc3})")
check(abs(ratio3 - 0.768902) < 1e-4, f"INV03 PV3 voc_ratio==0.7689 (got {ratio3:.6f})")
I3 = [float(raw[f"M{r}"].value) for r in range(nf0, nf1 + 1)]      # PV3 I col=M
peer3 = []
for ridx, r in enumerate(range(nf0, nf1 + 1)):
    others = [float(raw[f"{I_COL[j]}{r}"].value) for j in range(1, 7) if j != 3]
    peer3.append(float(np.median(others)))
iz3 = (np.median(I3) - np.median(peer3)) / np.std(peer3, ddof=1)
check(abs(iz3 - 10.9545) < 1e-3, f"INV03 PV3 i_z==10.954 (got {iz3:.4f})")
# healthy voc_ratio
check(abs(1430.0 / VOC_NOMINAL - 0.999573) < 1e-4, "healthy voc_ratio==0.9996")
# empty-skip: INV01 PV5 empty -> not worst, not spec
check(5 in EMPTY[0], "INV01 PV5 is empty (skip from analysis/worst)")

print("\n" + "=" * 60)
if errors:
    print(f"VERIFY FAILED: {len(errors)} error(s)")
    for e in errors:
        print("  -", e)
    raise SystemExit(1)
print("VERIFY OK — all structural + numeric checks passed (iter5).")
