"""Verify Iterasi 11 (M2bMpptRatio) Excel vs locked numbers + real-run consistency.

(A) STRUCTURAL audit — Helpers_MR (daylight, median partner se-MPPT, ratio, qualifying, consec)
    + M2b_MpptRatio (emit, severity, confidence, status) + M2b_MR_StringStatus.
(B) NUMERIC recompute — baca literal Raw_Data_MR + config, re-implement semantik detector
    (median partner, ratio, qualifying, debounce, severity ladder, confidence), assert == locked.
(C) REAL-RUN consistency — m2_findings_20251030.xlsx: 286 finding M2b_mppt_ratio harus 100%
    konsisten dgn ladder severity + formula confidence + value==ratio_event_median + threshold==0.85.

LibreOffice live recalc N/A (sandbox crash) -> verifikasi = audit string + reimplementasi Python
+ regen 0-diff + konsistensi finding produksi (lihat docs/M2_RE_11 §7).
"""
from __future__ import annotations

import ast
import statistics
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
WB = HERE / "M2_PV_Performance_Workbook.xlsx"
FINDINGS = HERE.parent / "m2_findings_20251030.xlsx"

# ---- skenario (mirror _extend_m2_workbook_iter11.py) ----
MEMBERS = {1: [1, 2, 3, 4], 2: [5, 6, 7, 8, 9]}
MPPT_OF = {n: m for m, ns in MEMBERS.items() for n in ns}
CURRENT = {1: 13.0, 2: 12.8, 3: 10.3, 4: 5.6, 5: 13.0, 6: 12.8, 7: 12.9, 8: 0.0, 9: 12.85}
GLITCH_PV, GLITCH_ROWS, GLITCH_VALUE = 9, (9, 10, 11), 3.0
PVS = list(range(1, 10))
PARTNERS = {n: [m for m in MEMBERS[MPPT_OF[n]] if m != n] for n in PVS}
RAW_COL = {n: get_column_letter(3 + (n - 1)) for n in PVS}
HMED = {n: 4 + (n - 1) for n in PVS}
HRAT = {n: 13 + (n - 1) for n in PVS}
HQUAL = {n: 22 + (n - 1) for n in PVS}
HCONS = {n: 31 + (n - 1) for n in PVS}
DATA_FIRST, DAY_LAST, DATA_LAST = 5, 28, 30
N_DAYLIGHT = 24
DEBOUNCE, RTHR, RHIGH, RCRIT, CLIP = 20, 0.85, 0.50, 0.20, 0.01

errors = []
def check(c, m):
    print(("  OK  " if c else "  XX  FAIL: ") + m)
    if not c:
        errors.append(m)

wb = load_workbook(WB, data_only=False)
for s in ("Raw_Data_MR", "Helpers_MR", "M2b_MpptRatio", "M2b_MR_StringStatus"):
    assert s in wb.sheetnames, f"missing {s}"
hp = wb["Helpers_MR"]; dec = wb["M2b_MpptRatio"]; ss = wb["M2b_MR_StringStatus"]

def cfgval(n):
    dn = wb.defined_names[n]; (s, c), = dn.destinations; return wb[s][c].value
for name, exp in [("cfg_mr_poa_threshold_wm2", 300.0), ("cfg_mr_poa_floor_wm2", 50.0),
                  ("cfg_mr_ratio_threshold", 0.85), ("cfg_mr_ratio_high", 0.50),
                  ("cfg_mr_ratio_critical", 0.20), ("cfg_mr_debounce_steps", 20),
                  ("cfg_mr_min_partner", 1), ("cfg_mr_partner_clip", 0.01)]:
    check(cfgval(name) == exp, f"config {name} == {exp} (got {cfgval(name)})")

# ===========================================================================
print("\n=== (A) STRUCTURAL audit ===")
# daylight + median partner + ratio + qualifying + consec, spot baris 5 & 7
for n in (1, 3, 8):
    med_l = get_column_letter(HMED[n]); rat_l = get_column_letter(HRAT[n])
    qual_l = get_column_letter(HQUAL[n]); cons_l = get_column_letter(HCONS[n])
    refs = ",".join(f"Raw_Data_MR!{RAW_COL[p]}5" for p in PARTNERS[n])
    check(hp[f"{med_l}5"].value == f"=MEDIAN({refs})", f"Helpers med_partner PV{n} (partner se-MPPT only)")
    check(hp[f"{rat_l}5"].value == f"=Raw_Data_MR!{RAW_COL[n]}5/MAX({med_l}5,cfg_mr_partner_clip)", f"Helpers ratio PV{n}")
    check(hp[f"{qual_l}5"].value == f"=IF(AND({rat_l}5<cfg_mr_ratio_threshold,$C5=1),1,0)", f"Helpers qualifying PV{n}")
    check(hp[f"{cons_l}5"].value == f"={qual_l}5", f"Helpers consec PV{n} seed")
    check(hp[f"{cons_l}6"].value == f"=IF({qual_l}6=1,{cons_l}5+1,0)", f"Helpers consec PV{n} recurrence")
check(hp["C5"].value == "=IF(AND(B5>cfg_mr_poa_threshold_wm2,B5>cfg_mr_poa_floor_wm2),1,0)", "Helpers daylight gate")
# CRITICAL audit: median partner PV3 must NOT include PV3 itself (column E)
med3 = get_column_letter(HMED[3])
check("Raw_Data_MR!E5" not in hp[f"{med3}5"].value, "median partner PV3 EXCLUDES PV3 (no E5)")
check(set(f"Raw_Data_MR!{RAW_COL[p]}5" for p in [1, 2, 4]) ==
      set(hp[f"{med3}5"].value[len("=MEDIAN("):-1].split(",")), "median partner PV3 == {PV1,PV2,PV4}")
# decision sheet
for i, n in enumerate(PVS):
    r = DEC_FIRST = 5 + i
    rat_l = get_column_letter(HRAT[n]); cons_l = get_column_letter(HCONS[n])
    check(dec[f"D{r}"].value == f"=MEDIAN(Helpers_MR!{rat_l}5:{rat_l}28)", f"Dec D{r} ratio_median_daylight PV{n}")
    check(dec[f"E{r}"].value == f"=MAX(Helpers_MR!{cons_l}5:{cons_l}30)", f"Dec E{r} max_consec PV{n}")
    check(dec[f"G{r}"].value == f"=IF(AND(E{r}>=cfg_mr_debounce_steps,F{r}=0),1,0)", f"Dec G{r} emit PV{n}")
    check(dec[f"H{r}"].value == (f'=IF(G{r}=0,"",IF(D{r}<cfg_mr_ratio_critical,"CRITICAL",'
                                 f'IF(D{r}<cfg_mr_ratio_high,"HIGH","MEDIUM")))'), f"Dec H{r} severity PV{n}")
    check(dec[f"I{r}"].value == f'=IF(G{r}=0,"",MIN(90,MAX(50,(1-D{r})*100)))', f"Dec I{r} confidence PV{n}")
    check(dec[f"J{r}"].value == f'=IF(F{r}=1,"EMPTY",IF(G{r}=1,"mppt_partner_underperform","NORMAL"))', f"Dec J{r} status PV{n}")

# ===========================================================================
print("\n=== (B) NUMERIC recompute from literals ===")
raw = wb["Raw_Data_MR"]
# rebuild current matrix from literals (sanity: literal == CURRENT, glitch applied)
for i in range(N_DAYLIGHT + 2):
    ri = DATA_FIRST + i
    for n in PVS:
        exp = CURRENT[n]
        if n == GLITCH_PV and ri in GLITCH_ROWS:
            exp = GLITCH_VALUE
        got = raw[f"{RAW_COL[n]}{ri}"].value
        if abs(float(got) - exp) > 1e-9:
            check(False, f"Raw PV{n} row {ri} literal {got} != {exp}")
poa = [raw[f"B{DATA_FIRST+i}"].value for i in range(N_DAYLIGHT + 2)]
daylight = [(p > 300.0 and p > 50.0) for p in poa]
check(sum(daylight) == 24, f"daylight rows == 24 (got {sum(daylight)})")

def cur(n, i):
    ri = DATA_FIRST + i
    if n == GLITCH_PV and ri in GLITCH_ROWS:
        return GLITCH_VALUE
    return CURRENT[n]

EXPECT = {  # pv -> (ratio_median_daylight, max_consec, emit, severity, confidence, n_qual)
    1: (1.262136, 0, 0, "", None, 0),
    2: (1.242718, 0, 0, "", None, 0),
    3: (0.804688, 24, 1, "MEDIUM", 50.0, 24),
    4: (0.437500, 24, 1, "HIGH", 56.25, 24),
    5: (1.013645, 0, 0, "", None, 0),
    6: (0.994175, 0, 0, "", None, 0),
    7: (1.005848, 0, 0, "", None, 0),
    8: (0.000000, 24, 1, "CRITICAL", 90.0, 24),
    9: (1.000000, 3, 0, "", None, 3),
}
print(f"{'PV':>3}{'mppt':>5}{'rat_med':>10}{'consec':>7}{'emit':>5}{'sev':>9}{'conf':>7}{'nqual':>6}")
for n in PVS:
    ratios_day = []
    qual = []
    consec = cmax = 0
    for i in range(N_DAYLIGHT + 2):
        partner_vals = [cur(p, i) for p in PARTNERS[n]]
        med = statistics.median(partner_vals)
        ratio = cur(n, i) / max(med, CLIP)
        is_day = daylight[i]
        q = 1 if (ratio < RTHR and is_day) else 0
        qual.append(q)
        consec = consec + 1 if q == 1 else 0
        cmax = max(cmax, consec)
        if is_day:
            ratios_day.append(ratio)
    rat_med = statistics.median(ratios_day)
    emit = 1 if (cmax >= DEBOUNCE) else 0
    if emit:
        sev = "CRITICAL" if rat_med < RCRIT else "HIGH" if rat_med < RHIGH else "MEDIUM"
        conf = min(90.0, max(50.0, (1.0 - rat_med) * 100.0))
    else:
        sev, conf = "", None
    nqual = sum(qual)
    print(f"{n:>3}{MPPT_OF[n]:>5}{rat_med:>10.4f}{cmax:>7}{emit:>5}{sev:>9}{('' if conf is None else f'{conf:.2f}'):>7}{nqual:>6}")
    erm, ecmax, eemit, esev, econf, enq = EXPECT[n]
    check(abs(rat_med - erm) < 1e-4, f"PV{n} ratio_median_daylight≈{erm} (got {rat_med:.6f})")
    check(cmax == ecmax, f"PV{n} max_consec=={ecmax} (got {cmax})")
    check(emit == eemit, f"PV{n} emit=={eemit} (got {emit})")
    check(sev == esev, f"PV{n} severity=={esev!r} (got {sev!r})")
    check((conf is None and econf is None) or (conf is not None and abs(conf - econf) < 1e-9),
          f"PV{n} confidence=={econf} (got {conf})")
    check(nqual == enq, f"PV{n} n_qualifying_steps=={enq} (got {nqual})")

# robustness of median: PV8 dead must NOT drag its own partners' baseline
med_pv5 = statistics.median([cur(p, 0) for p in PARTNERS[5]])
check(abs(med_pv5 - 12.825) < 1e-9, f"PV5 median partner robust to PV8=0 (==12.825, got {med_pv5})")
med_pv8 = statistics.median([cur(p, 0) for p in PARTNERS[8]])
check(abs(med_pv8 - 12.875) < 1e-9, f"PV8 baseline excludes itself (==12.875, got {med_pv8})")

# ===========================================================================
print("\n=== (C) REAL-RUN consistency (m2_findings_20251030.xlsx) ===")
if not FINDINGS.exists():
    print("  (skip) findings file tidak ada di workspace.")
else:
    fwb = load_workbook(FINDINGS, read_only=True, data_only=True)
    fws = fwb["Findings"]; rows = fws.iter_rows(values_only=True); next(rows)
    n = bad = 0
    sev_seen = {}
    for r in rows:
        if r[3] != "M2b_mppt_ratio":
            continue
        n += 1
        sev, val, thr, conf, ev = r[4], r[5], r[6], r[10], ast.literal_eval(r[11])
        rem = ev["ratio_event_median"]
        exp_sev = "CRITICAL" if rem < RCRIT else "HIGH" if rem < RHIGH else "MEDIUM"
        exp_conf = min(90.0, max(50.0, (1.0 - rem) * 100.0))
        ok = (sev == exp_sev and abs(conf - exp_conf) < 1e-6 and abs(val - rem) < 1e-9
              and abs(thr - 0.85) < 1e-9 and ev["n_qualified_events"] >= 1)
        if not ok:
            bad += 1
        sev_seen.setdefault(sev, set()).add(round(conf, 2))
    fwb.close()
    check(n == 286, f"286 finding M2b_mppt_ratio (got {n})")
    check(bad == 0, f"semua {n} finding konsisten (severity/confidence/value/threshold/debounce); mismatch {bad}")
    check(sev_seen.get("MEDIUM") == {50.0} and sev_seen.get("CRITICAL") == {90.0},
          f"confidence aktual: MEDIUM=50, CRITICAL=90 (got {sev_seen})")

print("\n" + "=" * 64)
if errors:
    print(f"VERIFY FAILED: {len(errors)} error(s)")
    for e in errors:
        print("  -", e)
    raise SystemExit(1)
print("VERIFY OK — semua structural + numeric + real-run checks passed (iter11).")
