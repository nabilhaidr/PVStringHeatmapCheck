"""Verify Iterasi 8 (M2aShading) Excel vs proto_iter8 locked numbers.

(A) STRUCTURAL audit — formula Helpers_SH (inv_total, CV_ts), SH_Hourly (cv_hour, pr_proxy,
    suspicious), M2a_Shading (median, threshold, n_am, asymmetry, fault_type, severity).
(B) NUMERIC recompute — baca literal Raw_Data_SH, re-implement CV_ts(STDEVP/AVG) → median per jam,
    pr_proxy, threshold, suspicious, asimetri AM/PM, klasifikasi, severity. Assert == proto.
"""
from __future__ import annotations
import numpy as np
from openpyxl import load_workbook

WB = "M2_PV_Performance_Workbook.xlsx"
HOURS = [8, 9, 10, 11, 12, 13, 14, 15]
NPV, NTS = 6, 6
DF = 5
errors = []
def check(c, m):
    print(("  OK  " if c else "  XX  FAIL: ") + m)
    if not c: errors.append(m)

wb = load_workbook(WB, data_only=False)
for s in ("Raw_Data_SH", "Helpers_SH", "SH_Hourly", "M2a_Shading"):
    assert s in wb.sheetnames, f"missing {s}"
raw = wb["Raw_Data_SH"]; hp = wb["Helpers_SH"]; shh = wb["SH_Hourly"]; dc = wb["M2a_Shading"]

def cfgval(n):
    dn = wb.defined_names[n]; (s, c), = dn.destinations; return wb[s][c].value
POA_THR = cfgval("cfg_sh_poa_threshold"); CV_MULT = cfgval("cfg_sh_cv_mult")
PR_MULT = cfgval("cfg_sh_pr_mult"); AMPM = cfgval("cfg_sh_am_pm_split")
ASYM_THR = cfgval("cfg_sh_asymmetry_thr")
print("config:", dict(cv_mult=CV_MULT, pr_mult=PR_MULT, am_pm=AMPM, asym_thr=ASYM_THR))
check(CV_MULT == 0.5 and PR_MULT == 0.85, "multipliers 0.5/0.85")
check(AMPM == 12.0 and ASYM_THR == 0.5, "am_pm 12 / asym_thr 0.5")

# ===========================================================================
print("\n=== (A) STRUCTURAL audit ===")
for ri in (5, 10, 47, 52):
    check(hp[f"D{ri}"].value == f"=SUM(Raw_Data_SH!E{ri}:J{ri})", f"Helpers D{ri} inv_total")
    check(hp[f"E{ri}"].value == f"=STDEVP(Raw_Data_SH!E{ri}:J{ri})/AVERAGE(Raw_Data_SH!E{ri}:J{ri})", f"Helpers E{ri} CV_ts")
for k, h in enumerate(HOURS):
    r = 5 + k; b0 = DF + k*NTS; b1 = b0 + NTS - 1
    check(shh[f"C{r}"].value == f"=MEDIAN(Helpers_SH!E{b0}:E{b1})", f"SH_Hourly C{r} cv_hour (h{h})")
    check(shh[f"D{r}"].value == f"=AVERAGE(Helpers_SH!D{b0}:D{b1})/MAX(AVERAGE(Helpers_SH!C{b0}:C{b1}),0.000001)", f"SH_Hourly D{r} pr_proxy")
    check(shh[f"H{r}"].value == f"=IF(AND(C{r}<M2a_Shading!$B$7,D{r}<M2a_Shading!$B$8),1,0)", f"SH_Hourly H{r} suspicious")
check(dc["B7"].value == "=cfg_sh_cv_mult*B5", "M2a cv_threshold")
check(dc["B11"].value == "=SUMPRODUCT(SH_Hourly!H5:H12*(SH_Hourly!A5:A12<cfg_sh_am_pm_split))", "M2a n_am SUMPRODUCT")
check(dc["B13"].value == "=ABS(B11-B12)/MAX(B11+B12,1)", "M2a asymmetry")
check(dc["B17"].value == '=IF(B10=0,"NORMAL",IF(B16>=0.6,"CRITICAL",IF(B16>=0.4,"HIGH",IF(B16>=0.2,"MEDIUM","INFO"))))', "M2a severity")

# ===========================================================================
print("\n=== (B) NUMERIC recompute from literals ===")
def cv_pop(arr):
    a = np.array([x for x in arr if x is not None and np.isfinite(x) and x > 0])
    return np.std(a)/a.mean() if a.size >= 2 and a.mean() > 0 else np.nan

# read per-ts
ts = []   # (hour, poa, inv_total, cv_ts)
for k, h in enumerate(HOURS):
    for ti in range(NTS):
        ri = DF + k*NTS + ti
        poa = float(raw[f"D{ri}"].value)
        powers = [float(raw[f"{chr(ord('E')+p)}{ri}"].value) for p in range(NPV)]
        ts.append((h, poa, sum(powers), cv_pop(powers)))

hourly = {}
for h in HOURS:
    rs = [t for t in ts if t[0] == h]
    cv_h = float(np.median([t[3] for t in rs]))
    pr_h = float(np.mean([t[2] for t in rs])) / max(float(np.mean([t[1] for t in rs])), 1e-6)
    hourly[h] = (cv_h, pr_h)
cv_med = float(np.median([hourly[h][0] for h in HOURS]))
pr_med = float(np.median([hourly[h][1] for h in HOURS]))
cv_thr = CV_MULT*cv_med; pr_thr = PR_MULT*pr_med
n_am = n_pm = n_susp = 0
for h in HOURS:
    if hourly[h][0] < cv_thr and hourly[h][1] < pr_thr:
        n_susp += 1
        if h < AMPM: n_am += 1
        else: n_pm += 1
asym = abs(n_am-n_pm)/max(n_am+n_pm, 1)
fault = "no_shading" if n_susp == 0 else ("shading_uniform" if asym < ASYM_THR else ("shading_morning" if n_am > n_pm else "shading_afternoon"))
frac = n_susp/len(HOURS); score = frac*0.7 + asym*0.3
sev = "NORMAL" if n_susp == 0 else ("CRITICAL" if score >= 0.6 else "HIGH" if score >= 0.4 else "MEDIUM" if score >= 0.2 else "INFO")
conf = 50 + asym*50

print(f"{'hour':>5}{'cv':>10}{'pr':>10}{'suspicious':>12}")
for h in HOURS:
    susp = hourly[h][0] < cv_thr and hourly[h][1] < pr_thr
    print(f"{h:>5}{hourly[h][0]:>10.5f}{hourly[h][1]:>10.5f}{str(susp):>12}")
print(f"cv_med={cv_med:.5f} cv_thr={cv_thr:.5f} | pr_med={pr_med:.5f} pr_thr={pr_thr:.5f}")
print(f"n_susp={n_susp} n_am={n_am} n_pm={n_pm} asym={asym:.3f} fault={fault} score={score:.4f} sev={sev} conf={conf:.1f}")

check(abs(hourly[8][0] - 0.012910) < 1e-5, f"cv shaded≈0.01291 (got {hourly[8][0]:.5f})")
check(abs(hourly[11][0] - 0.064550) < 1e-5, f"cv normal≈0.06455 (got {hourly[11][0]:.5f})")
check(abs(cv_thr - 0.032275) < 1e-5, f"cv_thr≈0.032275 (got {cv_thr:.5f})")
check(abs(pr_thr - 0.017) < 1e-6, f"pr_thr≈0.017 (got {pr_thr:.5f})")
check(n_susp == 3 and n_am == 3 and n_pm == 0, f"n_susp/am/pm 3/3/0 (got {n_susp}/{n_am}/{n_pm})")
check(abs(asym - 1.0) < 1e-9, f"asymmetry==1.0 (got {asym})")
check(fault == "shading_morning", f"fault==shading_morning (got {fault})")
check(sev == "HIGH", f"severity==HIGH (got {sev})")
check(abs(conf - 100.0) < 1e-9, f"confidence==100 (got {conf})")
# STDEVP parity (population std)
fn = np.array([1.00, 1.10, 0.90, 1.05, 0.95, 1.00])
check(np.isclose(np.std(fn), np.sqrt(np.mean((fn-fn.mean())**2))), "STDEVP == sqrt(mean dev^2)")

print("\n" + "="*60)
if errors:
    print(f"VERIFY FAILED: {len(errors)} error(s)")
    for e in errors:
        print("  -", e)
    raise SystemExit(1)
print("VERIFY OK — all structural + numeric checks passed (iter8).")
