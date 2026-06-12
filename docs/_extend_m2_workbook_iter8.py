"""Extend M2 PV Performance Workbook — Iterasi 8: M2aShading.

Input : docs/M2_PV_Performance_Workbook.xlsx (output Iterasi 7, 33 sheet)
Output: same file, +4 sheet baru:
  34. Raw_Data_SH   — daya per-PV + POA per timestamp (1 inverter, 8 jam × 6 ts)
  35. Helpers_SH    — per timestamp: inv_total, CV_ts = STDEVP/AVERAGE antar PV
  36. SH_Hourly     — per jam: cv_hour (MEDIAN CV_ts), pr_proxy, suspicious flag, AM/PM
  37. M2a_Shading   — median refs + threshold + n_am/n_pm + asimetri + klasifikasi + severity

Reverse-engineering shading.py:
  CV_ts = std_pop(daya antar PV)/mean ; cv_hour = MEDIAN(CV_ts) ; pr_proxy = mean(inv)/mean(POA).
  cv_thr=0.5·median(cv_hour) ; pr_thr=0.85·median(pr_proxy). suspicious = cv<cv_thr AND pr<pr_thr.
  asymmetry=|n_am−n_pm|/max(n_am+n_pm,1) ; <0.5 uniform, n_am>n_pm morning, else afternoon.
  severity score = frac·0.7 + asym·0.3 (≥.6 CRIT, ≥.4 HIGH, ≥.2 MED). confidence=50+asym·50.
Demo (angka EXACT proto_iter8.py): jam 8/9/10 shaded (CV seragam rendah + PR rendah) → shading_morning HIGH.
"""
from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule

INPUT = Path(__file__).parent / "M2_PV_Performance_Workbook.xlsx"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=13, bold=True, color="305496")
NOTE_FONT = Font(italic=True, size=9, color="808080")
SEV_FILL = {
    "CRITICAL": PatternFill("solid", fgColor="E06666"),
    "HIGH":     PatternFill("solid", fgColor="F6B26B"),
    "MEDIUM":   PatternFill("solid", fgColor="FFE599"),
    "INFO":     PatternFill("solid", fgColor="D9D9D9"),
    "NORMAL":   PatternFill("solid", fgColor="B6D7A8"),
}
SHADE_FILL = PatternFill("solid", fgColor="FCE5CD")   # shaded-hour tint


def set_header(ws, row, headers, start=1):
    for ci, h in enumerate(headers, start=start):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def title_note(ws, title, notes):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    for i, nt in enumerate(notes, start=2):
        ws.cell(row=i, column=1, value=nt).font = NOTE_FONT


wb = load_workbook(INPUT)
existing = list(wb.sheetnames)
assert len(existing) == 33, f"Expected 33 sheets from iter7, got {len(existing)}"
print(f"Loaded {len(existing)} sheets.")

# ===========================================================================
# README — insert row 13 (iter8) before "Cara membaca"
# ===========================================================================
ws = wb["README"]
assert ws.cell(row=13, column=1).value == "Cara membaca", "README layout berubah; abort"
ws.insert_rows(13)
ws.cell(row=13, column=1, value="8").border = BORDER
ws.cell(row=13, column=2, value="2026-05-30").border = BORDER
ws.cell(row=13, column=3, value="M2aShading").border = BORDER
ws.cell(row=13, column=4, value="Raw_Data_SH, Helpers_SH, SH_Hourly, M2a_Shading").border = BORDER
for rr in range(14, 22):
    if ws.cell(row=rr, column=1).value:
        ws.cell(row=rr, column=1).font = NOTE_FONT
ws.cell(row=13, column=1).font = Font(name="Calibri", size=11)

# ===========================================================================
# Config — append m2a_shading params + named cells (start row 59)
# ===========================================================================
ws = wb["Config"]
r = 4; last_cfg = 4
while ws.cell(row=r, column=1).value is not None:
    last_cfg = r; r += 1
start = last_cfg + 1
assert start == 59, f"Expected Config append at row 59, got {start}"
new_cfg = [
    ("m2a_shading", "poa_threshold_wm2", 100.0, "gate daylight (W/m²)", "cfg_sh_poa_threshold"),
    ("m2a_shading", "cv_low_multiplier", 0.5, "cv_hour < ini·median(cv) → CV rendah (uniform)", "cfg_sh_cv_mult"),
    ("m2a_shading", "pr_low_multiplier", 0.85, "pr_hour < ini·median(pr) → underperform", "cfg_sh_pr_mult"),
    ("m2a_shading", "am_pm_split_hour", 12.0, "batas AM/PM untuk asimetri diurnal", "cfg_sh_am_pm_split"),
    ("m2a_shading", "asymmetry_threshold", 0.5, "asimetri < ini → uniform (soiling/awan), bukan terrain", "cfg_sh_asymmetry_thr"),
    ("m2a_shading", "min_hours_for_analysis", 4, "min jam untuk median referensi", "cfg_sh_min_hours"),
    ("m2a_shading", "min_samples_per_hour", 5, "min sampel (PV×ts) per jam", "cfg_sh_min_samp_hour"),
]
for i, (sec, key, val, note, name) in enumerate(new_cfg):
    ri = start + i
    ws.cell(row=ri, column=1, value=sec).border = BORDER
    ws.cell(row=ri, column=2, value=key).border = BORDER
    c = ws.cell(row=ri, column=3, value=val); c.border = BORDER
    c.fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(row=ri, column=4, value=note).border = BORDER
    if name not in wb.defined_names:
        wb.defined_names[name] = DefinedName(name, attr_text=f"Config!$C${ri}")
print(f"Config: appended {len(new_cfg)} m2a_shading rows ({start}..{start+len(new_cfg)-1}) + named cells.")

# ===========================================================================
# Synthetic data (matches proto_iter8.py EXACTLY)
# ===========================================================================
HOURS = [8, 9, 10, 11, 12, 13, 14, 15]
SHADED = {8, 9, 10}
NPV, NTS = 6, 6
f_normal = np.array([1.00, 1.10, 0.90, 1.05, 0.95, 1.00])
f_shaded = np.array([1.00, 1.02, 0.98, 1.01, 0.99, 1.00])
m_ts = np.array([0.97, 0.99, 1.00, 1.01, 1.00, 1.03])
POA_H = {8: 350.0, 9: 550.0, 10: 720.0, 11: 850.0, 12: 900.0, 13: 820.0, 14: 650.0, 15: 450.0}
PR_LEVEL = {h: (0.010 if h in SHADED else 0.020) for h in HOURS}
records = []   # (hour, ts_idx, Start Time, POA_ts, [P1..P6])
for h in HOURS:
    f = f_shaded if h in SHADED else f_normal
    for ti in range(NTS):
        poa_ts = POA_H[h] * m_ts[ti]
        powers = PR_LEVEL[h] * POA_H[h] * m_ts[ti] * f / NPV
        t = datetime(2026, 5, 14, h, ti * 10)
        records.append((h, ti, t, float(poa_ts), [float(x) for x in powers]))
N = len(records)            # 48
DF, DL = 5, 5 + N - 1       # 5..52

# ===========================================================================
# Sheet 34: Raw_Data_SH
# ===========================================================================
ws = wb.create_sheet("Raw_Data_SH")
title_note(ws, "Raw_Data_SH — daya per-PV (kW) + POA per timestamp (WB05-INV01)", [
    "1 inverter, 6 PV, 8 jam (08:00–15:00) × 6 timestamp/jam = 48 baris. Jam 8/9/10 di-shade (di-highlight): daya turun seragam → CV rendah + PR rendah.",
    "Daya per-PV = Σ source PV{n} Power(kW) (atau V·I/1000). Ganti dengan data Huawei aktual (paste over).",
])
PCOL = {pv: get_column_letter(4 + pv) for pv in range(1, NPV + 1)}   # E..J power
set_header(ws, 4, ["inverter_id", "hour", "Start Time", "POA (W/m²)"]
          + [f"PV{pv} Power(kW)" for pv in range(1, NPV + 1)])
for idx, (h, ti, t, poa, powers) in enumerate(records):
    ri = DF + idx
    ws.cell(row=ri, column=1, value="WB05-INV01").border = BORDER
    ws.cell(row=ri, column=2, value=h).border = BORDER
    c = ws.cell(row=ri, column=3, value=t); c.number_format = "hh:mm"; c.border = BORDER
    ws.cell(row=ri, column=4, value=poa).number_format = "0.0"; ws.cell(row=ri, column=4).border = BORDER
    for pv in range(1, NPV + 1):
        cc = ws.cell(row=ri, column=4 + pv, value=powers[pv-1]); cc.number_format = "0.0000"; cc.border = BORDER
        if h in SHADED:
            cc.fill = SHADE_FILL
ws.column_dimensions["A"].width = 13
ws.column_dimensions["C"].width = 11

# ===========================================================================
# Sheet 35: Helpers_SH (per timestamp)
# ===========================================================================
ws = wb.create_sheet("Helpers_SH")
title_note(ws, "Helpers_SH — per timestamp: inv_total + CV antar PV", [
    "inv_total = Σ daya PV. CV_ts = STDEVP(daya PV)/AVERAGE(daya PV) — variasi antar string (uniform shading → CV kecil).",
    "Production pakai hanya daya finite & >0 (di sini semua positif). cv_hour = MEDIAN(CV_ts) dihitung di SH_Hourly.",
])
set_header(ws, 4, ["hour", "Start Time", "POA", "inv_total (kW)", "CV_ts"])
for idx in range(N):
    ri = DF + idx
    ws.cell(row=ri, column=1, value=f"=Raw_Data_SH!B{ri}")
    ws.cell(row=ri, column=2, value=f"=Raw_Data_SH!C{ri}").number_format = "hh:mm"
    ws.cell(row=ri, column=3, value=f"=Raw_Data_SH!D{ri}").number_format = "0.0"
    ws.cell(row=ri, column=4, value=f"=SUM(Raw_Data_SH!E{ri}:J{ri})").number_format = "0.000"
    ws.cell(row=ri, column=5, value=f"=STDEVP(Raw_Data_SH!E{ri}:J{ri})/AVERAGE(Raw_Data_SH!E{ri}:J{ri})").number_format = "0.00000"
    for ci in range(1, 6):
        ws.cell(row=ri, column=ci).border = BORDER
ws.column_dimensions["B"].width = 11

# ===========================================================================
# Sheet 36: SH_Hourly (per jam)
# ===========================================================================
ws = wb.create_sheet("SH_Hourly")
title_note(ws, "SH_Hourly — agregat per jam: cv_hour, pr_proxy, suspicious", [
    "cv_hour = MEDIAN(CV_ts dlm jam). pr_proxy = AVERAGE(inv_total)/AVERAGE(POA) dlm jam. "
    "suspicious = cv_hour < cv_threshold DAN pr_proxy < pr_threshold (threshold di M2a_Shading).",
])
set_header(ws, 4, ["hour", "AM/PM", "cv_hour", "pr_proxy", "mean_POA", "mean_inv", "n_samples", "suspicious"])
HRDF = 5   # SH_Hourly first hour row
for k, h in enumerate(HOURS):
    r = HRDF + k
    b0 = DF + k * NTS         # block first row in Helpers/Raw (6 rows)
    b1 = b0 + NTS - 1
    ws.cell(row=r, column=1, value=h)
    ws.cell(row=r, column=2, value=f'=IF(A{r}<cfg_sh_am_pm_split,"AM","PM")')
    ws.cell(row=r, column=3, value=f"=MEDIAN(Helpers_SH!E{b0}:E{b1})").number_format = "0.00000"
    ws.cell(row=r, column=4, value=f"=AVERAGE(Helpers_SH!D{b0}:D{b1})/MAX(AVERAGE(Helpers_SH!C{b0}:C{b1}),0.000001)").number_format = "0.00000"
    ws.cell(row=r, column=5, value=f"=AVERAGE(Helpers_SH!C{b0}:C{b1})").number_format = "0.0"
    ws.cell(row=r, column=6, value=f"=AVERAGE(Helpers_SH!D{b0}:D{b1})").number_format = "0.000"
    ws.cell(row=r, column=7, value=f"=COUNT(Helpers_SH!E{b0}:E{b1})*{NPV}")
    ws.cell(row=r, column=8, value=f"=IF(AND(C{r}<M2a_Shading!$B$7,D{r}<M2a_Shading!$B$8),1,0)")
    for ci in range(1, 9):
        ws.cell(row=r, column=ci).border = BORDER
    if h in SHADED:
        ws.cell(row=r, column=1).fill = SHADE_FILL
HRDL = HRDF + len(HOURS) - 1   # 12

# ===========================================================================
# Sheet 37: M2a_Shading (keputusan)
# ===========================================================================
ws = wb.create_sheet("M2a_Shading")
title_note(ws, "M2a_Shading — keputusan per inverter (median ref + asimetri diurnal)", [
    "cv_median/pr_median = MEDIAN cv_hour/pr_proxy lintas jam. threshold = mult·median. "
    "asimetri AM/PM membedakan terrain shadow (asimetris) vs soiling/awan (uniform). severity = frac·0.7 + asim·0.3.",
    "IKN: lat −0.99, panel hadap utara → performa AM≈PM saat normal; suspicious terkonsentrasi AM/PM = sinyal shading terrain.",
])
set_header(ws, 4, ["metric", "value"])
rows = [
    ("inverter_id", '="WB05-INV01"'),
    ("cv_median", "=MEDIAN(SH_Hourly!C5:C12)"),
    ("pr_median", "=MEDIAN(SH_Hourly!D5:D12)"),
    ("cv_threshold", "=cfg_sh_cv_mult*B6"),
    ("pr_threshold", "=cfg_sh_pr_mult*B7"),
    ("total_hours", "=COUNT(SH_Hourly!C5:C12)"),
    ("n_suspicious", "=SUM(SH_Hourly!H5:H12)"),
    ("n_am", "=SUMPRODUCT(SH_Hourly!H5:H12*(SH_Hourly!A5:A12<cfg_sh_am_pm_split))"),
    ("n_pm", "=B11-B12"),
    ("asymmetry", "=ABS(B12-B13)/MAX(B12+B13,1)"),
    ("fault_type", '=IF(B11=0,"no_shading",IF(B14<cfg_sh_asymmetry_thr,"shading_uniform",IF(B12>B13,"shading_morning","shading_afternoon")))'),
    ("frac_suspicious", "=B11/B10"),
    ("score", "=B16*0.7+B14*0.3"),
    ("severity", '=IF(B11=0,"NORMAL",IF(B17>=0.6,"CRITICAL",IF(B17>=0.4,"HIGH",IF(B17>=0.2,"MEDIUM","INFO"))))'),
    ("confidence", "=50+B14*50"),
]
# rows start at 5: B5 inverter, B6 cv_median, B7 cv? -- map carefully
# layout: r5 inverter_id, r6 cv_median, r7 pr_median, r8 cv_threshold, r9 pr_threshold,
# but SH_Hourly references M2a_Shading!$B$7 (cv_thr) & $B$8 (pr_thr). Adjust order so cv_thr=B7, pr_thr=B8.
rows = [
    ("cv_median", "=MEDIAN(SH_Hourly!C5:C12)"),          # B5
    ("pr_median", "=MEDIAN(SH_Hourly!D5:D12)"),          # B6
    ("cv_threshold", "=cfg_sh_cv_mult*B5"),              # B7
    ("pr_threshold", "=cfg_sh_pr_mult*B6"),              # B8
    ("total_hours", "=COUNT(SH_Hourly!C5:C12)"),         # B9
    ("n_suspicious", "=SUM(SH_Hourly!H5:H12)"),          # B10
    ("n_am", "=SUMPRODUCT(SH_Hourly!H5:H12*(SH_Hourly!A5:A12<cfg_sh_am_pm_split))"),  # B11
    ("n_pm", "=B10-B11"),                                # B12
    ("asymmetry", "=ABS(B11-B12)/MAX(B11+B12,1)"),       # B13
    ("fault_type", '=IF(B10=0,"no_shading",IF(B13<cfg_sh_asymmetry_thr,"shading_uniform",IF(B11>B12,"shading_morning","shading_afternoon")))'),  # B14
    ("frac_suspicious", "=B10/B9"),                      # B15
    ("score", "=B15*0.7+B13*0.3"),                       # B16
    ("severity", '=IF(B10=0,"NORMAL",IF(B16>=0.6,"CRITICAL",IF(B16>=0.4,"HIGH",IF(B16>=0.2,"MEDIUM","INFO"))))'),  # B17
    ("confidence", "=50+B13*50"),                        # B18
]
fmt = {"cv_median": "0.00000", "pr_median": "0.00000", "cv_threshold": "0.00000",
       "pr_threshold": "0.00000", "asymmetry": "0.000", "frac_suspicious": "0.000",
       "score": "0.0000", "confidence": "0.0"}
for i, (label, formula) in enumerate(rows):
    r = 5 + i
    ws.cell(row=r, column=1, value=label).border = BORDER
    c = ws.cell(row=r, column=2, value=formula); c.border = BORDER
    if label in fmt:
        c.number_format = fmt[label]
for sev, fill in SEV_FILL.items():
    ws.conditional_formatting.add("B17:B17",
        CellIsRule(operator="equal", formula=[f'"{sev}"'], fill=fill))
ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 24

# ===========================================================================
# Save + integrity verify
# ===========================================================================
wb.save(INPUT)
wb2 = load_workbook(INPUT, data_only=False)
after = list(wb2.sheetnames)
new_sheets = ["Raw_Data_SH", "Helpers_SH", "SH_Hourly", "M2a_Shading"]
assert after[:33] == existing, "EXISTING SHEET ORDER CHANGED!"
assert after[33:] == new_sheets, f"New sheets mismatch: {after[33:]}"
print(f"\nSheets now: {len(after)} (was 33, +4)")
h = wb2["Helpers_SH"]; sh = wb2["SH_Hourly"]; d = wb2["M2a_Shading"]
print("Helpers CV_ts E5:", h["E5"].value)
print("SH_Hourly cv_hour C5:", sh["C5"].value, "| suspicious H5:", sh["H5"].value)
print("M2a_Shading cv_threshold B7:", d["B7"].value, "| fault_type B14:", d["B14"].value, "| severity B17:", d["B17"].value)
print("OK — iter8 build complete.")
