"""Extend M2 PV Performance Workbook — Iterasi 4: M2bOpenCircuit.

Input : docs/M2_PV_Performance_Workbook.xlsx (output Iterasi 3, 16 sheet)
Output: same file, +4 sheet baru:
  17. Raw_Data_OC        — time-series I per PV string (WB05-INV05) + POA
  18. Helpers_OC         — per-row I_q95 (PERCENTILE across siblings), ratio, qualifying, debounce counter
  19. M2b_OpenCircuit    — per-PV decision: max_consec >= debounce -> emit open_circuit (CRITICAL 95%)
  20. M2b_OC_StringStatus— replika artifact StringStatus Python

Sheet existing Iterasi 1-3 TIDAK diubah (kecuali append: README log row, Config m2b_open_circuit rows).
Cell-level formula reproducible. Skenario dummy:
  PV1,PV2 healthy ; PV3 GENUINE open-circuit (sustained) ; PV4 glitch 3-step (debounce gating) ; PV5 EMPTY.
Debounce = 20 (PRODUKSI value, ~100 menit @5-min). 24 daylight row + 2 twilight row.

Kunci reverse-engineering open_circuit.py:
  - I_q95 = quantile(0.95, axis=1) ACROSS siblings == Excel PERCENTILE(...,0.95) [linear interp, verified parity]
  - ratio = I_string / max(I_q95, 0.01)
  - qualifying = (ratio < 0.05) AND daylight
  - emit <=> ADA run qualifying konsekutif dengan panjang >= debounce  <=>  MAX(running_consec) >= debounce
    (count_debounced_events: n_events>0). Empty PV di-skip dari emit.
"""
from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule

INPUT = Path(__file__).parent / "M2_PV_Performance_Workbook.xlsx"

# --- Styling (mirror iter3) ---
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=13, bold=True, color="305496")
NOTE_FONT = Font(italic=True, size=9, color="808080")
SEV_FILL = {
    "CRITICAL":     PatternFill("solid", fgColor="E06666"),
    "open_circuit": PatternFill("solid", fgColor="E06666"),
    "NORMAL":       PatternFill("solid", fgColor="B6D7A8"),
    "EMPTY":        PatternFill("solid", fgColor="DDDDDD"),
}
GLITCH_FILL = PatternFill("solid", fgColor="FCE5CD")  # highlight glitch cells
FAULT_FILL = PatternFill("solid", fgColor="F4CCCC")   # highlight genuine fault col


def set_header(ws, row, headers):
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def title_note(ws, title, notes):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    for i, nt in enumerate(notes, start=2):
        ws.cell(row=i, column=1, value=nt).font = NOTE_FONT


wb = load_workbook(INPUT)
existing = list(wb.sheetnames)
assert len(existing) == 16, f"Expected 16 sheets from iter3, got {len(existing)}: {existing}"
print(f"Loaded {len(existing)} sheets.")

# ===========================================================================
# README iterasi log — append row 9 (Iterasi 4)
# ===========================================================================
ws = wb["README"]
ws.cell(row=9, column=1, value="4").border = BORDER
ws.cell(row=9, column=2, value="2026-05-29").border = BORDER
ws.cell(row=9, column=3, value="M2bOpenCircuit").border = BORDER
ws.cell(row=9, column=4,
        value="Raw_Data_OC, Helpers_OC, M2b_OpenCircuit, M2b_OC_StringStatus").border = BORDER

# ===========================================================================
# Config — append m2b_open_circuit thresholds + named cells (start row 32)
# ===========================================================================
ws = wb["Config"]
r = 4
last_cfg = 4
while ws.cell(row=r, column=1).value is not None:
    last_cfg = r
    r += 1
start = last_cfg + 1
assert start == 32, f"Expected Config append at row 32, got {start}"

new_cfg = [
    ("m2b_open_circuit", "poa_threshold_wm2",          700.0, "W/m²; daylight gate utama (config override; spec 4.2.3 tulis 200)", "cfg_oc_poa_threshold_wm2"),
    ("m2b_open_circuit", "poa_floor_wm2",              50.0,  "W/m²; hard floor sunset/twilight (sensor-lag protection)", "cfg_oc_poa_floor_wm2"),
    ("m2b_open_circuit", "i_ratio_threshold",          0.05,  "I_string/I_q95 < ini → qualifying (spec 4.2.3: <5%)", "cfg_oc_i_ratio_threshold"),
    ("m2b_open_circuit", "debounce_consecutive_steps", 20,    "qualifying harus ≥N step konsekutif → genuine event (~100min @5-min)", "cfg_oc_debounce_steps"),
    ("m2b_open_circuit", "confidence_pct",             95.0,  "confidence finding (spec 4.2.3)", "cfg_oc_confidence_pct"),
    ("m2b_open_circuit", "iq95_clip_floor_a",          0.01,  "clip(lower) I_q95 untuk hindari div-by-zero", "cfg_oc_iq95_clip"),
]
for i, (sec, key, val, note, name) in enumerate(new_cfg):
    ri = start + i
    ws.cell(row=ri, column=1, value=sec).border = BORDER
    ws.cell(row=ri, column=2, value=key).border = BORDER
    c = ws.cell(row=ri, column=3, value=val)
    c.border = BORDER
    c.fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(row=ri, column=4, value=note).border = BORDER
    if name not in wb.defined_names:
        wb.defined_names[name] = DefinedName(name, attr_text=f"Config!$C${ri}")
print(f"Config: appended {len(new_cfg)} m2b_open_circuit rows ({start}..{start+len(new_cfg)-1}) + named cells.")

# ===========================================================================
# Synthetic data (matches proto_iter4.py EXACTLY)
# ===========================================================================
day_times = [datetime(2026, 5, 14, 12, 0) + timedelta(minutes=5 * i) for i in range(24)]
twi_times = [datetime(2026, 5, 14, 18, 10), datetime(2026, 5, 14, 18, 20)]
times = day_times + twi_times                     # 26 rows
poa = [900.0] * 24 + [120.0] * 2                  # twilight 120 < 700 → gated
N = len(times)                                     # 26
DAY_LAST = 5 + 24 - 1                              # last daylight Excel row = 28

PV = {
    1: [13.0] * 24 + [0.0] * 2,                   # healthy
    2: [12.8] * 24 + [0.0] * 2,                   # healthy
    3: [0.10] * 24 + [0.0] * 2,                   # GENUINE open-circuit (sustained)
    4: [12.9] * 24 + [0.0] * 2,                   # healthy ...
    5: [0.0] * 26,                                # EMPTY slot (Huawei reports 0)
}
GLITCH_IDX = [4, 5, 6]                             # daylight idx → Excel rows 9,10,11
for gi in GLITCH_IDX:
    PV[4][gi] = 0.05                              # 3-step glitch

# ===========================================================================
# Sheet 17: Raw_Data_OC
# ===========================================================================
ws = wb.create_sheet("Raw_Data_OC")
title_note(ws, "Raw_Data_OC — time-series arus per PV string (WB05-INV05)", [
    "Dummy 1 inverter, 5 PV string. PV1/PV2 healthy · PV3 GENUINE open-circuit · PV4 glitch 3-step · PV5 EMPTY.",
    "24 baris daylight (12:00–13:55 @5-min, POA=900) + 2 baris twilight (18:10/18:20, POA=120 → gated). Ganti dengan data Huawei aktual (paste over).",
])
set_header(ws, 4, ["Start Time", "POA (W/m²)",
                   "PV1 input current(A)", "PV2 input current(A)",
                   "PV3 input current(A)", "PV4 input current(A)", "PV5 input current(A)"])
for idx in range(N):
    ri = 5 + idx
    c = ws.cell(row=ri, column=1, value=times[idx]); c.number_format = "yyyy-mm-dd hh:mm:ss"; c.border = BORDER
    ws.cell(row=ri, column=2, value=poa[idx]).border = BORDER
    for k in range(1, 6):
        cc = ws.cell(row=ri, column=2 + k, value=PV[k][idx])
        cc.number_format = "0.000"
        cc.border = BORDER
        if k == 3:
            cc.fill = FAULT_FILL
        if k == 4 and idx in GLITCH_IDX:
            cc.fill = GLITCH_FILL
ws.column_dimensions["A"].width = 20
for col in "BCDEFG":
    ws.column_dimensions[col].width = 13

# ===========================================================================
# Sheet 18: Helpers_OC
# ===========================================================================
ws = wb.create_sheet("Helpers_OC")
title_note(ws, "Helpers_OC — jembatan open_circuit.py (per-row mekanik)", [
    "I_q95 = PERCENTILE(arus PV1..PV5 baris itu, 0.95) ≡ quantile(0.95,axis=1). ratio = I/MAX(I_q95,0.01). qualifying = (ratio<thr) DAN daylight.",
    "consec = running counter qualifying konsekutif (reset saat gagal). emit <=> MAX(consec) ≥ debounce. Production juga AND solar_elev>5° & inverter-shutdown (ephemeris, tak direplika statis).",
])
hdr = ["Start Time", "POA", "daylight",
       "I_q95",
       "ratio PV1", "ratio PV2", "ratio PV3", "ratio PV4", "ratio PV5",
       "qual PV1", "qual PV2", "qual PV3", "qual PV4", "qual PV5",
       "consec PV1", "consec PV2", "consec PV3", "consec PV4", "consec PV5"]
set_header(ws, 4, hdr)
# Column letters: A time, B POA, C daylight, D I_q95,
#   E..I ratio PV1..5, J..N qual PV1..5, O..S consec PV1..5
RAW_I_COL = {1: "C", 2: "D", 3: "E", 4: "F", 5: "G"}   # Raw_Data_OC current cols
RATIO_COL = {1: "E", 2: "F", 3: "G", 4: "H", 5: "I"}
QUAL_COL = {1: "J", 2: "K", 3: "L", 4: "M", 5: "N"}
CONSEC_COL = {1: "O", 2: "P", 3: "Q", 4: "R", 5: "S"}
for idx in range(N):
    ri = 5 + idx
    ws.cell(row=ri, column=1, value=f"=Raw_Data_OC!A{ri}").number_format = "yyyy-mm-dd hh:mm:ss"
    ws.cell(row=ri, column=2, value=f"=Raw_Data_OC!B{ri}").number_format = "0"
    ws.cell(row=ri, column=3, value=f"=IF(AND(B{ri}>cfg_oc_poa_threshold_wm2,B{ri}>cfg_oc_poa_floor_wm2),1,0)")
    ws.cell(row=ri, column=4, value=f"=PERCENTILE(Raw_Data_OC!C{ri}:G{ri},0.95)").number_format = "0.000"
    for k in range(1, 6):
        rcol = RAW_I_COL[k]
        # ratio
        cc = ws.cell(row=ri, column=4 + k,
                     value=f"=Raw_Data_OC!{rcol}{ri}/MAX($D{ri},cfg_oc_iq95_clip)")
        cc.number_format = "0.0000"
        # qualifying
        ratc = RATIO_COL[k]
        ws.cell(row=ri, column=9 + k,
                value=f"=IF(AND({ratc}{ri}<cfg_oc_i_ratio_threshold,$C{ri}=1),1,0)")
        # consec
        qc = QUAL_COL[k]
        if idx == 0:
            ws.cell(row=ri, column=14 + k, value=f"={qc}{ri}")
        else:
            cc2 = CONSEC_COL[k]
            ws.cell(row=ri, column=14 + k,
                    value=f"=IF({qc}{ri}=1,{cc2}{ri-1}+1,0)")
    for ci in range(1, 20):
        ws.cell(row=ri, column=ci).border = BORDER
ws.column_dimensions["A"].width = 19

# ===========================================================================
# Sheet 19: M2b_OpenCircuit (decision per-PV)
# ===========================================================================
ws = wb.create_sheet("M2b_OpenCircuit")
title_note(ws, "M2b_OpenCircuit — keputusan per-PV (open_circuit.py run())", [
    "emit = (MAX consec ≥ debounce) DAN bukan EMPTY. ratio_median/I_median = MEDIAN over baris daylight (rows 5..28, blok kontigu).",
    "PV3 sustained → run=24 ≥ 20 → CRITICAL conf 95%. PV4 glitch run=3 < 20 → suppressed. PV5 EMPTY → skip (q95 ikut 0-nya, harmless).",
])
set_header(ws, 4, ["PV", "I_median (daylight)", "I_q95_median (daylight)",
                   "ratio_median (daylight)", "MAX consec", "empty_by_design",
                   "emit", "status", "severity", "confidence", "message"])
for k in range(1, 6):
    r = 4 + k
    rcol = RAW_I_COL[k]                # Raw current col
    ratc = RATIO_COL[k]               # Helpers ratio col
    cons = CONSEC_COL[k]              # Helpers consec col
    empty_flag = 1 if k == 5 else 0
    ws.cell(row=r, column=1, value=f"PV{k}")
    ws.cell(row=r, column=2, value=f"=MEDIAN(Raw_Data_OC!{rcol}5:{rcol}{DAY_LAST})").number_format = "0.000"
    ws.cell(row=r, column=3, value=f"=MEDIAN(Helpers_OC!D5:D{DAY_LAST})").number_format = "0.000"
    ws.cell(row=r, column=4, value=f"=MEDIAN(Helpers_OC!{ratc}5:{ratc}{DAY_LAST})").number_format = "0.0000"
    ws.cell(row=r, column=5, value=f"=MAX(Helpers_OC!{cons}5:{cons}{4+N})")
    ws.cell(row=r, column=6, value=empty_flag)
    ws.cell(row=r, column=7, value=f"=IF(AND(E{r}>=cfg_oc_debounce_steps,F{r}=0),1,0)")
    ws.cell(row=r, column=8, value=f'=IF(F{r}=1,"EMPTY",IF(G{r}=1,"open_circuit","NORMAL"))')
    ws.cell(row=r, column=9, value=f'=IF(G{r}=1,"CRITICAL","")')
    ws.cell(row=r, column=10, value=f"=IF(G{r}=1,cfg_oc_confidence_pct,\"\")").number_format = "0"
    ws.cell(row=r, column=11,
            value=(f'=IF(G{r}=1,"Open-circuit suspect "&A{r}&": I/I_q95 median="&TEXT(D{r},"0.000")'
                   f'&" (run "&E{r}&" >= debounce "&cfg_oc_debounce_steps&")",'
                   f'IF(F{r}=1,"EMPTY slot — skip per EmptyPVMap",""))'))
    for ci in range(1, 12):
        ws.cell(row=r, column=ci).border = BORDER
# conditional format status col H (rows 5..9)
ws.conditional_formatting.add("H5:H9",
    CellIsRule(operator="equal", formula=['"open_circuit"'], fill=SEV_FILL["open_circuit"]))
ws.conditional_formatting.add("H5:H9",
    CellIsRule(operator="equal", formula=['"NORMAL"'], fill=SEV_FILL["NORMAL"]))
ws.conditional_formatting.add("H5:H9",
    CellIsRule(operator="equal", formula=['"EMPTY"'], fill=SEV_FILL["EMPTY"]))
ws.column_dimensions["A"].width = 6
for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
    ws.column_dimensions[col].width = 13
ws.column_dimensions["K"].width = 52

# ===========================================================================
# Sheet 20: M2b_OC_StringStatus (replika artifact Python)
# ===========================================================================
ws = wb.create_sheet("M2b_OC_StringStatus")
title_note(ws, "M2b_OC_StringStatus — replika artifact StringStatus (open_circuit.py)", [
    "Mirror self.artifacts['StringStatus']: per-PV status + median stats + debounce counts. EMPTY rows dari top-up EmptyPVMap.",
])
set_header(ws, 4, ["poa_source", "inverter_id", "wb_id", "pv_string", "status",
                   "i_string_median_daylight", "i_q95_median_daylight", "ratio_median_daylight",
                   "n_qualifying_steps", "n_debounced_events", "emitted_finding", "daylight_samples"])
for k in range(1, 6):
    r = 4 + k
    dec = 4 + k                        # decision sheet row
    qc = QUAL_COL[k]
    ws.cell(row=r, column=1, value="dummy_single")
    ws.cell(row=r, column=2, value="WB05-INV05")
    ws.cell(row=r, column=3, value="WB05")
    ws.cell(row=r, column=4, value=f"PV{k}")
    ws.cell(row=r, column=5, value=f"=M2b_OpenCircuit!H{dec}")
    ws.cell(row=r, column=6, value=f"=M2b_OpenCircuit!B{dec}").number_format = "0.000"
    ws.cell(row=r, column=7, value=f"=M2b_OpenCircuit!C{dec}").number_format = "0.000"
    ws.cell(row=r, column=8, value=f"=M2b_OpenCircuit!D{dec}").number_format = "0.0000"
    ws.cell(row=r, column=9, value=f"=SUM(Helpers_OC!{qc}5:{qc}{4+N})")
    ws.cell(row=r, column=10, value=f"=M2b_OpenCircuit!G{dec}")
    ws.cell(row=r, column=11, value=f"=IF(M2b_OpenCircuit!G{dec}=1,TRUE,FALSE)")
    ws.cell(row=r, column=12, value=f"=SUM(Helpers_OC!$C$5:$C${4+N})")
    for ci in range(1, 13):
        ws.cell(row=r, column=ci).border = BORDER
ws.conditional_formatting.add("E5:E9",
    CellIsRule(operator="equal", formula=['"open_circuit"'], fill=SEV_FILL["open_circuit"]))
ws.conditional_formatting.add("E5:E9",
    CellIsRule(operator="equal", formula=['"NORMAL"'], fill=SEV_FILL["NORMAL"]))
ws.conditional_formatting.add("E5:E9",
    CellIsRule(operator="equal", formula=['"EMPTY"'], fill=SEV_FILL["EMPTY"]))
for col in ["A", "B", "F", "G", "H"]:
    ws.column_dimensions[col].width = 14

# ===========================================================================
# Save + integrity verify
# ===========================================================================
wb.save(INPUT)

wb2 = load_workbook(INPUT, data_only=False)
after = list(wb2.sheetnames)
new_sheets = ["Raw_Data_OC", "Helpers_OC", "M2b_OpenCircuit", "M2b_OC_StringStatus"]
assert after[:16] == existing, "EXISTING SHEET ORDER CHANGED!"
assert after[16:] == new_sheets, f"New sheets mismatch: {after[16:]}"
print(f"\nSheets now: {len(after)} (was 16, +4)")
print("New:", new_sheets)
# spot-check a few formulas survived
h = wb2["Helpers_OC"]
print("Helpers_OC D5 (I_q95):", h["D5"].value)
print("Helpers_OC Q5,Q6 (consec PV3):", h["Q5"].value, "|", h["Q6"].value)
d = wb2["M2b_OpenCircuit"]
print("Decision PV3 emit G7:", d["G7"].value, "| status H7:", d["H7"].value)
print("Decision PV5 empty F9:", d["F9"].value, "| status H9:", d["H9"].value)
print("OK — iter4 build complete.")
