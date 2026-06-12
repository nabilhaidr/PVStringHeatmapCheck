"""Extend M2 PV Performance Workbook — Iterasi 9: M2aSoiling (SKELETON).

Input : docs/M2_PV_Performance_Workbook.xlsx (output Iterasi 8, 37 sheet)
Output: same file, +4 sheet baru:
  38. Raw_Data_SO   — daily energy + insolation (12 hari demo)
  39. Helpers_SO    — PR_daily = E/(H·capacity) + avg_daily_kwh
  40. SO_Economics  — kalkulator cleaning ROI: p_loss, daily_loss, payback, severity (LIVE)
  41. SO_Summary    — status data-sufficiency + caveat SRR

== CAVEAT (baca docs/M2_RE_09) ==
soiling.py = SKELETON (config IKN enabled=FALSE). Inti estimasi soiling ratio = rdtools.soiling_srr()
= Monte-Carlo 1000-rep = BLACK BOX, TIDAK bisa formula Excel (rdtools juga tak tersedia di sandbox).
Maka 'soiling_ratio (sr)' = INPUT terdokumentasi (output SRR). SEMUA hilir sr direproduksi PENUH & live:
  PR_daily=E/(H·kWp) ; p_loss=1−sr ; daily_loss=avg_kwh·tariff·p_loss ; payback=cost/daily_loss ;
  severity(_severity_from_economics) ; recommend=payback<thr. Gate: n_days<min_days → insufficient_data.
Angka EXACT proto_iter9.py (4 skenario sr: CRITICAL/HIGH/MEDIUM/INFO).
"""
from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule

INPUT = Path(__file__).parent / "M2_PV_Performance_Workbook.xlsx"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=13, bold=True, color="305496")
NOTE_FONT = Font(italic=True, size=9, color="808080")
SEV_FILL = {
    "CRITICAL": PatternFill("solid", fgColor="E06666"),
    "HIGH":     PatternFill("solid", fgColor="F6B26B"),
    "MEDIUM":   PatternFill("solid", fgColor="FFE599"),
    "INFO":     PatternFill("solid", fgColor="D9D9D9"),
}
INPUT_FILL = PatternFill("solid", fgColor="FCE5CD")   # SRR sr = input (black box)


def set_header(ws, row, headers, start=1):
    for ci, h in enumerate(headers, start=start):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def title_note(ws, title, notes):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    for i, nt in enumerate(notes, start=2):
        ws.cell(row=i, column=1, value=nt).font = NOTE_FONT


wb = load_workbook(INPUT)
existing = list(wb.sheetnames)
assert len(existing) == 37, f"Expected 37 sheets from iter8, got {len(existing)}"
print(f"Loaded {len(existing)} sheets.")

# ===========================================================================
# README — insert row 14 (iter9) before "Cara membaca"
# ===========================================================================
ws = wb["README"]
assert ws.cell(row=14, column=1).value == "Cara membaca", "README layout berubah; abort"
ws.insert_rows(14)
ws.cell(row=14, column=1, value="9").border = BORDER
ws.cell(row=14, column=2, value="2026-05-30").border = BORDER
ws.cell(row=14, column=3, value="M2aSoiling (skeleton)").border = BORDER
ws.cell(row=14, column=4, value="Raw_Data_SO, Helpers_SO, SO_Economics, SO_Summary  [SRR=input, ekonomi=live]").border = BORDER
for rr in range(15, 23):
    if ws.cell(row=rr, column=1).value:
        ws.cell(row=rr, column=1).font = NOTE_FONT
ws.cell(row=14, column=1).font = Font(name="Calibri", size=11)

# ===========================================================================
# Config — append m2a_soiling params + named cells (start row 66)
# ===========================================================================
ws = wb["Config"]
r = 4; last_cfg = 4
while ws.cell(row=r, column=1).value is not None:
    last_cfg = r; r += 1
start = last_cfg + 1
assert start == 66, f"Expected Config append at row 66, got {start}"
new_cfg = [
    ("m2a_soiling", "min_days", 90, "min window hari untuk SRR (skeleton gate)", "cfg_so_min_days"),
    ("m2a_soiling", "capacity_kwp", 71500.0, "kapasitas DC site PLTS-IKN (PR denominator)", "cfg_so_capacity_kwp"),
    ("m2a_soiling", "electricity_tariff_idr", 1500.0, "tarif PPA (IDR/kWh)", "cfg_so_tariff"),
    ("m2a_soiling", "cleaning_cost_idr", 50000000.0, "biaya 1 cleaning (DEMO; default 0 = user wajib isi)", "cfg_so_cleaning_cost"),
    ("m2a_soiling", "payback_threshold_days", 30.0, "payback < ini → recommend cleaning", "cfg_so_payback_thr"),
]
for i, (sec, key, val, note, name) in enumerate(new_cfg):
    ri = start + i
    ws.cell(row=ri, column=1, value=sec).border = BORDER
    ws.cell(row=ri, column=2, value=key).border = BORDER
    c = ws.cell(row=ri, column=3, value=val); c.border = BORDER
    c.fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(row=ri, column=4, value=note).border = BORDER
    if name not in wb.defined_names:
        wb.defined_names[name] = DefinedName(name, attr_text=f"Config!$C${ri}")
print(f"Config: appended {len(new_cfg)} m2a_soiling rows ({start}..{start+len(new_cfg)-1}) + named cells.")

# ===========================================================================
# Synthetic data (matches proto_iter9.py EXACTLY)
# ===========================================================================
energy_daily = [310000, 318000, 305000, 322000, 315000, 308000,
                320000, 312000, 317000, 309000, 314000, 321000]
insol_daily = [5.40, 5.50, 5.30, 5.60, 5.50, 5.35,
               5.55, 5.45, 5.50, 5.40, 5.45, 5.60]
NDAY = len(energy_daily)   # 12
SR_SCEN = [0.85, 0.92, 0.97, 0.995]
N_DAYS_ASSUMED = 120

# ===========================================================================
# Sheet 38: Raw_Data_SO
# ===========================================================================
ws = wb.create_sheet("Raw_Data_SO")
title_note(ws, "Raw_Data_SO — energi & insolasi harian site (12 hari demo)", [
    "energy_daily (kWh) = Σ daya inverter × Δt (Riemann). insolation_daily (kWh/m²) = Σ POA/1000 × Δt. "
    "Production butuh ≥90 hari (SRR); 12 hari di sini hanya ilustrasi metrik PR harian.",
    "Bangun dari pv_pipeline.baseline.BaselineAccumulator (simpan data NORMAL harian). Ganti dengan data aktual.",
])
set_header(ws, 4, ["date", "energy_daily (kWh)", "insolation_daily (kWh/m²)"])
base = datetime(2026, 2, 1)
for i in range(NDAY):
    ri = 5 + i
    c = ws.cell(row=ri, column=1, value=base + timedelta(days=i)); c.number_format = "yyyy-mm-dd"; c.border = BORDER
    ws.cell(row=ri, column=2, value=float(energy_daily[i])).number_format = "#,##0"; ws.cell(row=ri, column=2).border = BORDER
    ws.cell(row=ri, column=3, value=float(insol_daily[i])).number_format = "0.00"; ws.cell(row=ri, column=3).border = BORDER
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 20

# ===========================================================================
# Sheet 39: Helpers_SO
# ===========================================================================
ws = wb.create_sheet("Helpers_SO")
title_note(ws, "Helpers_SO — PR harian (IEC 61724-1) + avg_daily_kwh", [
    "PR_daily = energy_daily / (insolation_daily × capacity_kwp). avg_daily_kwh = AVERAGE(energy) (proxy tail-30 hari).",
])
set_header(ws, 4, ["date", "energy_daily", "insolation_daily", "PR_daily"])
DF, DL = 5, 5 + NDAY - 1   # 5..16
for i in range(NDAY):
    ri = DF + i
    ws.cell(row=ri, column=1, value=f"=Raw_Data_SO!A{ri}").number_format = "yyyy-mm-dd"
    ws.cell(row=ri, column=2, value=f"=Raw_Data_SO!B{ri}").number_format = "#,##0"
    ws.cell(row=ri, column=3, value=f"=Raw_Data_SO!C{ri}").number_format = "0.00"
    ws.cell(row=ri, column=4, value=f"=Raw_Data_SO!B{ri}/(Raw_Data_SO!C{ri}*cfg_so_capacity_kwp)").number_format = "0.0000"
    for ci in range(1, 5):
        ws.cell(row=ri, column=ci).border = BORDER
# stats block
ws.cell(row=DL + 2, column=1, value="avg_daily_kwh").border = BORDER
ws.cell(row=DL + 2, column=2, value=f"=AVERAGE(B{DF}:B{DL})").number_format = "#,##0"; ws.cell(row=DL+2, column=2).border = BORDER
ws.cell(row=DL + 3, column=1, value="PR_mean").border = BORDER
ws.cell(row=DL + 3, column=2, value=f"=AVERAGE(D{DF}:D{DL})").number_format = "0.0000"; ws.cell(row=DL+3, column=2).border = BORDER
ws.cell(row=DL + 4, column=1, value="n_days_sample").border = BORDER
ws.cell(row=DL + 4, column=2, value=f"=COUNT(B{DF}:B{DL})").border = BORDER
AVG_KWH_CELL = f"Helpers_SO!$B${DL+2}"
ws.column_dimensions["A"].width = 14

# ===========================================================================
# Sheet 40: SO_Economics (kalkulator cleaning ROI)
# ===========================================================================
ws = wb.create_sheet("SO_Economics")
title_note(ws, "SO_Economics — kalkulator cleaning ROI (hilir SRR; LIVE)", [
    "soiling_ratio (sr) = OUTPUT rdtools.soiling_srr() Monte-Carlo — INPUT di sini (di-highlight), TIDAK dihitung Excel.",
    "Hilir sr SEMUA live: p_loss=1−sr; daily_loss=avg_daily_kwh·tarif·p_loss; payback=biaya/daily_loss; severity & recommend.",
])
# input/ref block
ws.cell(row=4, column=1, value="avg_daily_kwh (dari Helpers)").border = BORDER
ws.cell(row=4, column=2, value=f"={AVG_KWH_CELL}").number_format = "#,##0"; ws.cell(row=4, column=2).border = BORDER
ws.cell(row=5, column=1, value="n_days_assumed").border = BORDER
ws.cell(row=5, column=2, value=N_DAYS_ASSUMED).border = BORDER
ws.cell(row=6, column=1, value="data_status (gate)").border = BORDER
ws.cell(row=6, column=2, value='=IF(B5<cfg_so_min_days,"insufficient_data","ok")').border = BORDER
ws.cell(row=7, column=1, value="cleaning_cost_idr / tariff / payback_thr").border = BORDER
ws.cell(row=7, column=2, value="=cfg_so_cleaning_cost&\" / \"&cfg_so_tariff&\" / \"&cfg_so_payback_thr").border = BORDER

# scenario table
HR = 9
set_header(ws, HR, ["soiling_ratio (sr) [INPUT]", "p_loss", "daily_loss_idr",
                    "payback_days", "recommend", "fault_type", "severity", "confidence"])
SC_DF = HR + 1   # 10
for k, sr in enumerate(SR_SCEN):
    r = SC_DF + k
    cs = ws.cell(row=r, column=1, value=sr); cs.number_format = "0.000"; cs.fill = INPUT_FILL; cs.border = BORDER
    ws.cell(row=r, column=2, value=f"=1-A{r}").number_format = "0.0000"                          # p_loss
    ws.cell(row=r, column=3, value=f"=$B$4*cfg_so_tariff*B{r}").number_format = "#,##0"           # daily_loss
    ws.cell(row=r, column=4, value=f"=IF(AND(C{r}>0,cfg_so_cleaning_cost>0),cfg_so_cleaning_cost/C{r},1E+99)").number_format = "0.000"  # payback
    ws.cell(row=r, column=5, value=f'=IF(D{r}<cfg_so_payback_thr,"YES","no")')                    # recommend
    ws.cell(row=r, column=6, value=f'=IF(E{r}="YES","cleaning_recommended","soiling_detected")')  # fault_type
    ws.cell(row=r, column=7, value=(                                                              # severity
        f'=IF(AND(B{r}>=0.1,D{r}<cfg_so_payback_thr/3),"CRITICAL",'
        f'IF(AND(B{r}>=0.05,D{r}<cfg_so_payback_thr),"HIGH",'
        f'IF(AND(B{r}>=0.02,D{r}<2*cfg_so_payback_thr),"MEDIUM","INFO")))'))
    ws.cell(row=r, column=8, value=f"=50+A{r}*50").number_format = "0.0"                          # confidence
    for ci in range(1, 9):
        ws.cell(row=r, column=ci).border = BORDER
SC_DL = SC_DF + len(SR_SCEN) - 1
for sev, fill in SEV_FILL.items():
    ws.conditional_formatting.add(f"G{SC_DF}:G{SC_DL}",
        CellIsRule(operator="equal", formula=[f'"{sev}"'], fill=fill))
ws.column_dimensions["A"].width = 22
ws.column_dimensions["C"].width = 16
ws.column_dimensions["F"].width = 22

# ===========================================================================
# Sheet 41: SO_Summary
# ===========================================================================
ws = wb.create_sheet("SO_Summary")
title_note(ws, "SO_Summary — status & caveat (mirror EconomicAnalysis)", [
    "Skeleton: dengan data < min_days, run() emit 'insufficient_data' (INFO) dan SKIP SRR. "
    "Vocab fault: insufficient_data, insufficient_dependency, rdtools_error, soiling_detected, cleaning_recommended.",
])
set_header(ws, 4, ["item", "value"])
rows = [
    ("n_days_assumed", "=SO_Economics!B5"),
    ("min_days_required", "=cfg_so_min_days"),
    ("status", "=SO_Economics!B6"),
    ("avg_daily_kwh", f"={AVG_KWH_CELL}"),
    ("n_cleaning_recommended", f'=COUNTIF(SO_Economics!E{SC_DF}:E{SC_DL},"YES")'),
    ("SRR engine", '="rdtools.soiling_srr (Monte-Carlo 1000 reps) — NOT Excel-reproducible"'),
    ("note", '="soiling_ratio adalah INPUT (output SRR); semua ekonomi hilir = live & verified"'),
]
for i, (label, formula) in enumerate(rows):
    r = 5 + i
    ws.cell(row=r, column=1, value=label).border = BORDER
    c = ws.cell(row=r, column=2, value=formula); c.border = BORDER
    if label in ("avg_daily_kwh",):
        c.number_format = "#,##0"
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 60

# ===========================================================================
# Save + integrity verify
# ===========================================================================
wb.save(INPUT)
wb2 = load_workbook(INPUT, data_only=False)
after = list(wb2.sheetnames)
new_sheets = ["Raw_Data_SO", "Helpers_SO", "SO_Economics", "SO_Summary"]
assert after[:37] == existing, "EXISTING SHEET ORDER CHANGED!"
assert after[37:] == new_sheets, f"New sheets mismatch: {after[37:]}"
print(f"\nSheets now: {len(after)} (was 37, +4)")
h = wb2["Helpers_SO"]; e = wb2["SO_Economics"]
print("Helpers PR_daily D5:", h["D5"].value)
print("SO_Economics p_loss B10:", e["B10"].value, "| payback D10:", e["D10"].value)
print("SO_Economics severity G10:", e["G10"].value)
print("OK — iter9 build complete.")
