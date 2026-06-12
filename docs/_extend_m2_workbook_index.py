"""Extend M2 PV Performance Workbook — Ringkasan: sheet M2_Index (peta detektor).

Input : docs/M2_PV_Performance_Workbook.xlsx (41 detektor-sheet)
Output: same file, sheet M2_Index (tab pertama) = peta detektor → sinyal → fault_type →
        severity → status reproduksibilitas Excel + inventaris sheet per iterasi.

Regen-safe: kalau M2_Index sudah ada, HAPUS lalu rebuild (sehingga update mudah).
README log + Config TIDAK diubah. 41 detektor-sheet harus tetap utuh.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

INPUT = Path(__file__).parent / "M2_PV_Performance_Workbook.xlsx"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="305496")
SUB_FONT = Font(name="Calibri", size=12, bold=True, color="305496")
NOTE_FONT = Font(italic=True, size=9, color="808080")
STATUS_FILL = {
    "PENUH": PatternFill("solid", fgColor="B6D7A8"),
    "PENUH*": PatternFill("solid", fgColor="D9EAD3"),
    "HILIR PENUH": PatternFill("solid", fgColor="FFE599"),
    "APPROKSIMASI": PatternFill("solid", fgColor="F6B26B"),
    "INPUT-ONLY": PatternFill("solid", fgColor="D9D9D9"),
}


def set_header(ws, row, headers):
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


wb = load_workbook(INPUT)
if "M2_Index" in wb.sheetnames:
    del wb["M2_Index"]
    print("removed existing M2_Index (rebuild).")
existing = list(wb.sheetnames)
assert len(existing) == 41, f"Expected 41 detektor-sheet, got {len(existing)}"
print(f"Loaded {len(existing)} sheets.")

ws = wb.create_sheet("M2_Index", 0)   # tab pertama (navigasi)
ws.sheet_properties.tabColor = "305496"

ws.cell(row=1, column=1, value="M2 PV Performance — INDEKS Famili Detektor").font = TITLE_FONT
ws.cell(row=2, column=1, value=("Peta tiap detektor → sinyal → fault_type → severity → status reproduksibilitas Excel. "
                                "Workbook 41 detektor-sheet (di luar indeks ini). Detail: docs/M2_RE_0X_*.md + M2_Family_Summary.")).font = NOTE_FONT

# ---- Table 1: detector map (8 aktif + 1 ML skeleton) ----
ws.cell(row=4, column=1, value="1. Peta Detektor (8 aktif + 1 ML skeleton)").font = SUB_FONT
set_header(ws, 5, ["#", "Detector", "Modul", "Sinyal utama", "fault_type",
                   "Severity (ringkas)", "Sheet keputusan", "Status Excel"])
DETECTORS = [
    (1, "M2eAvailability", "availability.py", "uptime inverter & string (downtime menit)",
     "(severity-based)", "<90 CRIT · <95 HIGH · <97 MED · <99 INFO · ≥99 NORMAL", "M2e_Availability", "PENUH"),
    (2, "M2bPeerZScore", "peer_zscore.py", "R_str=V/I → z-score peer + voc_ratio",
     "high_R", "|z|>3.5 HIGH · |z|>2.5 & voc_ratio>0.95 MED", "M2b_PeerZScore", "PENUH"),
    (3, "M2bOpenCircuit", "open_circuit.py", "I/I_q95 < 5% (across siblings) + debounce 20",
     "open_circuit", "CRITICAL (conf 95%)", "M2b_OpenCircuit", "PENUH"),
    (4, "M2bGroundFault", "ground_fault.py", "V_to_ground absolute/adaptive + voc_ratio & I_z",
     "ground_fault", "spec+(abs|adp)=90 · spec/abs+adp=80 · abs=70 · adp=60", "M2c_GroundFault", "PENUH*"),
    (5, "M2IForest", "iforest.py", "IsolationForest 5-fitur (V,I,V_dev,I_dev,R)",
     "iforest_anomaly", "kuartil rank flagged: CRIT/HIGH/MED/INFO", "IF_Anomaly", "APPROKSIMASI"),
    (6, "M2aLowIrradiance", "m2a/low_irradiance.py", "OLS slope PR_proxy vs POA (band low/mid)",
     "low_irradiance / general_underperform", "|slope_low|·r² → CRIT/HIGH/MED", "M2a_LowIrradiance", "PENUH"),
    (7, "M2aShading", "m2a/shading.py", "CV antar-string per jam + PR + asimetri AM/PM",
     "shading_morning/afternoon/uniform", "0.7·frac + 0.3·asym → CRIT/HIGH/MED", "M2a_Shading", "PENUH"),
    (8, "M2aSoiling", "m2a/soiling.py (skeleton)", "rdtools SRR (Monte-Carlo) → ekonomi payback",
     "soiling_detected / cleaning_recommended / insufficient_data", "(p_loss,payback) → CRIT/HIGH/MED/INFO", "SO_Economics", "HILIR PENUH"),
    (9, "M2bIntermittent (LSTM-AE)", "lstm_ae.py (skeleton, enabled=False)", "LSTM Autoencoder; reconstruction error window 24-jam (96×15-min)",
     "intermittent", "MEDIUM (conf 70)", "(tak ada — input-only)", "INPUT-ONLY"),
]
for i, row in enumerate(DETECTORS):
    r = 6 + i
    for ci, val in enumerate(row, start=1):
        c = ws.cell(row=r, column=ci, value=val); c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
    sc = ws.cell(row=r, column=8)
    if sc.value in STATUS_FILL:
        sc.fill = STATUS_FILL[sc.value]

# ---- Legend status Excel ----
lr = 6 + len(DETECTORS) + 1
ws.cell(row=lr, column=1, value="Legenda Status Excel").font = SUB_FONT
legend = [
    ("PENUH", "Direproduksi eksak sebagai formula live (verified vs source, selisih ~0)."),
    ("PENUH*", "Per-inverter math live & eksak; SATU input representatif: fleet V_gnd median/std (ground_fault)."),
    ("HILIR PENUH", "Semua hilir live & eksak; satu nilai = INPUT black-box: soiling_ratio dari SRR Monte-Carlo."),
    ("APPROKSIMASI", "Struktur faithful tapi SKOR proxy (MAD), BUKAN skor IsolationForest asli — menandai sampel berbeda."),
    ("INPUT-ONLY", "Jaringan PyTorch terlatih = black box; TAK ADA sheet & tak ada lapisan hilir bermakna (LSTM-AE). Lihat M2_RE_10."),
]
for i, (tag, desc) in enumerate(legend):
    r = lr + 1 + i
    c = ws.cell(row=r, column=1, value=tag); c.border = BORDER
    if tag in STATUS_FILL:
        c.fill = STATUS_FILL[tag]
    d = ws.cell(row=r, column=2, value=desc); d.border = BORDER
    d.alignment = Alignment(vertical="top", wrap_text=True)

# ---- Table 2: sheet inventory per iterasi ----
ir = lr + len(legend) + 2
ws.cell(row=ir, column=1, value="2. Inventaris Sheet & Dokumen (per iterasi)").font = SUB_FONT
set_header(ws, ir + 1, ["Iterasi", "Detector", "Dokumen RE", "Sheet"])
INVENTORY = [
    ("—", "Inti (shared)", "—", "README, Config"),
    ("2", "M2eAvailability", "(Iterasi 2)", "Raw_Data, EmptyPVMap, Helpers_M2e, M2e_Availability, M2e_AllStrings, Findings_Summary"),
    ("3", "M2bPeerZScore", "M2_RE_03", "PanelSpec, Raw_Data_M2b, Meteo_Dummy, Helpers_M2b, M2b_PeerZScore, M2b_StringStatus, M2b_StatComparison, Hampel_Preprocessing"),
    ("4", "M2bOpenCircuit", "M2_RE_04", "Raw_Data_OC, Helpers_OC, M2b_OpenCircuit, M2b_OC_StringStatus"),
    ("5", "M2bGroundFault", "M2_RE_05", "Raw_Data_GF, Helpers_GF, GF_StringMetrics, M2c_GroundFault, M2c_GF_StringStatus"),
    ("6", "M2IForest", "M2_RE_06", "Raw_Data_IF, Features_IF, IF_Anomaly, IF_Summary"),
    ("7", "M2aLowIrradiance", "M2_RE_07", "Raw_Data_LI, Helpers_LI, M2a_LowIrradiance, LI_Summary"),
    ("8", "M2aShading", "M2_RE_08", "Raw_Data_SH, Helpers_SH, SH_Hourly, M2a_Shading"),
    ("9", "M2aSoiling", "M2_RE_09", "Raw_Data_SO, Helpers_SO, SO_Economics, SO_Summary"),
    ("10", "M2bIntermittent (LSTM-AE)", "M2_RE_10", "(tak ada sheet — input-only; jaringan PyTorch terlatih)"),
]
for i, row in enumerate(INVENTORY):
    r = ir + 2 + i
    for ci, val in enumerate(row, start=1):
        c = ws.cell(row=r, column=ci, value=val); c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)

# ---- column widths ----
widths = {"A": 9, "B": 19, "C": 24, "D": 34, "E": 30, "F": 40, "G": 18, "H": 16}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# ===========================================================================
wb.save(INPUT)
wb2 = load_workbook(INPUT)
after = list(wb2.sheetnames)
assert after[0] == "M2_Index", "M2_Index harus jadi sheet pertama"
assert set(after[1:]) == set(existing), "Sheet lama berubah!"
ix = wb2["M2_Index"]
print(f"Sheets now: {len(after)} (M2_Index first tab). Detector rows: {len(DETECTORS)}")
print("Row 14 (LSTM-AE):", [ix.cell(14, c).value for c in (2, 5, 8)])
print("OK — M2_Index updated (9 detektor: 8 aktif + LSTM-AE input-only).")
