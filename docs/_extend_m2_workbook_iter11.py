"""Extend M2 PV Performance Workbook — Iterasi 11: M2bMpptRatio.

Input : docs/M2_PV_Performance_Workbook.xlsx (output iter9 + M2_Index, 42 sheet)
Output: same file, +4 sheet baru:
  Raw_Data_MR   — arus per-string 1 inverter (2 MPPT group SUN2000-330) + POA (26 timestep)
  Helpers_MR    — daylight gate + median partner se-MPPT + ratio + qualifying + running-consec
  M2b_MpptRatio — keputusan per string: ratio_median, max_consec, emit, severity, confidence, status (LIVE)
  M2b_MR_StringStatus — replika artifact StringStatus Python

Catatan desain (docs/M2_RE_11): peer = sibling se-MPPT (mppt_map), sinyal RATIO arus (bukan z-score).
  ratio = I_string / MAX(median(partner se-MPPT), 0.01)
  qualifying = ratio < 0.85 & daylight ; emit bila run konsekutif >= debounce 20
  severity: rem<0.20 CRIT | <0.50 HIGH | else MED ; confidence = min(90,max(50,(1-rem)*100))
Angka EXACT di-lock & diverifikasi verify_iter11.py. Idempotent: re-run = rebuild 4 sheet (regen 0-diff).

Layout MPPT (persis config/strings.yaml SUN2000-330KTL-H1): MPPT1=[1,2,3,4], MPPT2=[5,6,7,8,9].
Skenario: PV3 MEDIUM(0.80) · PV4 HIGH(0.44) · PV8 CRITICAL(0.0 string mati) · PV9 glitch(3 langkah) · sisanya NORMAL.
"""
from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule

INPUT = Path(__file__).parent / "M2_PV_Performance_Workbook.xlsx"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=13, bold=True, color="305496")
NOTE_FONT = Font(italic=True, size=9, color="808080")
SEV_FILL = {
    "CRITICAL": PatternFill("solid", fgColor="E06666"),
    "HIGH":     PatternFill("solid", fgColor="F6B26B"),
    "MEDIUM":   PatternFill("solid", fgColor="FFE599"),
    "NORMAL":   PatternFill("solid", fgColor="B6D7A8"),
    "EMPTY":    PatternFill("solid", fgColor="D9D9D9"),
}
STATUS_FILL = {
    "mppt_partner_underperform": PatternFill("solid", fgColor="E06666"),
    "NORMAL":                    PatternFill("solid", fgColor="B6D7A8"),
    "EMPTY":                     PatternFill("solid", fgColor="D9D9D9"),
}
NEW_SHEETS = ["Raw_Data_MR", "Helpers_MR", "M2b_MpptRatio", "M2b_MR_StringStatus"]


def set_header(ws, row, headers, start=1):
    for ci, h in enumerate(headers, start=start):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def title_note(ws, title, notes):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    for i, nt in enumerate(notes, start=2):
        ws.cell(row=i, column=1, value=nt).font = NOTE_FONT


# ============================================================================
# Skenario sintetis (single source of truth; verify_iter11 baca ulang literal)
# ============================================================================
INVERTER_ID = "WB05-INV05"
WB_ID = "WB05"
POA_SOURCE = "pyranometer_per_ws"
MEMBERS = {1: [1, 2, 3, 4], 2: [5, 6, 7, 8, 9]}     # SUN2000-330: MPPT1, MPPT2
MPPT_OF = {n: m for m, ns in MEMBERS.items() for n in ns}
CURRENT = {1: 13.0, 2: 12.8, 3: 10.3, 4: 5.6, 5: 13.0, 6: 12.8, 7: 12.9, 8: 0.0, 9: 12.85}
GLITCH_PV = 9
GLITCH_ROWS = (9, 10, 11)        # 3 langkah konsekutif (< debounce) -> harus ditelan
GLITCH_VALUE = 3.0
N_DAYLIGHT = 24                  # 12:00..13:55 @ 5-min
POA_DAY, POA_TWI = 900.0, 120.0
DATA_FIRST = 5                   # baris data pertama
DAY_LAST = DATA_FIRST + N_DAYLIGHT - 1          # 28
DATA_LAST = DAY_LAST + 2                         # 30 (+2 twilight)
PVS = list(range(1, 10))

# Kolom Raw_Data_MR: A=timestamp, B=POA, C..K = PV1..PV9
RAW_COL = {n: get_column_letter(3 + (n - 1)) for n in PVS}    # PV1=C ... PV9=K
# Kolom Helpers_MR: A ts, B POA, C daylight, lalu blok median/ratio/qual/consec (9 tiap blok)
HMED = {n: 4 + (n - 1) for n in PVS}     # D..L
HRAT = {n: 13 + (n - 1) for n in PVS}    # M..U
HQUAL = {n: 22 + (n - 1) for n in PVS}   # V..AD
HCONS = {n: 31 + (n - 1) for n in PVS}   # AE..AM
PARTNERS = {n: [m for m in MEMBERS[MPPT_OF[n]] if m != n] for n in PVS}


# ============================================================================
wb = load_workbook(INPUT)

# Idempotent: hapus 4 sheet MR bila sudah ada (rebuild -> regen 0-diff).
for s in NEW_SHEETS:
    if s in wb.sheetnames:
        del wb[s]
        print(f"  removed existing {s} (rebuild).")

existing = list(wb.sheetnames)
assert existing[0] == "M2_Index", f"M2_Index harus tab pertama, got {existing[0]!r}"
print(f"Loaded {len(existing)} sheets (M2_Index first).")

# ---------------------------------------------------------------------------
# README — insert iter11 row sebelum "Cara membaca" (idempotent)
# ---------------------------------------------------------------------------
ws = wb["README"]
cara_row = None
already = False
for r in range(1, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if v == "Cara membaca":
        cara_row = r
    if isinstance(v, str) and v.strip() == "11":
        already = True
if already:
    print("  README: iter11 row already present (skip).")
elif cara_row is None:
    print("  [WARN] README 'Cara membaca' tak ditemukan; README tidak diubah.")
else:
    ws.insert_rows(cara_row)
    ws.cell(row=cara_row, column=1, value="11").border = BORDER
    ws.cell(row=cara_row, column=1).font = Font(name="Calibri", size=11)
    ws.cell(row=cara_row, column=2, value="2026-06-13").border = BORDER
    ws.cell(row=cara_row, column=3, value="M2bMpptRatio").border = BORDER
    ws.cell(row=cara_row, column=4,
            value="Raw_Data_MR, Helpers_MR, M2b_MpptRatio, M2b_MR_StringStatus").border = BORDER
    print(f"  README: inserted iter11 row at {cara_row}.")

# ---------------------------------------------------------------------------
# Config — append m2b_mppt_ratio params + named cells cfg_mr_* (idempotent)
# ---------------------------------------------------------------------------
ws = wb["Config"]
new_cfg = [
    ("m2b_mppt_ratio", "poa_threshold_wm2", 300.0, "moderate sun gate", "cfg_mr_poa_threshold_wm2"),
    ("m2b_mppt_ratio", "poa_floor_wm2", 50.0, "hard floor twilight", "cfg_mr_poa_floor_wm2"),
    ("m2b_mppt_ratio", "ratio_threshold", 0.85, "I<85% median partner -> qualifying", "cfg_mr_ratio_threshold"),
    ("m2b_mppt_ratio", "ratio_high", 0.50, "event_median<0.50 -> HIGH", "cfg_mr_ratio_high"),
    ("m2b_mppt_ratio", "ratio_critical", 0.20, "event_median<0.20 -> CRITICAL", "cfg_mr_ratio_critical"),
    ("m2b_mppt_ratio", "debounce_consecutive_steps", 20, "~100 menit @ 5-min", "cfg_mr_debounce_steps"),
    ("m2b_mppt_ratio", "min_partner_strings", 1, "grup 2-string tetap dianalisis", "cfg_mr_min_partner"),
    ("m2b_mppt_ratio", "partner_clip", 0.01, "clip(lower=0.01) divide-by-zero guard", "cfg_mr_partner_clip"),
]
if "cfg_mr_ratio_threshold" in wb.defined_names:
    print("  Config: cfg_mr_* already present (skip).")
else:
    r = 4; last_cfg = 4
    while ws.cell(row=r, column=1).value is not None:
        last_cfg = r; r += 1
    start = last_cfg + 1
    for i, (sec, key, val, note, name) in enumerate(new_cfg):
        ri = start + i
        ws.cell(row=ri, column=1, value=sec).border = BORDER
        ws.cell(row=ri, column=2, value=key).border = BORDER
        c = ws.cell(row=ri, column=3, value=val); c.border = BORDER
        c.fill = PatternFill("solid", fgColor="FFF2CC")
        ws.cell(row=ri, column=4, value=note).border = BORDER
        wb.defined_names[name] = DefinedName(name, attr_text=f"Config!$C${ri}")
    print(f"  Config: appended {len(new_cfg)} m2b_mppt_ratio rows ({start}..{start+len(new_cfg)-1}) + named cells.")

# ===========================================================================
# Sheet: Raw_Data_MR
# ===========================================================================
ws = wb.create_sheet("Raw_Data_MR")
title_note(ws, "Raw_Data_MR — arus per-string 1 inverter (2 MPPT group) + POA", [
    "Model SUN2000-330KTL-H1 (WB03-10): MPPT1=[PV1-4], MPPT2=[PV5-9] — persis config/strings.yaml.",
    "Arus konstan per string (sun di-faktorkan keluar; ratio invarian POA). PV9 glitch 3 langkah (baris 9-11).",
    "Skenario: PV3 MEDIUM · PV4 HIGH · PV8 CRITICAL (string mati, BUKAN slot kosong) · PV9 glitch · sisanya NORMAL.",
])
set_header(ws, 4, ["timestamp", "POA (W/m^2)"] + [f"PV{n} input current(A)" for n in PVS])
base = datetime(2026, 5, 14, 12, 0)
for i in range(N_DAYLIGHT + 2):
    ri = DATA_FIRST + i
    if i < N_DAYLIGHT:
        ts = base + timedelta(minutes=5 * i); poa = POA_DAY
    else:
        ts = datetime(2026, 5, 14, 18, 10 + 10 * (i - N_DAYLIGHT)); poa = POA_TWI
    c = ws.cell(row=ri, column=1, value=ts); c.number_format = "hh:mm"; c.border = BORDER
    ws.cell(row=ri, column=2, value=poa).border = BORDER
    for n in PVS:
        val = CURRENT[n]
        if n == GLITCH_PV and ri in GLITCH_ROWS:
            val = GLITCH_VALUE
        cc = ws.cell(row=ri, column=3 + (n - 1), value=float(val))
        cc.number_format = "0.00"; cc.border = BORDER
ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 12
for n in PVS:
    ws.column_dimensions[RAW_COL[n]].width = 11

# ===========================================================================
# Sheet: Helpers_MR
# ===========================================================================
ws = wb.create_sheet("Helpers_MR")
title_note(ws, "Helpers_MR — daylight + median partner se-MPPT + ratio + qualifying + running-consec", [
    "daylight = (POA>cfg_mr_poa_threshold) & (POA>cfg_mr_poa_floor)  [sub-gate POA saja; ephemeris di luar sheet].",
    "median partner = MEDIAN(cell partner se-MPPT)  -> meniru median(axis=1) atas sibling (mppt_ratio.py baris 274-276).",
    "ratio = I / MAX(median_partner, cfg_mr_partner_clip) ; qualifying = ratio<threshold & daylight ; consec = run True.",
])
hdr = ["timestamp", "POA", "daylight"]
hdr += [f"med_partner_PV{n}" for n in PVS]
hdr += [f"ratio_PV{n}" for n in PVS]
hdr += [f"qual_PV{n}" for n in PVS]
hdr += [f"consec_PV{n}" for n in PVS]
set_header(ws, 4, hdr)
for i in range(N_DAYLIGHT + 2):
    ri = DATA_FIRST + i
    ws.cell(row=ri, column=1, value=f"=Raw_Data_MR!A{ri}").number_format = "hh:mm"
    ws.cell(row=ri, column=2, value=f"=Raw_Data_MR!B{ri}")
    ws.cell(row=ri, column=3,
            value=f"=IF(AND(B{ri}>cfg_mr_poa_threshold_wm2,B{ri}>cfg_mr_poa_floor_wm2),1,0)")
    for n in PVS:
        med_l = get_column_letter(HMED[n]); rat_l = get_column_letter(HRAT[n])
        qual_l = get_column_letter(HQUAL[n]); cons_l = get_column_letter(HCONS[n])
        refs = ",".join(f"Raw_Data_MR!{RAW_COL[p]}{ri}" for p in PARTNERS[n])
        ws.cell(row=ri, column=HMED[n], value=f"=MEDIAN({refs})").number_format = "0.000"
        ws.cell(row=ri, column=HRAT[n],
                value=f"=Raw_Data_MR!{RAW_COL[n]}{ri}/MAX({med_l}{ri},cfg_mr_partner_clip)").number_format = "0.0000"
        ws.cell(row=ri, column=HQUAL[n],
                value=f"=IF(AND({rat_l}{ri}<cfg_mr_ratio_threshold,$C{ri}=1),1,0)")
        if ri == DATA_FIRST:
            ws.cell(row=ri, column=HCONS[n], value=f"={qual_l}{ri}")
        else:
            ws.cell(row=ri, column=HCONS[n],
                    value=f"=IF({qual_l}{ri}=1,{cons_l}{ri-1}+1,0)")
    for ci in range(1, HCONS[9] + 1):
        ws.cell(row=ri, column=ci).border = BORDER
ws.column_dimensions["A"].width = 9
ws.freeze_panes = "D5"

# ===========================================================================
# Sheet: M2b_MpptRatio (keputusan per string, baris 5..13 = PV1..PV9)
# ===========================================================================
ws = wb.create_sheet("M2b_MpptRatio")
title_note(ws, "M2b_MpptRatio — keputusan per string (LIVE)", [
    "value = ratio_median_daylight ; emit bila max_consec >= cfg_mr_debounce_steps DAN bukan EMPTY.",
    "severity: value<critical CRIT | <high HIGH | else MED. confidence = min(90,max(50,(1-value)*100)).",
    "Untuk fault sustained ber-ratio konstan, ratio_event_median == ratio_median_daylight (lihat M2_RE_11 §5.3).",
])
set_header(ws, 4, ["pv_string", "mppt", "partner_strings", "ratio_median_daylight",
                   "max_consec", "empty_by_design", "emit", "severity", "confidence", "status"])
DEC_FIRST = 5
for i, n in enumerate(PVS):
    r = DEC_FIRST + i
    rat_l = get_column_letter(HRAT[n]); cons_l = get_column_letter(HCONS[n])
    partner_txt = ", ".join(f"PV{m}" for m in PARTNERS[n])
    ws.cell(row=r, column=1, value=f"PV{n}")
    ws.cell(row=r, column=2, value=MPPT_OF[n])
    ws.cell(row=r, column=3, value=partner_txt)
    ws.cell(row=r, column=4,
            value=f"=MEDIAN(Helpers_MR!{rat_l}{DATA_FIRST}:{rat_l}{DAY_LAST})").number_format = "0.0000"
    ws.cell(row=r, column=5,
            value=f"=MAX(Helpers_MR!{cons_l}{DATA_FIRST}:{cons_l}{DATA_LAST})")
    ws.cell(row=r, column=6, value=0)   # empty_by_design (tak ada slot kosong di demo)
    ws.cell(row=r, column=7,
            value=f"=IF(AND(E{r}>=cfg_mr_debounce_steps,F{r}=0),1,0)")
    ws.cell(row=r, column=8,
            value=(f'=IF(G{r}=0,"",IF(D{r}<cfg_mr_ratio_critical,"CRITICAL",'
                   f'IF(D{r}<cfg_mr_ratio_high,"HIGH","MEDIUM")))'))
    ws.cell(row=r, column=9,
            value=f'=IF(G{r}=0,"",MIN(90,MAX(50,(1-D{r})*100)))').number_format = "0.0#"
    ws.cell(row=r, column=10,
            value=f'=IF(F{r}=1,"EMPTY",IF(G{r}=1,"mppt_partner_underperform","NORMAL"))')
    for ci in range(1, 11):
        ws.cell(row=r, column=ci).border = BORDER
DEC_LAST = DEC_FIRST + len(PVS) - 1
for sev, fill in SEV_FILL.items():
    ws.conditional_formatting.add(f"H{DEC_FIRST}:H{DEC_LAST}",
        CellIsRule(operator="equal", formula=[f'"{sev}"'], fill=fill))
for st, fill in STATUS_FILL.items():
    ws.conditional_formatting.add(f"J{DEC_FIRST}:J{DEC_LAST}",
        CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill))
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 19
ws.column_dimensions["J"].width = 26

# ===========================================================================
# Sheet: M2b_MR_StringStatus (replika artifact Python)
# ===========================================================================
ws = wb.create_sheet("M2b_MR_StringStatus")
title_note(ws, "M2b_MR_StringStatus — replika artifact StringStatus (mppt_ratio.py baris 341-355)", [
    "Kolom = field artifact Python. ratio_event_median == ratio_median_daylight untuk fault sustained (M2_RE_11 §5.3).",
])
set_header(ws, 4, ["poa_source", "inverter_id", "wb_id", "pv_string", "mppt", "status",
                   "partner_strings", "ratio_median_daylight", "ratio_event_median",
                   "n_qualifying_steps", "n_debounced_events", "emitted_finding", "daylight_samples"])
for i, n in enumerate(PVS):
    r = DEC_FIRST + i
    drow = DEC_FIRST + i        # baris pasangan di M2b_MpptRatio
    qual_l = get_column_letter(HQUAL[n])
    ws.cell(row=r, column=1, value=POA_SOURCE)
    ws.cell(row=r, column=2, value=INVERTER_ID)
    ws.cell(row=r, column=3, value=WB_ID)
    ws.cell(row=r, column=4, value=f"PV{n}")
    ws.cell(row=r, column=5, value=MPPT_OF[n])
    ws.cell(row=r, column=6, value=f"=M2b_MpptRatio!J{drow}")
    ws.cell(row=r, column=7, value=f"=M2b_MpptRatio!C{drow}")
    ws.cell(row=r, column=8, value=f"=M2b_MpptRatio!D{drow}").number_format = "0.0000"
    ws.cell(row=r, column=9, value=f"=M2b_MpptRatio!D{drow}").number_format = "0.0000"
    ws.cell(row=r, column=10, value=f"=SUM(Helpers_MR!{qual_l}{DATA_FIRST}:{qual_l}{DATA_LAST})")
    ws.cell(row=r, column=11, value=f"=M2b_MpptRatio!G{drow}")
    ws.cell(row=r, column=12, value=f"=M2b_MpptRatio!G{drow}")
    ws.cell(row=r, column=13, value=f"=SUM(Helpers_MR!C{DATA_FIRST}:C{DATA_LAST})")
    for ci in range(1, 14):
        ws.cell(row=r, column=ci).border = BORDER
for st, fill in STATUS_FILL.items():
    ws.conditional_formatting.add(f"F{DEC_FIRST}:F{DEC_LAST}",
        CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill))
ws.column_dimensions["A"].width = 18
ws.column_dimensions["G"].width = 18
ws.column_dimensions["F"].width = 24

# ===========================================================================
# Save + integrity verify
# ===========================================================================
wb.save(INPUT)
wb2 = load_workbook(INPUT, data_only=False)
after = list(wb2.sheetnames)
assert after[:len(existing)] == existing, "EXISTING SHEET ORDER CHANGED!"
assert after[len(existing):] == NEW_SHEETS, f"New sheets mismatch: {after[len(existing):]}"
assert after[0] == "M2_Index", "M2_Index bukan tab pertama lagi!"
print(f"\nSheets now: {len(after)} (was {len(existing)}, +4)")
h = wb2["Helpers_MR"]; d = wb2["M2b_MpptRatio"]
print("Helpers med_partner_PV3 (D5):", h["D5"].value)
print("Helpers ratio_PV3 (M5):", h["M5"].value)
print("Decision PV3 emit (G7) / severity (H7):", d["G7"].value, "/", d["H7"].value)
print("Decision PV8 severity (H12) / confidence (I12):", d["H12"].value, "/", d["I12"].value)
print("OK — iter11 build complete.")
