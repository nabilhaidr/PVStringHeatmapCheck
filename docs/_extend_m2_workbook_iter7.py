"""Extend M2 PV Performance Workbook — Iterasi 7: M2aLowIrradiance.

Input : docs/M2_PV_Performance_Workbook.xlsx (output Iterasi 6, 29 sheet)
Output: same file, +4 sheet baru:
  30. Raw_Data_LI       — POA + P_inv per (inverter, sampel)
  31. Helpers_LI        — pr_proxy=P_inv/POA, in_low/in_mid band flags
  32. M2a_LowIrradiance — regresi OLS dua band via SUMPRODUCT (slope/intercept/r2) + klasifikasi + severity
  33. LI_Summary        — hitung per klasifikasi

Reverse-engineering low_irradiance.py:
  pr_proxy = P_inv / POA ; band low=[50,250], mid=[300,800] (POA W/m²)
  OLS y=a+bx (linear_regression_slope) == Excel SLOPE/INTERCEPT/RSQ.
  Excel reproduksi via SUMPRODUCT(mask*…) (LibreOffice-safe, tanpa array formula):
     slope = (Sxy_raw − Sx·Sy/n)/(Sxx_raw − Sx²/n) ; r2 = Sxy²/(Sxx·Syy).
  classify: low<0 & mid>=0 → low_irradiance_underperform ; low<0 & mid<0 → general_underperform ; else normal.
  severity: score=|slope_low|·clamp(r2_low); ≥8e-4 CRIT, ≥4e-4 HIGH, ≥1e-4 MED. emit kalau class≠normal & r2_low≥0.3.
Demo 2 inverter (angka EXACT proto_iter7.py): INV01 low_irradiance_underperform HIGH; INV02 general_underperform CRITICAL.
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule

INPUT = Path(__file__).parent / "M2_PV_Performance_Workbook.xlsx"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=13, bold=True, color="305496")
NOTE_FONT = Font(italic=True, size=9, color="808080")
SEV_FILL = {
    "CRITICAL": PatternFill("solid", fgColor="E06666"),
    "HIGH":     PatternFill("solid", fgColor="F6B26B"),
    "MEDIUM":   PatternFill("solid", fgColor="FFE599"),
    "INFO":     PatternFill("solid", fgColor="D9D9D9"),
}
LOW_FILL = PatternFill("solid", fgColor="DDEBF7")    # low band rows tint


def set_header(ws, row, headers, start=1):
    for ci, h in enumerate(headers, start=start):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def title_note(ws, title, notes):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    for i, nt in enumerate(notes, start=2):
        ws.cell(row=i, column=1, value=nt).font = NOTE_FONT


wb = load_workbook(INPUT)
existing = list(wb.sheetnames)
assert len(existing) == 29, f"Expected 29 sheets from iter6, got {len(existing)}"
print(f"Loaded {len(existing)} sheets.")

# ===========================================================================
# README — insert row 12 (iter7) before "Cara membaca"
# ===========================================================================
ws = wb["README"]
assert ws.cell(row=12, column=1).value == "Cara membaca", "README layout berubah; abort"
ws.insert_rows(12)
ws.cell(row=12, column=1, value="7").border = BORDER
ws.cell(row=12, column=2, value="2026-05-30").border = BORDER
ws.cell(row=12, column=3, value="M2aLowIrradiance").border = BORDER
ws.cell(row=12, column=4, value="Raw_Data_LI, Helpers_LI, M2a_LowIrradiance, LI_Summary").border = BORDER
for rr in range(13, 21):
    if ws.cell(row=rr, column=1).value:
        ws.cell(row=rr, column=1).font = NOTE_FONT
ws.cell(row=12, column=1).font = Font(name="Calibri", size=11)

# ===========================================================================
# Config — append m2a_low_irradiance params + named cells (start row 52)
# ===========================================================================
ws = wb["Config"]
r = 4; last_cfg = 4
while ws.cell(row=r, column=1).value is not None:
    last_cfg = r; r += 1
start = last_cfg + 1
assert start == 52, f"Expected Config append at row 52, got {start}"
new_cfg = [
    ("m2a_low_irradiance", "poa_low_min", 50.0, "band low POA min (W/m²)", "cfg_li_poa_low_min"),
    ("m2a_low_irradiance", "poa_low_max", 250.0, "band low POA max", "cfg_li_poa_low_max"),
    ("m2a_low_irradiance", "poa_mid_min", 300.0, "band mid POA min (cross-check soiling)", "cfg_li_poa_mid_min"),
    ("m2a_low_irradiance", "poa_mid_max", 800.0, "band mid POA max", "cfg_li_poa_mid_max"),
    ("m2a_low_irradiance", "slope_threshold", 0.0, "slope_low < ini → flag (default 0 = slope negatif)", "cfg_li_slope_threshold"),
    ("m2a_low_irradiance", "r_squared_min", 0.3, "r2_low ≥ ini supaya finding di-emit", "cfg_li_r2_min"),
    ("m2a_low_irradiance", "min_low_samples", 30, "min sampel band low untuk regresi valid", "cfg_li_min_low_samples"),
]
for i, (sec, key, val, note, name) in enumerate(new_cfg):
    ri = start + i
    ws.cell(row=ri, column=1, value=sec).border = BORDER
    ws.cell(row=ri, column=2, value=key).border = BORDER
    c = ws.cell(row=ri, column=3, value=val); c.border = BORDER
    c.fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(row=ri, column=4, value=note).border = BORDER
    if name not in wb.defined_names:
        wb.defined_names[name] = DefinedName(name, attr_text=f"Config!$C${ri}")
print(f"Config: appended {len(new_cfg)} m2a_low_irradiance rows ({start}..{start+len(new_cfg)-1}) + named cells.")

# ===========================================================================
# Synthetic data (matches proto_iter7.py EXACTLY)
# ===========================================================================
NL, NM = 32, 32
POA_low = np.linspace(60.0, 245.0, NL)
POA_mid = np.linspace(310.0, 790.0, NM)
noise_low = 0.003 * np.sin(np.linspace(0, 6.0, NL))
noise_mid = 0.003 * np.sin(np.linspace(0, 6.0, NM))
DESIGN = [
    ("WB05-INV01", dict(a_low=0.160, b_low=-0.00050, a_mid=0.100, b_mid=+0.00002)),
    ("WB05-INV02", dict(a_low=0.280, b_low=-0.00090, a_mid=0.150, b_mid=-0.00010)),
]
records = []  # (inverter, POA, P_inv, band)
for inv, d in DESIGN:
    for p, nz in zip(POA_low, noise_low):
        pr = d["a_low"] + d["b_low"] * p + nz
        records.append((inv, float(p), float(pr * p), "low"))
    for p, nz in zip(POA_mid, noise_mid):
        pr = d["a_mid"] + d["b_mid"] * p + nz
        records.append((inv, float(p), float(pr * p), "mid"))
N = len(records)            # 128
DF, DL = 5, 5 + N - 1       # 5..132

# ===========================================================================
# Sheet 30: Raw_Data_LI
# ===========================================================================
ws = wb.create_sheet("Raw_Data_LI")
title_note(ws, "Raw_Data_LI — POA & daya inverter (P_inv) per sampel", [
    "2 inverter WB05. P_inv = Σ daya per-PV (kW); di sini disediakan langsung (production menjumlah PV{n} Power). "
    "Tiap inverter: 32 sampel band-low (POA 60–245) + 32 band-mid (POA 310–790) = ≥ min 30/band.",
    "pr_proxy = P_inv / POA dihitung di Helpers_LI. Ganti dengan data Huawei aktual (paste over).",
])
set_header(ws, 4, ["inverter_id", "sample", "POA (W/m²)", "P_inv (kW)", "band"])
for idx, (inv, poa, pinv, band) in enumerate(records):
    ri = DF + idx
    ws.cell(row=ri, column=1, value=inv).border = BORDER
    ws.cell(row=ri, column=2, value=idx).border = BORDER
    ws.cell(row=ri, column=3, value=poa).number_format = "0.0"; ws.cell(row=ri, column=3).border = BORDER
    ws.cell(row=ri, column=4, value=pinv).number_format = "0.000"; ws.cell(row=ri, column=4).border = BORDER
    cb = ws.cell(row=ri, column=5, value=band); cb.border = BORDER
    if band == "low":
        cb.fill = LOW_FILL
ws.column_dimensions["A"].width = 13

# ===========================================================================
# Sheet 31: Helpers_LI
# ===========================================================================
ws = wb.create_sheet("Helpers_LI")
title_note(ws, "Helpers_LI — pr_proxy + band membership", [
    "pr_proxy = P_inv / POA. in_low = 1 jika POA∈[50,250]; in_mid = 1 jika POA∈[300,800] (band dari Config).",
    "Dipakai SUMPRODUCT regresi di M2a_LowIrradiance (mask per inverter × band).",
])
set_header(ws, 4, ["inverter_id", "POA", "P_inv", "pr_proxy", "in_low", "in_mid"])
for idx in range(N):
    ri = DF + idx
    ws.cell(row=ri, column=1, value=f"=Raw_Data_LI!A{ri}")
    ws.cell(row=ri, column=2, value=f"=Raw_Data_LI!C{ri}").number_format = "0.0"
    ws.cell(row=ri, column=3, value=f"=Raw_Data_LI!D{ri}").number_format = "0.000"
    ws.cell(row=ri, column=4, value=f"=Raw_Data_LI!D{ri}/Raw_Data_LI!C{ri}").number_format = "0.00000"
    ws.cell(row=ri, column=5, value=f"=IF(AND(B{ri}>=cfg_li_poa_low_min,B{ri}<=cfg_li_poa_low_max),1,0)")
    ws.cell(row=ri, column=6, value=f"=IF(AND(B{ri}>=cfg_li_poa_mid_min,B{ri}<=cfg_li_poa_mid_max),1,0)")
    for ci in range(1, 7):
        ws.cell(row=ri, column=ci).border = BORDER
ws.column_dimensions["A"].width = 13

# ===========================================================================
# Sheet 32: M2a_LowIrradiance — regresi via SUMPRODUCT + keputusan
# ===========================================================================
ws = wb.create_sheet("M2a_LowIrradiance")
title_note(ws, "M2a_LowIrradiance — regresi OLS dua band (SUMPRODUCT) + klasifikasi", [
    "Per inverter: slope/intercept/r2 band low & mid via SUMPRODUCT(mask·…) — identik OLS linear_regression_slope ≡ Excel SLOPE/INTERCEPT/RSQ.",
    "classify: slope_low<0 & slope_mid≥0 → low_irradiance_underperform; slope_low<0 & slope_mid<0 → general_underperform. severity dari |slope_low|·r2_low. Kolom P–AA = sum antara SUMPRODUCT.",
])
hdr = ["inverter_id", "slope_low", "intercept_low", "r2_low", "n_low",
       "slope_mid", "r2_mid", "n_mid", "classification", "severity",
       "emit", "confidence", "message", "score"]
set_header(ws, 4, hdr)
# intermediate sum headers (P-AA = 16..27)
for ci, h in enumerate(["n_lo", "Sx_lo", "Sy_lo", "Sxx_lo", "Sxy_lo", "Syy_lo",
                        "n_mi", "Sx_mi", "Sy_mi", "Sxx_mi", "Sxy_mi", "Syy_mi"], start=16):
    c = ws.cell(row=4, column=ci, value=h); c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = BORDER
HR = f"Helpers_LI!$A${DF}:$A${DL}"   # inverter range
HB = f"Helpers_LI!$B${DF}:$B${DL}"   # POA
HD = f"Helpers_LI!$D${DF}:$D${DL}"   # pr_proxy
HE = f"Helpers_LI!$E${DF}:$E${DL}"   # in_low
HF = f"Helpers_LI!$F${DF}:$F${DL}"   # in_mid
for k, (inv, _d) in enumerate(DESIGN):
    r = 5 + k
    inv_match = f"({HR}=$A{r})"
    # --- intermediate SUMPRODUCT sums (low band: mask = inv_match * in_low) ---
    ws.cell(row=r, column=16, value=f"=SUMPRODUCT({inv_match}*{HE})")                       # P n_lo
    ws.cell(row=r, column=17, value=f"=SUMPRODUCT({inv_match}*{HE}*{HB})")                  # Q Sx_lo
    ws.cell(row=r, column=18, value=f"=SUMPRODUCT({inv_match}*{HE}*{HD})")                  # R Sy_lo
    ws.cell(row=r, column=19, value=f"=SUMPRODUCT({inv_match}*{HE}*{HB}*{HB})")             # S Sxx_lo
    ws.cell(row=r, column=20, value=f"=SUMPRODUCT({inv_match}*{HE}*{HB}*{HD})")             # T Sxy_lo
    ws.cell(row=r, column=21, value=f"=SUMPRODUCT({inv_match}*{HE}*{HD}*{HD})")             # U Syy_lo
    ws.cell(row=r, column=22, value=f"=SUMPRODUCT({inv_match}*{HF})")                       # V n_mi
    ws.cell(row=r, column=23, value=f"=SUMPRODUCT({inv_match}*{HF}*{HB})")                  # W Sx_mi
    ws.cell(row=r, column=24, value=f"=SUMPRODUCT({inv_match}*{HF}*{HD})")                  # X Sy_mi
    ws.cell(row=r, column=25, value=f"=SUMPRODUCT({inv_match}*{HF}*{HB}*{HB})")             # Y Sxx_mi
    ws.cell(row=r, column=26, value=f"=SUMPRODUCT({inv_match}*{HF}*{HB}*{HD})")             # Z Sxy_mi
    ws.cell(row=r, column=27, value=f"=SUMPRODUCT({inv_match}*{HF}*{HD}*{HD})")             # AA Syy_mi
    # --- regression outputs (centered sums) ---
    # slope_low = (T - Q*R/P)/(S - Q*Q/P)
    ws.cell(row=r, column=2, value=f"=(T{r}-Q{r}*R{r}/P{r})/(S{r}-Q{r}*Q{r}/P{r})").number_format = "0.000000"
    ws.cell(row=r, column=3, value=f"=R{r}/P{r}-B{r}*Q{r}/P{r}").number_format = "0.0000"           # intercept_low
    ws.cell(row=r, column=4, value=f"=(T{r}-Q{r}*R{r}/P{r})^2/((S{r}-Q{r}*Q{r}/P{r})*(U{r}-R{r}*R{r}/P{r}))").number_format = "0.0000"
    ws.cell(row=r, column=5, value=f"=P{r}")                                                        # n_low
    ws.cell(row=r, column=6, value=f"=(Z{r}-W{r}*X{r}/V{r})/(Y{r}-W{r}*W{r}/V{r})").number_format = "0.000000"
    ws.cell(row=r, column=7, value=f"=(Z{r}-W{r}*X{r}/V{r})^2/((Y{r}-W{r}*W{r}/V{r})*(AA{r}-X{r}*X{r}/V{r}))").number_format = "0.0000"
    ws.cell(row=r, column=8, value=f"=V{r}")                                                        # n_mid
    # score (col N=14)
    ws.cell(row=r, column=14, value=f"=ABS(B{r}-cfg_li_slope_threshold)*MAX(0,MIN(1,D{r}))").number_format = "0.000000"
    # classification (I=9)
    ws.cell(row=r, column=9, value=(
        f'=IF(E{r}<cfg_li_min_low_samples,"insufficient_data",'
        f'IF(AND(B{r}<cfg_li_slope_threshold,F{r}>=cfg_li_slope_threshold),"low_irradiance_underperform",'
        f'IF(AND(B{r}<cfg_li_slope_threshold,F{r}<cfg_li_slope_threshold),"general_underperform","normal")))'))
    # severity (J=10)
    ws.cell(row=r, column=10, value=(
        f'=IF(B{r}>=cfg_li_slope_threshold,"INFO",'
        f'IF(N{r}>=0.0008,"CRITICAL",IF(N{r}>=0.0004,"HIGH",IF(N{r}>=0.0001,"MEDIUM","INFO"))))'))
    # emit (K=11)
    ws.cell(row=r, column=11, value=(
        f'=IF(AND(I{r}<>"normal",I{r}<>"insufficient_data",D{r}>=cfg_li_r2_min),1,0)'))
    # confidence (L=12)
    ws.cell(row=r, column=12, value=f"=50+D{r}*50").number_format = "0.0"
    # message (M=13)
    ws.cell(row=r, column=13, value=(
        f'=IF(K{r}=1,"Low-irradiance underperformance ("&I{r}&"): slope_low="&TEXT(B{r},"0.000000")'
        f'&" (r2="&TEXT(D{r},"0.000")&", n="&E{r}&"); slope_mid="&TEXT(F{r},"0.000000"),"")'))
    for ci in range(1, 28):
        ws.cell(row=r, column=ci).border = BORDER
for sev, fill in SEV_FILL.items():
    ws.conditional_formatting.add("J5:J6",
        CellIsRule(operator="equal", formula=[f'"{sev}"'], fill=fill))
ws.column_dimensions["A"].width = 13
ws.column_dimensions["I"].width = 26
ws.column_dimensions["M"].width = 50

# ===========================================================================
# Sheet 33: LI_Summary
# ===========================================================================
ws = wb.create_sheet("LI_Summary")
title_note(ws, "LI_Summary — hitung per klasifikasi (mirror LowIrradianceSummary)", [
    "Aggregat klasifikasi inverter. slope_threshold & r2_min dari Config.",
])
set_header(ws, 4, ["classification", "count"])
cats = ["low_irradiance_underperform", "general_underperform", "normal", "insufficient_data"]
for i, cat in enumerate(cats):
    rr = 5 + i
    ws.cell(row=rr, column=1, value=cat).border = BORDER
    ws.cell(row=rr, column=2, value=f'=COUNTIF(M2a_LowIrradiance!$I$5:$I$6,"{cat}")').border = BORDER
ws.cell(row=10, column=1, value="n_emit").border = BORDER
ws.cell(row=10, column=2, value="=SUM(M2a_LowIrradiance!K5:K6)").border = BORDER
ws.column_dimensions["A"].width = 28

# ===========================================================================
# Save + integrity verify
# ===========================================================================
wb.save(INPUT)
wb2 = load_workbook(INPUT, data_only=False)
after = list(wb2.sheetnames)
new_sheets = ["Raw_Data_LI", "Helpers_LI", "M2a_LowIrradiance", "LI_Summary"]
assert after[:29] == existing, "EXISTING SHEET ORDER CHANGED!"
assert after[29:] == new_sheets, f"New sheets mismatch: {after[29:]}"
print(f"\nSheets now: {len(after)} (was 29, +4)")
d = wb2["M2a_LowIrradiance"]
print("INV01 slope_low B5:", d["B5"].value)
print("INV01 classification I5:", d["I5"].value)
print("INV02 severity J6:", d["J6"].value)
print("OK — iter7 build complete.")
