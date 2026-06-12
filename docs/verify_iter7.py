"""Verify Iterasi 7 (M2aLowIrradiance) Excel vs proto_iter7 locked numbers.

(A) STRUCTURAL audit — formula string Helpers_LI (pr_proxy, in_low/mid) + M2a_LowIrradiance
    (SUMPRODUCT sums, slope/r2 via centered sums, classification, severity, emit).
(B) NUMERIC recompute — baca literal Raw_Data_LI, re-implement pr_proxy + regresi OLS dua band
    (lewat bentuk SUMPRODUCT yang sama), assert slope/intercept/r2/classification/severity == proto.
"""
from __future__ import annotations
import numpy as np
from openpyxl import load_workbook

WB = "M2_PV_Performance_Workbook.xlsx"
DF, DL = 5, 132
errors = []
def check(c, m):
    print(("  OK  " if c else "  XX  FAIL: ") + m)
    if not c: errors.append(m)

wb = load_workbook(WB, data_only=False)
for s in ("Raw_Data_LI", "Helpers_LI", "M2a_LowIrradiance", "LI_Summary"):
    assert s in wb.sheetnames, f"missing {s}"
raw = wb["Raw_Data_LI"]; hp = wb["Helpers_LI"]; dc = wb["M2a_LowIrradiance"]

def cfgval(n):
    dn = wb.defined_names[n]; (s, c), = dn.destinations; return wb[s][c].value
LOWMIN = cfgval("cfg_li_poa_low_min"); LOWMAX = cfgval("cfg_li_poa_low_max")
MIDMIN = cfgval("cfg_li_poa_mid_min"); MIDMAX = cfgval("cfg_li_poa_mid_max")
SLOPE_THR = cfgval("cfg_li_slope_threshold"); R2MIN = cfgval("cfg_li_r2_min")
MINLOW = cfgval("cfg_li_min_low_samples")
print("config:", dict(low=(LOWMIN, LOWMAX), mid=(MIDMIN, MIDMAX), thr=SLOPE_THR, r2min=R2MIN, minlow=MINLOW))
check((LOWMIN, LOWMAX) == (50, 250), "band low [50,250]")
check((MIDMIN, MIDMAX) == (300, 800), "band mid [300,800]")
check(SLOPE_THR == 0.0 and R2MIN == 0.3 and MINLOW == 30, "thresholds 0/0.3/30")

# ===========================================================================
print("\n=== (A) STRUCTURAL audit ===")
for ri in (5, 68, 69, 132):
    check(hp[f"D{ri}"].value == f"=Raw_Data_LI!D{ri}/Raw_Data_LI!C{ri}", f"Helpers D{ri} pr_proxy")
    check(hp[f"E{ri}"].value == f"=IF(AND(B{ri}>=cfg_li_poa_low_min,B{ri}<=cfg_li_poa_low_max),1,0)", f"Helpers E{ri} in_low")
    check(hp[f"F{ri}"].value == f"=IF(AND(B{ri}>=cfg_li_poa_mid_min,B{ri}<=cfg_li_poa_mid_max),1,0)", f"Helpers F{ri} in_mid")
for r in (5, 6):
    check(dc[f"P{r}"].value == f"=SUMPRODUCT((Helpers_LI!$A$5:$A$132=$A{r})*Helpers_LI!$E$5:$E$132)", f"Dec P{r} n_lo")
    check(dc[f"T{r}"].value == f"=SUMPRODUCT((Helpers_LI!$A$5:$A$132=$A{r})*Helpers_LI!$E$5:$E$132*Helpers_LI!$B$5:$B$132*Helpers_LI!$D$5:$D$132)", f"Dec T{r} Sxy_lo")
    check(dc[f"B{r}"].value == f"=(T{r}-Q{r}*R{r}/P{r})/(S{r}-Q{r}*Q{r}/P{r})", f"Dec B{r} slope_low")
    check(dc[f"D{r}"].value == f"=(T{r}-Q{r}*R{r}/P{r})^2/((S{r}-Q{r}*Q{r}/P{r})*(U{r}-R{r}*R{r}/P{r}))", f"Dec D{r} r2_low")
    check(dc[f"I{r}"].value.startswith('=IF(E'+str(r)+'<cfg_li_min_low_samples,"insufficient_data"'), f"Dec I{r} classification")
    check(dc[f"K{r}"].value == f'=IF(AND(I{r}<>"normal",I{r}<>"insufficient_data",D{r}>=cfg_li_r2_min),1,0)', f"Dec K{r} emit")

# ===========================================================================
print("\n=== (B) NUMERIC recompute from literals ===")
# read raw
rows = []
for ri in range(DF, DL + 1):
    inv = raw[f"A{ri}"].value; poa = float(raw[f"C{ri}"].value); pinv = float(raw[f"D{ri}"].value)
    rows.append((inv, poa, pinv))

def reg_band(inv, lo, hi):
    xs, ys = [], []
    for (i2, poa, pinv) in rows:
        if i2 == inv and lo <= poa <= hi:
            xs.append(poa); ys.append(pinv / poa)
    x = np.array(xs); y = np.array(ys); n = x.size
    xb, yb = x.mean(), y.mean()
    Sxy = np.sum(x*y) - n*xb*yb; Sxx = np.sum(x*x) - n*xb*xb; Syy = np.sum(y*y) - n*yb*yb
    slope = Sxy/Sxx; intercept = yb - slope*xb; r2 = Sxy**2/(Sxx*Syy)
    return slope, intercept, r2, n

def classify(sl, sm, thr=0.0):
    lo = sl < thr; mi = sm < thr
    if lo and not mi: return "low_irradiance_underperform"
    if lo and mi: return "general_underperform"
    return "normal"

def severity(sl, r2, thr=0.0):
    if sl >= thr: return "INFO"
    score = abs(sl-thr) * max(0.0, min(1.0, r2))
    return "CRITICAL" if score >= 0.0008 else "HIGH" if score >= 0.0004 else "MEDIUM" if score >= 0.0001 else "INFO"

expect = {
    "WB05-INV01": dict(sl=-0.000531, r2l=0.9980, sm=+0.000008, cls="low_irradiance_underperform", sev="HIGH"),
    "WB05-INV02": dict(sl=-0.000931, r2l=0.9993, sm=-0.000112, cls="general_underperform", sev="CRITICAL"),
}
print(f"{'inverter':<13}{'slope_low':>12}{'r2_low':>9}{'n_low':>6}{'slope_mid':>12}{'class':>28}{'sev':>9}{'conf':>7}")
for inv in ("WB05-INV01", "WB05-INV02"):
    sl, il, r2l, nl = reg_band(inv, LOWMIN, LOWMAX)
    sm, im, r2m, nm = reg_band(inv, MIDMIN, MIDMAX)
    cls = classify(sl, sm); sev = severity(sl, r2l); conf = 50 + r2l*50
    print(f"{inv:<13}{sl:>12.6f}{r2l:>9.4f}{nl:>6}{sm:>12.6f}{cls:>28}{sev:>9}{conf:>7.1f}")
    e = expect[inv]
    check(abs(sl - e["sl"]) < 5e-6, f"{inv} slope_low≈{e['sl']} (got {sl:.6f})")
    check(abs(r2l - e["r2l"]) < 1e-3, f"{inv} r2_low≈{e['r2l']} (got {r2l:.4f})")
    check((sm >= 0) == (e["sm"] >= 0), f"{inv} slope_mid sign matches ({sm:.6f})")
    check(nl == 32, f"{inv} n_low==32 (got {nl})")
    check(cls == e["cls"], f"{inv} classification=={e['cls']}")
    check(sev == e["sev"], f"{inv} severity=={e['sev']}")
    check(50 < conf <= 100, f"{inv} confidence in (50,100]")

print("\n" + "="*60)
if errors:
    print(f"VERIFY FAILED: {len(errors)} error(s)")
    for e in errors:
        print("  -", e)
    raise SystemExit(1)
print("VERIFY OK — all structural + numeric checks passed (iter7).")
