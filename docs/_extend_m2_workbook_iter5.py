"""Extend M2 PV Performance Workbook — Iterasi 5: M2bGroundFault (label user "M2c").

Input : docs/M2_PV_Performance_Workbook.xlsx (output Iterasi 4, 20 sheet)
Output: same file, +5 sheet baru:
  21. Raw_Data_GF        — time-series V/I per PV + V_to_ground, 4 inverter (dawn+noon)
  22. Helpers_GF         — per-row daylight, voc_cand (Voc estimator), peer_median_I, abs(V_gnd)
  23. GF_StringMetrics   — per (inv,string): voc_actual, voc_ratio, i_z, spec_flag
  24. M2c_GroundFault    — per inverter: triple-signal (abs/adaptive/spec) → confidence → severity
  25. M2c_GF_StringStatus— replika artifact StringStatus (per-PV, fan-out inverter status)

Sheet existing Iterasi 1-4 TIDAK diubah (kecuali append: README log, Config m2b_ground_fault).
Cell-level formula reproducible. Angka EXACT dari proto_iter5.py (deterministic).

Reverse-engineering ground_fault.py (triple-signal cross-check):
  1. absolute : MAX|V_to_ground|(daylight) > 50 V
  2. adaptive : |median(V_gnd) - fleet_median| / max(fleet_std,0.01) > 3   [fleet = INPUT representatif]
  3. spec_4.2.3 : ADA string voc_ratio<0.85 AND i_z(peer)>2.0
  confidence: spec+(abs|adp)=90 ; spec=80 ; abs+adp=80 ; abs=70 ; adp=60 ; severity CRITICAL kalau >=80.
  voc_actual = MEDIAN(V saat |I|<0.5 & V>10) ; voc_string_nominal = voc_string_26_calc (reuse iter3).
Skenario: INV01 NORMAL · INV02 abs+adp(80) · INV03 spec(80) · INV04 spec+abs+adp(90). Empty PV5 di INV01/02.
"""
from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule

INPUT = Path(__file__).parent / "M2_PV_Performance_Workbook.xlsx"

# --- Styling (mirror iter4) ---
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=13, bold=True, color="305496")
NOTE_FONT = Font(italic=True, size=9, color="808080")
SEV_FILL = {
    "ground_fault": PatternFill("solid", fgColor="E06666"),
    "NORMAL":       PatternFill("solid", fgColor="B6D7A8"),
    "EMPTY":        PatternFill("solid", fgColor="DDDDDD"),
}
FAULT_FILL = PatternFill("solid", fgColor="F4CCCC")   # ground-fault string cells
VGND_FILL = PatternFill("solid", fgColor="FCE5CD")    # elevated V_to_ground cells
EMPTY_FILL = PatternFill("solid", fgColor="DDDDDD")   # empty slot cells


def set_header(ws, row, headers):
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def title_note(ws, title, notes):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    for i, nt in enumerate(notes, start=2):
        ws.cell(row=i, column=1, value=nt).font = NOTE_FONT


wb = load_workbook(INPUT)
existing = list(wb.sheetnames)
assert len(existing) == 20, f"Expected 20 sheets from iter4, got {len(existing)}: {existing}"
print(f"Loaded {len(existing)} sheets.")

# ===========================================================================
# README iterasi log — insert row 10 (iter5) before "Cara membaca" section
# ===========================================================================
ws = wb["README"]
assert ws.cell(row=10, column=1).value == "Cara membaca", "README layout berubah; abort"
ws.insert_rows(10)
ws.cell(row=10, column=1, value="5").border = BORDER
ws.cell(row=10, column=2, value="2026-05-30").border = BORDER
ws.cell(row=10, column=3, value="M2bGroundFault").border = BORDER
ws.cell(row=10, column=4,
        value="Raw_Data_GF, Helpers_GF, GF_StringMetrics, M2c_GroundFault, M2c_GF_StringStatus").border = BORDER
# Re-apply NOTE_FONT to shifted how-to rows (now 11..17) — insert_rows doesn't move styles.
for rr in range(11, 18):
    if ws.cell(row=rr, column=1).value:
        ws.cell(row=rr, column=1).font = NOTE_FONT
ws.cell(row=10, column=1).font = Font(name="Calibri", size=11)  # log row, not a note

# ===========================================================================
# Config — append m2b_ground_fault thresholds + fleet inputs + named cells (start 38)
# ===========================================================================
ws = wb["Config"]
r = 4
last_cfg = 4
while ws.cell(row=r, column=1).value is not None:
    last_cfg = r
    r += 1
start = last_cfg + 1
assert start == 38, f"Expected Config append at row 38, got {start}"

new_cfg = [
    ("m2b_ground_fault", "poa_threshold_wm2",       200.0, "W/m²; daylight gate (selaras spec 4.2.3 & default kode)", "cfg_gf_poa_threshold_wm2"),
    ("m2b_ground_fault", "poa_floor_wm2",           50.0,  "W/m²; hard floor sunset/twilight", "cfg_gf_poa_floor_wm2"),
    ("m2b_ground_fault", "v_to_ground_abs_threshold_v", 50.0, "MAX|V_PV-ground| > ini → trigger 'absolute'", "cfg_gf_v_abs_threshold"),
    ("m2b_ground_fault", "adaptive_z_threshold",    3.0,   "|median-fleet|/std > ini → trigger 'adaptive'", "cfg_gf_adaptive_z_threshold"),
    ("m2b_ground_fault", "voc_ratio_threshold",     0.85,  "voc_ratio < ini (+i_z>2) → trigger 'spec_4.2.3'", "cfg_gf_voc_ratio_threshold"),
    ("m2b_ground_fault", "i_high_z_threshold",      2.0,   "peer I z-score > ini → arus abnormal tinggi (spec)", "cfg_gf_i_high_z_threshold"),
    ("m2b_ground_fault", "fleet_v_gnd_median_v",    0.0,   "INPUT: median V_gnd fleet (~200 inv IKN); demo tak reproduksi fleet penuh", "cfg_gf_fleet_v_gnd_median"),
    ("m2b_ground_fault", "fleet_v_gnd_std_v",       8.0,   "INPUT: std V_gnd fleet; lihat catatan limitasi §5", "cfg_gf_fleet_v_gnd_std"),
]
for i, (sec, key, val, note, name) in enumerate(new_cfg):
    ri = start + i
    ws.cell(row=ri, column=1, value=sec).border = BORDER
    ws.cell(row=ri, column=2, value=key).border = BORDER
    c = ws.cell(row=ri, column=3, value=val)
    c.border = BORDER
    c.fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(row=ri, column=4, value=note).border = BORDER
    if name not in wb.defined_names:
        wb.defined_names[name] = DefinedName(name, attr_text=f"Config!$C${ri}")
print(f"Config: appended {len(new_cfg)} m2b_ground_fault rows ({start}..{start+len(new_cfg)-1}) + named cells.")

# ===========================================================================
# Synthetic data (matches proto_iter5.py EXACTLY, deterministic)
# ===========================================================================
PV_MAX = 6
DELTA = [0.1, -0.1, 0.1, -0.1, 0.1, -0.1]   # noon current jitter (median 0, sample-std 0.109545)
INVS = [
    # (inverter_id, v_gnd, fault_pv, empty_set)
    ("WB05-INV01", 2.0, None, {5}),
    ("WB05-INV02", -80.0, None, {5}),
    ("WB05-INV03", 1.0, 3, set()),
    ("WB05-INV04", -90.0, 3, set()),
]
dawn_times = [datetime(2026, 5, 14, 6, 0) + timedelta(minutes=5 * i) for i in range(3)]
noon_times = [datetime(2026, 5, 14, 12, 0) + timedelta(minutes=5 * i) for i in range(6)]
PER = 9                                       # rows per inverter (3 dawn + 6 noon)
N = len(INVS) * PER                           # 36 rows

def inv_block(i):          # Raw/Helpers data rows for inverter idx i (0-based)
    first = 5 + i * PER
    return first, first + 2, first + 3, first + 8   # dawn_first, dawn_last, noon_first, noon_last

# Build per-row records
records = []  # each: dict(time, inv, poa, vgnd, V[1..6], I[1..6], fault_pv, empties)
for i, (inv_id, vgnd, fault_pv, empties) in enumerate(INVS):
    for j in range(PER):
        is_noon = j >= 3
        t = noon_times[j - 3] if is_noon else dawn_times[j]
        poa = 900.0 if is_noon else 100.0
        V = {}
        I = {}
        for pv in range(1, PV_MAX + 1):
            if pv in empties:
                V[pv], I[pv] = 0.0, 0.0
            elif is_noon:
                d = DELTA[j - 3]
                if fault_pv is not None and pv == fault_pv:
                    V[pv], I[pv] = 1100.0, round(14.2 + d, 4)
                else:
                    V[pv], I[pv] = 1180.0, round(13.0 + d, 4)
            else:  # dawn (open-circuit: V~Voc, I~0)
                V[pv] = 1100.0 if (fault_pv is not None and pv == fault_pv) else 1430.0
                I[pv] = 0.2
        records.append(dict(time=t, inv=inv_id, poa=poa, vgnd=vgnd, V=V, I=I,
                            fault_pv=fault_pv, empties=empties))

# ===========================================================================
# Sheet 21: Raw_Data_GF
# ===========================================================================
ws = wb.create_sheet("Raw_Data_GF")
title_note(ws, "Raw_Data_GF — V/I per PV string + V_to_ground (4 inverter WB05)", [
    "INV01 NORMAL · INV02 V_to_ground tinggi (abs+adaptive) · INV03 PV3 ground-fault (spec: Voc↓ + I↑) · INV04 V_gnd tinggi + PV3 spec.",
    "Tiap inverter: 3 baris dawn (06:00-06:10, I≈0 → Voc estimate) + 6 baris noon (12:00-12:25, POA=900 daylight). Empty slot PV5 (INV01/02) = 0.",
    "Ganti dengan data Huawei aktual (paste over). V_to_ground = kolom 'Voltage between PV– and the ground(V)'.",
])
V_COL = {k: get_column_letter(4 + k) for k in range(1, PV_MAX + 1)}   # E..J voltage
I_COL = {k: get_column_letter(10 + k) for k in range(1, PV_MAX + 1)}  # K..P current
hdr = ["Start Time", "Inverter_ID", "POA (W/m²)", "V_to_ground (V)"] \
    + [f"PV{k} input voltage(V)" for k in range(1, PV_MAX + 1)] \
    + [f"PV{k} input current(A)" for k in range(1, PV_MAX + 1)]
set_header(ws, 4, hdr)
for idx, rec in enumerate(records):
    ri = 5 + idx
    c = ws.cell(row=ri, column=1, value=rec["time"]); c.number_format = "yyyy-mm-dd hh:mm:ss"; c.border = BORDER
    ws.cell(row=ri, column=2, value=rec["inv"]).border = BORDER
    ws.cell(row=ri, column=3, value=rec["poa"]).border = BORDER
    cv = ws.cell(row=ri, column=4, value=rec["vgnd"]); cv.number_format = "0.0"; cv.border = BORDER
    if abs(rec["vgnd"]) > 50:
        cv.fill = VGND_FILL
    for k in range(1, PV_MAX + 1):
        cV = ws.cell(row=ri, column=4 + k, value=rec["V"][k]); cV.number_format = "0.0"; cV.border = BORDER
        cI = ws.cell(row=ri, column=10 + k, value=rec["I"][k]); cI.number_format = "0.000"; cI.border = BORDER
        if k in rec["empties"]:
            cV.fill = EMPTY_FILL; cI.fill = EMPTY_FILL
        elif rec["fault_pv"] == k:
            cV.fill = FAULT_FILL; cI.fill = FAULT_FILL
ws.column_dimensions["A"].width = 19
ws.column_dimensions["B"].width = 13

# ===========================================================================
# Sheet 22: Helpers_GF (per-row mechanics)
# ===========================================================================
ws = wb.create_sheet("Helpers_GF")
title_note(ws, "Helpers_GF — jembatan ground_fault.py (per-row)", [
    "daylight = (POA>thr & POA>floor). voc_cand = IF(|I|<0.5 & V>10, V, '') → Voc estimator (median nanti). "
    "peer_med_I PVk = MEDIAN arus 5 string LAIN (per baris). abs_Vgnd = |V_to_ground|.",
    "Production AND solar_elev>5° & inverter-shutdown (ephemeris, tak direplika statis); di sini daylight = baris noon.",
])
# Helpers columns: A inv, B time, C POA, D daylight, E V_gnd, F-K voc_cand, L-Q peer_med_I, R abs_Vgnd
VOC_CAND_COL = {k: get_column_letter(5 + k) for k in range(1, PV_MAX + 1)}   # F..K
PEER_COL = {k: get_column_letter(11 + k) for k in range(1, PV_MAX + 1)}      # L..Q
ABS_VGND_COL = "R"
hdr = ["Inverter_ID", "Start Time", "POA", "daylight", "V_gnd"] \
    + [f"voc_cand PV{k}" for k in range(1, PV_MAX + 1)] \
    + [f"peer_med_I PV{k}" for k in range(1, PV_MAX + 1)] + ["abs_Vgnd"]
set_header(ws, 4, hdr)
for idx in range(N):
    ri = 5 + idx
    ws.cell(row=ri, column=1, value=f"=Raw_Data_GF!B{ri}")
    ws.cell(row=ri, column=2, value=f"=Raw_Data_GF!A{ri}").number_format = "yyyy-mm-dd hh:mm:ss"
    ws.cell(row=ri, column=3, value=f"=Raw_Data_GF!C{ri}").number_format = "0"
    ws.cell(row=ri, column=4, value=f"=IF(AND(C{ri}>cfg_gf_poa_threshold_wm2,C{ri}>cfg_gf_poa_floor_wm2),1,0)")
    ws.cell(row=ri, column=5, value=f"=Raw_Data_GF!D{ri}").number_format = "0.0"
    for k in range(1, PV_MAX + 1):
        # voc_cand
        ws.cell(row=ri, column=5 + k,
                value=(f"=IF(AND(ABS(Raw_Data_GF!{I_COL[k]}{ri})<cfg_i_threshold_a,"
                       f"Raw_Data_GF!{V_COL[k]}{ri}>cfg_min_voc_v),Raw_Data_GF!{V_COL[k]}{ri},\"\")")
                ).number_format = "0.0"
        # peer_med_I = MEDIAN of other 5 currents
        others = ",".join(f"Raw_Data_GF!{I_COL[j]}{ri}" for j in range(1, PV_MAX + 1) if j != k)
        ws.cell(row=ri, column=11 + k, value=f"=MEDIAN({others})").number_format = "0.000"
    ws.cell(row=ri, column=18, value=f"=ABS(E{ri})").number_format = "0.0"
    for ci in range(1, 19):
        ws.cell(row=ri, column=ci).border = BORDER
ws.column_dimensions["A"].width = 13
ws.column_dimensions["B"].width = 19

# ===========================================================================
# Sheet 23: GF_StringMetrics (per inv,string)
# ===========================================================================
ws = wb.create_sheet("GF_StringMetrics")
title_note(ws, "GF_StringMetrics — per (inverter, PV string): Voc & peer-I z-score", [
    "voc_actual = MEDIAN(voc_cand) (hanya baris dawn I<0.5 non-blank). voc_ratio = voc_actual / voc_string_26_calc (1430.61V @30°C). "
    "i_z = (I_median_daylight − peer_med) / peer_std. spec_flag = voc_ratio<0.85 AND i_z>2 (string bukan-empty).",
])
set_header(ws, 4, ["Inverter_ID", "PV", "is_empty", "voc_actual (V)", "voc_string_nominal (V)",
                   "voc_ratio", "I_median_daylight", "peer_med_daylight", "peer_std_daylight",
                   "i_z", "spec_flag"])
SM_FIRST = 5
def sm_row(i, k):   # StringMetrics row for inverter idx i, PV k
    return SM_FIRST + i * PV_MAX + (k - 1)
for i, (inv_id, vgnd, fault_pv, empties) in enumerate(INVS):
    df0, df1, nf0, nf1 = inv_block(i)
    for k in range(1, PV_MAX + 1):
        r = sm_row(i, k)
        is_empty = 1 if k in empties else 0
        vcol = VOC_CAND_COL[k]
        pcol = PEER_COL[k]
        icol = I_COL[k]
        ws.cell(row=r, column=1, value=inv_id)
        ws.cell(row=r, column=2, value=f"PV{k}")
        ws.cell(row=r, column=3, value=is_empty)
        # voc_actual: median of voc_cand over inverter's 9 rows (dawn non-blank)
        ws.cell(row=r, column=4,
                value=f"=IFERROR(MEDIAN(Helpers_GF!{vcol}{df0}:{vcol}{nf1}),\"\")").number_format = "0.0"
        ws.cell(row=r, column=5, value="=voc_string_26_calc").number_format = "0.00"
        ws.cell(row=r, column=6, value=f"=IFERROR(D{r}/E{r},\"\")").number_format = "0.0000"
        # I_median over noon (daylight) rows
        ws.cell(row=r, column=7,
                value=f"=MEDIAN(Raw_Data_GF!{icol}{nf0}:{icol}{nf1})").number_format = "0.000"
        ws.cell(row=r, column=8,
                value=f"=MEDIAN(Helpers_GF!{pcol}{nf0}:{pcol}{nf1})").number_format = "0.000"
        ws.cell(row=r, column=9,
                value=f"=IF(STDEV(Helpers_GF!{pcol}{nf0}:{pcol}{nf1})=0,0.01,STDEV(Helpers_GF!{pcol}{nf0}:{pcol}{nf1}))").number_format = "0.000000"
        ws.cell(row=r, column=10, value=f"=IFERROR((G{r}-H{r})/I{r},\"\")").number_format = "0.000"
        ws.cell(row=r, column=11,
                value=f"=IF(C{r}=1,0,IF(AND(F{r}<cfg_gf_voc_ratio_threshold,J{r}>cfg_gf_i_high_z_threshold),1,0))")
        for ci in range(1, 12):
            ws.cell(row=r, column=ci).border = BORDER
ws.column_dimensions["A"].width = 13

# ===========================================================================
# Sheet 24: M2c_GroundFault (per-inverter decision)
# ===========================================================================
ws = wb.create_sheet("M2c_GroundFault")
title_note(ws, "M2c_GroundFault — keputusan per-inverter (ground_fault.py run())", [
    "Triple-signal: absolute(MAX|V_gnd|>50) · adaptive(|med-fleet|/std>3) · spec(ADA string voc_ratio<0.85 & i_z>2). "
    "confidence: spec+(abs|adp)=90 · spec=80 · abs+adp=80 · abs=70 · adp=60. severity CRITICAL kalau conf≥80 else HIGH.",
    "Finding per-inverter; pv_string = worst (voc_ratio terendah). value = |V_gnd_median|. fleet median/std = INPUT (Config).",
])
set_header(ws, 4, ["Inverter_ID", "V_gnd_max_abs", "V_gnd_median", "adaptive_z",
                   "flag_absolute", "flag_adaptive", "flag_spec", "triggered_by",
                   "confidence", "severity", "worst_PV", "worst_voc_ratio", "status", "message"])
DEC_FIRST = 5
for i, (inv_id, vgnd, fault_pv, empties) in enumerate(INVS):
    r = DEC_FIRST + i
    df0, df1, nf0, nf1 = inv_block(i)
    sm0 = sm_row(i, 1)
    sm1 = sm_row(i, PV_MAX)
    ws.cell(row=r, column=1, value=inv_id)
    ws.cell(row=r, column=2, value=f"=MAX(Helpers_GF!{ABS_VGND_COL}{nf0}:{ABS_VGND_COL}{nf1})").number_format = "0.0"
    ws.cell(row=r, column=3, value=f"=MEDIAN(Raw_Data_GF!D{nf0}:D{nf1})").number_format = "0.0"
    ws.cell(row=r, column=4, value=f"=ABS(C{r}-cfg_gf_fleet_v_gnd_median)/MAX(cfg_gf_fleet_v_gnd_std,0.01)").number_format = "0.000"
    ws.cell(row=r, column=5, value=f"=IF(B{r}>cfg_gf_v_abs_threshold,1,0)")
    ws.cell(row=r, column=6, value=f"=IF(D{r}>cfg_gf_adaptive_z_threshold,1,0)")
    ws.cell(row=r, column=7, value=f"=IF(SUM(GF_StringMetrics!K{sm0}:K{sm1})>0,1,0)")
    ws.cell(row=r, column=8, value=f'=MID(IF(E{r}=1,"+absolute","")&IF(F{r}=1,"+adaptive","")&IF(G{r}=1,"+spec_4.2.3",""),2,99)')
    ws.cell(row=r, column=9,
            value=(f"=IF(AND(G{r}=1,OR(E{r}=1,F{r}=1)),90,"
                   f"IF(G{r}=1,80,IF(AND(E{r}=1,F{r}=1),80,IF(E{r}=1,70,IF(F{r}=1,60,0)))))")).number_format = "0"
    ws.cell(row=r, column=10, value=f'=IF(I{r}>=80,"CRITICAL",IF(I{r}>0,"HIGH","-"))')
    ws.cell(row=r, column=11, value=f"=INDEX(GF_StringMetrics!B{sm0}:B{sm1},MATCH(L{r},GF_StringMetrics!F{sm0}:F{sm1},0))")
    ws.cell(row=r, column=12, value=f"=MIN(GF_StringMetrics!F{sm0}:F{sm1})").number_format = "0.0000"
    ws.cell(row=r, column=13, value=f'=IF(OR(E{r}=1,F{r}=1,G{r}=1),"ground_fault","NORMAL")')
    ws.cell(row=r, column=14,
            value=(f'=IF(M{r}="ground_fault","Ground-fault suspect ("&H{r}&"): |V_gnd_med|="&TEXT(ABS(C{r}),"0.0")'
                   f'&"V worst "&K{r}&" voc_ratio="&TEXT(L{r},"0.000")&" conf "&I{r}&"%","")'))
    for ci in range(1, 15):
        ws.cell(row=r, column=ci).border = BORDER
ws.conditional_formatting.add(f"M5:M{DEC_FIRST+len(INVS)-1}",
    CellIsRule(operator="equal", formula=['"ground_fault"'], fill=SEV_FILL["ground_fault"]))
ws.conditional_formatting.add(f"M5:M{DEC_FIRST+len(INVS)-1}",
    CellIsRule(operator="equal", formula=['"NORMAL"'], fill=SEV_FILL["NORMAL"]))
ws.column_dimensions["A"].width = 13
ws.column_dimensions["H"].width = 26
ws.column_dimensions["N"].width = 60

# ===========================================================================
# Sheet 25: M2c_GF_StringStatus (replika artifact)
# ===========================================================================
ws = wb.create_sheet("M2c_GF_StringStatus")
title_note(ws, "M2c_GF_StringStatus — replika artifact StringStatus (ground_fault.py)", [
    "Mirror self.artifacts['StringStatus']: per-PV status (fan-out status inverter), is_worst_string, V_gnd stats, confidence. "
    "Empty slot → EMPTY (top-up EmptyPVMap). Status string = status inverter (semua string inverter ter-flag dapat 'ground_fault').",
])
set_header(ws, 4, ["poa_source", "inverter_id", "wb_id", "pv_string", "status", "is_worst_string",
                   "v_gnd_median", "v_gnd_max_abs", "adaptive_z", "triggered_by", "confidence", "voc_ratio"])
SS_FIRST = 5
for i, (inv_id, vgnd, fault_pv, empties) in enumerate(INVS):
    dec = DEC_FIRST + i
    for k in range(1, PV_MAX + 1):
        r = SS_FIRST + i * PV_MAX + (k - 1)
        smr = sm_row(i, k)
        is_empty = k in empties
        ws.cell(row=r, column=1, value="dummy_single")
        ws.cell(row=r, column=2, value=inv_id)
        ws.cell(row=r, column=3, value="WB05")
        ws.cell(row=r, column=4, value=f"PV{k}")
        if is_empty:
            ws.cell(row=r, column=5, value="EMPTY")
        else:
            ws.cell(row=r, column=5, value=f"=M2c_GroundFault!M{dec}")
        ws.cell(row=r, column=6, value=f'=IF(M2c_GroundFault!K{dec}="PV{k}",TRUE,FALSE)')
        ws.cell(row=r, column=7, value=f"=M2c_GroundFault!C{dec}").number_format = "0.0"
        ws.cell(row=r, column=8, value=f"=M2c_GroundFault!B{dec}").number_format = "0.0"
        ws.cell(row=r, column=9, value=f"=M2c_GroundFault!D{dec}").number_format = "0.000"
        ws.cell(row=r, column=10, value=f"=M2c_GroundFault!H{dec}")
        ws.cell(row=r, column=11, value=f"=M2c_GroundFault!I{dec}").number_format = "0"
        ws.cell(row=r, column=12, value=f"=IFERROR(GF_StringMetrics!F{smr},\"\")").number_format = "0.0000"
        for ci in range(1, 13):
            ws.cell(row=r, column=ci).border = BORDER
last_ss = SS_FIRST + len(INVS) * PV_MAX - 1
ws.conditional_formatting.add(f"E5:E{last_ss}",
    CellIsRule(operator="equal", formula=['"ground_fault"'], fill=SEV_FILL["ground_fault"]))
ws.conditional_formatting.add(f"E5:E{last_ss}",
    CellIsRule(operator="equal", formula=['"NORMAL"'], fill=SEV_FILL["NORMAL"]))
ws.conditional_formatting.add(f"E5:E{last_ss}",
    CellIsRule(operator="equal", formula=['"EMPTY"'], fill=SEV_FILL["EMPTY"]))
ws.column_dimensions["B"].width = 13
ws.column_dimensions["J"].width = 24

# ===========================================================================
# Save + integrity verify
# ===========================================================================
wb.save(INPUT)

wb2 = load_workbook(INPUT, data_only=False)
after = list(wb2.sheetnames)
new_sheets = ["Raw_Data_GF", "Helpers_GF", "GF_StringMetrics", "M2c_GroundFault", "M2c_GF_StringStatus"]
assert after[:20] == existing, "EXISTING SHEET ORDER CHANGED!"
assert after[20:] == new_sheets, f"New sheets mismatch: {after[20:]}"
print(f"\nSheets now: {len(after)} (was 20, +5)")
print("New:", new_sheets)
sm = wb2["GF_StringMetrics"]
print("StringMetrics INV03 PV3 voc_ratio F19:", sm["F19"].value)
print("StringMetrics INV03 PV3 i_z J19:", sm["J19"].value, "| spec_flag K19:", sm["K19"].value)
d = wb2["M2c_GroundFault"]
print("Decision INV02 conf I6:", d["I6"].value)
print("Decision INV04 trig H8:", d["H8"].value)
print("Decision INV04 worst K8:", d["K8"].value)
print("OK — iter5 build complete.")
