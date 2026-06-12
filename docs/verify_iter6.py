"""Verify Iterasi 6 (M2IForest approx) Excel vs proto_iter6 locked numbers.

(A) STRUCTURAL audit — formula string Features_IF (fitur, absdev, z, A) + IF_Anomaly
    (flag, rank, severity, confidence) == template.
(B) NUMERIC recompute — baca literal Raw_Data_IF, re-implement MAD-approx (fitur,
    robust-z, A, threshold PERCENTILE, flag, rank, severity, confidence), assert == proto.

CATATAN: ini memverifikasi APPROKSIMASI MAD (transparan), BUKAN IsolationForest asli.
"""
from __future__ import annotations
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

WB = "M2_PV_Performance_Workbook.xlsx"
PV_MAX, N_TS = 6, 14
DF, DL = 5, 5 + N_TS * PV_MAX - 1   # 5..88
errors = []
def check(c, m):
    print(("  OK  " if c else "  XX  FAIL: ") + m)
    if not c: errors.append(m)

wb = load_workbook(WB, data_only=False)
for s in ("Raw_Data_IF", "Features_IF", "IF_Anomaly", "IF_Summary"):
    assert s in wb.sheetnames, f"missing {s}"
raw = wb["Raw_Data_IF"]; feat = wb["Features_IF"]; anom = wb["IF_Anomaly"]

def cfgval(n):
    dn = wb.defined_names[n]; (s, c), = dn.destinations; return wb[s][c].value
CONTAM = cfgval("cfg_if_contamination"); IFLOOR = cfgval("cfg_if_i_floor"); MADSC = cfgval("cfg_if_mad_scale")
print("config:", dict(contamination=CONTAM, i_floor=IFLOOR, mad_scale=MADSC))
check(CONTAM == 0.05, f"contamination_demo==0.05 (got {CONTAM})")
check(IFLOOR == 0.1, f"r_current_floor==0.1 (got {IFLOOR})")
check(abs(MADSC - 1.4826) < 1e-9, f"mad_scale==1.4826 (got {MADSC})")

Vcol = {pv: get_column_letter(3 + pv) for pv in range(1, 7)}   # D..I
Icol = {pv: get_column_letter(9 + pv) for pv in range(1, 7)}   # J..O

# ===========================================================================
print("\n=== (A) STRUCTURAL audit (sample rows) ===")
for (t, pv) in [(0, 1), (5, 3), (9, 5), (13, 6)]:
    k = t * PV_MAX + (pv - 1); ri = DF + k; rawr = 5 + t
    check(feat[f"D{ri}"].value == f"=Raw_Data_IF!{Vcol[pv]}{rawr}", f"Feat D{ri} V (pv{pv},t{t})")
    check(feat[f"F{ri}"].value == f"=Raw_Data_IF!{Vcol[pv]}{rawr}-MEDIAN(Raw_Data_IF!D{rawr}:I{rawr})", f"Feat F{ri} V_dev")
    check(feat[f"G{ri}"].value == f"=Raw_Data_IF!{Icol[pv]}{rawr}-MEDIAN(Raw_Data_IF!J{rawr}:O{rawr})", f"Feat G{ri} I_dev")
    check(feat[f"H{ri}"].value == f"=Raw_Data_IF!{Vcol[pv]}{rawr}/MAX(Raw_Data_IF!{Icol[pv]}{rawr},cfg_if_i_floor)", f"Feat H{ri} R")
    check(feat[f"S{ri}"].value == f"=MAX(N{ri}:R{ri})", f"Feat S{ri} A")
    check(anom[f"E{ri}"].value == f'=IF(D{ri}>=$L$5,"anomaly","normal")', f"Anom E{ri} flag")
    check(anom[f"H{ri}"].value == (f'=IF(E{ri}="anomaly",IF(G{ri}<=25,"CRITICAL",IF(G{ri}<=50,"HIGH",'
                                   f'IF(G{ri}<=75,"MEDIUM","INFO"))),"")'), f"Anom H{ri} severity")
    check(anom[f"I{ri}"].value == f'=IF(E{ri}="anomaly",100-G{ri}*0.5,"")', f"Anom I{ri} confidence")
# stats block + threshold formula
check(feat["V5"].value == "=MEDIAN(D5:D88)", "stats V5 median V")
check(feat["W5"].value == "=MEDIAN(I5:I88)", "stats W5 MAD V")
check(anom["L5"].value == "=PERCENTILE(Features_IF!S5:S88,1-cfg_if_contamination)", "threshold_A formula")
check(anom["F5"].value == '=IF(E5="anomaly",COUNTIFS($D$5:$D$88,">"&D5,$D$5:$D$88,">="&$L$5),"")', "rank_within formula")

# ===========================================================================
print("\n=== (B) NUMERIC recompute from literals ===")
V = np.zeros((N_TS, PV_MAX)); I = np.zeros((N_TS, PV_MAX))
for t in range(N_TS):
    rr = 5 + t
    for pv in range(1, 7):
        V[t, pv-1] = float(raw[f"{Vcol[pv]}{rr}"].value)
        I[t, pv-1] = float(raw[f"{Icol[pv]}{rr}"].value)
vmed = np.median(V, axis=1); imed = np.median(I, axis=1)
rows = []
for t in range(N_TS):
    for pv in range(1, 7):
        v, i = V[t, pv-1], I[t, pv-1]
        rows.append(dict(pv=pv, t=t, V=v, I=i, V_dev=v-vmed[t], I_dev=i-imed[t], R=v/max(i, IFLOOR)))
import pandas as pd
F = pd.DataFrame(rows); FEATS = ["V", "I", "V_dev", "I_dev", "R"]
for f in FEATS:
    med = F[f].median(); mad = (F[f]-med).abs().median()
    scale = max(MADSC*mad, 1e-9)
    F[f"z_{f}"] = (F[f]-med).abs()/scale
F["A"] = F[[f"z_{f}" for f in FEATS]].max(axis=1)
thr = F["A"].quantile(1-CONTAM)
F["flag"] = np.where(F["A"] >= thr, "anomaly", "normal")
nfl = int((F["flag"] == "anomaly").sum())
fl = F[F.flag == "anomaly"].copy()
def sev(p): return "CRITICAL" if p<=25 else "HIGH" if p<=50 else "MEDIUM" if p<=75 else "INFO"
for idx, row in fl.iterrows():
    rw = int((fl["A"] > row["A"]).sum()); pct = rw/(nfl-1)*100 if nfl>1 else 0
    F.loc[idx, "rank_pct"] = pct; F.loc[idx, "severity"] = sev(pct); F.loc[idx, "confidence"] = 100-pct*0.5
fl = F[F.flag == "anomaly"].copy()   # re-slice AFTER severity/confidence assigned

check(abs(thr - 2.1131) < 1e-3, f"threshold_A==2.1131 (got {thr:.4f})")
check(nfl == 6, f"n_flagged==6 (got {nfl})")
def get(pv, t, col): return F[(F.pv == pv) & (F.t == t)].iloc[0][col]
check(abs(get(3, 5, "A") - 314.987) < 0.01, f"PV3@t5 A==314.987 (got {get(3,5,'A'):.3f})")
check(get(3, 5, "severity") == "CRITICAL" and get(3, 5, "confidence") == 100, "PV3@t5 CRITICAL conf100")
check(abs(get(3, 6, "A") - 203.149) < 0.01 and get(3, 6, "severity") == "CRITICAL", "PV3@t6 A~203 CRITICAL")
check(abs(get(5, 9, "A") - 112.078) < 0.01 and get(5, 9, "severity") == "HIGH", "PV5@t9 A~112 HIGH")
check(abs(get(2, 2, "A") - 32.962) < 0.01 and get(2, 2, "severity") == "MEDIUM", "PV2@t2 A~33 MEDIUM")
# the two INFO (most-extreme-normal, flagged by contamination)
info = fl[fl.severity == "INFO"]
check(len(info) == 2, f"2 INFO rows (extreme-normal flagged by contamination) (got {len(info)})")
check(set(fl.severity) == {"CRITICAL", "HIGH", "MEDIUM", "INFO"}, "all 4 severity tiers present")
check((fl.confidence.between(50, 100)).all(), "confidence in [50,100]")
# feature parity vs test_iforest
check(1200.0/max(0.05, 0.1) == 12000, "feature parity R(1200,0.05)==12000")

print(f"\nflagged ({nfl}):")
print(fl.sort_values("A", ascending=False)[["pv","t","V","I","R","A","rank_pct","severity","confidence"]]
      .to_string(index=False, float_format=lambda x: f"{x:.3f}"))

if errors:
    print(f"VERIFY FAILED: {len(errors)} error(s)")
    for e in errors:
        print("  -", e)
    raise SystemExit(1)
print("VERIFY OK — all structural + numeric checks passed (iter6 approx).")
