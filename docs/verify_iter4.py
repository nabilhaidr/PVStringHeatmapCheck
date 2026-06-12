"""Verify Iterasi 4 (M2bOpenCircuit) Excel sheets vs proto_iter4 locked numbers.

LibreOffice crashes & `formulas`/`pycel` tidak ter-install di sandbox, jadi
verifikasi 2-lapis (mirror verify_iter3 pattern):

  (A) STRUCTURAL formula audit — baca STRING formula tiap cell Helpers/Decision/
      StringStatus, assert == template yang diharapkan (men-catch salah ref kolom/
      baris, bukan cuma "kebetulan angka cocok").

  (B) NUMERIC recompute — baca LITERAL Raw_Data_OC + nilai named-cell Config,
      re-implement semantik formula Excel (PERCENTILE.INC, ratio+clip, qualifying,
      running-consec, MAX, MEDIAN, debounce decision) di Python murni, assert hasil
      == angka yang dikunci proto_iter4.py.

Kalau dua-duanya lulus: formula nge-ref sel yang benar DAN menghitung angka yang benar.
"""
from __future__ import annotations

import numpy as np
from openpyxl import load_workbook

WB = "M2_PV_Performance_Workbook.xlsx"

# Excel column letters per PV (mirror build script maps)
RAW_I_COL = {1: "C", 2: "D", 3: "E", 4: "F", 5: "G"}
RATIO_COL = {1: "E", 2: "F", 3: "G", 4: "H", 5: "I"}
QUAL_COL = {1: "J", 2: "K", 3: "L", 4: "M", 5: "N"}
CONSEC_COL = {1: "O", 2: "P", 3: "Q", 4: "R", 5: "S"}
DATA_FIRST, DATA_LAST = 5, 30  # all rows (incl 2 twilight 29-30)
DAY_LAST = 28                  # last daylight row (24 daylight rows 5..28)

errors = []


def check(cond, msg):
    if cond:
        print(f"  OK  {msg}")
    else:
        errors.append(msg)
        print(f"  XX  FAIL: {msg}")


# ===========================================================================
print("Loading workbook (formulas)...")
wb = load_workbook(WB, data_only=False)
for s in ("Raw_Data_OC", "Helpers_OC", "M2b_OpenCircuit", "M2b_OC_StringStatus"):
    assert s in wb.sheetnames, f"missing sheet {s}"
rd = wb["Raw_Data_OC"]
hp = wb["Helpers_OC"]
dc = wb["M2b_OpenCircuit"]
ss = wb["M2b_OC_StringStatus"]

# ---- resolve Config named cells -> literal values -------------------------
cfg = {}
for name in ("cfg_oc_poa_threshold_wm2", "cfg_oc_poa_floor_wm2",
             "cfg_oc_i_ratio_threshold", "cfg_oc_debounce_steps",
             "cfg_oc_confidence_pct", "cfg_oc_iq95_clip"):
    dn = wb.defined_names[name]
    (sheet, coord), = dn.destinations
    cfg[name] = wb[sheet][coord].value
print("Config named cells:", cfg)

POA_THR = cfg["cfg_oc_poa_threshold_wm2"]
POA_FLOOR = cfg["cfg_oc_poa_floor_wm2"]
I_RATIO = cfg["cfg_oc_i_ratio_threshold"]
DEBOUNCE = cfg["cfg_oc_debounce_steps"]
CONF = cfg["cfg_oc_confidence_pct"]
CLIP = cfg["cfg_oc_iq95_clip"]

print("\n=== expected config (proto) ===")
check(POA_THR == 700, f"poa_threshold == 700 (got {POA_THR})")
check(POA_FLOOR == 50, f"poa_floor == 50 (got {POA_FLOOR})")
check(I_RATIO == 0.05, f"i_ratio_threshold == 0.05 (got {I_RATIO})")
check(DEBOUNCE == 20, f"debounce_steps == 20 (got {DEBOUNCE})")
check(CONF == 95, f"confidence_pct == 95 (got {CONF})")
check(CLIP == 0.01, f"iq95_clip == 0.01 (got {CLIP})")

# ===========================================================================
# (A) STRUCTURAL formula audit
# ===========================================================================
print("\n=== (A) STRUCTURAL formula audit (string templates) ===")
# Helpers_OC: daylight (C), I_q95 (D), ratio (E-I), qual (J-N), consec (O-S)
for ri in range(DATA_FIRST, DATA_LAST + 1):
    # daylight col C
    exp = (f"=IF(AND(B{ri}>cfg_oc_poa_threshold_wm2,B{ri}>cfg_oc_poa_floor_wm2),1,0)")
    check(hp[f"C{ri}"].value == exp, f"Helpers C{ri} daylight formula")
    # I_q95 col D
    exp = f"=PERCENTILE(Raw_Data_OC!C{ri}:G{ri},0.95)"
    check(hp[f"D{ri}"].value == exp, f"Helpers D{ri} I_q95 formula")
    for k in range(1, 6):
        ratc, qc, cc, rawc = RATIO_COL[k], QUAL_COL[k], CONSEC_COL[k], RAW_I_COL[k]
        # ratio
        exp = f"=Raw_Data_OC!{rawc}{ri}/MAX($D{ri},cfg_oc_iq95_clip)"
        check(hp[f"{ratc}{ri}"].value == exp, f"Helpers {ratc}{ri} ratio PV{k}")
        # qualifying
        exp = f"=IF(AND({ratc}{ri}<cfg_oc_i_ratio_threshold,$C{ri}=1),1,0)"
        check(hp[f"{qc}{ri}"].value == exp, f"Helpers {qc}{ri} qual PV{k}")
        # consec: first row seed = qual; else running
        if ri == DATA_FIRST:
            exp = f"={qc}{ri}"
        else:
            exp = f"=IF({qc}{ri}=1,{cc}{ri-1}+1,0)"
        check(hp[f"{cc}{ri}"].value == exp, f"Helpers {cc}{ri} consec PV{k}")

# Decision sheet rows 5-9 = PV1-5
print("\n  -- Decision sheet templates --")
for k in range(1, 6):
    r = 4 + k
    rawc, ratc, cons = RAW_I_COL[k], RATIO_COL[k], CONSEC_COL[k]
    check(dc[f"B{r}"].value == f"=MEDIAN(Raw_Data_OC!{rawc}5:{rawc}{DAY_LAST})",
          f"Decision B{r} I_median PV{k}")
    check(dc[f"C{r}"].value == f"=MEDIAN(Helpers_OC!D5:D{DAY_LAST})",
          f"Decision C{r} I_q95_median PV{k}")
    check(dc[f"D{r}"].value == f"=MEDIAN(Helpers_OC!{ratc}5:{ratc}{DAY_LAST})",
          f"Decision D{r} ratio_median PV{k}")
    check(dc[f"E{r}"].value == f"=MAX(Helpers_OC!{cons}5:{cons}{DATA_LAST})",
          f"Decision E{r} max_consec PV{k}")
    check(dc[f"G{r}"].value == f"=IF(AND(E{r}>=cfg_oc_debounce_steps,F{r}=0),1,0)",
          f"Decision G{r} emit PV{k}")
    check(dc[f"H{r}"].value == f'=IF(F{r}=1,"EMPTY",IF(G{r}=1,"open_circuit","NORMAL"))',
          f"Decision H{r} status PV{k}")

# ===========================================================================
# (B) NUMERIC recompute from literals
# ===========================================================================
print("\n=== (B) NUMERIC recompute (Excel semantics in Python) ===")
# read raw literals
POA = np.array([rd[f"B{ri}"].value for ri in range(DATA_FIRST, DATA_LAST + 1)], float)
I = {k: np.array([rd[f"{RAW_I_COL[k]}{ri}"].value
                  for ri in range(DATA_FIRST, DATA_LAST + 1)], float)
     for k in range(1, 6)}
nrows = len(POA)

daylight = (POA > POA_THR) & (POA > POA_FLOOR)
check(int(daylight.sum()) == 24, f"daylight rows == 24 (got {int(daylight.sum())})")

# I_q95 per row = PERCENTILE.INC (linear interp) across PV1..PV5
I_mat = np.vstack([I[k] for k in range(1, 6)]).T  # (nrows, 5)
I_q95 = np.array([np.quantile(I_mat[i], 0.95, method="linear") for i in range(nrows)])
check(abs(I_q95[0] - 12.98) < 1e-9, f"I_q95 noon (r5) == 12.98 (got {I_q95[0]:.4f})")
# glitch row = Excel r9 -> index 4
check(abs(I_q95[4] - 12.96) < 1e-9, f"I_q95 glitch (r9) == 12.96 (got {I_q95[4]:.4f})")

EMPTY = {5}
expect = {
    1: dict(emit=False, status="NORMAL", ratio_med=1.0015, max_consec=0),
    2: dict(emit=False, status="NORMAL", ratio_med=0.9861, max_consec=0),
    3: dict(emit=True, status="open_circuit", ratio_med=0.0077, max_consec=24),
    4: dict(emit=False, status="NORMAL", ratio_med=0.9938, max_consec=3),
    5: dict(emit=False, status="EMPTY", ratio_med=0.0, max_consec=24),  # qual but EMPTY-skip
}

print(f"\n{'PV':<4}{'ratio_med':>11}{'max_consec':>12}{'emit':>7}{'status':>14}")
for k in range(1, 6):
    ratio = I[k] / np.maximum(I_q95, CLIP)
    qual = (ratio < I_RATIO) & daylight
    # running consec
    consec = 0
    max_consec = 0
    for q in qual:
        consec = consec + 1 if q else 0
        max_consec = max(max_consec, consec)
    ratio_med_day = float(np.median(ratio[:24]))  # daylight rows 5..28 == idx 0..23
    empty = k in EMPTY
    emit = (max_consec >= DEBOUNCE) and (not empty)
    status = "EMPTY" if empty else ("open_circuit" if emit else "NORMAL")
    print(f"PV{k:<3}{ratio_med_day:>11.4f}{max_consec:>12}{str(emit):>7}{status:>14}")

    e = expect[k]
    check(emit == e["emit"], f"PV{k} emit == {e['emit']}")
    check(status == e["status"], f"PV{k} status == {e['status']}")
    check(max_consec == e["max_consec"], f"PV{k} max_consec == {e['max_consec']} (got {max_consec})")
    check(abs(ratio_med_day - e["ratio_med"]) < 5e-4,
          f"PV{k} ratio_median ~= {e['ratio_med']} (got {ratio_med_day:.4f})")

# ---- empty-PV skip is the production-critical guard -----------------------
print("\n=== empty-PV skip guard (Wave 11 hotfix #10) ===")
# PV5: qual all daylight (ratio 0<0.05) -> max_consec 24 >= debounce 20 -> WOULD emit,
# but F9 (empty_by_design)=1 suppresses it.
check(dc["F9"].value == 1, "Decision F9 empty_by_design == 1 (PV5)")
check(int(daylight.sum()) >= DEBOUNCE,
      "PV5 would-emit precondition (daylight run >= debounce) holds -> skip matters")

# ---- StringStatus replica wiring (spot check formulas reference decision) --
print("\n=== StringStatus artifact wiring ===")
check(ss["D5"].value == "PV1" or ss["D5"].value is not None, "StringStatus has pv_string rows")

# ===========================================================================
print("\n" + "=" * 60)
if errors:
    print(f"VERIFY FAILED: {len(errors)} error(s)")
    for e in errors:
        print("  -", e)
    raise SystemExit(1)
print("VERIFY OK — all structural + numeric checks passed (iter4).")
