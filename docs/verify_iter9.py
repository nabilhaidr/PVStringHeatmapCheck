"""Verify Iterasi 9 (M2aSoiling) Excel vs proto_iter9 locked numbers.

(A) STRUCTURAL audit — Helpers_SO (PR_daily, avg_daily_kwh) + SO_Economics (p_loss,
    daily_loss, payback, recommend, fault_type, severity, confidence).
(B) NUMERIC recompute — baca literal Raw_Data_SO + config, re-implement PR_daily + ekonomi
    (p_loss, daily_loss, payback, severity ladder), assert == proto.

CATATAN: memverifikasi metrik PR + ekonomi (LIVE). soiling_ratio = INPUT (SRR Monte-Carlo,
BUKAN direproduksi). Hilir sr = faithful.
"""
from __future__ import annotations
import numpy as np
from openpyxl import load_workbook

WB = "M2_PV_Performance_Workbook.xlsx"
NDAY = 12
DF, DL = 5, 16
SR_SCEN = [0.85, 0.92, 0.97, 0.995]
SC_DF = 10
errors = []
def check(c, m):
    print(("  OK  " if c else "  XX  FAIL: ") + m)
    if not c: errors.append(m)

wb = load_workbook(WB, data_only=False)
for s in ("Raw_Data_SO", "Helpers_SO", "SO_Economics", "SO_Summary"):
    assert s in wb.sheetnames, f"missing {s}"
raw = wb["Raw_Data_SO"]; hp = wb["Helpers_SO"]; ec = wb["SO_Economics"]

def cfgval(n):
    dn = wb.defined_names[n]; (s, c), = dn.destinations; return wb[s][c].value
MIN_DAYS = cfgval("cfg_so_min_days"); CAP = cfgval("cfg_so_capacity_kwp")
TARIFF = cfgval("cfg_so_tariff"); COST = cfgval("cfg_so_cleaning_cost"); PB_THR = cfgval("cfg_so_payback_thr")
print("config:", dict(min_days=MIN_DAYS, capacity=CAP, tariff=TARIFF, cost=COST, payback_thr=PB_THR))
check(MIN_DAYS == 90 and CAP == 71500 and TARIFF == 1500, "min_days 90 / cap 71500 / tariff 1500")
check(COST == 50000000 and PB_THR == 30, "cleaning_cost 50M / payback_thr 30")

# ===========================================================================
print("\n=== (A) STRUCTURAL audit ===")
for ri in (5, 10, 16):
    check(hp[f"D{ri}"].value == f"=Raw_Data_SO!B{ri}/(Raw_Data_SO!C{ri}*cfg_so_capacity_kwp)", f"Helpers D{ri} PR_daily")
check(hp["B18"].value == "=AVERAGE(B5:B16)", "Helpers avg_daily_kwh")
check(ec["B6"].value == '=IF(B5<cfg_so_min_days,"insufficient_data","ok")', "Economics data_status gate")
for k in range(4):
    r = SC_DF + k
    check(ec[f"B{r}"].value == f"=1-A{r}", f"Econ B{r} p_loss")
    check(ec[f"C{r}"].value == f"=$B$4*cfg_so_tariff*B{r}", f"Econ C{r} daily_loss")
    check(ec[f"D{r}"].value == f"=IF(AND(C{r}>0,cfg_so_cleaning_cost>0),cfg_so_cleaning_cost/C{r},1E+99)", f"Econ D{r} payback")
    check(ec[f"E{r}"].value == f'=IF(D{r}<cfg_so_payback_thr,"YES","no")', f"Econ E{r} recommend")
    check(ec[f"G{r}"].value == (f'=IF(AND(B{r}>=0.1,D{r}<cfg_so_payback_thr/3),"CRITICAL",'
                                f'IF(AND(B{r}>=0.05,D{r}<cfg_so_payback_thr),"HIGH",'
                                f'IF(AND(B{r}>=0.02,D{r}<2*cfg_so_payback_thr),"MEDIUM","INFO")))'), f"Econ G{r} severity")

# ===========================================================================
print("\n=== (B) NUMERIC recompute from literals ===")
energy = [float(raw[f"B{DF+i}"].value) for i in range(NDAY)]
insol = [float(raw[f"C{DF+i}"].value) for i in range(NDAY)]
pr_daily = [energy[i]/(insol[i]*CAP) for i in range(NDAY)]
avg_kwh = float(np.mean(energy))
check(abs(pr_daily[0] - 0.80290) < 1e-4, f"PR_daily[0]≈0.8029 (got {pr_daily[0]:.5f})")
check(abs(avg_kwh - 314250.0) < 1e-6, f"avg_daily_kwh==314250 (got {avg_kwh})")

def severity(p_loss, pb, thr=30.0):
    if p_loss >= 0.10 and pb < thr/3: return "CRITICAL"
    if p_loss >= 0.05 and pb < thr: return "HIGH"
    if p_loss >= 0.02 and pb < 2*thr: return "MEDIUM"
    return "INFO"

expect = {0.85: ("CRITICAL", 0.15, 70706250, 0.707), 0.92: ("HIGH", 0.08, 37710000, 1.326),
          0.97: ("MEDIUM", 0.03, 14141250, 3.536), 0.995: ("INFO", 0.005, 2356875, 21.215)}
print(f"{'sr':>7}{'p_loss':>9}{'daily_loss':>14}{'payback':>10}{'recommend':>11}{'severity':>10}")
for sr in SR_SCEN:
    p_loss = 1.0 - sr
    dl = avg_kwh * TARIFF * p_loss
    pb = COST/dl if (dl > 0 and COST > 0) else 1e99
    recommend = pb < PB_THR
    sev = severity(p_loss, pb, PB_THR)
    print(f"{sr:>7.3f}{p_loss:>9.4f}{dl:>14,.0f}{pb:>10.3f}{str(recommend):>11}{sev:>10}")
    esev, ep, edl, epb = expect[sr]
    check(abs(p_loss - ep) < 1e-9, f"sr={sr} p_loss=={ep}")
    check(abs(dl - edl) < 1.0, f"sr={sr} daily_loss≈{edl:,.0f} (got {dl:,.0f})")
    check(abs(pb - epb) < 1e-2, f"sr={sr} payback≈{epb} (got {pb:.3f})")
    check(sev == esev, f"sr={sr} severity=={esev} (got {sev})")
    check(recommend is True, f"sr={sr} recommend (payback<30)")
# data sufficiency gate
check((120 < MIN_DAYS) is False, "n_days=120 >= min_days 90 -> ok (gate path documented)")
check((60 < MIN_DAYS) is True, "n_days=60 < 90 -> insufficient_data (gate would block)")

print("\n" + "="*60)
if errors:
    print(f"VERIFY FAILED: {len(errors)} error(s)")
    for e in errors:
        print("  -", e)
    raise SystemExit(1)
print("VERIFY OK — all structural + numeric checks passed (iter9, economics live).")
