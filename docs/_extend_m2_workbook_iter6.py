"""Extend M2 PV Performance Workbook — Iterasi 6: M2IForest (APPROKSIMASI).

Input : docs/M2_PV_Performance_Workbook.xlsx (output Iterasi 5, 25 sheet)
Output: same file, +4 sheet baru:
  26. Raw_Data_IF   — time-series V/I per PV (1 inverter, 14 daylight ts)
  27. Features_IF   — fitur live (V, I, V_dev, I_dev, R) + robust-z (MAD) + skor A
  28. IF_Anomaly    — flag top-contamination%, rank, severity (kuartil), confidence
  29. IF_Summary    — ringkasan per inverter

== CAVEAT PENTING (baca docs/M2_RE_06) ==
Detector ASLI = sklearn IsolationForest (100 pohon trained, contamination 0.01, seed 42).
Skor iForest = path-length ensemble = BLACK BOX TERLATIH, TIDAK bisa jadi formula Excel.
sklearn juga tak tersedia di sandbox. Maka workbook ini APPROKSIMASI TRANSPARAN:
  - Fitur (V,I,V_dev,I_dev,R) = IDENTIK build_feature_matrix (faithful, live).
  - Skor anomali A = MAX robust-z(MAD) lintas 5 fitur (BUKAN path-length iForest).
  - Struktur faithful: per-inverter, flag fraksi contamination teratas, severity kuartil
    (_severity_from_quartile), confidence = 100 - pct*0.5.
Approx ini akan menandai SAMPEL BERBEDA dari iForest asli — pakai untuk transparansi logika,
bukan sebagai pengganti detektor produksi. Angka EXACT dari proto_iter6.py.
"""
from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule

INPUT = Path(__file__).parent / "M2_PV_Performance_Workbook.xlsx"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=13, bold=True, color="305496")
NOTE_FONT = Font(italic=True, size=9, color="808080")
SEV_FILL = {
    "CRITICAL": PatternFill("solid", fgColor="E06666"),
    "HIGH":     PatternFill("solid", fgColor="F6B26B"),
    "MEDIUM":   PatternFill("solid", fgColor="FFE599"),
    "INFO":     PatternFill("solid", fgColor="D9D9D9"),
}
ANOM_FILL = PatternFill("solid", fgColor="F4CCCC")   # injected anomaly cells


def set_header(ws, row, headers, start=1):
    for ci, h in enumerate(headers, start=start):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def title_note(ws, title, notes):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    for i, nt in enumerate(notes, start=2):
        ws.cell(row=i, column=1, value=nt).font = NOTE_FONT


wb = load_workbook(INPUT)
existing = list(wb.sheetnames)
assert len(existing) == 25, f"Expected 25 sheets from iter5, got {len(existing)}"
print(f"Loaded {len(existing)} sheets.")

# ===========================================================================
# README iterasi log — insert row 11 (iter6) before "Cara membaca"
# ===========================================================================
ws = wb["README"]
assert ws.cell(row=11, column=1).value == "Cara membaca", "README layout berubah; abort"
ws.insert_rows(11)
ws.cell(row=11, column=1, value="6").border = BORDER
ws.cell(row=11, column=2, value="2026-05-30").border = BORDER
ws.cell(row=11, column=3, value="M2IForest (approx)").border = BORDER
ws.cell(row=11, column=4,
        value="Raw_Data_IF, Features_IF, IF_Anomaly, IF_Summary  [APPROKSIMASI MAD — bukan true iForest]").border = BORDER
for rr in range(12, 20):
    if ws.cell(row=rr, column=1).value:
        ws.cell(row=rr, column=1).font = NOTE_FONT
ws.cell(row=11, column=1).font = Font(name="Calibri", size=11)

# ===========================================================================
# Config — append m2_iforest params + named cells (start row 46)
# ===========================================================================
ws = wb["Config"]
r = 4; last_cfg = 4
while ws.cell(row=r, column=1).value is not None:
    last_cfg = r; r += 1
start = last_cfg + 1
assert start == 46, f"Expected Config append at row 46, got {start}"
new_cfg = [
    ("m2_iforest", "contamination_demo", 0.05, "fraksi anomali demo (default produksi 0.01)", "cfg_if_contamination"),
    ("m2_iforest", "r_current_floor_a", 0.1, "R = V / MAX(I, ini)  (proxy resistansi)", "cfg_if_i_floor"),
    ("m2_iforest", "mad_scale", 1.4826, "MAD → sigma (konsistensi normal): z = |x-med|/(1.4826·MAD)", "cfg_if_mad_scale"),
    ("m2_iforest", "real_contamination", 0.01, "DOKUMENTASI: contamination iForest asli (sklearn)", None),
    ("m2_iforest", "real_n_estimators", 100, "DOKUMENTASI: n_estimators iForest asli", None),
    ("m2_iforest", "real_random_state", 42, "DOKUMENTASI: random_state iForest asli (reproducible)", None),
]
for i, (sec, key, val, note, name) in enumerate(new_cfg):
    ri = start + i
    ws.cell(row=ri, column=1, value=sec).border = BORDER
    ws.cell(row=ri, column=2, value=key).border = BORDER
    c = ws.cell(row=ri, column=3, value=val); c.border = BORDER
    c.fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(row=ri, column=4, value=note).border = BORDER
    if name and name not in wb.defined_names:
        wb.defined_names[name] = DefinedName(name, attr_text=f"Config!$C${ri}")
print(f"Config: appended {len(new_cfg)} m2_iforest rows ({start}..{start+len(new_cfg)-1}) + named cells.")

# ===========================================================================
# Synthetic data (matches proto_iter6.py EXACTLY, deterministic)
# ===========================================================================
PV_MAX, N_TS = 6, 14
PV_OFF_V = {1: 0.6, 2: -0.6, 3: 0.3, 4: -0.3, 5: 0.9, 6: -0.9}
PV_OFF_I = {1: 0.05, 2: -0.05, 3: 0.03, 4: -0.03, 5: 0.07, 6: -0.07}
ramp_I = 0.2 * np.sin(np.linspace(0.2, np.pi - 0.2, N_TS))
Vmat = np.zeros((N_TS, PV_MAX)); Imat = np.zeros((N_TS, PV_MAX))
for t in range(N_TS):
    for j, pv in enumerate(range(1, PV_MAX + 1)):
        Vmat[t, j] = 1200.0 + PV_OFF_V[pv]
        Imat[t, j] = 10.0 + PV_OFF_I[pv] + ramp_I[t]
INJECT = {(3, 5): {"I": 3.0}, (3, 6): {"I": 4.0}, (5, 9): {"V": 1100.0}, (2, 2): {"I": 12.6}}
for (pv, t), ov in INJECT.items():
    j = pv - 1
    if "I" in ov: Imat[t, j] = ov["I"]
    if "V" in ov: Vmat[t, j] = ov["V"]
times = [datetime(2026, 5, 14, 12, 0) + timedelta(minutes=5 * t) for t in range(N_TS)]

# ===========================================================================
# Sheet 26: Raw_Data_IF
# ===========================================================================
ws = wb.create_sheet("Raw_Data_IF")
title_note(ws, "Raw_Data_IF — V/I per PV string (WB05-INV01, 14 ts daylight)", [
    "1 inverter, 6 PV string, 14 timestamp noon (POA=900). Sel anomali di-inject (di-highlight): PV3@12:25/12:30 arus drop, PV5@12:45 V drop, PV2@12:10 arus tinggi.",
    "Detector iForest ASLI train per inverter di SEMUA daylight sample. Ganti dengan data Huawei aktual (paste over).",
])
Vcol = {pv: get_column_letter(3 + pv) for pv in range(1, PV_MAX + 1)}   # D..I
Icol = {pv: get_column_letter(9 + pv) for pv in range(1, PV_MAX + 1)}   # J..O
hdr = ["t", "Start Time", "POA"] + [f"PV{pv} V" for pv in range(1, PV_MAX + 1)] \
      + [f"PV{pv} I" for pv in range(1, PV_MAX + 1)]
set_header(ws, 4, hdr)
for t in range(N_TS):
    ri = 5 + t
    ws.cell(row=ri, column=1, value=t).border = BORDER
    c = ws.cell(row=ri, column=2, value=times[t]); c.number_format = "hh:mm"; c.border = BORDER
    ws.cell(row=ri, column=3, value=900.0).border = BORDER
    for pv in range(1, PV_MAX + 1):
        j = pv - 1
        cV = ws.cell(row=ri, column=3 + pv, value=float(Vmat[t, j])); cV.number_format = "0.0"; cV.border = BORDER
        cI = ws.cell(row=ri, column=9 + pv, value=float(Imat[t, j])); cI.number_format = "0.000"; cI.border = BORDER
        if (pv, t) in INJECT:
            if "V" in INJECT[(pv, t)]: cV.fill = ANOM_FILL
            if "I" in INJECT[(pv, t)]: cI.fill = ANOM_FILL
ws.column_dimensions["B"].width = 13

# ===========================================================================
# Sheet 27: Features_IF (84 rows) + stats block (cols U-X)
# ===========================================================================
ws = wb.create_sheet("Features_IF")
title_note(ws, "Features_IF — fitur build_feature_matrix + robust-z (MAD)", [
    "Fitur IDENTIK detector: V, I, V_dev=V−MEDIAN(siblings@ts), I_dev=I−MEDIAN(siblings@ts), R=V/MAX(I,floor). 84 sampel = 6 PV × 14 ts.",
    "Skor A = MAX robust-z lintas 5 fitur, z=|x−median|/(1.4826·MAD) per fitur (stats kolom U-X). CATATAN: A ≈ proxy MAD, BUKAN skor IsolationForest asli (path-length 100 pohon).",
])
DF, DL = 5, 5 + N_TS * PV_MAX - 1   # 5 .. 88
hdr = ["pv", "t", "Start Time", "V", "I", "V_dev", "I_dev", "R",
       "absdev_V", "absdev_I", "absdev_Vdev", "absdev_Idev", "absdev_R",
       "z_V", "z_I", "z_Vdev", "z_Idev", "z_R", "A (max z)"]
set_header(ws, 4, hdr)
# stats block header (U-X = 21-24)
for ci, h in enumerate(["feature", "median", "MAD", "scale=1.4826·MAD"], start=21):
    c = ws.cell(row=4, column=ci, value=h); c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = BORDER
FEAT_DATACOL = {"V": "D", "I": "E", "V_dev": "F", "I_dev": "G", "R": "H"}
ABSDEV_COL = {"V": "I", "I": "J", "V_dev": "K", "I_dev": "L", "R": "M"}
Z_COL = {"V": "N", "I": "O", "V_dev": "P", "I_dev": "Q", "R": "R"}
FEAT_ORDER = ["V", "I", "V_dev", "I_dev", "R"]
# stats rows 5..9: U feature, V median, W MAD, X scale
for k, f in enumerate(FEAT_ORDER):
    sr = 5 + k
    ws.cell(row=sr, column=21, value=f).border = BORDER
    ws.cell(row=sr, column=22, value=f"=MEDIAN({FEAT_DATACOL[f]}{DF}:{FEAT_DATACOL[f]}{DL})").number_format = "0.0000"
    ws.cell(row=sr, column=23, value=f"=MEDIAN({ABSDEV_COL[f]}{DF}:{ABSDEV_COL[f]}{DL})").number_format = "0.0000"
    ws.cell(row=sr, column=24, value=f"=MAX(cfg_if_mad_scale*W{sr},0.000000001)").number_format = "0.000000"
    for ci in (21, 22, 23, 24):
        ws.cell(row=sr, column=ci).border = BORDER
MED_CELL = {f: f"$V${5+k}" for k, f in enumerate(FEAT_ORDER)}
SCALE_CELL = {f: f"$X${5+k}" for k, f in enumerate(FEAT_ORDER)}
# data rows
for t in range(N_TS):
    rawr = 5 + t
    for pv in range(1, PV_MAX + 1):
        k = t * PV_MAX + (pv - 1)
        ri = DF + k
        ws.cell(row=ri, column=1, value=pv)
        ws.cell(row=ri, column=2, value=t)
        ws.cell(row=ri, column=3, value=f"=Raw_Data_IF!B{rawr}").number_format = "hh:mm"
        # features
        ws.cell(row=ri, column=4, value=f"=Raw_Data_IF!{Vcol[pv]}{rawr}").number_format = "0.0"
        ws.cell(row=ri, column=5, value=f"=Raw_Data_IF!{Icol[pv]}{rawr}").number_format = "0.000"
        ws.cell(row=ri, column=6, value=f"=Raw_Data_IF!{Vcol[pv]}{rawr}-MEDIAN(Raw_Data_IF!D{rawr}:I{rawr})").number_format = "0.000"
        ws.cell(row=ri, column=7, value=f"=Raw_Data_IF!{Icol[pv]}{rawr}-MEDIAN(Raw_Data_IF!J{rawr}:O{rawr})").number_format = "0.000"
        ws.cell(row=ri, column=8, value=f"=Raw_Data_IF!{Vcol[pv]}{rawr}/MAX(Raw_Data_IF!{Icol[pv]}{rawr},cfg_if_i_floor)").number_format = "0.00"
        # absdev (cols I-M = 9-13)
        for fi, f in enumerate(FEAT_ORDER):
            ws.cell(row=ri, column=9 + fi,
                    value=f"=ABS({FEAT_DATACOL[f]}{ri}-{MED_CELL[f]})").number_format = "0.000"
        # z (cols N-R = 14-18)
        for fi, f in enumerate(FEAT_ORDER):
            ws.cell(row=ri, column=14 + fi,
                    value=f"={ABSDEV_COL[f]}{ri}/{SCALE_CELL[f]}").number_format = "0.000"
        # A = MAX(z) (col S = 19)
        ws.cell(row=ri, column=19, value=f"=MAX(N{ri}:R{ri})").number_format = "0.000"
        for ci in range(1, 20):
            ws.cell(row=ri, column=ci).border = BORDER

# ===========================================================================
# Sheet 28: IF_Anomaly (decision, 84 rows)
# ===========================================================================
ws = wb.create_sheet("IF_Anomaly")
title_note(ws, "IF_Anomaly — flag contamination teratas + severity (kuartil) + confidence", [
    "threshold_A = PERCENTILE(A, 1−contamination). flag = A ≥ threshold. rank dalam flagged-set (A desc) → severity kuartil (≤25 CRITICAL, ≤50 HIGH, ≤75 MEDIUM, >75 INFO) — faithful _severity_from_quartile. confidence = 100−pct·0.5.",
    "Severity & confidence FAITHFUL ke detector; tapi 'A' adalah skor MAD-approx, bukan IsolationForest. Sampel yang ditandai bisa BERBEDA dari iForest asli.",
])
set_header(ws, 4, ["pv", "t", "Start Time", "A", "flag", "rank_within_flagged",
                   "rank_pct", "severity", "confidence"])
# stats block (K-L)
ws.cell(row=4, column=11, value="metric").fill = HEADER_FILL; ws.cell(row=4, column=11).font = HEADER_FONT
ws.cell(row=4, column=12, value="value").fill = HEADER_FILL; ws.cell(row=4, column=12).font = HEADER_FONT
ws.cell(row=5, column=11, value="threshold_A").border = BORDER
ws.cell(row=5, column=12, value=f"=PERCENTILE(Features_IF!S{DF}:S{DL},1-cfg_if_contamination)").number_format = "0.0000"
ws.cell(row=6, column=11, value="n_flagged").border = BORDER
ws.cell(row=6, column=12, value="=COUNTIF(E5:E88,\"anomaly\")")
ws.cell(row=7, column=11, value="n_samples").border = BORDER
ws.cell(row=7, column=12, value="=COUNTA(A5:A88)")
for rr in (5, 6, 7):
    ws.cell(row=rr, column=12).border = BORDER
THR, NFL = "$L$5", "$L$6"
for t in range(N_TS):
    for pv in range(1, PV_MAX + 1):
        k = t * PV_MAX + (pv - 1)
        ri = 5 + k
        fr = DF + k   # Features_IF row
        ws.cell(row=ri, column=1, value=pv)
        ws.cell(row=ri, column=2, value=t)
        ws.cell(row=ri, column=3, value=f"=Features_IF!C{fr}").number_format = "hh:mm"
        ws.cell(row=ri, column=4, value=f"=Features_IF!S{fr}").number_format = "0.000"
        ws.cell(row=ri, column=5, value=f'=IF(D{ri}>={THR},"anomaly","normal")')
        ws.cell(row=ri, column=6, value=f'=IF(E{ri}="anomaly",COUNTIFS($D$5:$D$88,">"&D{ri},$D$5:$D$88,">="&{THR}),"")')
        ws.cell(row=ri, column=7, value=f'=IF(E{ri}="anomaly",IF({NFL}>1,F{ri}/({NFL}-1)*100,0),"")').number_format = "0.0"
        ws.cell(row=ri, column=8,
                value=(f'=IF(E{ri}="anomaly",IF(G{ri}<=25,"CRITICAL",IF(G{ri}<=50,"HIGH",'
                       f'IF(G{ri}<=75,"MEDIUM","INFO"))),"")'))
        ws.cell(row=ri, column=9, value=f'=IF(E{ri}="anomaly",100-G{ri}*0.5,"")').number_format = "0"
        for ci in range(1, 10):
            ws.cell(row=ri, column=ci).border = BORDER
for sev, fill in SEV_FILL.items():
    ws.conditional_formatting.add("H5:H88",
        CellIsRule(operator="equal", formula=[f'"{sev}"'], fill=fill))

# ===========================================================================
# Sheet 29: IF_Summary
# ===========================================================================
ws = wb.create_sheet("IF_Summary")
title_note(ws, "IF_Summary — ringkasan per inverter (approx)", [
    "Mirror artifact AnomalySummary. n_flagged ≈ contamination × n_samples (PERCENTILE boundary + ties).",
])
set_header(ws, 4, ["inverter_id", "n_samples", "n_flagged", "flagged_pct",
                   "threshold_A", "max_A", "min_A"])
ws.cell(row=5, column=1, value="WB05-INV01")
ws.cell(row=5, column=2, value="=IF_Anomaly!L7")
ws.cell(row=5, column=3, value="=IF_Anomaly!L6")
ws.cell(row=5, column=4, value="=IF_Anomaly!L6/IF_Anomaly!L7*100").number_format = "0.00"
ws.cell(row=5, column=5, value="=IF_Anomaly!L5").number_format = "0.0000"
ws.cell(row=5, column=6, value=f"=MAX(Features_IF!S{DF}:S{DL})").number_format = "0.000"
ws.cell(row=5, column=7, value=f"=MIN(Features_IF!S{DF}:S{DL})").number_format = "0.000"
for ci in range(1, 8):
    ws.cell(row=5, column=ci).border = BORDER
ws.column_dimensions["A"].width = 13

# ===========================================================================
# Save + integrity verify
# ===========================================================================
wb.save(INPUT)
wb2 = load_workbook(INPUT, data_only=False)
after = list(wb2.sheetnames)
new_sheets = ["Raw_Data_IF", "Features_IF", "IF_Anomaly", "IF_Summary"]
assert after[:25] == existing, "EXISTING SHEET ORDER CHANGED!"
assert after[25:] == new_sheets, f"New sheets mismatch: {after[25:]}"
print(f"\nSheets now: {len(after)} (was 25, +4)")
f = wb2["Features_IF"]
print("Features_IF A col (PV3@t5 row):", f"S{5 + 5*6 + 2}", "=", f[f"S{5+5*6+2}"].value)
print("Features_IF stats V5 (median V):", f["V5"].value, "| X5 (scale V):", f["X5"].value)
a = wb2["IF_Anomaly"]
print("IF_Anomaly threshold L5:", a["L5"].value)
print("IF_Anomaly flag E5:", a["E5"].value, "| severity H5:", a["H5"].value)
print("OK — iter6 build complete.")
